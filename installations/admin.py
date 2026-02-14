from django.contrib import admin
from django.db.models import Count, Q, Sum
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.utils.translation import gettext_lazy as _

from accounts.models import SystemSettings
from core.admin_mixins import SoftDeleteAdminMixin
from manufacturing.models import ManufacturingOrder

from . import admin_filters
from .models import (
    CustomerDebt,
    Driver,
    InstallationAnalytics,
    InstallationArchive,
    InstallationEventLog,
    InstallationPayment,
    InstallationSchedule,
    InstallationStatusLog,
    InstallationTeam,
)
from .models import ManufacturingOrder as InstallationManufacturingOrder
from .models import (
    ModificationErrorAnalysis,
    ModificationErrorType,
    ModificationImage,
    ModificationReport,
    ModificationRequest,
    ReceiptMemo,
    Technician,
    Vehicle,
)


# فلتر مخصص لنوع الطلب - للاستخدام مع ManufacturingOrder
class OrderTypeFilter(admin.SimpleListFilter):
    title = "نوع الطلب"
    parameter_name = "order_type"

    def lookups(self, request, model_admin):
        return [
            ("installation", "تركيب"),
            ("custom", "تفصيل"),
            ("accessory", "اكسسوار"),
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(order_type=self.value())
        return queryset


# فلتر مخصص لنوع الطلب - للاستخدام مع InstallationSchedule (من Order)
class InstallationOrderTypeFilter(admin.SimpleListFilter):
    title = "نوع الطلب"
    parameter_name = "order_type"

    def lookups(self, request, model_admin):
        return [
            ("installation", "تركيب"),
            ("custom", "تفصيل"),
            ("accessory", "اكسسوار"),
        ]

    def queryset(self, request, queryset):
        if self.value():
            # فلترة حسب نوع الطلب من Order
            if self.value() == "installation":
                return queryset.filter(order__selected_types__icontains="installation")
            elif self.value() == "custom":
                return queryset.filter(order__selected_types__icontains="custom")
            elif self.value() == "accessory":
                return queryset.filter(order__selected_types__icontains="accessory")
        return queryset


# فلتر مخصص لحالة التركيب
class InstallationStatusFilter(admin.SimpleListFilter):
    title = "حالة التركيب"
    parameter_name = "installation_status"

    def lookups(self, request, model_admin):
        return [
            ("needs_scheduling", "يحتاج جدولة"),
            ("scheduled", "مجدول"),
            ("in_installation", "قيد التركيب"),
            ("completed", "مكتمل"),
            ("cancelled", "ملغي"),
            ("modification_required", "يحتاج تعديل"),
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(order__installation_status=self.value())
        return queryset


# فلتر مخصص لحالة المصنع
class ManufacturingStatusFilter(admin.SimpleListFilter):
    title = "حالة المصنع"
    parameter_name = "manufacturing_status"

    def lookups(self, request, model_admin):
        return [
            ("pending_approval", "قيد الموافقة"),
            ("pending", "قيد الانتظار"),
            ("in_progress", "قيد التصنيع"),
            ("ready_install", "جاهز للتركيب"),
            ("completed", "مكتمل"),
            ("delivered", "تم التسليم"),
            ("rejected", "مرفوض"),
            ("cancelled", "ملغي"),
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(status=self.value())
        return queryset


def currency_format(amount):
    """تنسيق المبلغ مع عملة النظام"""
    try:
        settings = SystemSettings.get_settings()
        symbol = settings.currency_symbol
        formatted_amount = f"{amount:,.2f}"
        return f"{formatted_amount} {symbol}"
    except Exception:
        return f"{amount:,.2f} ر.س"


@admin.register(CustomerDebt)
class CustomerDebtAdmin(admin.ModelAdmin):
    list_per_page = 50
    list_display = [
        "customer_name",
        "order_number",
        "branch_name",
        "salesperson_name",
        "debt_amount_formatted",
        "payment_status",
        "days_overdue",
        "payment_date",
        "created_at",
        "payment_receiver",
        "is_paid",
    ]
    list_filter = [
        "is_paid",
        admin_filters.BranchFilter,
        admin_filters.SalespersonFilter,
        admin_filters.DebtAmountRangeFilter,
        admin_filters.OverdueFilter,
        admin_filters.PaymentMethodFilter,
        admin_filters.CustomerTypeFilter,
        admin_filters.OrderTypeFilter,
        admin_filters.DebtAgeFilter,
        ("created_at", admin.DateFieldListFilter),
        ("payment_date", admin.DateFieldListFilter),
    ]
    search_fields = [
        "customer__name",
        "customer__phone",
        "order__order_number",
        "order__customer__branch__name",
        "order__salesperson__name",
        "order__salesperson__user__first_name",
        "order__salesperson__user__last_name",
        "payment_receiver_name",
    ]
    list_editable = ["is_paid"]
    ordering = ["-created_at"]
    actions = [
        "mark_as_paid",
        "delete_selected_debts",
        "export_to_excel",
        "print_debts_report",
    ]

    # إضافة إمكانية الترتيب لجميع الأعمدة
    sortable_by = [
        "customer__name",
        "order__order_number",
        "debt_amount",
        "is_paid",
        "payment_date",
        "created_at",
    ]

    def get_queryset(self, request):
        """تحسين الاستعلام لتقليل عدد استعلامات قاعدة البيانات"""
        return (
            super()
            .get_queryset(request)
            .select_related(
                "customer", "order", "order__customer__branch", "order__salesperson"
            )
        )

    def has_delete_permission(self, request, obj=None):
        """السماح بالحذف للمدراء والمستخدمين المصرح لهم"""
        # السماح للمدراء (superuser)
        if request.user.is_superuser:
            return True
        # التحقق من الصلاحية العادية
        return request.user.has_perm("installations.delete_customerdebt")

    def get_deleted_objects(self, objs, request):
        """
        تجاوز فحص الصلاحيات المتقدم للحذف
        """
        # إذا كان المستخدم مدير، نسمح بالحذف مباشرة
        if request.user.is_superuser:
            from django.contrib.admin.utils import NestedObjects
            from django.db import router

            collector = NestedObjects(using=router.db_for_write(self.model))
            collector.collect(objs)

            def format_callback(obj):
                return str(obj)

            to_delete = collector.nested(format_callback)
            protected = []
            model_count = {self.model._meta.verbose_name_plural: len(objs)}

            return to_delete, model_count, set(), protected

        # للمستخدمين الآخرين، استخدم السلوك الافتراضي
        return super().get_deleted_objects(objs, request)

    def customer_name(self, obj):
        """عرض اسم العميل مع رقم الهاتف"""
        phone = obj.customer.phone or "لا يوجد"
        return f"{obj.customer.name}\n{phone}"

    customer_name.short_description = "العميل"
    customer_name.admin_order_field = "customer__name"

    def order_number(self, obj):
        """عرض رقم الطلب كرابط"""
        from django.urls import reverse
        from django.utils.html import format_html

        url = reverse("admin:orders_order_change", args=[obj.order.id])
        return format_html(
            '<a href="{}" target="_blank">{}</a>', url, obj.order.order_number
        )

    order_number.short_description = "رقم الطلب"
    order_number.admin_order_field = "order__order_number"

    def branch_name(self, obj):
        """عرض اسم الفرع"""
        return (
            obj.order.customer.branch.name if obj.order.customer.branch else "غير محدد"
        )

    branch_name.short_description = "الفرع"
    branch_name.admin_order_field = "order__customer__branch__name"

    def salesperson_name(self, obj):
        """عرض اسم البائع"""
        if obj.order.salesperson:
            return obj.order.salesperson.name
        return "غير محدد"

    salesperson_name.short_description = "البائع"
    salesperson_name.admin_order_field = "order__salesperson__name"

    def debt_amount_formatted(self, obj):
        """عرض مبلغ المديونية مع التنسيق"""
        return currency_format(obj.debt_amount)

    debt_amount_formatted.short_description = "مبلغ المديونية"
    debt_amount_formatted.admin_order_field = "debt_amount"

    def payment_status(self, obj):
        """عرض حالة الدفع مع الألوان"""
        from django.utils.html import format_html

        if obj.is_paid:
            return format_html(
                '<span style="color: green; font-weight: bold;">✓ مدفوع</span>'
            )
        else:
            # تحديد إذا كانت متأخرة
            from django.utils import timezone

            days_diff = (timezone.now() - obj.created_at).days
            if days_diff > 30:
                return format_html(
                    '<span style="color: red; font-weight: bold;">⚠ متأخر</span>'
                )
            else:
                return format_html(
                    '<span style="color: orange; font-weight: bold;">⏳ غير مدفوع</span>'
                )

    payment_status.short_description = "حالة الدفع"

    def days_overdue(self, obj):
        """عرض عدد الأيام المتأخرة"""
        if obj.is_paid:
            return "-"
        from django.utils import timezone

        days = (timezone.now() - obj.created_at).days
        if days > 30:
            return f"{days} يوم"
        return f"{days} يوم"

    days_overdue.short_description = "أيام التأخير"

    def payment_receiver(self, obj):
        """عرض اسم مستلم الدفع"""
        return obj.payment_receiver_name or "-"

    payment_receiver.short_description = "مستلم الدفع"

    def save_model(self, request, obj, form, change):
        """حفظ النموذج مع تحديث معلومات الدفع وتسجيل الدفعة في الطلب"""
        if change and "is_paid" in form.changed_data:
            if obj.is_paid and not obj.payment_date:
                obj.payment_date = timezone.now()
                obj.payment_receiver_name = (
                    request.user.get_full_name() or request.user.username
                )
                if obj.notes:
                    obj.notes += f' - تم التسديد بواسطة {obj.payment_receiver_name} في {timezone.now().strftime("%Y-%m-%d %H:%M")}'
                else:
                    obj.notes = f'تم التسديد بواسطة {obj.payment_receiver_name} في {timezone.now().strftime("%Y-%m-%d %H:%M")}'

                # تحديث الطلب بالدفعة الجديدة
                if obj.order:
                    # إنشاء دفعة جديدة في جدول الدفعات
                    from orders.models import Payment

                    Payment.objects.create(
                        order=obj.order,
                        amount=obj.debt_amount,
                        payment_method="cash",  # افتراضي
                        payment_date=obj.payment_date or timezone.now(),
                        notes=f"إغلاق مديونية تلقائي من لوحة التحكم بواسطة {obj.payment_receiver_name}",
                        created_by=request.user,
                    )

                    # إنشاء ملاحظة في الطلب
                    from orders.models import OrderNote

                    OrderNote.objects.create(
                        order=obj.order,
                        note_type="payment",
                        title="تسديد مديونية",
                        content=f"تم تسديد مديونية بمبلغ {obj.debt_amount} ج.م بواسطة {obj.payment_receiver_name} من لوحة التحكم وتسجيل دفعة تلقائية",
                        created_by=request.user,
                    )

        super().save_model(request, obj, form, change)

        # إرسال إشعار نجاح
        if change and "is_paid" in form.changed_data and obj.is_paid:
            self.message_user(
                request,
                f"تم تسديد مديونية العميل {obj.customer.name} بمبلغ {obj.debt_amount} ج.م بنجاح وتحديث الطلب.",
                level="SUCCESS",
            )

    # الإجراءات المخصصة
    def mark_as_paid(self, request, queryset):
        """تسديد المديونيات المحددة مع تحديث الطلبات"""
        from django.utils import timezone

        from orders.models import OrderNote

        updated = 0
        for debt in queryset.filter(is_paid=False):
            debt.is_paid = True
            debt.payment_date = timezone.now()
            debt.payment_receiver_name = (
                request.user.get_full_name() or request.user.username
            )
            debt.notes += f' - تم التسديد بواسطة {debt.payment_receiver_name} في {timezone.now().strftime("%Y-%m-%d %H:%M")}'

            # تحديث الطلب بالدفعة الجديدة
            if debt.order:
                # إنشاء دفعة جديدة في جدول الدفعات
                from orders.models import Payment

                Payment.objects.create(
                    order=debt.order,
                    amount=debt.debt_amount,
                    payment_method="cash",  # افتراضي
                    payment_date=debt.payment_date,
                    notes=f"إغلاق مديونية تلقائي (إجراء جماعي) من لوحة التحكم بواسطة {debt.payment_receiver_name}",
                    created_by=request.user,
                )

                # إنشاء ملاحظة في الطلب
                OrderNote.objects.create(
                    order=debt.order,
                    note_type="payment",
                    title="تسديد مديونية (إجراء جماعي)",
                    content=f"تم تسديد مديونية بمبلغ {debt.debt_amount} ج.م بواسطة {debt.payment_receiver_name} من لوحة التحكم (إجراء جماعي) وتسجيل دفعة تلقائية",
                    created_by=request.user,
                )

            debt.save()
            updated += 1

        self.message_user(
            request, f"تم تسديد {updated} مديونية بنجاح وتحديث الطلبات المرتبطة."
        )

    mark_as_paid.short_description = "تسديد المديونيات المحددة"

    def delete_selected_debts(self, request, queryset):
        """حذف المديونيات المحددة (للمدراء فقط)"""
        if not request.user.is_superuser:
            self.message_user(request, "ليس لديك صلاحية حذف المديونيات.", level="ERROR")
            return

        count = queryset.count()
        queryset.delete()
        self.message_user(request, f"تم حذف {count} مديونية بنجاح.")

    delete_selected_debts.short_description = "حذف المديونيات المحددة"

    def export_to_excel(self, request, queryset):
        """تصدير المديونيات إلى Excel"""
        import openpyxl
        from django.http import HttpResponse
        from django.utils import timezone
        from openpyxl.styles import Alignment, Font, PatternFill

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير المديونيات"

        # تنسيق الرأس
        headers = [
            "اسم العميل",
            "رقم الهاتف",
            "رقم الطلب",
            "الفرع",
            "البائع",
            "مبلغ المديونية",
            "حالة الدفع",
            "تاريخ الإنشاء",
            "تاريخ الدفع",
            "مستلم الدفع",
            "أيام التأخير",
            "ملاحظات",
        ]

        # كتابة الرأس
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # كتابة البيانات
        for row, debt in enumerate(
            queryset.select_related(
                "customer", "order", "order__customer__branch", "order__salesperson"
            ),
            2,
        ):
            ws.cell(row=row, column=1, value=debt.customer.name)
            ws.cell(row=row, column=2, value=debt.customer.phone or "لا يوجد")
            ws.cell(row=row, column=3, value=debt.order.order_number)
            ws.cell(
                row=row,
                column=4,
                value=(
                    debt.order.customer.branch.name
                    if debt.order.customer.branch
                    else "غير محدد"
                ),
            )

            salesperson = ""
            if debt.order.salesperson:
                salesperson = debt.order.salesperson.name
            ws.cell(row=row, column=5, value=salesperson or "غير محدد")

            ws.cell(row=row, column=6, value=float(debt.debt_amount))
            ws.cell(row=row, column=7, value="مدفوع" if debt.is_paid else "غير مدفوع")
            ws.cell(row=row, column=8, value=debt.created_at.strftime("%Y-%m-%d"))
            ws.cell(
                row=row,
                column=9,
                value=(
                    debt.payment_date.strftime("%Y-%m-%d") if debt.payment_date else ""
                ),
            )
            ws.cell(row=row, column=10, value=debt.payment_receiver_name or "")

            # حساب أيام التأخير
            if not debt.is_paid:
                days_overdue = (timezone.now() - debt.created_at).days
                ws.cell(row=row, column=11, value=days_overdue)
            else:
                ws.cell(row=row, column=11, value=0)

            ws.cell(row=row, column=12, value=debt.notes)

        # تنسيق الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # إعداد الاستجابة
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f'تقرير_المديونيات_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    export_to_excel.short_description = "تصدير إلى Excel"

    def print_debts_report(self, request, queryset):
        """طباعة تقرير المديونيات"""
        from django.http import HttpResponse
        from django.shortcuts import render
        from django.template.loader import get_template
        from django.utils import timezone

        # حساب الإحصائيات
        total_debts = queryset.count()
        total_amount = sum(debt.debt_amount for debt in queryset)
        paid_debts = queryset.filter(is_paid=True).count()
        unpaid_debts = queryset.filter(is_paid=False).count()
        overdue_debts = queryset.filter(
            is_paid=False, created_at__lt=timezone.now() - timezone.timedelta(days=30)
        ).count()

        context = {
            "debts": queryset.select_related(
                "customer", "order", "order__customer__branch", "order__salesperson"
            ),
            "stats": {
                "total_debts": total_debts,
                "total_amount": total_amount,
                "paid_debts": paid_debts,
                "unpaid_debts": unpaid_debts,
                "overdue_debts": overdue_debts,
            },
            "report_date": timezone.now(),
            "generated_by": request.user.get_full_name() or request.user.username,
        }

        template = get_template("admin/debt_print_report.html")
        html = template.render(context)

        response = HttpResponse(html, content_type="text/html")
        return response

    print_debts_report.short_description = "طباعة تقرير المديونيات"

    def get_readonly_fields(self, request, obj=None):
        """تحديد الحقول للقراءة فقط"""
        readonly_fields = ["created_at", "updated_at"]
        if obj and obj.is_paid:
            readonly_fields.extend(["debt_amount", "customer", "order"])
        return readonly_fields

    def get_fieldsets(self, request, obj=None):
        """تنظيم الحقول في مجموعات"""
        fieldsets = [
            (
                "معلومات المديونية",
                {"fields": ("customer", "order", "debt_amount", "notes")},
            ),
            (
                "معلومات الدفع",
                {
                    "fields": (
                        "is_paid",
                        "payment_date",
                        "payment_receiver_name",
                        "payment_receipt_number",
                    ),
                    "classes": ("collapse",),
                },
            ),
            (
                "معلومات النظام",
                {"fields": ("created_at", "updated_at"), "classes": ("collapse",)},
            ),
        ]
        return fieldsets

    def has_delete_permission(self, request, obj=None):
        """منع حذف المديونيات المدفوعة"""
        if obj and obj.is_paid:
            return False
        return super().has_delete_permission(request, obj)

    class Media:
        css = {"all": ("admin/css/debt_admin.css",)}
        js = ("admin/js/debt_admin.js",)


@admin.register(Technician)
class TechnicianAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = ["name", "phone", "specialization", "is_active", "created_at"]
    list_filter = ["is_active", "specialization", "created_at"]
    search_fields = ["name", "phone", "specialization"]
    list_editable = ["is_active"]
    ordering = ["name"]


@admin.register(Driver)
class DriverAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "name",
        "phone",
        "license_number",
        "vehicle_number",
        "is_active",
        "created_at",
    ]
    list_filter = ["is_active", "created_at"]
    search_fields = ["name", "phone", "license_number", "vehicle_number"]
    list_editable = ["is_active"]
    ordering = ["name"]


@admin.register(Vehicle)
class VehicleAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_per_page = 50
    list_display = [
        "name",
        "plate_number",
        "model",
        "vehicle_type",
        "status",
        "created_at",
    ]
    list_filter = ["status", "vehicle_type", "created_at"]
    search_fields = ["name", "plate_number", "chassis_number", "model"]
    list_editable = ["status"]
    ordering = ["name"]


@admin.register(InstallationTeam)
class InstallationTeamAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = ["name", "driver", "technicians_count", "is_active", "created_at"]
    list_filter = ["is_active", "created_at"]
    search_fields = ["name"]
    list_editable = ["is_active"]
    filter_horizontal = ["technicians"]
    ordering = ["name"]

    def technicians_count(self, obj):
        return obj.technicians.count()

    technicians_count.short_description = "عدد الفنيين"


@admin.register(InstallationSchedule)
class InstallationScheduleAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "installation_code",
        "customer_name",
        "scheduled_date",
        "scheduled_time",
        "team",
        "status_display",
        "created_at",
    ]

    def get_queryset(self, request):
        """تحسين الاستعلامات لتقليل N+1 queries"""
        return (
            super()
            .get_queryset(request)
            .select_related("order__customer", "team", "created_by")
            .only(
                "id",
                "installation_code",
                "status",
                "scheduled_date",
                "scheduled_time",
                "created_at",
                "order__id",
                "order__customer__name",
                "team__id",
                "team__name",
            )
        )

    list_filter = [
        "status",
        InstallationOrderTypeFilter,  # فلتر مخصص لنوع الطلب (من Order)
        "scheduled_date",
        "team",
        "created_at",
    ]
    search_fields = ["order__order_number", "order__customer__name"]
    list_editable = ["team"]
    date_hierarchy = "scheduled_date"
    ordering = ["-scheduled_date", "-scheduled_time"]

    # إضافة إجراءات مجمعة لتغيير حالة التركيب
    actions = [
        "mark_status_scheduled",
        "mark_status_in_installation",
        "mark_status_completed",
        "mark_status_cancelled",
        "mark_status_modification_required",
    ]

    fieldsets = (
        ("معلومات الطلب", {"fields": ("order", "status")}),
        ("جدولة التركيب", {"fields": ("team", "scheduled_date", "scheduled_time")}),
        ("ملاحظات", {"fields": ("notes",), "classes": ("collapse",)}),
    )

    def status_display(self, obj):
        """عرض حالة التركيب بألوان"""
        colors = {
            "needs_scheduling": "#ffc107",  # أصفر
            "scheduled": "#17a2b8",  # أزرق فاتح
            "in_installation": "#007bff",  # أزرق
            "completed": "#28a745",  # أخضر
            "cancelled": "#dc3545",  # أحمر
            "modification_required": "#fd7e14",  # برتقالي
            "modification_in_progress": "#6f42c1",  # بنفسجي
            "modification_completed": "#20c997",  # أخضر فاتح
        }
        color = colors.get(obj.status, "#6c757d")
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 12px; font-weight: bold; font-size: 11px; white-space: nowrap;">{}</span>',
            color,
            obj.get_status_display(),
        )

    status_display.short_description = "الحالة"

    def customer_name(self, obj):
        if obj.order and obj.order.customer:
            return obj.order.customer.name
        return "-"

    customer_name.short_description = "اسم العميل"

    def get_urls(self):
        """إضافة URLs مخصصة للوصول للتركيبات باستخدام الكود"""
        urls = super().get_urls()
        custom_urls = [
            path(
                "by-code/<str:installation_code>/",
                self.admin_site.admin_view(self.installation_by_code_view),
                name="installations_installationschedule_by_code",
            ),
        ]
        return custom_urls + urls

    def installation_by_code_view(self, request, installation_code):
        """عرض التركيب باستخدام الكود وإعادة التوجيه لصفحة التحرير"""
        try:
            # البحث باستخدام order_number إذا كان يحتوي على رقم الطلب
            if installation_code.endswith("-T"):
                base_code = installation_code[:-2]  # إزالة '-T'
                if base_code.startswith("#"):
                    # البحث باستخدام ID مباشرة
                    installation_id = base_code[1:]  # إزالة '#'
                    installation = InstallationSchedule.objects.get(id=installation_id)
                else:
                    # البحث باستخدام order_number
                    installation = InstallationSchedule.objects.get(
                        order__order_number=base_code
                    )
            else:
                # محاولة البحث المباشر بالكود
                installation = InstallationSchedule.objects.get(id=installation_code)

            return HttpResponseRedirect(
                reverse(
                    "admin:installations_installationschedule_change",
                    args=[installation.pk],
                )
            )
        except (InstallationSchedule.DoesNotExist, ValueError):
            self.message_user(
                request, f"التركيب بكود {installation_code} غير موجود", level="error"
            )
            return HttpResponseRedirect(
                reverse("admin:installations_installationschedule_changelist")
            )

    def installation_code(self, obj):
        """عرض رقم طلب التركيب الموحد مع روابط محسنة - تحديث للاستخدام الكود في admin"""
        code = obj.installation_code

        try:
            # رابط عرض التركيب في الواجهة
            view_url = reverse("installations:installation_detail_by_code", args=[code])
            # رابط تحرير التركيب في لوحة التحكم باستخدام الكود
            admin_url = reverse(
                "admin:installations_installationschedule_by_code",
                kwargs={"installation_code": code},
            )

            return format_html(
                "<strong>{}</strong><br/>"
                '<a href="{}" target="_blank" title="عرض في الواجهة">'
                '<span style="color: #0073aa;">👁️ عرض</span></a> | '
                '<a href="{}" title="تحرير في لوحة التحكم">'
                '<span style="color: #d63638;">✏️ تحرير</span></a>',
                code,
                view_url,
                admin_url,
            )
        except Exception:
            return code

    installation_code.short_description = "رقم طلب التركيب"

    # إجراءات التحديث المجمع لحالة التركيب
    def mark_status_scheduled(self, request, queryset):
        """تغيير حالة التركيب إلى مجدول"""
        updated = queryset.update(status="scheduled")
        self.message_user(request, f'✅ تم تغيير حالة {updated} تركيب إلى "مجدول"')

    mark_status_scheduled.short_description = "تغيير الحالة إلى مجدول"

    def mark_status_in_installation(self, request, queryset):
        """تغيير حالة التركيب إلى قيد التركيب"""
        updated = queryset.update(status="in_installation")
        self.message_user(
            request, f'✅ تم تغيير حالة {updated} تركيب إلى "قيد التركيب"'
        )

    mark_status_in_installation.short_description = "تغيير الحالة إلى قيد التركيب"

    def mark_status_completed(self, request, queryset):
        """تغيير حالة التركيب إلى مكتمل"""
        updated = queryset.update(status="completed")
        self.message_user(request, f'✅ تم تغيير حالة {updated} تركيب إلى "مكتمل"')

    mark_status_completed.short_description = "تغيير الحالة إلى مكتمل"

    def mark_status_cancelled(self, request, queryset):
        """تغيير حالة التركيب إلى ملغي"""
        updated = queryset.update(status="cancelled")
        self.message_user(request, f'✅ تم تغيير حالة {updated} تركيب إلى "ملغي"')

    mark_status_cancelled.short_description = "تغيير الحالة إلى ملغي"

    def mark_status_modification_required(self, request, queryset):
        """تغيير حالة التركيب إلى يحتاج تعديل"""
        updated = queryset.update(status="modification_required")
        self.message_user(
            request, f'✅ تم تغيير حالة {updated} تركيب إلى "يحتاج تعديل"'
        )

    mark_status_modification_required.short_description = "تغيير الحالة إلى يحتاج تعديل"

    def get_queryset(self, request):
        """تحسين الاستعلامات لتحسين الأداء"""
        return (
            super()
            .get_queryset(request)
            .select_related("order", "order__customer", "order__branch", "team")
        )


@admin.register(ModificationRequest)
class ModificationRequestAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "installation",
        "customer",
        "modification_type",
        "priority",
        "estimated_cost_formatted",
        "customer_approval",
        "created_at",
    ]
    list_filter = ["priority", "customer_approval", "created_at"]
    search_fields = [
        "installation__order__order_number",
        "customer__name",
        "modification_type",
    ]
    list_editable = ["priority", "customer_approval"]
    ordering = ["-created_at"]

    def estimated_cost_formatted(self, obj):
        return currency_format(obj.estimated_cost)

    estimated_cost_formatted.short_description = "التكلفة المتوقعة"


@admin.register(ModificationImage)
class ModificationImageAdmin(admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = ["modification", "image_preview", "description", "uploaded_at"]
    list_filter = ["uploaded_at"]
    search_fields = ["modification__installation__order__order_number", "description"]
    ordering = ["-uploaded_at"]

    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="max-height: 50px; max-width: 50px;" />',
                obj.image.url,
            )
        return "-"

    image_preview.short_description = "معاينة الصورة"


# إنشاء admin مخصص لأوامر التصنيع في قسم التركيبات
# سيتم تسجيله يدوياً في نهاية الملف لتجنب التكرار
class InstallationManufacturingOrderAdmin(admin.ModelAdmin):
    list_per_page = 50
    list_display = [
        "manufacturing_code",
        "order_number",
        "customer_name",
        "order_type_display",
        "status",
        "installation_status_display",
        "order_date",
        "expected_delivery_date",
        "production_line",
    ]
    list_filter = [
        ManufacturingStatusFilter,  # فلتر مخصص لحالة المصنع
        OrderTypeFilter,  # فلتر مخصص لنوع الطلب (من ManufacturingOrder)
        InstallationStatusFilter,  # فلتر مخصص لحالة التركيب
        "order_date",
        "expected_delivery_date",
        ("order__branch", admin.RelatedFieldListFilter),
        ("order__salesperson", admin.RelatedFieldListFilter),
        ("production_line", admin.RelatedFieldListFilter),
    ]

    # إضافة روابط سريعة للفلترة
    def changelist_view(self, request, extra_context=None):
        if extra_context is None:
            extra_context = {}

        # إضافة روابط سريعة للحالات المختلفة
        extra_context["quick_filters"] = [
            {
                "name": "يحتاج جدولة",
                "url": "?order__installation_status__exact=needs_scheduling",
                "count": self.get_queryset(request)
                .filter(order__installation_status="needs_scheduling")
                .count(),
            },
            {
                "name": "مجدول",
                "url": "?order__installation_status__exact=scheduled",
                "count": self.get_queryset(request)
                .filter(order__installation_status="scheduled")
                .count(),
            },
            {
                "name": "قيد التركيب",
                "url": "?order__installation_status__exact=in_installation",
                "count": self.get_queryset(request)
                .filter(order__installation_status="in_installation")
                .count(),
            },
            {
                "name": "مكتمل",
                "url": "?order__installation_status__exact=completed",
                "count": self.get_queryset(request)
                .filter(order__installation_status="completed")
                .count(),
            },
            {
                "name": "ملغي",
                "url": "?order__installation_status__exact=cancelled",
                "count": self.get_queryset(request)
                .filter(order__installation_status="cancelled")
                .count(),
            },
            {
                "name": "يحتاج تعديل",
                "url": "?order__installation_status__exact=modification_required",
                "count": self.get_queryset(request)
                .filter(order__installation_status="modification_required")
                .count(),
            },
        ]

        # مسح cache إذا كان هناك تحديث
        if (
            request.GET.get("updated")
            or request.GET.get("_refresh")
            or request.GET.get("_nocache")
        ):
            from django.core.cache import cache

            cache.clear()

            # مسح cache الـ ORM أيضاً
            from django.db import connection

            if hasattr(connection, "queries_log"):
                connection.queries_log.clear()

        # إضافة header لمنع caching في المتصفح دائماً
        response = super().changelist_view(request, extra_context)
        if (
            request.GET.get("updated")
            or request.GET.get("_refresh")
            or request.GET.get("_nocache")
        ):
            response["Cache-Control"] = "no-cache, no-store, must-revalidate, max-age=0"
            response["Pragma"] = "no-cache"
            response["Expires"] = "0"
            response["Last-Modified"] = ""
            response["ETag"] = ""

        return response

    search_fields = [
        "id",
        "contract_number",
        "invoice_number",
        "order__order_number",
        "order__customer__name",
        "order__customer__phone",
    ]
    list_editable = ["status", "production_line"]
    ordering = ["-order_date"]
    date_hierarchy = "order_date"

    # إضافة إجراءات مجمعة لتغيير الحالة
    actions = [
        # إجراءات حالة أوامر التصنيع
        "mark_manufacturing_pending_approval",
        "mark_manufacturing_pending",
        "mark_manufacturing_in_progress",
        "mark_manufacturing_ready_install",
        "mark_manufacturing_completed",
        "mark_manufacturing_delivered",
        "mark_manufacturing_rejected",
        "mark_manufacturing_cancelled",
        # إجراءات حالة التركيب
        "mark_installation_needs_scheduling",
        "mark_installation_scheduled",
        "mark_installation_in_progress",
        "mark_installation_completed",
        "mark_installation_cancelled",
        "mark_installation_modification_required",
    ]

    fieldsets = (
        (
            "معلومات الطلب",
            {"fields": ("order", "order_type", "contract_number", "invoice_number")},
        ),
        (
            "حالة التصنيع",
            {
                "fields": (
                    "status",
                    "production_line",
                    "order_date",
                    "expected_delivery_date",
                )
            },
        ),
        (
            "ملفات",
            {"fields": ("contract_file", "inspection_file"), "classes": ("collapse",)},
        ),
        ("ملاحظات", {"fields": ("notes",), "classes": ("collapse",)}),
    )

    def changelist_view(self, request, extra_context=None):
        """تخصيص عرض قائمة التغييرات لضمان إعادة التحميل الصحيح"""
        if extra_context is None:
            extra_context = {}

        # مسح cache إذا كان هناك تحديث
        if (
            request.GET.get("updated")
            or request.GET.get("_refresh")
            or request.GET.get("_nocache")
        ):
            from django.core.cache import cache

            cache.clear()

            # مسح cache الـ ORM أيضاً
            from django.db import connection

            if hasattr(connection, "queries_log"):
                connection.queries_log.clear()

        # إضافة header لمنع caching في المتصفح دائماً
        response = super().changelist_view(request, extra_context)
        if (
            request.GET.get("updated")
            or request.GET.get("_refresh")
            or request.GET.get("_nocache")
        ):
            response["Cache-Control"] = "no-cache, no-store, must-revalidate, max-age=0"
            response["Pragma"] = "no-cache"
            response["Expires"] = "0"
            response["Last-Modified"] = ""
            response["ETag"] = ""

        return response

    def get_queryset(self, request):
        """تخصيص الـ queryset لضمان الحصول على أحدث البيانات"""
        queryset = super().get_queryset(request)

        # إذا كان هناك تحديث، تجنب الـ cache
        if (
            request.GET.get("updated")
            or request.GET.get("_refresh")
            or request.GET.get("_nocache")
        ):
            # إعادة تقييم الـ queryset لضمان الحصول على أحدث البيانات
            queryset = queryset.all()

        return queryset

    def order_number(self, obj):
        """عرض رقم الطلب مع رابط"""
        if obj.order:
            return format_html(
                '<a href="{}" target="_blank">{}</a>',
                reverse("admin:orders_order_change", args=[obj.order.pk]),
                obj.order.order_number,
            )
        return "-"

    order_number.short_description = "رقم الطلب"
    order_number.admin_order_field = "order__order_number"

    def customer_name(self, obj):
        """عرض اسم العميل"""
        if obj.order and obj.order.customer:
            return obj.order.customer.name
        return "-"

    customer_name.short_description = "اسم العميل"
    customer_name.admin_order_field = "order__customer__name"

    def order_type_display(self, obj):
        """عرض نوع الطلب بألوان"""
        colors = {
            "installation": "#007bff",  # أزرق
            "custom": "#28a745",  # أخضر
            "accessory": "#ffc107",  # أصفر
        }
        color = colors.get(obj.order_type, "#6c757d")
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 12px; font-weight: bold; font-size: 11px; white-space: nowrap;">{}</span>',
            color,
            obj.get_order_type_display(),
        )

    order_type_display.short_description = "نوع الطلب"

    def status(self, obj):
        """عرض حالة التصنيع بألوان"""
        colors = {
            "pending_approval": "#ffc107",  # أصفر
            "pending": "#6c757d",  # رمادي
            "in_progress": "#007bff",  # أزرق
            "ready_install": "#17a2b8",  # أزرق فاتح
            "completed": "#28a745",  # أخضر
            "delivered": "#20c997",  # أخضر فاتح
            "rejected": "#dc3545",  # أحمر
            "cancelled": "#6c757d",  # رمادي
        }
        color = colors.get(obj.status, "#6c757d")
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 12px; font-weight: bold; font-size: 11px; white-space: nowrap;">{}</span>',
            color,
            obj.get_status_display(),
        )

    status.short_description = "الحالة"

    def installation_status_display(self, obj):
        """عرض حالة التركيب بألوان"""
        if not obj.order:
            return "-"

        colors = {
            "needs_scheduling": "#fd7e14",  # برتقالي
            "scheduled": "#17a2b8",  # أزرق فاتح
            "in_installation": "#007bff",  # أزرق
            "completed": "#28a745",  # أخضر
            "cancelled": "#6c757d",  # رمادي
            "modification_required": "#ffc107",  # أصفر
            "modification_in_progress": "#e83e8c",  # وردي
            "modification_completed": "#20c997",  # أخضر فاتح
        }
        color = colors.get(obj.order.installation_status, "#6c757d")
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 12px; font-weight: bold; font-size: 11px; white-space: nowrap;">{}</span>',
            color,
            obj.order.get_installation_status_display(),
        )

    installation_status_display.short_description = "حالة التركيب"
    installation_status_display.admin_order_field = "order__installation_status"

    # إجراءات مجمعة لتغيير الحالة
    def mark_as_pending(self, request, queryset):
        """تغيير الحالة إلى في الانتظار"""
        updated = queryset.update(status="pending")
        self.message_user(
            request, f'تم تغيير حالة {updated} أمر تصنيع إلى "في الانتظار"'
        )

    mark_as_pending.short_description = "تغيير الحالة إلى في الانتظار"

    def mark_as_in_progress(self, request, queryset):
        """تغيير الحالة إلى قيد التصنيع"""
        updated = queryset.update(status="in_progress")
        self.message_user(
            request, f'تم تغيير حالة {updated} أمر تصنيع إلى "قيد التصنيع"'
        )

    mark_as_in_progress.short_description = "تغيير الحالة إلى قيد التصنيع"

    def mark_as_ready_install(self, request, queryset):
        """تغيير الحالة إلى جاهز للتركيب"""
        updated = queryset.update(status="ready_install")
        self.message_user(
            request, f'تم تغيير حالة {updated} أمر تصنيع إلى "جاهز للتركيب"'
        )

    mark_as_ready_install.short_description = "تغيير الحالة إلى جاهز للتركيب"

    def mark_as_completed(self, request, queryset):
        """تغيير الحالة إلى مكتمل"""
        updated = queryset.update(status="completed")
        self.message_user(request, f'تم تغيير حالة {updated} أمر تصنيع إلى "مكتمل"')

    mark_as_completed.short_description = "تغيير الحالة إلى مكتمل"

    def mark_as_delivered(self, request, queryset):
        """تغيير الحالة إلى تم التسليم"""
        updated = queryset.update(status="delivered")
        self.message_user(
            request, f'تم تغيير حالة {updated} أمر تصنيع إلى "تم التسليم"'
        )

    mark_as_delivered.short_description = "تغيير الحالة إلى تم التسليم"

    def get_queryset(self, request):
        """تحسين الاستعلامات لتحسين الأداء - عرض جميع أوامر التصنيع"""
        return (
            super()
            .get_queryset(request)
            .select_related(
                "order", "order__customer", "order__branch", "production_line"
            )
        )

    # دالة مساعدة لتحديث حالة التركيب
    def _update_installation_status(
        self, request, queryset, new_status, status_display
    ):
        """دالة مساعدة لتحديث حالة التركيب"""
        from django.db import transaction

        from installations.models import InstallationSchedule

        with transaction.atomic():
            # الحصول على الطلبات
            orders = [order.order for order in queryset if order.order]
            order_ids = [order.id for order in orders]

            if order_ids:
                # تحديث حالة التركيب في الطلبات مباشرة
                from orders.models import Order

                Order.objects.filter(id__in=order_ids).update(
                    installation_status=new_status
                )

                # تحديث أو إنشاء InstallationSchedule
                existing_installations = InstallationSchedule.objects.filter(
                    order_id__in=order_ids
                )
                existing_order_ids = set(
                    existing_installations.values_list("order_id", flat=True)
                )

                # إنشاء InstallationSchedule للطلبات التي لا تملكها
                new_installations = []
                for order in orders:
                    if order.id not in existing_order_ids:
                        new_installations.append(
                            InstallationSchedule(order=order, status=new_status)
                        )

                if new_installations:
                    InstallationSchedule.objects.bulk_create(new_installations)

                # تحديث الحالات الموجودة
                updated_count = existing_installations.update(status=new_status)
                total_updated = len(new_installations) + updated_count

                # مسح cache
                from django.core.cache import cache

                cache.clear()
            else:
                total_updated = 0

        self.message_user(
            request,
            f'✅ تم تغيير حالة التركيب لـ {total_updated} طلب إلى "{status_display}"',
        )
        return total_updated

    # إجراءات تغيير حالة التركيب - محسنة للأداء
    def mark_installation_needs_scheduling(self, request, queryset):
        """تغيير حالة التركيب إلى بحاجة جدولة"""
        self._update_installation_status(
            request, queryset, "needs_scheduling", "بحاجة جدولة"
        )

    mark_installation_needs_scheduling.short_description = (
        "تغيير حالة التركيب إلى بحاجة جدولة"
    )

    def mark_installation_scheduled(self, request, queryset):
        """تغيير حالة التركيب إلى مجدول"""
        self._update_installation_status(request, queryset, "scheduled", "مجدول")

    mark_installation_scheduled.short_description = "تغيير حالة التركيب إلى مجدول"

    def mark_installation_in_progress(self, request, queryset):
        """تغيير حالة التركيب إلى قيد التركيب"""
        self._update_installation_status(
            request, queryset, "in_installation", "قيد التركيب"
        )

    mark_installation_in_progress.short_description = (
        "تغيير حالة التركيب إلى قيد التركيب"
    )

    def mark_installation_completed(self, request, queryset):
        """تغيير حالة التركيب إلى مكتمل"""
        self._update_installation_status(request, queryset, "completed", "مكتمل")

    mark_installation_completed.short_description = "تغيير حالة التركيب إلى مكتمل"

    def mark_installation_cancelled(self, request, queryset):
        """تغيير حالة التركيب إلى ملغي"""
        self._update_installation_status(request, queryset, "cancelled", "ملغي")

    mark_installation_cancelled.short_description = "تغيير حالة التركيب إلى ملغي"

    def mark_installation_modification_required(self, request, queryset):
        """تغيير حالة التركيب إلى يحتاج تعديل"""
        self._update_installation_status(
            request, queryset, "modification_required", "يحتاج تعديل"
        )

    mark_installation_modification_required.short_description = (
        "تغيير حالة التركيب إلى يحتاج تعديل"
    )

    # إجراءات تحديث حالات أوامر التصنيع
    def mark_manufacturing_pending_approval(self, request, queryset):
        """تغيير حالة أمر التصنيع إلى قيد الموافقة"""
        updated = queryset.update(status="pending_approval")
        self.message_user(
            request, f'✅ تم تغيير حالة {updated} أمر تصنيع إلى "قيد الموافقة"'
        )

    mark_manufacturing_pending_approval.short_description = (
        "تغيير حالة التصنيع إلى قيد الموافقة"
    )

    def mark_manufacturing_pending(self, request, queryset):
        """تغيير حالة أمر التصنيع إلى قيد الانتظار"""
        updated = queryset.update(status="pending")
        self.message_user(
            request, f'✅ تم تغيير حالة {updated} أمر تصنيع إلى "قيد الانتظار"'
        )

    mark_manufacturing_pending.short_description = "تغيير حالة التصنيع إلى قيد الانتظار"

    def mark_manufacturing_in_progress(self, request, queryset):
        """تغيير حالة أمر التصنيع إلى قيد التصنيع"""
        updated = queryset.update(status="in_progress")
        self.message_user(
            request, f'✅ تم تغيير حالة {updated} أمر تصنيع إلى "قيد التصنيع"'
        )

    mark_manufacturing_in_progress.short_description = (
        "تغيير حالة التصنيع إلى قيد التصنيع"
    )

    def mark_manufacturing_ready_install(self, request, queryset):
        """تغيير حالة أمر التصنيع إلى جاهز للتركيب"""
        updated = queryset.update(status="ready_install")
        self.message_user(
            request, f'✅ تم تغيير حالة {updated} أمر تصنيع إلى "جاهز للتركيب"'
        )

    mark_manufacturing_ready_install.short_description = (
        "تغيير حالة التصنيع إلى جاهز للتركيب"
    )

    def mark_manufacturing_completed(self, request, queryset):
        """تغيير حالة أمر التصنيع إلى مكتمل"""
        updated = queryset.update(status="completed")
        self.message_user(request, f'✅ تم تغيير حالة {updated} أمر تصنيع إلى "مكتمل"')

    mark_manufacturing_completed.short_description = "تغيير حالة التصنيع إلى مكتمل"

    def mark_manufacturing_delivered(self, request, queryset):
        """تغيير حالة أمر التصنيع إلى تم التسليم"""
        updated = queryset.update(status="delivered")
        self.message_user(
            request, f'✅ تم تغيير حالة {updated} أمر تصنيع إلى "تم التسليم"'
        )

    mark_manufacturing_delivered.short_description = "تغيير حالة التصنيع إلى تم التسليم"

    def mark_manufacturing_rejected(self, request, queryset):
        """تغيير حالة أمر التصنيع إلى مرفوض"""
        updated = queryset.update(status="rejected")
        self.message_user(request, f'✅ تم تغيير حالة {updated} أمر تصنيع إلى "مرفوض"')

    mark_manufacturing_rejected.short_description = "تغيير حالة التصنيع إلى مرفوض"

    def mark_manufacturing_cancelled(self, request, queryset):
        """تغيير حالة أمر التصنيع إلى ملغي"""
        updated = queryset.update(status="cancelled")
        self.message_user(request, f'✅ تم تغيير حالة {updated} أمر تصنيع إلى "ملغي"')

    mark_manufacturing_cancelled.short_description = "تغيير حالة التصنيع إلى ملغي"


# إزالة تسجيل نموذج أوامر التعديل - سنستخدم admin مخصص لـ manufacturing.models.ManufacturingOrder
# @admin.register(InstallationManufacturingOrder)
class ModificationManufacturingOrderAdmin(admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "order_number",
        "customer_name",
        "order_type",
        "status",
        "assigned_to",
        "estimated_completion",
        "created_at",
    ]
    list_filter = [
        "order_type",
        "status",
        "assigned_to",
        "created_at",
        (
            "modification_request__installation__order__order_date",
            admin.DateFieldListFilter,
        ),
        (
            "modification_request__installation__order__branch",
            admin.RelatedFieldListFilter,
        ),
    ]
    search_fields = [
        "modification_request__installation__order__order_number",
        "modification_request__installation__order__customer__name",
        "modification_request__installation__order__customer__phone",
        "modification_request__installation__order__order_number",
    ]
    list_editable = ["status", "assigned_to"]
    ordering = ["-created_at"]
    date_hierarchy = "created_at"

    # إضافة إجراءات مجمعة لتغيير الحالة
    actions = [
        "mark_as_pending",
        "mark_as_in_progress",
        "mark_as_completed",
        "mark_as_delivered",
    ]

    fieldsets = (
        ("معلومات الطلب", {"fields": ("modification_request", "order_type")}),
        (
            "حالة التصنيع",
            {"fields": ("status", "assigned_to", "estimated_completion_date")},
        ),
        (
            "ملاحظات",
            {"fields": ("description", "manager_notes"), "classes": ("collapse",)},
        ),
    )

    def order_number(self, obj):
        """عرض رقم الطلب مع رابط"""
        if (
            obj.modification_request
            and obj.modification_request.installation
            and obj.modification_request.installation.order
        ):
            order = obj.modification_request.installation.order
            return format_html(
                '<a href="{}" target="_blank">{}</a>',
                reverse("admin:orders_order_change", args=[order.pk]),
                order.order_number,
            )
        return "-"

    order_number.short_description = "رقم الطلب"
    order_number.admin_order_field = (
        "modification_request__installation__order__order_number"
    )

    def customer_name(self, obj):
        """عرض اسم العميل"""
        if (
            obj.modification_request
            and obj.modification_request.installation
            and obj.modification_request.installation.order
            and obj.modification_request.installation.order.customer
        ):
            return obj.modification_request.installation.order.customer.name
        return "-"

    customer_name.short_description = "اسم العميل"
    customer_name.admin_order_field = (
        "modification_request__installation__order__customer__name"
    )

    def status(self, obj):
        """عرض حالة التصنيع بألوان"""
        colors = {
            "pending": "#ffc107",  # أصفر
            "in_progress": "#007bff",  # أزرق
            "completed": "#28a745",  # أخضر
            "delivered": "#20c997",  # أخضر فاتح
            "cancelled": "#dc3545",  # أحمر
        }
        color = colors.get(obj.status, "#6c757d")
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 12px; font-weight: bold; font-size: 11px; white-space: nowrap;">{}</span>',
            color,
            obj.get_status_display(),
        )

    status.short_description = "الحالة"

    def estimated_completion(self, obj):
        """عرض تاريخ الإكمال المتوقع"""
        if obj.estimated_completion_date:
            return obj.estimated_completion_date.strftime("%Y-%m-%d")
        return "-"

    estimated_completion.short_description = "تاريخ الإكمال المتوقع"
    estimated_completion.admin_order_field = "estimated_completion_date"

    # إجراءات مجمعة لتغيير الحالة
    def mark_as_pending(self, request, queryset):
        """تغيير الحالة إلى في الانتظار"""
        updated = queryset.update(status="pending")
        self.message_user(
            request, f'تم تغيير حالة {updated} أمر تصنيع إلى "في الانتظار"'
        )

    mark_as_pending.short_description = "تغيير الحالة إلى في الانتظار"

    def mark_as_in_progress(self, request, queryset):
        """تغيير الحالة إلى قيد التنفيذ"""
        updated = queryset.update(status="in_progress")
        self.message_user(
            request, f'تم تغيير حالة {updated} أمر تصنيع إلى "قيد التنفيذ"'
        )

    mark_as_in_progress.short_description = "تغيير الحالة إلى قيد التنفيذ"

    def mark_as_completed(self, request, queryset):
        """تغيير الحالة إلى مكتمل"""
        updated = queryset.update(status="completed")
        self.message_user(request, f'تم تغيير حالة {updated} أمر تصنيع إلى "مكتمل"')

    mark_as_completed.short_description = "تغيير الحالة إلى مكتمل"

    def mark_as_delivered(self, request, queryset):
        """تغيير الحالة إلى تم التسليم"""
        updated = queryset.update(status="delivered")
        self.message_user(
            request, f'تم تغيير حالة {updated} أمر تصنيع إلى "تم التسليم"'
        )

    mark_as_delivered.short_description = "تغيير الحالة إلى تم التسليم"

    def get_queryset(self, request):
        """تحسين الاستعلامات لتحسين الأداء"""
        return (
            super()
            .get_queryset(request)
            .select_related(
                "modification_request__installation__order__customer",
                "modification_request__installation__order__branch",
                "assigned_to",
            )
            .prefetch_related("modification_request__installation")
        )


@admin.register(ModificationReport)
class ModificationReportAdmin(admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "modification_request",
        "manufacturing_order",
        "created_by",
        "created_at",
    ]
    list_filter = ["created_at", "created_by"]
    search_fields = [
        "modification_request__installation__order__order_number",
        "description",
    ]
    ordering = ["-created_at"]


@admin.register(ReceiptMemo)
class ReceiptMemoAdmin(admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "installation",
        "receipt_image_preview",
        "customer_signature",
        "amount_received_formatted",
        "created_at",
    ]
    list_filter = ["customer_signature", "created_at"]
    search_fields = ["installation__order__order_number"]
    ordering = ["-created_at"]

    def receipt_image_preview(self, obj):
        if obj.receipt_image:
            return format_html(
                '<img src="{}" style="max-height: 50px; max-width: 50px;" />',
                obj.receipt_image.url,
            )
        return "-"

    receipt_image_preview.short_description = "صورة المذكرة"

    def amount_received_formatted(self, obj):
        return currency_format(obj.amount_received)

    amount_received_formatted.short_description = "المبلغ المستلم"


@admin.register(InstallationPayment)
class InstallationPaymentAdmin(admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "installation",
        "payment_type",
        "amount_formatted",
        "payment_method",
        "created_at",
    ]
    list_filter = ["payment_type", "payment_method", "created_at"]
    search_fields = ["installation__order__order_number", "receipt_number"]
    ordering = ["-created_at"]

    def amount_formatted(self, obj):
        return currency_format(obj.amount)

    amount_formatted.short_description = "المبلغ"


@admin.register(InstallationArchive)
class InstallationArchiveAdmin(admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "installation",
        "completion_date",
        "archived_by_display",
        "archive_notes_short",
    ]
    list_filter = ["completion_date", "archived_by"]
    search_fields = [
        "installation__order__order_number",
        "archive_notes",
        "archived_by__username",
        "archived_by__first_name",
        "archived_by__last_name",
        "archived_by__email",
    ]
    ordering = ["-completion_date"]
    readonly_fields = ["completion_date", "archived_by"]

    def archive_notes_short(self, obj):
        """عرض مختصر لملاحظات الأرشفة"""
        if obj.archive_notes:
            return (
                obj.archive_notes[:50] + "..."
                if len(obj.archive_notes) > 50
                else obj.archive_notes
            )
        return "-"

    archive_notes_short.short_description = "ملاحظات الأرشفة"

    def archived_by_display(self, obj):
        """عرض اسم المستخدم الذي أرشف التركيب"""
        if obj.archived_by:
            full_name = obj.archived_by.get_full_name()
            username = obj.archived_by.username
            if full_name:
                return format_html('<span title="{}">{}</span>', username, full_name)
            return username
        return "-"

    archived_by_display.short_description = "أرشف بواسطة"
    archived_by_display.admin_order_field = "archived_by__username"


@admin.register(InstallationStatusLog)
class InstallationStatusLogAdmin(admin.ModelAdmin):
    """إدارة سجل تغيير حالات التركيبات"""

    list_per_page = 50
    list_display = [
        "installation",
        "old_status_display",
        "new_status_display",
        "changed_by_display",
        "created_at",
        "reason_short",
    ]
    list_filter = ["created_at", "changed_by", "old_status", "new_status"]
    search_fields = [
        "installation__order__order_number",
        "reason",
        "notes",
        "changed_by__username",
        "changed_by__first_name",
        "changed_by__last_name",
        "changed_by__email",
    ]
    ordering = ["-created_at"]
    readonly_fields = [
        "installation",
        "old_status",
        "new_status",
        "changed_by",
        "reason",
        "notes",
        "created_at",
    ]

    def old_status_display(self, obj):
        """عرض حالة التركيب القديمة بشكل واضح"""
        return obj.get_old_status_display()

    old_status_display.short_description = "من حالة"
    old_status_display.admin_order_field = "old_status"

    def new_status_display(self, obj):
        """عرض حالة التركيب الجديدة بشكل واضح"""
        return obj.get_new_status_display()

    new_status_display.short_description = "إلى حالة"
    new_status_display.admin_order_field = "new_status"

    def changed_by_display(self, obj):
        """عرض اسم المستخدم الذي غير الحالة"""
        if obj.changed_by:
            full_name = obj.changed_by.get_full_name()
            username = obj.changed_by.username
            if full_name:
                return format_html('<span title="{}">{}</span>', username, full_name)
            return username
        return "-"

    changed_by_display.short_description = "غير بواسطة"
    changed_by_display.admin_order_field = "changed_by__username"

    def reason_short(self, obj):
        """عرض مختصر لسبب التغيير"""
        if obj.reason:
            return obj.reason[:30] + "..." if len(obj.reason) > 30 else obj.reason
        return "-"

    reason_short.short_description = "السبب"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(InstallationEventLog)
class InstallationEventLogAdmin(admin.ModelAdmin):
    """إدارة سجل أحداث التركيبات"""

    list_per_page = 50
    list_display = [
        "installation",
        "event_type_display",
        "description_short",
        "user_display",
        "created_at",
    ]
    list_filter = ["created_at", "user", "event_type"]
    search_fields = [
        "installation__order__order_number",
        "description",
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__email",
    ]
    ordering = ["-created_at"]
    readonly_fields = [
        "installation",
        "event_type",
        "description",
        "user",
        "metadata",
        "created_at",
    ]

    def event_type_display(self, obj):
        """عرض نوع الحدث بشكل واضح"""
        return obj.get_event_type_display()

    event_type_display.short_description = "نوع الحدث"
    event_type_display.admin_order_field = "event_type"

    def user_display(self, obj):
        """عرض اسم المستخدم الذي قام بالحدث"""
        if obj.user:
            full_name = obj.user.get_full_name()
            username = obj.user.username
            if full_name:
                return format_html('<span title="{}">{}</span>', username, full_name)
            return username
        return "-"

    user_display.short_description = "المستخدم"
    user_display.admin_order_field = "user__username"

    def description_short(self, obj):
        """عرض مختصر لوصف الحدث"""
        if obj.description:
            return (
                obj.description[:50] + "..."
                if len(obj.description) > 50
                else obj.description
            )
        return "-"

    description_short.short_description = "الوصف"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(InstallationAnalytics)
class InstallationAnalyticsAdmin(admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "month",
        "total_installations",
        "completed_installations",
        "total_customers",
        "total_modifications",
        "completion_rate",
        "modification_rate",
    ]
    list_filter = ["month"]
    search_fields = ["month"]
    readonly_fields = ["completion_rate", "modification_rate"]

    def save_model(self, request, obj, form, change):
        obj.calculate_rates()
        super().save_model(request, obj, form, change)


@admin.register(ModificationErrorAnalysis)
class ModificationErrorAnalysisAdmin(admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = [
        "modification_request",
        "error_type",
        "cost_impact_formatted",
        "time_impact_hours",
        "created_at",
    ]
    list_filter = ["error_type", "created_at"]
    search_fields = [
        "modification_request__installation__order__order_number",
        "error_description",
    ]
    readonly_fields = ["created_at", "updated_at"]

    def cost_impact_formatted(self, obj):
        return currency_format(obj.cost_impact)

    cost_impact_formatted.short_description = "التأثير المالي"


@admin.register(ModificationErrorType)
class ModificationErrorTypeAdmin(admin.ModelAdmin):
    list_per_page = 50  # عرض 50 صف كافتراضي
    list_display = ["name", "description", "is_active", "created_at"]
    list_filter = ["is_active", "created_at"]
    search_fields = ["name", "description"]
    readonly_fields = ["created_at", "updated_at"]
    list_editable = ["is_active"]


# تسجيل admin مخصص لأوامر التصنيع في قسم التركيبات
# إزالة admin الأصلي من manufacturing
try:
    admin.site.unregister(ManufacturingOrder)
except admin.sites.NotRegistered:
    pass

# تسجيل admin مخصص في قسم التركيبات
admin.site.register(ManufacturingOrder, InstallationManufacturingOrderAdmin)

# تخصيص عنوان لوحة الإدارة
admin.site.site_header = "إدارة التركيبات - نظام الخواجه"
admin.site.site_title = "إدارة التركيبات"
admin.site.index_title = "مرحباً بك في إدارة التركيبات"
