"""
عروض التصدير والطباعة
"""
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import datetime, date
import json

from .services.pdf_export import PDFExportService, PrintService
from .services.calendar_service import CalendarService
from .models_new import InstallationNew, InstallationTeamNew


@login_required
@require_http_methods(["GET"])
def export_daily_schedule_pdf(request):
    """تصدير جدول التركيبات اليومي كـ PDF"""
    
    try:
        # الحصول على المعاملات
        date_str = request.GET.get('date')
        branch_name = request.GET.get('branch', '')
        
        if not date_str:
            return JsonResponse({
                'success': False,
                'error': 'التاريخ مطلوب'
            })
        
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'تنسيق التاريخ غير صحيح'
            })
        
        # تصدير PDF
        response = PDFExportService.export_daily_schedule(
            target_date, 
            branch_name if branch_name else None
        )
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@require_http_methods(["GET"])
def export_technician_report_pdf(request):
    """تصدير تقرير أداء فني كـ PDF"""
    
    try:
        technician_id = request.GET.get('technician_id')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if not all([technician_id, start_date_str, end_date_str]):
            return JsonResponse({
                'success': False,
                'error': 'جميع المعاملات مطلوبة'
            })
        
        try:
            technician_id = int(technician_id)
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'تنسيق البيانات غير صحيح'
            })
        
        # تصدير PDF
        response = PDFExportService.export_technician_report(
            technician_id, start_date, end_date
        )
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@require_http_methods(["GET"])
def export_monthly_summary_pdf(request):
    """تصدير الملخص الشهري كـ PDF"""
    
    try:
        year = request.GET.get('year')
        month = request.GET.get('month')
        branch_name = request.GET.get('branch', '')
        
        if not all([year, month]):
            return JsonResponse({
                'success': False,
                'error': 'السنة والشهر مطلوبان'
            })
        
        try:
            year = int(year)
            month = int(month)
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'تنسيق السنة أو الشهر غير صحيح'
            })
        
        # تصدير PDF
        response = PDFExportService.export_monthly_summary(
            year, month, 
            branch_name if branch_name else None
        )
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@require_http_methods(["GET"])
def print_daily_schedule(request):
    """طباعة جدول التركيبات اليومي"""
    
    try:
        # الحصول على المعاملات
        date_str = request.GET.get('date')
        branch_name = request.GET.get('branch', '')
        
        if not date_str:
            target_date = timezone.now().date()
        else:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'error': 'تنسيق التاريخ غير صحيح'
                })
        
        # إنشاء HTML للطباعة
        html_content = PrintService.generate_printable_html(
            target_date,
            branch_name if branch_name else None
        )
        
        return HttpResponse(html_content, content_type='text/html')
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@require_http_methods(["GET"])
def export_installations_excel(request):
    """تصدير التركيبات كـ Excel"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # الحصول على المعاملات
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        branch_name = request.GET.get('branch', '')
        status = request.GET.get('status', '')
        
        # تحديد التواريخ الافتراضية
        if not start_date_str:
            start_date = timezone.now().date().replace(day=1)  # بداية الشهر
        else:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        
        if not end_date_str:
            end_date = timezone.now().date()
        else:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # استعلام التركيبات
        installations = InstallationNew.objects.filter(
            scheduled_date__range=[start_date, end_date]
        ).select_related('team', 'order')
        
        if branch_name:
            installations = installations.filter(branch_name=branch_name)
        
        if status:
            installations = installations.filter(status=status)
        
        # إنشاء ملف Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "تقرير التركيبات"
        
        # تنسيق العناوين
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # رؤوس الأعمدة
        headers = [
            'رقم التركيب', 'اسم العميل', 'رقم الهاتف', 'الفرع',
            'عدد الشبابيك', 'تاريخ الطلب', 'تاريخ الجدولة',
            'وقت البدء', 'وقت الانتهاء', 'الفريق', 'الحالة',
            'الأولوية', 'نوع المكان', 'ملاحظات'
        ]
        
        # كتابة رؤوس الأعمدة
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # كتابة البيانات
        for row, installation in enumerate(installations, 2):
            data = [
                installation.id,
                installation.customer_name,
                installation.customer_phone,
                installation.branch_name,
                installation.windows_count,
                installation.order_date.strftime('%Y-%m-%d') if installation.order_date else '',
                installation.scheduled_date.strftime('%Y-%m-%d') if installation.scheduled_date else '',
                installation.scheduled_time_start.strftime('%H:%M') if installation.scheduled_time_start else '',
                installation.scheduled_time_end.strftime('%H:%M') if installation.scheduled_time_end else '',
                installation.team.name if installation.team else 'غير محدد',
                installation.get_status_display(),
                installation.get_priority_display(),
                installation.get_location_type_display(),
                installation.notes
            ]
            
            for col, value in enumerate(data, 1):
                worksheet.cell(row=row, column=col, value=value)
        
        # تنسيق العرض التلقائي للأعمدة
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].auto_size = True
        
        # إعداد الاستجابة
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"installations_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # حفظ الملف
        workbook.save(response)
        
        return response
        
    except ImportError:
        return JsonResponse({
            'success': False,
            'error': 'مكتبة openpyxl غير متاحة. يرجى تثبيتها لتصدير Excel'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@require_http_methods(["GET"])
def export_team_performance_pdf(request):
    """تصدير تقرير أداء الفرق كـ PDF"""
    
    try:
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if not all([start_date_str, end_date_str]):
            return JsonResponse({
                'success': False,
                'error': 'تواريخ البداية والنهاية مطلوبة'
            })
        
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'تنسيق التاريخ غير صحيح'
            })
        
        # الحصول على بيانات الفرق
        teams = InstallationTeamNew.objects.filter(is_active=True)
        teams_data = []
        
        for team in teams:
            team_installations = InstallationNew.objects.filter(
                team=team,
                scheduled_date__range=[start_date, end_date]
            )
            
            total_installations = team_installations.count()
            total_windows = sum(inst.windows_count for inst in team_installations)
            completed_installations = team_installations.filter(status='completed').count()
            
            teams_data.append({
                'team': team,
                'total_installations': total_installations,
                'total_windows': total_windows,
                'completed_installations': completed_installations,
                'completion_rate': (completed_installations / total_installations * 100) if total_installations > 0 else 0
            })
        
        # إنشاء PDF (تنفيذ مبسط)
        from django.template.loader import render_to_string
        from django.http import HttpResponse
        
        context = {
            'teams_data': teams_data,
            'start_date': start_date,
            'end_date': end_date,
            'generated_at': timezone.now()
        }
        
        # يمكن استخدام قالب HTML وتحويله لـ PDF
        html_content = render_to_string('installations/reports/team_performance.html', context)
        
        response = HttpResponse(html_content, content_type='text/html')
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@require_http_methods(["POST"])
def generate_custom_report(request):
    """إنشاء تقرير مخصص"""
    
    try:
        data = json.loads(request.body)
        
        report_type = data.get('report_type')
        filters = data.get('filters', {})
        export_format = data.get('format', 'pdf')  # pdf, excel, html
        
        if not report_type:
            return JsonResponse({
                'success': False,
                'error': 'نوع التقرير مطلوب'
            })
        
        # تطبيق الفلاتر
        installations = InstallationNew.objects.all()
        
        if filters.get('start_date'):
            start_date = datetime.strptime(filters['start_date'], '%Y-%m-%d').date()
            installations = installations.filter(scheduled_date__gte=start_date)
        
        if filters.get('end_date'):
            end_date = datetime.strptime(filters['end_date'], '%Y-%m-%d').date()
            installations = installations.filter(scheduled_date__lte=end_date)
        
        if filters.get('branch'):
            installations = installations.filter(branch_name=filters['branch'])
        
        if filters.get('status'):
            installations = installations.filter(status=filters['status'])
        
        if filters.get('priority'):
            installations = installations.filter(priority=filters['priority'])
        
        # إنشاء التقرير حسب النوع والتنسيق
        if report_type == 'daily_schedule':
            if export_format == 'pdf':
                target_date = datetime.strptime(filters.get('date', timezone.now().date().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
                return PDFExportService.export_daily_schedule(target_date, filters.get('branch'))
            elif export_format == 'html':
                target_date = datetime.strptime(filters.get('date', timezone.now().date().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
                return PrintService.generate_printable_html(target_date, filters.get('branch'))
        
        # المزيد من أنواع التقارير يمكن إضافتها هنا
        
        return JsonResponse({
            'success': False,
            'error': 'نوع التقرير غير مدعوم'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
