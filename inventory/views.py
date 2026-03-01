from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Case, Count, F, OuterRef, Q, Subquery, Sum, When
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView

from accounts.models import SystemSettings

from .forms import ProductForm
from .inventory_utils import (
    get_cached_dashboard_stats,
    get_cached_product_list,
    get_cached_stock_level,
    invalidate_product_cache,
)
from .models import Category, Product, PurchaseOrder, StockAlert, StockTransaction


# ===== API Ù„ØªØ¨Ø¯ÙŠÙ„ ÙˆØ¶Ø¹ Ø¹Ø±Ø¶ Ø§Ù„Ø³Ø¹Ø± =====
@login_required
@require_POST
def toggle_price_display_mode(request):
    """ØªØ¨Ø¯ÙŠÙ„ ÙˆØ¶Ø¹ Ø¹Ø±Ø¶ Ø§Ù„Ø³Ø¹Ø± Ø¨ÙŠÙ† Ù‚Ø·Ø§Ø¹ÙŠ ÙˆØ¬Ù…Ù„Ø© ÙˆØ­ÙØ¸Ù‡ ÙÙŠ Ø§Ù„Ø¬Ù„Ø³Ø©"""
    mode = request.POST.get("mode", "retail")
    if mode in ["retail", "wholesale"]:
        request.session["price_display_mode"] = mode
        return JsonResponse({"success": True, "mode": mode})
    return JsonResponse({"success": False, "error": "Invalid mode"}, status=400)


class InventoryDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„ØªØ®Ø²ÙŠÙ† Ø§Ù„Ù…Ø¤Ù‚Øª Ù„Ù„Ø¥Ø­ØµØ§Ø¦ÙŠØ§Øª - Ù…Ø­Ø³Ù† Ù„Ù„ØºØ§ÙŠØ©
        stats = get_cached_dashboard_stats()
        context.update(stats)

        # Ø¥Ø¶Ø§ÙØ© active_menu Ù„Ù„Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø¬Ø§Ù†Ø¨ÙŠØ©
        context["active_menu"] = "dashboard"

        # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ù…Ù†Ø®ÙØ¶Ø© Ø§Ù„Ù…Ø®Ø²ÙˆÙ† - Ù…Ø­Ø³Ù† Ø¬Ø¯Ø§Ù‹
        # Ø§Ø³ØªØ®Ø¯Ø§Ù… only() ÙˆØªØ­Ø¯ÙŠØ¯ 5 ÙÙ‚Ø· Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† 10
        latest_balance = (
            StockTransaction.objects.filter(product=OuterRef("pk"))
            .order_by("-transaction_date")
            .values("running_balance")[:1]
        )

        low_stock_products = (
            Product.objects.annotate(current_stock_level=Subquery(latest_balance))
            .filter(
                current_stock_level__gt=0, current_stock_level__lte=F("minimum_stock")
            )
            .select_related("category")
            .only("id", "name", "code", "minimum_stock", "category__name")[:5]
        )  # Ù…Ù† 10 Ø¥Ù„Ù‰ 5

        context["low_stock_products"] = [
            {
                "product": p,
                "current_stock": p.current_stock_level or 0,
                "status": "Ù…Ø®Ø²ÙˆÙ† Ù…Ù†Ø®ÙØ¶",
                "is_available": (p.current_stock_level or 0) > 0,
            }
            for p in low_stock_products
        ]

        # Ø¢Ø®Ø± Ø­Ø±ÙƒØ§Øª Ø§Ù„Ù…Ø®Ø²ÙˆÙ† - Ù…Ø­Ø³Ù‘Ù†
        recent_transactions = (
            StockTransaction.objects.select_related("product", "created_by")
            .only(
                "id",
                "product__name",
                "transaction_type",
                "quantity",
                "date",
                "created_by__username",
            )
            .order_by("-date")[:5]
        )  # Ù…Ù† 10 Ø¥Ù„Ù‰ 5

        context["recent_transactions"] = recent_transactions

        # Ø¹Ø¯Ø¯ Ø·Ù„Ø¨Ø§Øª Ø§Ù„Ø´Ø±Ø§Ø¡ Ø§Ù„Ù…Ø¹Ù„Ù‚Ø© - Ø§Ø³ØªØ®Ø¯Ø§Ù… count ÙÙ‚Ø·
        context["pending_purchase_orders"] = PurchaseOrder.objects.filter(
            status__in=["draft", "pending"]
        ).count()

        # Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø±Ø³Ù… Ø§Ù„Ø¨ÙŠØ§Ù†ÙŠ - Ù…Ø­Ø³Ù‘Ù† Ø¬Ø¯Ø§Ù‹
        category_stats = (
            Category.objects.annotate(product_count=Count("products"))
            .filter(product_count__gt=0)
            .only("id", "name")
            .order_by("-product_count")[:5]
        )  # Ù…Ù† 10 Ø¥Ù„Ù‰ 5

        stock_by_category = []
        for cat in category_stats:
            # Ø­Ø³Ø§Ø¨ Ù…Ø¨Ø³Ù‘Ø· Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… aggregate ÙÙ‚Ø·
            total_stock = (
                StockTransaction.objects.filter(product__category=cat).aggregate(
                    total=Sum("running_balance")
                )["total"]
                or 0
            )

            stock_by_category.append(
                {
                    "name": cat.name,
                    "stock": int(total_stock),
                    "product_count": cat.product_count,
                }
            )

        context["stock_by_category"] = stock_by_category

        # Ø¨ÙŠØ§Ù†Ø§Øª Ø­Ø±ÙƒØ© Ø§Ù„Ù…Ø®Ø²ÙˆÙ† - Ù…Ø­Ø³Ù‘Ù† (3 Ø£ÙŠØ§Ù… ÙÙ‚Ø·)
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=2)  # Ù…Ù† 6 Ø¥Ù„Ù‰ 2

        from django.db.models.functions import TruncDate

        stock_movements = (
            StockTransaction.objects.filter(date__date__range=[start_date, end_date])
            .annotate(date_only=TruncDate("date"))
            .values("date_only", "transaction_type")
            .annotate(total=Sum("quantity"))
            .order_by("date_only", "transaction_type")
        )

        # ØªÙ†Ø¸ÙŠÙ… Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
        dates = []
        stock_in = []
        stock_out = []

        movement_data = {}
        for movement in stock_movements:
            date_str = str(movement["date_only"])
            if date_str not in movement_data:
                movement_data[date_str] = {"in": 0, "out": 0}
            movement_data[date_str][movement["transaction_type"]] = movement["total"]

        current_date = start_date
        while current_date <= end_date:
            date_str = str(current_date)
            dates.append(current_date)
            stock_in.append(movement_data.get(date_str, {}).get("in", 0))
            stock_out.append(movement_data.get(date_str, {}).get("out", 0))
            current_date += timedelta(days=1)

        context["stock_movement_dates"] = dates
        context["stock_movement_in"] = stock_in
        context["stock_movement_out"] = stock_out

        # Ø§Ù„ØªÙ†Ø¨ÙŠÙ‡Ø§Øª - Ù…Ø­Ø³Ù‘Ù†
        context["alerts_count"] = StockAlert.objects.filter(status="active").count()
        context["recent_alerts"] = (
            StockAlert.objects.filter(status="active")
            .select_related("product")
            .only("id", "product__name", "alert_type", "created_at")
            .order_by("-created_at")[:3]
        )  # Ù…Ù† 5 Ø¥Ù„Ù‰ 3

        context["current_year"] = timezone.now().year

        return context


@login_required
def product_list(request):
    # Ø§Ù„Ø¨Ø­Ø« ÙˆØ§Ù„ØªØµÙÙŠØ©
    search_query = request.GET.get("search", "")
    category_id = request.GET.get("category", "")
    filter_type = request.GET.get("filter", "")
    sort_by = request.GET.get("sort", "-created_at")

    # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ù…Ø¹ Ø­Ø³Ø§Ø¨ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ
    # Ø§Ø³ØªØ®Ø¯Ø§Ù… Subquery Ù„Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø¢Ø®Ø± Ø±ØµÙŠØ¯ Ù…Ù† Ø¬Ø¯ÙˆÙ„ StockTransaction
    from django.db.models import IntegerField, OuterRef, Subquery
    from django.db.models.functions import Coalesce

    latest_balance = (
        StockTransaction.objects.filter(product=OuterRef("pk"))
        .order_by("-transaction_date", "-id")
        .values("running_balance")[:1]
    )

    products = (
        Product.objects.select_related("category")
        .annotate(
            current_stock_calc=Coalesce(
                Subquery(latest_balance, output_field=IntegerField()), 0
            )
        )
        .only("id", "name", "code", "price", "category", "created_at", "minimum_stock")
    )

    # ØªØ·Ø¨ÙŠÙ‚ Ø§Ù„Ø¨Ø­Ø«
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query)
            | Q(code__icontains=search_query)
            | Q(description__icontains=search_query)
        )

    # Ù…Ù„Ø§Ø­Ø¸Ø©: Ù„Ø§ Ù†Ø·Ø¨Ù‚ ÙÙ„ØªØ± Ø§Ù„Ø³Ù†Ø© Ø¹Ù„Ù‰ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ù„Ø£Ù†Ù‡Ø§ Ø¯Ø§Ø¦Ù…Ø© ÙˆÙ„ÙŠØ³Øª Ù…Ø±ØªØ¨Ø·Ø© Ø¨Ø³Ù†Ø© Ù…Ø¹ÙŠÙ†Ø©

    # ØªØ·Ø¨ÙŠÙ‚ ÙÙ„ØªØ± Ø§Ù„ÙØ¦Ø©
    if category_id:
        products = products.filter(category_id=category_id)

    # ØªØ·Ø¨ÙŠÙ‚ ÙÙ„Ø§ØªØ± Ø®Ø§ØµØ©
    if filter_type == "low_stock":
        # Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ø§Ù„ØªÙŠ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ Ø£Ù‚Ù„ Ù…Ù† Ø§Ù„Ø­Ø¯ Ø§Ù„Ø£Ø¯Ù†Ù‰
        products = products.filter(current_stock_calc__lt=F('minimum_stock'))
    elif filter_type == "out_of_stock":
        # Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ø§Ù„ØªÙŠ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† = 0
        products = products.filter(current_stock_calc=0)
    elif filter_type == "in_stock":
        # Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ø§Ù„Ù…ØªÙˆÙØ±Ø© ÙÙŠ Ø§Ù„Ù…Ø®Ø²ÙˆÙ†
        products = products.filter(current_stock_calc__gt=0)

    # ØªØ·Ø¨ÙŠÙ‚ Ø§Ù„ØªØ±ØªÙŠØ¨
    valid_sort_fields = ['name', 'code', 'price', 'created_at', 'minimum_stock']
    sort_field = sort_by.lstrip("-")
    
    if sort_field in valid_sort_fields:
        products = products.order_by(sort_by)
    elif sort_field == 'current_stock':
        # ØªØ±ØªÙŠØ¨ Ø­Ø³Ø¨ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ
        products = products.order_by(sort_by.replace('current_stock', 'current_stock_calc'))
    else:
        products = products.order_by("-created_at")

    # Ø§Ù„ØµÙØ­Ø§Øª - Ø²ÙŠØ§Ø¯Ø© Ø§Ù„Ø­Ø¬Ù… Ø§Ù„Ø§ÙØªØ±Ø§Ø¶ÙŠ
    page_size = request.GET.get("page_size", "50")  # Ù…Ù† 20 Ø¥Ù„Ù‰ 50
    try:
        page_size = int(page_size)
        if page_size > 200:  # Ù…Ù† 100 Ø¥Ù„Ù‰ 200
            page_size = 200
        elif page_size < 1:
            page_size = 50
    except Exception:
        page_size = 50

    paginator = Paginator(products, page_size)
    page_number = request.GET.get("page", "1")

    # Ø¥ØµÙ„Ø§Ø­ Ù…Ø´ÙƒÙ„Ø© pagination - ØªØ¨Ø³ÙŠØ· Ø§Ù„Ù…Ù†Ø·Ù‚
    try:
        page_number = int(page_number) if page_number else 1
    except (ValueError, TypeError):
        page_number = 1

    page_obj = paginator.get_page(page_number)

    # Ø§Ù„ØªÙ†Ø¨ÙŠÙ‡Ø§Øª - Ù…Ø­Ø³Ù‘Ù†
    from .models import StockAlert

    alerts_count = StockAlert.objects.filter(status="active").count()
    recent_alerts = (
        StockAlert.objects.filter(status="active")
        .select_related("product")
        .only("id", "product__name", "alert_type", "created_at")
        .order_by("-created_at")[:5]
    )

    from datetime import datetime

    current_year = datetime.now().year

    # ===== Ù†ÙˆØ¹ Ø§Ù„Ø³Ø¹Ø± Ø§Ù„Ù…Ø¹Ø±ÙˆØ¶ (Ù‚Ø·Ø§Ø¹ÙŠ/Ø¬Ù…Ù„Ø©) - Ù…Ø­ÙÙˆØ¸ ÙÙŠ Ø§Ù„Ø¬Ù„Ø³Ø© =====
    price_display_mode = request.session.get("price_display_mode", "retail")

    context = {
        "page_obj": page_obj,
        "categories": Category.objects.only("id", "name"),
        "search_query": search_query,
        "selected_category": category_id,
        "selected_filter": filter_type,
        "sort_by": sort_by,
        "active_menu": "products",
        "alerts_count": alerts_count,
        "recent_alerts": recent_alerts,
        "current_year": current_year,
        "page_size": page_size,
        "paginator": paginator,
        "page_number": page_number,
        "price_display_mode": price_display_mode,
    }

    return render(request, "inventory/product_list_new_icons.html", context)


@login_required
def product_create(request):
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            try:
                # Ø­ÙØ¸ Ø§Ù„Ù…Ù†ØªØ¬
                product = form.save()

                # Ø¥Ø¶Ø§ÙØ© Ø§Ù„ÙƒÙ…ÙŠØ© Ø§Ù„Ø­Ø§Ù„ÙŠØ© Ø¥Ø°Ø§ ØªÙ… ØªØ­Ø¯ÙŠØ¯Ù‡Ø§
                initial_quantity = form.cleaned_data.get("initial_quantity", 0)
                warehouse = form.cleaned_data.get("warehouse")

                if initial_quantity > 0 and warehouse:
                    StockTransaction.objects.create(
                        product=product,
                        warehouse=warehouse,
                        transaction_type="in",
                        reason="initial_stock",
                        quantity=initial_quantity,
                        reference="Ø¥Ø¶Ø§ÙØ© Ù…Ù†ØªØ¬ Ø¬Ø¯ÙŠØ¯",
                        notes="Ø§Ù„ÙƒÙ…ÙŠØ© Ø§Ù„Ø§Ø¨ØªØ¯Ø§Ø¦ÙŠØ© Ø¹Ù†Ø¯ Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ù…Ù†ØªØ¬",
                        created_by=request.user,
                        transaction_date=timezone.now(),
                    )

                # Ø¥Ø¹Ø§Ø¯Ø© ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø°Ø§ÙƒØ±Ø© Ø§Ù„Ù…Ø¤Ù‚ØªØ© Ù„Ù„Ù…Ù†ØªØ¬Ø§Øª
                invalidate_product_cache(product.id)

                success_msg = "ØªÙ… Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ù…Ù†ØªØ¬ Ø¨Ù†Ø¬Ø§Ø­."
                if initial_quantity > 0:
                    success_msg += f" ØªÙ… Ø¥Ø¶Ø§ÙØ© {initial_quantity} ÙˆØ­Ø¯Ø© Ø¥Ù„Ù‰ Ù…Ø³ØªÙˆØ¯Ø¹ {warehouse.name}."

                messages.success(request, success_msg)
                return redirect("inventory:product_list")

            except Exception as e:
                messages.error(request, f"Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ù…Ù†ØªØ¬: {str(e)}")
    else:
        form = ProductForm()

    return render(request, "inventory/product_form.html", {"form": form})


@login_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            try:
                # Ø­ÙØ¸ Ø§Ù„Ù…Ù†ØªØ¬
                product = form.save()

                # Ù…Ù„Ø§Ø­Ø¸Ø©: ÙÙŠ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„ØŒ Ù„Ø§ Ù†Ø¶ÙŠÙ ÙƒÙ…ÙŠØ© Ø¬Ø¯ÙŠØ¯Ø© ØªÙ„Ù‚Ø§Ø¦ÙŠØ§Ù‹
                # ÙŠÙ…ÙƒÙ† Ù„Ù„Ù…Ø³ØªØ®Ø¯Ù… Ø¥Ø¶Ø§ÙØ© ÙƒÙ…ÙŠØ© Ù…Ù† Ø®Ù„Ø§Ù„ ØµÙØ­Ø© Ø­Ø±ÙƒØ§Øª Ø§Ù„Ù…Ø®Ø²ÙˆÙ†

                # Ø¥Ø¹Ø§Ø¯Ø© ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø°Ø§ÙƒØ±Ø© Ø§Ù„Ù…Ø¤Ù‚ØªØ© Ù„Ù„Ù…Ù†ØªØ¬Ø§Øª
                invalidate_product_cache(product.id)
                messages.success(request, "ØªÙ… ØªØ­Ø¯ÙŠØ« Ø§Ù„Ù…Ù†ØªØ¬ Ø¨Ù†Ø¬Ø§Ø­.")
                return redirect("inventory:product_list")

            except Exception as e:
                messages.error(request, f"Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ ØªØ­Ø¯ÙŠØ« Ø§Ù„Ù…Ù†ØªØ¬: {str(e)}")
    else:
        form = ProductForm(instance=product)

    return render(
        request, "inventory/product_form.html", {"form": form, "product": product}
    )


@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        try:
            product.delete()
            # Ø¥Ø¹Ø§Ø¯Ø© ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø°Ø§ÙƒØ±Ø© Ø§Ù„Ù…Ø¤Ù‚ØªØ© Ù„Ù„Ù…Ù†ØªØ¬Ø§Øª
            invalidate_product_cache(product.id)
            messages.success(request, "ØªÙ… Ø­Ø°Ù Ø§Ù„Ù…Ù†ØªØ¬ Ø¨Ù†Ø¬Ø§Ø­.")
        except Exception as e:
            messages.error(request, "Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø­Ø°Ù Ø§Ù„Ù…Ù†ØªØ¬.")
        return redirect("inventory:product_list")

    return render(
        request, "inventory/product_confirm_delete.html", {"product": product}
    )


@login_required
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)

    # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ Ù…Ù† property (ÙŠØ­Ø³Ø¨ Ù…Ù† Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹Ø§Øª)
    current_stock = product.current_stock

    # Ø§Ø³ØªØ®Ø¯Ø§Ù… select_related Ù„Ù„Ù…Ø¹Ø§Ù…Ù„Ø§Øª
    transactions = product.transactions.select_related(
        "created_by", "warehouse"
    ).order_by("-transaction_date", "-id")

    # Ø­Ø³Ø§Ø¨ Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„ÙˆØ§Ø±Ø¯ ÙˆØ§Ù„ØµØ§Ø¯Ø±
    from django.db.models import Sum

    transactions_in = product.transactions.filter(transaction_type="in")
    transactions_out = product.transactions.filter(transaction_type="out")

    transactions_in_total = (
        transactions_in.aggregate(total=Sum("quantity"))["total"] or 0
    )
    transactions_out_total = (
        transactions_out.aggregate(total=Sum("quantity"))["total"] or 0
    )

    # Ø¥Ø¹Ø¯Ø§Ø¯ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø±Ø³Ù… Ø§Ù„Ø¨ÙŠØ§Ù†ÙŠ - Ø§Ø³ØªØ®Ø¯Ø§Ù… running_balance
    from datetime import timedelta

    from django.utils import timezone

    from .models import Warehouse

    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=29)

    # Ø­Ø³Ø§Ø¨ Ø§Ù„Ø±ØµÙŠØ¯ Ù„ÙƒÙ„ ÙŠÙˆÙ… Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… running_balance Ù…Ù† Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹Ø§Øª
    transaction_dates = []
    transaction_balances = []

    # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹Ø§Øª Ø§Ù„Ù†Ø´Ø·Ø©
    warehouses = Warehouse.objects.filter(is_active=True)

    # Ù„ÙƒÙ„ ÙŠÙˆÙ…ØŒ Ø§Ø­Ø³Ø¨ Ù…Ø¬Ù…ÙˆØ¹ Ø§Ù„Ø±ØµÙŠØ¯ Ù…Ù† Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹Ø§Øª
    current_date = start_date
    while current_date <= end_date:
        daily_total = 0

        # Ù„ÙƒÙ„ Ù…Ø³ØªÙˆØ¯Ø¹ØŒ Ø§Ø­ØµÙ„ Ø¹Ù„Ù‰ Ø¢Ø®Ø± Ø±ØµÙŠØ¯ Ø­ØªÙ‰ Ù‡Ø°Ø§ Ø§Ù„ÙŠÙˆÙ…
        for warehouse in warehouses:
            last_trans = (
                product.transactions.filter(
                    warehouse=warehouse, transaction_date__date__lte=current_date
                )
                .order_by("-transaction_date", "-id")
                .first()
            )

            if last_trans:
                daily_total += last_trans.running_balance

        transaction_dates.append(current_date)
        transaction_balances.append(float(daily_total))

        current_date += timedelta(days=1)

    # Ø¬Ù„Ø¨ Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ù†Ø¸Ø§Ù…
    system_settings = SystemSettings.get_settings()
    currency_symbol = system_settings.currency_symbol if system_settings else "Ø¬.Ù…"

    # Ø¥Ø¶Ø§ÙØ© Ø¹Ø¯Ø¯ Ø§Ù„ØªÙ†Ø¨ÙŠÙ‡Ø§Øª Ø§Ù„Ù†Ø´Ø·Ø©
    from .models import StockAlert

    alerts_count = StockAlert.objects.filter(status="active").count()

    # Ø¥Ø¶Ø§ÙØ© Ø¢Ø®Ø± Ø§Ù„ØªÙ†Ø¨ÙŠÙ‡Ø§Øª Ù„Ù„Ø¹Ø±Ø¶ ÙÙŠ Ø§Ù„Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ù…Ù†Ø³Ø¯Ù„Ø©
    recent_alerts = (
        StockAlert.objects.filter(status="active")
        .select_related("product")
        .order_by("-created_at")[:5]
    )

    # Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø­Ø§Ù„ÙŠØ© Ù„Ø´Ø±ÙŠØ· Ø§Ù„ØªØ°ÙŠÙŠÙ„
    from datetime import datetime

    current_year = datetime.now().year

    # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø­Ø³Ø¨ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹
    from django.db.models import Max

    from .models import StockTransaction, Warehouse

    warehouses_stock = []
    warehouses = Warehouse.objects.filter(is_active=True)

    for warehouse in warehouses:
        # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø¢Ø®Ø± Ø­Ø±ÙƒØ© Ù„Ù„Ù…Ù†ØªØ¬ ÙÙŠ Ù‡Ø°Ø§ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹
        last_transaction = (
            StockTransaction.objects.filter(product=product, warehouse=warehouse)
            .order_by("-transaction_date", "-id")
            .first()
        )

        if last_transaction:
            warehouses_stock.append(
                {
                    "warehouse": warehouse,
                    "stock": last_transaction.running_balance,
                    "last_update": last_transaction.transaction_date,
                }
            )

    context = {
        "product": product,
        "current_stock": current_stock,
        "stock_status": (
            "Ù†ÙØ° Ù…Ù† Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"
            if current_stock == 0
            else "Ù…Ø®Ø²ÙˆÙ† Ù…Ù†Ø®ÙØ¶" if current_stock <= product.minimum_stock else "Ù…ØªÙˆÙØ±"
        ),
        "transactions": transactions,
        "transactions_in_total": transactions_in_total,
        "transactions_out_total": transactions_out_total,
        "transaction_dates": transaction_dates,
        "transaction_balances": transaction_balances,
        "active_menu": "products",
        "alerts_count": alerts_count,
        "recent_alerts": recent_alerts,
        "current_year": current_year,
        "currency_symbol": currency_symbol,
        "warehouses_stock": warehouses_stock,
    }
    return render(request, "inventory/product_detail.html", context)


@login_required
def transaction_create(request, product_pk):
    product = get_object_or_404(Product, pk=product_pk)

    # Ø¥Ø¶Ø§ÙØ© Ø£Ù†ÙˆØ§Ø¹ Ø§Ù„Ù…Ø¹Ø§Ù…Ù„Ø§Øª ÙˆØ£Ø³Ø¨Ø§Ø¨Ù‡Ø§ Ù„Ù„Ù‚Ø§Ù„Ø¨
    transaction_types = [
        ("in", "ÙˆØ§Ø±Ø¯"),
        ("out", "ØµØ§Ø¯Ø±"),
    ]

    transaction_reasons = [
        ("purchase", "Ø´Ø±Ø§Ø¡"),
        ("return", "Ù…Ø±ØªØ¬Ø¹"),
        ("adjustment", "ØªØ³ÙˆÙŠØ©"),
        ("transfer", "Ù†Ù‚Ù„"),
        ("sale", "Ø¨ÙŠØ¹"),
        ("damage", "ØªØ§Ù„Ù"),
        ("other", "Ø£Ø®Ø±Ù‰"),
    ]

    # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ù…Ø³ØªÙˆÙ‰ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ Ù…Ù† property
    current_stock = product.current_stock

    # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹Ø§Øª Ø§Ù„Ù†Ø´Ø·Ø© Ù…Ø¹ Ù…Ø®Ø²ÙˆÙ†Ù‡Ø§
    from .models import Warehouse

    warehouses_list = []
    for warehouse in Warehouse.objects.filter(is_active=True):
        last_trans = (
            StockTransaction.objects.filter(product=product, warehouse=warehouse)
            .order_by("-transaction_date", "-id")
            .first()
        )

        warehouses_list.append(
            {
                "id": warehouse.id,
                "name": warehouse.name,
                "stock": last_trans.running_balance if last_trans else 0,
            }
        )

    if request.method == "POST":
        try:
            # Ø§Ø³ØªØ®Ø±Ø§Ø¬ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ù† Ø§Ù„Ù†Ù…ÙˆØ°Ø¬
            transaction_type = request.POST.get("transaction_type")
            reason = request.POST.get("reason")
            quantity = request.POST.get("quantity")
            warehouse_id = request.POST.get("warehouse")
            reference = request.POST.get("reference", "")
            notes = request.POST.get("notes", "")

            # Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø©
            if not all([transaction_type, reason, quantity, warehouse_id]):
                raise ValueError("Ø¬Ù…ÙŠØ¹ Ø§Ù„Ø­Ù‚ÙˆÙ„ Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø© ÙŠØ¬Ø¨ Ù…Ù„Ø¤Ù‡Ø§")

            # Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ØµØ­Ø© Ø§Ù„ÙƒÙ…ÙŠØ©
            try:
                quantity = Decimal(quantity)
                if quantity <= 0:
                    raise ValueError("ÙŠØ¬Ø¨ Ø£Ù† ØªÙƒÙˆÙ† Ø§Ù„ÙƒÙ…ÙŠØ© Ø£ÙƒØ¨Ø± Ù…Ù† ØµÙØ±")
            except (ValueError, TypeError, Exception):
                raise ValueError("Ø§Ù„ÙƒÙ…ÙŠØ© ÙŠØ¬Ø¨ Ø£Ù† ØªÙƒÙˆÙ† Ø±Ù‚Ù…Ø§Ù‹ ØµØ­ÙŠØ­Ø§Ù‹")

            # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹
            warehouse = get_object_or_404(Warehouse, pk=warehouse_id)

            # ØªØ­Ø¯ÙŠØ« ÙˆØ­Ø¯Ø© Ø§Ù„Ù‚ÙŠØ§Ø³ Ø¥Ø°Ø§ ØªÙ… ØªØºÙŠÙŠØ±Ù‡Ø§
            new_unit = request.POST.get("unit")
            if new_unit and new_unit != product.unit:
                # Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ø£Ù† Ø§Ù„ÙˆØ­Ø¯Ø© ØµØ§Ù„Ø­Ø©
                valid_units = [choice[0] for choice in Product.UNIT_CHOICES]
                if new_unit in valid_units:
                    product.unit = new_unit
                    product.save()
                    # ØªØ­Ø¯ÙŠØ« Ø§Ù„ÙƒØ§Ø¦Ù† ÙÙŠ Ø§Ù„Ø°Ø§ÙƒØ±Ø©
                    product.refresh_from_db()

            # Ø¥Ø¹Ø§Ø¯Ø© ÙØ­Øµ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ Ù‚Ø¨Ù„ ØªØ³Ø¬ÙŠÙ„ Ø§Ù„Ø­Ø±ÙƒØ©
            current_stock = product.current_stock

            # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹ Ø§Ù„Ù…Ø­Ø¯Ø¯
            warehouse_stock = 0
            last_trans = (
                StockTransaction.objects.filter(product=product, warehouse=warehouse)
                .order_by("-transaction_date", "-id")
                .first()
            )

            if last_trans:
                warehouse_stock = last_trans.running_balance

            # Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ØªÙˆÙØ± Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ù„Ù„Ø­Ø±ÙƒØ§Øª Ø§Ù„ØµØ§Ø¯Ø±Ø©
            if transaction_type == "out":
                if warehouse_stock <= 0:
                    raise ValueError(
                        f"Ù„Ø§ ÙŠÙˆØ¬Ø¯ Ù…Ø®Ø²ÙˆÙ† Ù…ØªØ§Ø­ Ù„Ù„ØµØ±Ù Ù…Ù† Ù…Ø³ØªÙˆØ¯Ø¹ {warehouse.name}"
                    )
                if quantity > warehouse_stock:
                    raise ValueError(
                        f"Ø§Ù„ÙƒÙ…ÙŠØ© Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø© ({quantity}) Ø£ÙƒØ¨Ø± Ù…Ù† Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ù…ØªØ§Ø­ ÙÙŠ {warehouse.name} ({warehouse_stock})"
                    )

            # Ø­Ø³Ø§Ø¨ running_balance
            if last_trans:
                previous_balance = last_trans.running_balance
            else:
                previous_balance = 0

            if transaction_type == "in":
                new_balance = previous_balance + quantity
            else:
                new_balance = previous_balance - quantity

            # Ø¥Ù†Ø´Ø§Ø¡ Ø­Ø±ÙƒØ© Ø§Ù„Ù…Ø®Ø²ÙˆÙ†
            transaction = StockTransaction.objects.create(
                product=product,
                warehouse=warehouse,
                transaction_type=transaction_type,
                reason=reason,
                quantity=quantity,
                running_balance=new_balance,
                reference=reference,
                notes=notes,
                created_by=request.user,
                transaction_date=timezone.now(),
            )

            # Ø¥Ø¹Ø§Ø¯Ø© ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø°Ø§ÙƒØ±Ø© Ø§Ù„Ù…Ø¤Ù‚ØªØ© Ù„Ù„Ù…Ù†ØªØ¬
            invalidate_product_cache(product.id)

            # Ø¥Ø¶Ø§ÙØ© Ø±Ø³Ø§Ù„Ø© Ù†Ø¬Ø§Ø­
            if transaction_type == "in":
                messages.success(
                    request,
                    f"ØªÙ… ØªØ³Ø¬ÙŠÙ„ Ø­Ø±ÙƒØ© ÙˆØ§Ø±Ø¯ Ø¨Ù†Ø¬Ø§Ø­ Ø¥Ù„Ù‰ {warehouse.name}. Ø§Ù„ÙƒÙ…ÙŠØ©: {quantity}",
                )
            else:
                messages.success(
                    request,
                    f"ØªÙ… ØªØ³Ø¬ÙŠÙ„ Ø­Ø±ÙƒØ© ØµØ§Ø¯Ø± Ø¨Ù†Ø¬Ø§Ø­ Ù…Ù† {warehouse.name}. Ø§Ù„ÙƒÙ…ÙŠØ©: {quantity}",
                )

            return redirect("inventory:product_detail", pk=product.pk)

        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ ØªØ³Ø¬ÙŠÙ„ Ø­Ø±ÙƒØ© Ø§Ù„Ù…Ø®Ø²ÙˆÙ†: {str(e)}")

    # Ø¥Ø¶Ø§ÙØ© Ø¹Ø¯Ø¯ Ø§Ù„ØªÙ†Ø¨ÙŠÙ‡Ø§Øª Ø§Ù„Ù†Ø´Ø·Ø©
    alerts_count = StockAlert.objects.filter(status="active").count()

    # Ø¥Ø¶Ø§ÙØ© Ø¢Ø®Ø± Ø§Ù„ØªÙ†Ø¨ÙŠÙ‡Ø§Øª Ù„Ù„Ø¹Ø±Ø¶ ÙÙŠ Ø§Ù„Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ù…Ù†Ø³Ø¯Ù„Ø©
    recent_alerts = (
        StockAlert.objects.filter(status="active")
        .select_related("product")
        .order_by("-created_at")[:5]
    )

    # Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ø³Ù†Ø© Ø§Ù„Ø­Ø§Ù„ÙŠØ© Ù„Ø´Ø±ÙŠØ· Ø§Ù„ØªØ°ÙŠÙŠÙ„
    current_year = datetime.now().year

    context = {
        "product": product,
        "transaction_types": transaction_types,
        "transaction_reasons": transaction_reasons,
        "current_stock": current_stock,
        "alerts_count": alerts_count,
        "recent_alerts": recent_alerts,
        "current_year": current_year,
        "unit_choices": Product.UNIT_CHOICES,
    }

    return render(request, "inventory/transaction_form_new.html", context)


# API Endpoints
from django.http import JsonResponse


@login_required
def product_api_detail(request, pk):
    try:
        product = get_object_or_404(Product, pk=pk)
        current_stock = get_cached_stock_level(product.id)

        data = {
            "id": product.id,
            "name": product.name,
            "code": product.code,
            "category": str(product.category),
            "description": product.description,
            "price": product.price,
            "minimum_stock": product.minimum_stock,
            "current_stock": current_stock,
        }
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({"error": "Ø§Ù„Ù…Ù†ØªØ¬ ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯"}, status=404)


@login_required
def product_api_list(request):
    product_type = request.GET.get("type")

    # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ù…Ù† Ø§Ù„Ø°Ø§ÙƒØ±Ø© Ø§Ù„Ù…Ø¤Ù‚ØªØ©
    products = get_cached_product_list(include_stock=True)

    # ØªØ·Ø¨ÙŠÙ‚ Ø§Ù„ÙÙ„ØªØ± Ø­Ø³Ø¨ Ø§Ù„Ù†ÙˆØ¹
    if product_type:
        if product_type == "fabric":
            products = [p for p in products if p.category.name == "Ø£Ù‚Ù…Ø´Ø©"]
        elif product_type == "accessory":
            products = [p for p in products if p.category.name == "Ø§ÙƒØ³Ø³ÙˆØ§Ø±Ø§Øª"]

    # ØªØ­ÙˆÙŠÙ„ Ø¥Ù„Ù‰ JSON
    data = [
        {
            "id": p.id,
            "name": p.name,
            "code": p.code,
            "category": str(p.category),
            "description": p.description,
            "price": p.price,
            "minimum_stock": p.minimum_stock,
            "current_stock": p.current_stock,
        }
        for p in products
    ]

    return JsonResponse(data, safe=False)


@login_required
def product_api_autocomplete(request):
    """
    API Ù„Ù„Ø¨Ø­Ø« Ø§Ù„Ø³Ø±ÙŠØ¹ Ø¹Ù† Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª (autocomplete) Ù…Ø¹ Ø§Ù„ØªØ®Ø²ÙŠÙ† Ø§Ù„Ù…Ø¤Ù‚Øª
    ÙŠÙ‚Ø¨Ù„ Ø¨Ø§Ø±Ø§Ù…ÙŠØªØ± ?query= ÙˆÙŠØ¹ÙŠØ¯ Ù‚Ø§Ø¦Ù…Ø© Ù…Ø®ØªØµØ±Ø© (id, name, code, price, current_stock)
    """
    query = request.GET.get("query", "").strip()

    # Ø¥Ø°Ø§ ÙƒØ§Ù† Ù‡Ù†Ø§Ùƒ Ø§Ø³ØªØ¹Ù„Ø§Ù…ØŒ Ø§Ø³ØªØ®Ø¯Ù… Ø§Ù„ØªØ®Ø²ÙŠÙ† Ø§Ù„Ù…Ø¤Ù‚Øª
    if query and len(query) >= 2:
        try:
            from orders.cache import search_products_cached

            results = search_products_cached(query)

            # ØªØ­Ø¯ÙŠØ¯ Ø§Ù„Ù†ØªØ§Ø¦Ø¬ Ø¥Ù„Ù‰ 10 Ø¹Ù†Ø§ØµØ± ÙÙ‚Ø·
            results = results[:10]

        except Exception as e:
            import logging

            logger = logging.getLogger(__name__)
            logger.error(f"Ø®Ø·Ø£ ÙÙŠ Ø§Ù„Ø¨Ø­Ø« Ø§Ù„Ù…Ø¤Ù‚Øª Ø¹Ù† Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª: {str(e)}")

            # Ø§Ù„Ø¹ÙˆØ¯Ø© Ù„Ù„Ø·Ø±ÙŠÙ‚Ø© Ø§Ù„ØªÙ‚Ù„ÙŠØ¯ÙŠØ© ÙÙŠ Ø­Ø§Ù„Ø© Ø§Ù„Ø®Ø·Ø£
            results = []
            products = Product.objects.filter(
                Q(name__icontains=query) | Q(code__icontains=query)
            ).select_related("category")[:10]

            for p in products:
                results.append(
                    {
                        "id": p.id,
                        "name": p.name,
                        "code": p.code,
                        "price": float(p.price),
                        "current_stock": p.current_stock,
                        "category": p.category.name if p.category else None,
                        "description": p.description,
                    }
                )
    else:
        # Ù„Ù„Ø§Ø³ØªØ¹Ù„Ø§Ù…Ø§Øª Ø§Ù„ÙØ§Ø±ØºØ© Ø£Ùˆ Ø§Ù„Ù‚ØµÙŠØ±Ø©ØŒ Ø¹Ø±Ø¶ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ø§Ù„Ø£ÙƒØ«Ø± Ø´ÙŠÙˆØ¹Ø§Ù‹
        results = []
        products = Product.objects.select_related("category").order_by("-id")[:10]

        for p in products:
            results.append(
                {
                    "id": p.id,
                    "name": p.name,
                    "code": p.code,
                    "price": float(p.price),
                    "current_stock": p.current_stock,
                    "category": p.category.name if p.category else None,
                    "description": p.description,
                }
            )

    return JsonResponse(results, safe=False)


from django.db.models import Count, Sum

# New API View
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.models import ActivityLog
from customers.models import Customer
from orders.models import Order

from .models import Product


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard_view(request):
    try:
        # Calculate statistics
        stats = {
            "totalCustomers": Customer.objects.count(),
            "totalOrders": Order.objects.count(),
            "inventoryValue": Product.objects.aggregate(total=Sum("price", default=0))[
                "total"
            ],
            "pendingInstallations": 0,
        }

        # Get recent activities
        activities = (
            ActivityLog.objects.select_related("user")
            .order_by("-timestamp")[:10]
            .values("id", "type", "description", "timestamp")
        )

        return Response({"stats": stats, "activities": list(activities)})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@login_required
def product_search_api(request):
    """API Ù„Ù„Ø¨Ø­Ø« Ø¹Ù† Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ù„Ù€ Select2"""
    query = request.GET.get("q", "").strip()
    page = int(request.GET.get("page", 1))
    page_size = 30

    # Ø§Ù„Ø¨Ø­Ø« ÙÙŠ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª
    products = Product.objects.all()

    if query:
        products = products.filter(
            Q(name__icontains=query)
            | Q(code__icontains=query)
            | Q(description__icontains=query)
        )

    # Ø­Ø³Ø§Ø¨ Ø§Ù„Ø¹Ø¯Ø¯ Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ
    total_count = products.count()

    # ØªØ·Ø¨ÙŠÙ‚ Ø§Ù„ØµÙØ­Ø§Øª
    start = (page - 1) * page_size
    end = start + page_size
    products = products[start:end]

    # ØªØ­Ø¶ÙŠØ± Ø§Ù„Ù†ØªØ§Ø¦Ø¬
    results = []
    for product in products:
        results.append(
            {
                "id": product.id,
                "text": f"{product.name} - {product.price} Ø¬.Ù…",
                "name": product.name,
                "price": float(product.price),
                "code": product.code or "",
                "current_stock": get_cached_stock_level(product.id),
            }
        )

    return JsonResponse(
        {
            "results": results,
            "pagination": {"more": end < total_count},
            "total_count": total_count,
        }
    )


@login_required
def barcode_scan_api(request):
    """
    API Ù„ÙØ­Øµ Ø§Ù„Ø¨Ø§Ø±ÙƒÙˆØ¯ ÙˆØ§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„Ù…Ù†ØªØ¬
    ÙŠØ³ØªÙ‚Ø¨Ù„ ÙƒÙˆØ¯ Ø§Ù„Ù…Ù†ØªØ¬ ÙˆÙŠØ±Ø¬Ø¹ ÙƒÙ„ Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª
    """
    barcode = request.GET.get("barcode", "").strip()

    if not barcode:
        return JsonResponse(
            {"success": False, "error": "Ù„Ù… ÙŠØªÙ… ØªÙˆÙÙŠØ± Ø±Ù…Ø² Ø§Ù„Ø¨Ø§Ø±ÙƒÙˆØ¯"}, status=400
        )

    try:
        # Ø§Ù„Ø¨Ø­Ø« Ø¹Ù† Ø§Ù„Ù…Ù†ØªØ¬ Ø¨ÙˆØ§Ø³Ø·Ø© Ø§Ù„ÙƒÙˆØ¯ Ø£ÙˆÙ„Ø§Ù‹ØŒ Ø«Ù… Ø¨Ø§Ù„Ø§Ø³Ù…
        try:
            product = Product.objects.select_related("category").get(code=barcode)
        except Product.DoesNotExist:
            # Ù…Ø­Ø§ÙˆÙ„Ø© Ø§Ù„Ø¨Ø­Ø« Ø¨Ø§Ù„Ø§Ø³Ù… (Ù…Ø·Ø§Ø¨Ù‚Ø© ØªØ§Ù…Ø© Ø£Ùˆ Ø¬Ø²Ø¦ÙŠØ©)
            products = Product.objects.select_related("category").filter(
                Q(name__iexact=barcode) | Q(name__icontains=barcode)
            )
            if products.exists():
                product = products.first()
            else:
                raise Product.DoesNotExist()

        # Ø­Ø³Ø§Ø¨ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ Ù…Ù† Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹Ø§Øª
        current_stock = product.current_stock

        # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø­Ø³Ø¨ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹
        from .models import Warehouse

        warehouses_stock = []
        warehouses = Warehouse.objects.filter(is_active=True)

        for warehouse in warehouses:
            last_transaction = (
                StockTransaction.objects.filter(product=product, warehouse=warehouse)
                .order_by("-transaction_date", "-id")
                .first()
            )

            if last_transaction and last_transaction.running_balance > 0:
                warehouses_stock.append(
                    {
                        "warehouse_id": warehouse.id,
                        "warehouse_name": warehouse.name,
                        "warehouse_code": warehouse.code,
                        "stock": float(last_transaction.running_balance),
                        "last_update": last_transaction.transaction_date.strftime(
                            "%Y-%m-%d %H:%M"
                        ),
                    }
                )

        # Ø¬Ù„Ø¨ Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ù†Ø¸Ø§Ù… Ù„Ù„Ø¹Ù…Ù„Ø©
        system_settings = SystemSettings.get_settings()
        currency_symbol = system_settings.currency_symbol if system_settings else "Ø¬.Ù…"

        # ØªØ¬Ù‡ÙŠØ² Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
        data = {
            "success": True,
            "product": {
                "id": product.id,
                "name": product.name,
                "code": product.code,
                "price": float(product.price),
                "currency": product.currency,
                "currency_symbol": currency_symbol,
                "unit": product.unit,
                "unit_display": product.get_unit_display(),
                "category": product.category.name if product.category else "ØºÙŠØ± Ù…ØµÙ†Ù",
                "description": product.description,
                "current_stock": float(current_stock),
                "minimum_stock": product.minimum_stock,
                "stock_status": product.stock_status,
                "is_available": product.is_available,
                "warehouses": warehouses_stock,
                "created_at": product.created_at.strftime("%Y-%m-%d"),
            },
        }

        return JsonResponse(data)

    except Product.DoesNotExist:
        return JsonResponse(
            {
                "success": False,
                "error": "Ø§Ù„Ù…Ù†ØªØ¬ ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯",
                "message": f"Ù„Ø§ ÙŠÙˆØ¬Ø¯ Ù…Ù†ØªØ¬ Ø¨ÙƒÙˆØ¯ Ø§Ù„Ø¨Ø§Ø±ÙƒÙˆØ¯: {barcode}",
            },
            status=404,
        )
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": "Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ù„Ø¨Ø­Ø« Ø¹Ù† Ø§Ù„Ù…Ù†ØªØ¬",
                "message": str(e),
            },
            status=500,
        )


# ==================== QR Code Generation APIs ====================


@login_required
def generate_single_qr_api(request, pk):
    """
    API Ù„ØªÙˆÙ„ÙŠØ¯ QR Ù„Ù…Ù†ØªØ¬ ÙˆØ§Ø­Ø¯
    POST /inventory/api/product/<pk>/generate-qr/
    """
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Method not allowed"}, status=405
        )

    product = get_object_or_404(Product, pk=pk)

    if not product.code:
        return JsonResponse(
            {"success": False, "error": "Ù„Ø§ ÙŠÙˆØ¬Ø¯ ÙƒÙˆØ¯ Ù„Ù„Ù…Ù†ØªØ¬ - Ù„Ø§ ÙŠÙ…ÙƒÙ† ØªÙˆÙ„ÙŠØ¯ QR"}
        )

    # Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ÙˆØ¬ÙˆØ¯ QR Ù…Ø³Ø¨Ù‚
    had_existing = bool(product.qr_code_base64)

    # ØªÙˆÙ„ÙŠØ¯ QR (Ù…Ø¹ Ø§Ù„Ø¥Ø¬Ø¨Ø§Ø± Ù„Ø¥Ø¹Ø§Ø¯Ø© Ø§Ù„ØªÙˆÙ„ÙŠØ¯)
    product.generate_qr(force=True)
    product.save(update_fields=["qr_code_base64"])

    return JsonResponse(
        {
            "success": True,
            "message": (
                "ØªÙ… Ø¥Ø¹Ø§Ø¯Ø© ØªÙˆÙ„ÙŠØ¯ Ø±Ù…Ø² QR Ø¨Ù†Ø¬Ø§Ø­"
                if had_existing
                else "ØªÙ… ØªÙˆÙ„ÙŠØ¯ Ø±Ù…Ø² QR Ø¨Ù†Ø¬Ø§Ø­"
            ),
            "qr_exists": True,
            "product_code": product.code,
        }
    )


@login_required
def generate_all_qr_api(request):
    """
    API Ù„ØªÙˆÙ„ÙŠØ¯ QR Ù„Ù„Ù…Ù†ØªØ¬Ø§Øª Ø§Ù„Ø£Ø³Ø§Ø³ÙŠØ© (Ø§Ù„Ù†Ø¸Ø§Ù… Ø§Ù„Ø¬Ø¯ÙŠØ¯)
    POST /inventory/api/generate-all-qr/
    """
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Method not allowed"}, status=405
        )

    from .models import BaseProduct

    # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ø§Ù„Ø£Ø³Ø§Ø³ÙŠØ© Ø§Ù„ØªÙŠ Ù„ÙŠØ³ Ù„Ù‡Ø§ QR ÙˆÙ„Ø¯ÙŠÙ‡Ø§ ÙƒÙˆØ¯
    base_products_no_qr = (
        BaseProduct.objects.filter(qr_code_base64__isnull=True)
        .exclude(code__isnull=True)
        .exclude(code="")
    )

    total = base_products_no_qr.count()

    if total == 0:
        return JsonResponse(
            {
                "success": True,
                "message": "Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ù„Ø¯ÙŠÙ‡Ø§ Ø±Ù…ÙˆØ² QR Ø¨Ø§Ù„ÙØ¹Ù„!",
                "generated": 0,
                "total": 0,
            }
        )

    generated = 0
    errors = 0
    error_list = []

    # Ù…Ø¹Ø§Ù„Ø¬Ø© Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª (Ø¨Ø­Ø¯ Ø£Ù‚ØµÙ‰ 500 Ù„ØªØ¬Ù†Ø¨ Ø§Ù„ØªÙˆÙ‚Ù)
    LIMIT = 500
    for bp in base_products_no_qr[:LIMIT]:
        try:
            if bp.generate_qr():
                # Ø§Ø³ØªØ®Ø¯Ø§Ù… update_fields Ù„Ù„ØªØ­Ø¯ÙŠØ« Ø§Ù„Ø³Ø±ÙŠØ¹ ÙˆØ§Ù„Ø¢Ù…Ù†
                BaseProduct.objects.filter(pk=bp.pk).update(
                    qr_code_base64=bp.qr_code_base64
                )
                generated += 1
        except Exception as e:
            errors += 1
            error_list.append(f"{bp.code}: {str(e)}")

    remaining = total - generated - errors

    return JsonResponse(
        {
            "success": True,
            "message": f"ØªÙ… ØªÙˆÙ„ÙŠØ¯ {generated} Ø±Ù…Ø² QR. Ø§Ù„Ø£Ø®Ø·Ø§Ø¡: {errors}. Ø§Ù„Ù…ØªØ¨Ù‚ÙŠ: {remaining}",
            "generated": generated,
            "errors": errors,
            "remaining": remaining,
            "total": total,
        }
    )


@login_required
def generate_qr_pdf_api(request):
    """
    API Ù„ØªÙˆÙ„ÙŠØ¯ ÙˆØªØ­Ù…ÙŠÙ„ Ù…Ù„Ù PDF Ù„Ù„Ù€ QR Codes
    GET /inventory/api/generate-qr-pdf/
    """
    import os

    from django.conf import settings
    from django.core.management import call_command

    try:
        # Define output relative to MEDIA_ROOT
        filename = "products_qr_catalog.pdf"
        relative_path = os.path.join("qr_codes", filename)
        full_path = os.path.join(settings.MEDIA_ROOT, relative_path)

        # Ensure directory exists
        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        # Run command
        call_command("generate_qr_pdf", output=full_path)

        # Construct URL
        file_url = os.path.join(settings.MEDIA_URL, relative_path)

        return JsonResponse(
            {"success": True, "message": "ØªÙ… ØªÙˆÙ„ÙŠØ¯ Ù…Ù„Ù PDF Ø¨Ù†Ø¬Ø§Ø­", "file_url": file_url}
        )
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


# ==================== Excel Export API ====================


@login_required
def export_products_excel(request):
    """
    ØªØµØ¯ÙŠØ± Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ø¥Ù„Ù‰ Ù…Ù„Ù Excel
    ÙŠØ­ØªÙˆÙŠ Ø¹Ù„Ù‰: Ø§Ù„ÙƒÙˆØ¯ØŒ Ø§Ø³Ù… Ø§Ù„Ù…Ù†ØªØ¬ØŒ Ø§Ù„ÙØ¦Ø©ØŒ Ø§Ù„Ø³Ø¹Ø±ØŒ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ
    """
    # ÙØ­Øµ Ø§Ù„ØµÙ„Ø§Ø­ÙŠØ© - ÙÙ‚Ø· Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…ÙˆÙ† Ø§Ù„Ø°ÙŠÙ† Ù„Ø¯ÙŠÙ‡Ù… can_export=True
    if not request.user.can_export:
        messages.error(request, "Ù„ÙŠØ³ Ù„Ø¯ÙŠÙƒ ØµÙ„Ø§Ø­ÙŠØ© ØªØµØ¯ÙŠØ± Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª")
        return redirect("inventory:product_list")

    from django.db.models import IntegerField, OuterRef, Subquery
    from django.db.models.functions import Coalesce
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Ø¥Ù†Ø´Ø§Ø¡ Ù…Ù„Ù Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª"

    # ØªÙ†Ø³ÙŠÙ‚ Ø§Ù„Ø¹Ù†Ø§ÙˆÙŠÙ†
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ø¹Ù†Ø§ÙˆÙŠÙ†
    headers = [
        "Code",
        "Product Name",
        "Category",
        "Price",
        "Wholesale Price",
        "Currency",
        "Unit",
        "Width",
        "Material",
        "Description",
        "Total Stock",
        "Warehouse",
        "Quantity",
        "Min Stock",
        "Date Added",
        "Last Updated",
        "Status",
    ]
    ws.append(headers)

    # ØªÙ†Ø³ÙŠÙ‚ ØµÙ Ø§Ù„Ø¹Ù†Ø§ÙˆÙŠÙ†
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª Ù…Ø¹ Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ
    latest_balance = (
        StockTransaction.objects.filter(product=OuterRef("pk"))
        .order_by("-transaction_date", "-id")
        .values("running_balance")[:1]
    )

    products = (
        Product.objects.select_related("category")
        .annotate(
            current_stock_calc=Coalesce(
                Subquery(latest_balance, output_field=IntegerField()), 0
            )
        )
        .order_by("code")
    )

    # ØªØ¬Ù‡ÙŠØ² Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹Ø§Øª Ø§Ù„Ù†Ø´Ø·Ø©
    from .models import Warehouse

    active_warehouses = list(Warehouse.objects.filter(is_active=True))

    # ØªØ­Ø³ÙŠÙ† Ø§Ù„Ø£Ø¯Ø§Ø¡: Ø¬Ù„Ø¨ Ø£Ø±ØµØ¯Ø© Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹Ø§Øª Ø¯ÙØ¹Ø© ÙˆØ§Ø­Ø¯Ø© Ù„ØªØ¬Ù†Ø¨ N+1 queries
    # Ø³Ù†Ù‚ÙˆÙ… Ø¨Ø¬Ù„Ø¨ Ø£Ø­Ø¯Ø« Ø­Ø±ÙƒØ© Ù„ÙƒÙ„ (Ù…Ù†ØªØ¬ØŒ Ù…Ø³ØªÙˆØ¯Ø¹)
    # Ù…Ù„Ø§Ø­Ø¸Ø©: Ù‡Ø°Ø§ ÙŠØªØ·Ù„Ø¨ Ø§Ø³ØªØ¹Ù„Ø§Ù…Ø§Ù‹ Ù…Ø¹Ù‚Ø¯Ø§Ù‹ Ù‚Ù„ÙŠÙ„Ø§Ù‹ØŒ Ù„Ø°Ø§ Ù„Ù„ØªØ¨Ø³ÙŠØ· ÙˆØ§Ù„Ø³Ø±Ø¹Ø© Ø³Ù†Ù‚ÙˆÙ… Ø¨Ø¬Ù„Ø¨ ÙƒÙ„ Ø§Ù„Ø£Ø±ØµØ¯Ø© Ø§Ù„Ø­Ø§Ù„ÙŠØ©
    # ÙˆÙ†Ù‚ÙˆÙ… Ø¨ØªØ¬Ù…ÙŠØ¹Ù‡Ø§ ÙÙŠ Ø§Ù„Ù‚Ø§Ù…ÙˆØ³.

    # 1. Ø¬Ù„Ø¨ Ù…Ø¹Ø±ÙØ§Øª Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª
    product_ids = [p.id for p in products]

    # 2. Ø¬Ù„Ø¨ Ø£Ø­Ø¯Ø« Ø§Ù„Ø­Ø±ÙƒØ§Øª Ù„ÙƒÙ„ Ù…Ù†ØªØ¬ ÙˆÙ…Ø³ØªÙˆØ¯Ø¹
    # Ù†Ø³ØªØ®Ø¯Ù… Subquery Ù„Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ ID Ø£Ø­Ø¯Ø« Ø­Ø±ÙƒØ©
    from django.db.models import Max

    latest_tx_ids = (
        StockTransaction.objects.filter(
            product_id__in=product_ids, warehouse__in=active_warehouses
        )
        .values("product_id", "warehouse_id")
        .annotate(max_id=Max("id"))
        .values("max_id")
    )

    stock_data = StockTransaction.objects.filter(id__in=latest_tx_ids).values(
        "product_id", "warehouse__name", "running_balance"
    )

    # 3. ØªØ­ÙˆÙŠÙ„ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø¥Ù„Ù‰ Ù‚Ø§Ù…ÙˆØ³ Ù„Ù„ÙˆØµÙˆÙ„ Ø§Ù„Ø³Ø±ÙŠØ¹
    # structure: {product_id: [{'name': warehouse_name, 'qty': balance}, ...]}
    product_stock_map = {}
    for item in stock_data:
        p_id = item["product_id"]
        w_name = item["warehouse__name"]
        balance = item["running_balance"]

        if balance > 0:
            if p_id not in product_stock_map:
                product_stock_map[p_id] = []
            product_stock_map[p_id].append({"name": w_name, "qty": balance})

    # Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
    for product in products:
        # ØªØ­Ø¯ÙŠØ¯ Ø§Ù„Ø­Ø§Ù„Ø©
        if product.current_stock_calc == 0:
            status = "Out of Stock"
        elif product.current_stock_calc <= product.minimum_stock:
            status = "Low Stock"
        else:
            status = "Available"

        # Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø£Ø³Ø§Ø³ÙŠØ© Ù„Ù„Ù…Ù†ØªØ¬
        base_row = [
            str(product.code or "-"),
            str(product.name),
            str(product.category.name if product.category else "Uncategorized"),
            float(product.price),
            float(product.wholesale_price),
            str(product.currency),
            str(product.get_unit_display()),
            str(product.width),
            str(product.material),
            str(product.description),
            product.current_stock_calc,
        ]

        # Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø®ØªØ§Ù…ÙŠØ©
        end_row = [
            product.minimum_stock,
            product.created_at.strftime("%Y-%m-%d"),
            product.updated_at.strftime("%Y-%m-%d %H:%M"),
            status,
        ]

        # ØªÙØ§ØµÙŠÙ„ Ø§Ù„Ù…Ø³ØªÙˆØ¯Ø¹Ø§Øª (ØµÙÙˆÙ Ù…ØªØ¹Ø¯Ø¯Ø©)
        wh_entries = product_stock_map.get(product.id, [])

        if wh_entries:
            for entry in wh_entries:
                # Ø¯Ù…Ø¬: Ø£Ø³Ø§Ø³ÙŠ + Ù…Ø³ØªÙˆØ¯Ø¹ + ÙƒÙ…ÙŠØ© + Ø®ØªØ§Ù…ÙŠ
                full_row = base_row + [entry["name"], entry["qty"]] + end_row
                ws.append(full_row)
        else:
            # ØµÙ ÙˆØ§Ø­Ø¯ Ø¨Ø¨ÙŠØ§Ù†Ø§Øª ÙØ§Ø±ØºØ© Ù„Ù„Ù…Ø³ØªÙˆØ¯Ø¹ Ø¥Ø°Ø§ Ù„Ù… ÙŠÙˆØ¬Ø¯ Ù…Ø®Ø²ÙˆÙ†
            full_row = base_row + ["-", 0] + end_row
            ws.append(full_row)

    # ØªÙ†Ø³ÙŠÙ‚ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø©
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)

        # ØªØ¹ÙŠÙŠÙ† Ø¹Ø±Ø¶ Ø§Ù„Ø¹Ù…ÙˆØ¯
        if col_num == 1:  # Ø§Ù„ÙƒÙˆØ¯
            ws.column_dimensions[column_letter].width = 15
        elif col_num == 2:  # Ø§Ø³Ù… Ø§Ù„Ù…Ù†ØªØ¬
            ws.column_dimensions[column_letter].width = 40
        elif col_num == 3:  # Ø§Ù„ÙØ¦Ø©
            ws.column_dimensions[column_letter].width = 20
        elif col_num == 10:  # Description
            ws.column_dimensions[column_letter].width = 50
        elif col_num == 12:  # Warehouse Name
            ws.column_dimensions[column_letter].width = 25
        elif col_num == 13:  # Quantity
            ws.column_dimensions[column_letter].width = 15
        elif col_num in [15, 16]:  # Dates
            ws.column_dimensions[column_letter].width = 20
        else:
            ws.column_dimensions[column_letter].width = 15

        # Ù…Ø­Ø§Ø°Ø§Ø© Ø§Ù„Ù†Øµ
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            if col_num in [4, 5, 6]:  # Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ø±Ù‚Ù…ÙŠØ©
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")

    # Ø¥Ø¹Ø¯Ø§Ø¯ Ø§Ù„Ø§Ø³ØªØ¬Ø§Ø¨Ø©
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ØªØ³Ù…ÙŠØ© Ø§Ù„Ù…Ù„Ù Ø¨Ø§Ù„ØªØ§Ø±ÙŠØ® Ø§Ù„Ø­Ø§Ù„ÙŠ
    from datetime import datetime

    filename = f"products_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Ø­ÙØ¸ Ø§Ù„Ù…Ù„Ù
    wb.save(response)

    return response


@login_required
def feature_not_implemented(request, *args, **kwargs):
    """
    Ø¹Ø±Ø¶ Ù…Ø¤Ù‚Øª Ù„Ù„Ù…ÙŠØ²Ø§Øª Ù‚ÙŠØ¯ Ø§Ù„ØªØ·ÙˆÙŠØ±.
    ÙŠÙØ³ØªØ®Ø¯Ù… Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø¥Ø¹Ø§Ø¯Ø© ØªÙˆØ¬ÙŠÙ‡ stub URLs Ø¥Ù„Ù‰ Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª.
    """
    messages.info(
        request,
        "ğŸš§ Ù‡Ø°Ù‡ Ø§Ù„Ù…ÙŠØ²Ø© Ù‚ÙŠØ¯ Ø§Ù„ØªØ·ÙˆÙŠØ± ÙˆØ³ØªÙƒÙˆÙ† Ù…ØªØ§Ø­Ø© Ù‚Ø±ÙŠØ¨Ø§Ù‹.",
    )
    # Ø§Ù„Ø±Ø¬ÙˆØ¹ Ø¥Ù„Ù‰ Ø§Ù„ØµÙØ­Ø© Ø§Ù„Ø³Ø§Ø¨Ù‚Ø© Ø£Ùˆ Ù„ÙˆØ­Ø© Ø§Ù„ØªØ­ÙƒÙ…
    referer = request.META.get("HTTP_REFERER")
    if referer:
        return redirect(referer)
    return redirect("inventory:dashboard")
