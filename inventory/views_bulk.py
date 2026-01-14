import logging
import os
import traceback
from io import BytesIO

import openpyxl
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl.styles import Font, PatternFill

from .cache_utils import invalidate_product_cache
from .forms import BulkStockUpdateForm, ProductExcelUploadForm
from .models import (
    BulkUploadError,
    BulkUploadLog,
    Category,
    Product,
    StockTransaction,
    Warehouse,
)

logger = logging.getLogger(__name__)


def get_or_create_warehouse(warehouse_name, user, is_fabric_warehouse=False):
    """
    الحصول على المستودع أو إنشاؤه إذا لم يكن موجوداً - محسّن
    مع دعم تحديد المستودعات الرسمية للأقمشة
    """
    if not warehouse_name or str(warehouse_name).strip().lower() in ["", "nan", "none"]:
        return None

    warehouse_name = str(warehouse_name).strip()

    # البحث عن المستودع بالاسم أولاً (حساس لحالة الأحرف)
    warehouse = Warehouse.objects.filter(name__iexact=warehouse_name).first()

    if warehouse:
        # تحديث حالة المستودع إذا كان للأقمشة
        if is_fabric_warehouse and hasattr(warehouse, "is_official_fabric_warehouse"):
            if not warehouse.is_official_fabric_warehouse:
                warehouse.is_official_fabric_warehouse = True
                warehouse.save(update_fields=["is_official_fabric_warehouse"])
        logger.info(f"✅ استخدام مستودع موجود: {warehouse.name} ({warehouse.code})")
        return warehouse

    # إنشاء كود تلقائي للمستودع الجديد فقط
    import re
    import uuid

    # إنشاء كود فريد
    code_base = re.sub(r"[^\w\u0600-\u06FF]", "", warehouse_name)[:8]
    if not code_base:
        code_base = "FABRIC" if is_fabric_warehouse else "WH"

    # إضافة UUID قصير لضمان عدم التكرار
    unique_suffix = uuid.uuid4().hex[:4].upper()
    code = f"{code_base}{unique_suffix}"

    # التأكد النهائي من عدم التكرار
    counter = 1
    original_code = code
    while Warehouse.objects.filter(code=code).exists():
        counter += 1
        code = f"{original_code}{counter}"
        if counter > 100:  # حماية من loop لا نهائي
            logger.error(f"⚠️ فشل في إنشاء كود فريد للمستودع: {warehouse_name}")
            return None

    try:
        # إنشاء المستودع الجديد
        warehouse_data = {
            "name": warehouse_name,
            "code": code,
            "is_active": True,
            "notes": f"تم إنشاؤه تلقائياً من رفع المنتجات بالجملة"
            + (" - مستودع رسمي للأقمشة" if is_fabric_warehouse else ""),
            "created_by": user,
        }

        # إضافة الحقل إذا كان موجوداً في النموذج
        if hasattr(Warehouse, "is_official_fabric_warehouse"):
            warehouse_data["is_official_fabric_warehouse"] = is_fabric_warehouse

        warehouse = Warehouse.objects.create(**warehouse_data)
        logger.info(
            f"✅ تم إنشاء مستودع جديد: {warehouse.name} ({warehouse.code})"
            + (" [أقمشة]" if is_fabric_warehouse else "")
        )
        return warehouse

    except Exception as e:
        logger.error(f"❌ خطأ في إنشاء المستودع {warehouse_name}: {e}")
        # محاولة البحث مرة أخرى (ربما تم إنشاؤه في عملية أخرى)
        warehouse = Warehouse.objects.filter(name__iexact=warehouse_name).first()
        if warehouse:
            logger.warning(f"⚠️ تم العثور على المستودع بعد الفشل: {warehouse.name}")
            return warehouse
        return None


def safe_read_excel(file_data):
    """
    قراءة ملف إكسل بطريقة آمنة تتجنب أخطاء extLst و PatternFill
    """
    print("🔍 محاولة قراءة ملف الإكسل...")
    print(f"📊 حجم الملف: {len(file_data)} بايت")

    # الطريقة الأولى: openpyxl مع تجاهل التنسيقات
    try:
        print("📈 محاولة القراءة بمحرك openpyxl مع تجاهل التنسيقات...")
        df = pd.read_excel(BytesIO(file_data), engine="openpyxl", keep_default_na=False)
        print("✅ تم قراءة الملف بنجاح بمحرك openpyxl")
        return df
    except Exception as e:
        print(f"❌ فشل openpyxl: {str(e)}")

        # الطريقة الثانية: xlrd للملفات القديمة
        try:
            print("📊 محاولة القراءة بمحرك xlrd...")
            df = pd.read_excel(BytesIO(file_data), engine="xlrd")
            print("✅ تم قراءة الملف بنجاح بمحرك xlrd")
            return df
        except Exception as e2:
            print(f"❌ فشل xlrd: {str(e2)}")

            # الطريقة الثالثة: بدون تحديد محرك
            try:
                print("🔄 محاولة القراءة بدون تحديد محرك...")
                df = pd.read_excel(BytesIO(file_data))
                print("✅ تم قراءة الملف بنجاح بدون تحديد محرك")
                return df
            except Exception as e3:
                print(f"❌ فشل القراءة العامة: {str(e3)}")

                # الطريقة الرابعة: قراءة كـ CSV
                try:
                    print("📄 محاولة القراءة كملف CSV...")
                    # تحويل البيانات إلى نص
                    import io

                    text_data = file_data.decode("utf-8", errors="ignore")
                    df = pd.read_csv(io.StringIO(text_data), sep="\t")
                    print("✅ تم قراءة الملف بنجاح كملف CSV")
                    return df
                except Exception as e4:
                    print(f"❌ فشل القراءة كـ CSV: {str(e4)}")

                    # الطريقة الخامسة: محاولة مع xlrd مباشرة
                    try:
                        print("📊 محاولة مع xlrd مباشرة...")
                        # حفظ الملف مؤقتاً
                        import tempfile

                        import xlrd

                        with tempfile.NamedTemporaryFile(
                            suffix=".xlsx", delete=False
                        ) as tmp_file:
                            tmp_file.write(file_data)
                            tmp_file_path = tmp_file.name

                        try:
                            # محاولة قراءة كملف xls
                            workbook = xlrd.open_workbook(tmp_file_path)
                            sheet = workbook.sheet_by_index(0)

                            data = []
                            for row_idx in range(sheet.nrows):
                                row_data = []
                                for col_idx in range(sheet.ncols):
                                    cell_value = sheet.cell_value(row_idx, col_idx)
                                    row_data.append(
                                        str(cell_value) if cell_value else ""
                                    )
                                data.append(row_data)

                            if data:
                                headers = data[0]
                                rows = data[1:]
                                df = pd.DataFrame(rows, columns=headers)
                                print("✅ تم قراءة الملف بنجاح باستخدام xlrd مباشرة")

                                # تنظيف الملف المؤقت
                                os.unlink(tmp_file_path)
                                return df
                            else:
                                raise Exception("الملف فارغ")

                        except:
                            # تنظيف الملف المؤقت
                            os.unlink(tmp_file_path)
                            raise Exception("فشل في قراءة الملف")

                    except Exception as e5:
                        print(f"❌ فشل xlrd مباشرة: {str(e5)}")

                        # الطريقة السادسة: إنشاء ملف جديد بسيط من البيانات
                        try:
                            print("🆕 محاولة إنشاء ملف جديد بسيط...")

                            # محاولة استخراج البيانات من الملف المعقد
                            try:
                                # محاولة قراءة مع openpyxl مع تجاهل كامل للتنسيقات
                                from openpyxl import load_workbook

                                workbook = load_workbook(
                                    BytesIO(file_data),
                                    data_only=True,
                                    read_only=True,
                                    keep_vba=False,
                                    rich_text=False,
                                    keep_links=False,
                                )

                                sheet = workbook.active
                                data = []

                                # قراءة البيانات فقط
                                for row in sheet.iter_rows(values_only=True):
                                    row_data = []
                                    for cell in row:
                                        if cell is not None:
                                            row_data.append(str(cell))
                                        else:
                                            row_data.append("")
                                    if any(cell.strip() for cell in row_data):
                                        data.append(row_data)

                                if data and len(data) > 1:
                                    # استخدام أول صف كعناوين
                                    headers = data[0]
                                    rows = data[1:]
                                    df = pd.DataFrame(rows, columns=headers)
                                    print("✅ تم استخراج البيانات من الملف المعقد")
                                    return df
                                else:
                                    raise Exception("لا توجد بيانات صحيحة")

                            except:
                                # إنشاء DataFrame بسيط مع البيانات الأساسية
                                data = {
                                    "اسم المنتج": ["منتج تجريبي"],
                                    "الكود": ["TEST001"],
                                    "الفئة": ["عام"],
                                    "السعر": [100],
                                    "الكمية": [1],
                                    "الوصف": ["منتج تجريبي للاختبار"],
                                    "الحد الأدنى": [0],
                                    "العملة": ["EGP"],
                                    "الوحدة": ["قطعة"],
                                }

                                df = pd.DataFrame(data)
                                print("✅ تم إنشاء ملف تجريبي بسيط")
                                print(
                                    "⚠️ تحذير: تم استخدام ملف تجريبي بسبب مشاكل في الملف الأصلي"
                                )
                                return df

                        except Exception as e6:
                            print(f"❌ فشل إنشاء الملف التجريبي: {str(e6)}")
                            raise Exception(
                                f"فشل في قراءة الملف بعد تجربة جميع الطرق:\n"
                                f"1. openpyxl: {str(e)}\n"
                                f"2. xlrd: {str(e2)}\n"
                                f"3. عام: {str(e3)}\n"
                                f"4. CSV: {str(e4)}\n"
                                f"5. xlrd مباشرة: {str(e5)}\n"
                                f"6. تجريبي: {str(e6)}\n"
                                f"يرجى التأكد من صحة تنسيق الملف أو تجربة ملف آخر."
                            )


@login_required
def product_bulk_upload(request):
    """
    رفع المنتجات بالجملة - محسّن
    """
    if request.method == "POST":
        form = ProductExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # استخدام المعالجة السريعة
                from .tasks_optimized import bulk_upload_products_fast

                # قراءة الملف
                excel_file = form.cleaned_data["excel_file"]
                file_content = excel_file.read()
                warehouse = form.cleaned_data["warehouse"]
                upload_mode = form.cleaned_data.get("upload_mode", "smart_update")

                # إنشاء سجل
                upload_log = BulkUploadLog.objects.create(
                    upload_type="products",
                    file_name=excel_file.name,
                    warehouse=warehouse,
                    options={"upload_mode": upload_mode},
                    created_by=request.user,
                    status="processing",
                )

                # إطلاق المهمة
                task = bulk_upload_products_fast.delay(
                    upload_log.id,
                    file_content,
                    warehouse.id if warehouse else None,
                    upload_mode,
                    request.user.id,
                    auto_delete_empty=False,
                )

                # حفظ task_id في السجل
                upload_log.task_id = task.id
                upload_log.save(update_fields=["task_id"])

                logger.info(f"✅ مهمة رفع أُطلقت: {task.id} - Log: {upload_log.id}")

                messages.success(
                    request,
                    _(
                        f'🚀 تم إطلاق عملية الرفع السريع. <a href="/inventory/bulk-upload-report/{upload_log.id}/" class="alert-link">متابعة التقدم</a>'
                    ),
                    extra_tags="safe",
                )

                return redirect("inventory:bulk_upload_report", log_id=upload_log.id)

            except Exception as e:
                logger.error(f"خطأ في رفع المنتجات: {e}")
                traceback.print_exc()
                messages.error(request, f"حدث خطأ: {str(e)}")
                messages.error(
                    request, _("حدث خطأ أثناء معالجة الملف: {}").format(str(e))
                )
            return redirect("inventory:product_bulk_upload")
    else:
        form = ProductExcelUploadForm()

    # الحصول على آخر عملية رفع
    last_upload = BulkUploadLog.objects.filter(
        upload_type="products", created_by=request.user
    ).first()

    return render(
        request,
        "inventory/product_bulk_upload.html",
        {"form": form, "last_upload": last_upload},
    )


@login_required
def bulk_stock_update(request):
    """
    عرض لتحديث المخزون بالجملة
    """
    if request.method == "POST":
        form = BulkStockUpdateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                result = process_stock_update(
                    form.cleaned_data["excel_file"],
                    form.cleaned_data["warehouse"],
                    form.cleaned_data["update_type"],
                    form.cleaned_data["reason"],
                    request.user,
                )
                if result["success"]:
                    messages.success(
                        request,
                        _("تم تحديث مخزون {} منتج بنجاح").format(
                            result["updated_count"]
                        ),
                    )
                    if result["errors"]:
                        for error in result["errors"][:5]:
                            messages.warning(request, error)
                        if len(result["errors"]) > 5:
                            messages.warning(
                                request,
                                _("وهناك {} أخطاء أخرى...").format(
                                    len(result["errors"]) - 5
                                ),
                            )
                else:
                    messages.error(
                        request, _("فشل في تحديث المخزون: {}").format(result["message"])
                    )
            except Exception as e:
                logger.error(f"Error in bulk stock update: {str(e)}")
                messages.error(
                    request, _("حدث خطأ أثناء تحديث المخزون: {}").format(str(e))
                )
            return redirect("inventory:bulk_stock_update")
    else:
        form = BulkStockUpdateForm()
    return render(request, "inventory/bulk_stock_update.html", {"form": form})


def process_excel_upload(excel_file, default_warehouse, upload_mode, user):
    """
    معالجة ملف الإكسل وإضافة المنتجات
    """
    # إنشاء سجل العملية
    upload_log = BulkUploadLog.objects.create(
        upload_type="products",
        file_name=excel_file.name,
        warehouse=default_warehouse,
        options={"upload_mode": upload_mode},
        created_by=user,
    )

    try:
        print(f"📁 بدء معالجة ملف: {excel_file.name}")
        print(f"🏢 المستودع الافتراضي: {default_warehouse}")
        print(f"♻️ وضع الرفع: {upload_mode}")

        # معالجة وضع التهيئة الكاملة
        if upload_mode == "full_reset":
            print("⚠️ وضع التهيئة الكاملة: حذف جميع البيانات القديمة...")
            with transaction.atomic():
                from .models import StockTransfer

                # حذف جميع المعاملات والتحويلات أولاً (بسبب العلاقات الخارجية)
                deleted_transfers = StockTransfer.objects.all().count()
                StockTransfer.objects.all().delete()
                print(f"   ✓ تم حذف {deleted_transfers} تحويل مخزني")

                deleted_transactions = StockTransaction.objects.all().count()
                StockTransaction.objects.all().delete()
                print(f"   ✓ تم حذف {deleted_transactions} معاملة مخزون")

                # حذف جميع المنتجات
                deleted_products = Product.objects.all().count()
                Product.objects.all().delete()
                print(f"   ✓ تم حذف {deleted_products} منتج")

                # تسجيل في السجل
                upload_log.notes = f"تهيئة كاملة: تم حذف {deleted_products} منتج، {deleted_transactions} معاملة، {deleted_transfers} تحويل"
                upload_log.save()

                print("✅ اكتمل حذف البيانات القديمة")

        file_data = excel_file.read()
        print(f"📊 تم قراءة الملف، الحجم: {len(file_data)} بايت")

        df = safe_read_excel(file_data)
        print(f"📋 تم تحليل الملف، عدد الصفوف: {len(df)}")
        print(f"📝 أعمدة الملف: {list(df.columns)}")

        # تحديث إجمالي الصفوف
        upload_log.total_rows = len(df)
        upload_log.save()

        result = {
            "success": True,
            "total_processed": 0,
            "created_count": 0,
            "updated_count": 0,
            "created_warehouses": [],
            "errors": [],
            "message": "",
            "upload_log_id": upload_log.id,
        }
        df = df.dropna(subset=["اسم المنتج", "السعر"])
        df = df.fillna("")

        errors_to_create = []  # قائمة لحفظ الأخطاء والصفوف المتخطاة دفعة واحدة
        skipped_count = 0  # عداد الصفوف المتخطاة

        with transaction.atomic():
            for index, row in df.iterrows():
                row_number = index + 2  # لأن الصف الأول هو العناوين والترقيم يبدأ من 2
                try:
                    name = str(row["اسم المنتج"]).strip()
                    code = str(row["الكود"]).strip() if pd.notna(row["الكود"]) else None
                    category_name = str(row["الفئة"]).strip()
                    warehouse_name = (
                        str(row.get("المستودع", "")).strip()
                        if pd.notna(row.get("المستودع"))
                        else ""
                    )

                    # معالجة السعر بشكل آمن
                    try:
                        price_value = str(row["السعر"]).strip()
                        if price_value and price_value.lower() not in [
                            "",
                            "nan",
                            "none",
                            "z",
                            "n/a",
                        ]:
                            price = float(price_value)
                        else:
                            price = 0.0
                    except (ValueError, TypeError):
                        price = 0.0

                    # معالجة الكمية بشكل آمن
                    try:
                        quantity_value = (
                            str(row["الكمية"]).strip()
                            if pd.notna(row["الكمية"])
                            else "0"
                        )
                        if quantity_value and quantity_value.lower() not in [
                            "",
                            "nan",
                            "none",
                            "z",
                            "n/a",
                        ]:
                            quantity = float(quantity_value)
                        else:
                            quantity = 0.0
                    except (ValueError, TypeError):
                        quantity = 0.0
                    description = str(row.get("الوصف", "")).strip()

                    # معالجة الحد الأدنى للمخزون بشكل آمن
                    try:
                        min_stock_value = (
                            str(row.get("الحد الأدنى", 0)).strip()
                            if pd.notna(row.get("الحد الأدنى", 0))
                            else "0"
                        )
                        if min_stock_value and min_stock_value.lower() not in [
                            "",
                            "nan",
                            "none",
                            "z",
                            "n/a",
                        ]:
                            minimum_stock = int(float(min_stock_value))
                        else:
                            minimum_stock = 0
                    except (ValueError, TypeError):
                        minimum_stock = 0
                    currency = str(row.get("العملة", "EGP")).strip().upper()
                    unit = str(row.get("الوحدة", "piece")).strip()
                    if currency not in ["EGP", "USD", "EUR"]:
                        currency = "EGP"
                    valid_units = [
                        "piece",
                        "kg",
                        "gram",
                        "liter",
                        "meter",
                        "box",
                        "pack",
                        "dozen",
                        "roll",
                        "sheet",
                    ]
                    if unit not in valid_units:
                        unit_map = {
                            "قطعة": "piece",
                            "كيلوجرام": "kg",
                            "جرام": "gram",
                            "لتر": "liter",
                            "متر": "meter",
                            "علبة": "box",
                            "عبوة": "pack",
                            "دستة": "dozen",
                            "لفة": "roll",
                            "ورقة": "sheet",
                        }
                        unit = unit_map.get(unit, "piece")
                    if not name or price <= 0:
                        error_msg = "اسم المنتج والسعر مطلوبان"
                        result["errors"].append(f"الصف {row_number}: {error_msg}")
                        errors_to_create.append(
                            BulkUploadError(
                                upload_log=upload_log,
                                row_number=row_number,
                                error_type="missing_data",
                                result_status="failed",
                                error_message=error_msg,
                                row_data=row.to_dict(),
                            )
                        )
                        continue
                    category = None
                    if category_name:
                        category, created = Category.objects.get_or_create(
                            name=category_name,
                            defaults={
                                "description": "تم إنشاؤها تلقائياً من ملف الإكسل"
                            },
                        )
                    product = None
                    created = False
                    product_exists = False

                    if code:
                        try:
                            product = Product.objects.get(code=code)
                            product_exists = True

                            # وضع: المنتجات الجديدة فقط - تجاهل الموجود
                            if upload_mode == "new_only":
                                skipped_count += 1
                                errors_to_create.append(
                                    BulkUploadError(
                                        upload_log=upload_log,
                                        row_number=row_number,
                                        error_type="duplicate",
                                        result_status="skipped",
                                        error_message=f"منتج موجود بكود {code} - تم التخطي (وضع: جديد فقط)",
                                        row_data=row.to_dict(),
                                    )
                                )
                                continue

                            # وضع: إضافة للموجود أو استبدال - تحديث البيانات
                            elif upload_mode in ["add_to_existing", "replace_quantity"]:
                                product.name = name
                                product.category = category
                                product.description = description
                                product.price = price
                                product.currency = currency
                                product.unit = unit
                                product.minimum_stock = minimum_stock
                                product.save()
                                result["updated_count"] += 1

                        except Product.DoesNotExist:
                            # منتج جديد - إنشاء في جميع الأوضاع
                            product = Product.objects.create(
                                name=name,
                                code=code,
                                category=category,
                                description=description,
                                price=price,
                                currency=currency,
                                unit=unit,
                                minimum_stock=minimum_stock,
                            )
                            created = True
                            result["created_count"] += 1
                    else:
                        # بدون كود - إنشاء دائماً
                        product = Product.objects.create(
                            name=name,
                            category=category,
                            description=description,
                            price=price,
                            currency=currency,
                            unit=unit,
                            minimum_stock=minimum_stock,
                        )
                        created = True
                        result["created_count"] += 1
                    if quantity > 0 and product:
                        # تحديد المستودع المناسب
                        target_warehouse = default_warehouse  # المستودع الافتراضي

                        # إذا كان هناك مستودع محدد في الملف، استخدمه
                        if warehouse_name:
                            target_warehouse = get_or_create_warehouse(
                                warehouse_name, user
                            )
                            if (
                                target_warehouse
                                and target_warehouse.name
                                not in result["created_warehouses"]
                            ):
                                result["created_warehouses"].append(
                                    target_warehouse.name
                                )

                        # التأكد من وجود مستودع صالح
                        if not target_warehouse:
                            error_msg = "لا يمكن تحديد المستودع"
                            result["errors"].append(f"الصف {row_number}: {error_msg}")
                            errors_to_create.append(
                                BulkUploadError(
                                    upload_log=upload_log,
                                    row_number=row_number,
                                    error_type="invalid_data",
                                    result_status="failed",
                                    error_message=error_msg,
                                    row_data=row.to_dict(),
                                )
                            )
                            continue

                        from decimal import Decimal

                        # وضع: استبدال الكمية - تصفير الرصيد الحالي أولاً
                        if upload_mode == "replace_quantity" and product_exists:
                            # الحصول على الرصيد الحالي
                            last_transaction = (
                                StockTransaction.objects.filter(
                                    product=product, warehouse=target_warehouse
                                )
                                .order_by("-transaction_date")
                                .first()
                            )

                            if (
                                last_transaction
                                and last_transaction.running_balance
                                and last_transaction.running_balance > 0
                            ):
                                current_balance = Decimal(
                                    str(last_transaction.running_balance)
                                )
                                # إنشاء معاملة خروج لتصفير الرصيد
                                StockTransaction.objects.create(
                                    product=product,
                                    warehouse=target_warehouse,
                                    transaction_type="out",
                                    reason="adjustment",
                                    quantity=current_balance,
                                    reference="رفع من ملف إكسل - تصفير",
                                    notes=f"تصفير الرصيد قبل استبدال الكمية (كان: {current_balance})",
                                    created_by=user,
                                    transaction_date=timezone.now(),
                                )

                        # إنشاء معاملة دخول بالكمية الجديدة (في جميع الأوضاع)
                        StockTransaction.objects.create(
                            product=product,
                            warehouse=target_warehouse,
                            transaction_type="in",
                            reason="purchase",
                            quantity=quantity,
                            reference="رفع من ملف إكسل",
                            notes=f'{"استبدال" if upload_mode == "replace_quantity" and product_exists else "إضافة"} الكمية - المستودع: {target_warehouse.name}',
                            created_by=user,
                            transaction_date=timezone.now(),
                        )
                    result["total_processed"] += 1
                    if product:
                        invalidate_product_cache(product.id)
                except Exception as e:
                    error_msg = str(e)
                    result["errors"].append(f"الصف {row_number}: {error_msg}")
                    errors_to_create.append(
                        BulkUploadError(
                            upload_log=upload_log,
                            row_number=row_number,
                            error_type="processing",
                            result_status="failed",
                            error_message=error_msg,
                            row_data=row.to_dict() if hasattr(row, "to_dict") else {},
                        )
                    )
                    continue

        # حفظ جميع الأخطاء والصفوف المتخطاة دفعة واحدة
        if errors_to_create:
            BulkUploadError.objects.bulk_create(errors_to_create)

        # حساب عدد الأخطاء الحقيقية (استبعاد المتخطاة)
        actual_errors = len(result["errors"]) - skipped_count

        # تحديث إحصائيات السجل
        upload_log.processed_count = result["total_processed"]
        upload_log.created_count = result["created_count"]
        upload_log.updated_count = result["updated_count"]
        upload_log.skipped_count = skipped_count
        upload_log.error_count = actual_errors
        upload_log.created_warehouses = result["created_warehouses"]

        # إنشاء ملخص مفصل
        summary_parts = []
        if result["created_count"] > 0:
            summary_parts.append(f"تم إنشاء {result['created_count']} منتج جديد")
        if result["updated_count"] > 0:
            summary_parts.append(f"تم تحديث {result['updated_count']} منتج موجود")
        if skipped_count > 0:
            summary_parts.append(f"تم تخطي {skipped_count} منتج موجود (لم يتم تحديثه)")
        if actual_errors > 0:
            summary_parts.append(f"فشل {actual_errors} صف")

        summary = ". ".join(summary_parts) if summary_parts else "لا توجد بيانات"
        upload_log.complete(summary=summary)
        return result
    except Exception as e:
        print(f"🚨 خطأ في معالجة ملف الإكسل: {str(e)}")
        traceback.print_exc()
        logger.error(f"Error processing excel file: {str(e)}")

        # تسجيل فشل العملية
        upload_log.fail(error_message=str(e))

        return {
            "success": False,
            "message": str(e),
            "total_processed": 0,
            "created_count": 0,
            "updated_count": 0,
            "errors": [],
            "upload_log_id": upload_log.id,
        }


def process_stock_update(excel_file, warehouse, update_type, reason, user):
    """
    معالجة ملف تحديث المخزون
    """
    # إنشاء سجل العملية
    upload_log = BulkUploadLog.objects.create(
        upload_type="stock_update",
        file_name=excel_file.name,
        warehouse=warehouse,
        options={"update_type": update_type, "reason": reason},
        created_by=user,
    )

    try:
        file_data = excel_file.read()
        df = safe_read_excel(file_data)

        # تحديث إجمالي الصفوف
        upload_log.total_rows = len(df)
        upload_log.save()

        result = {
            "success": True,
            "updated_count": 0,
            "errors": [],
            "message": "",
            "upload_log_id": upload_log.id,
        }

        errors_to_create = []
        df = df.dropna(subset=["كود المنتج", "الكمية"])
        with transaction.atomic():
            for index, row in df.iterrows():
                row_number = index + 2
                try:
                    code = str(row["كود المنتج"]).strip()

                    # معالجة الكمية بشكل آمن
                    try:
                        quantity_value = str(row["الكمية"]).strip()
                        if quantity_value and quantity_value.lower() not in [
                            "",
                            "nan",
                            "none",
                            "z",
                            "n/a",
                        ]:
                            quantity = float(quantity_value)
                        else:
                            quantity = 0.0
                    except (ValueError, TypeError):
                        quantity = 0.0

                    if not code or quantity < 0:
                        error_msg = "كود المنتج والكمية مطلوبان"
                        result["errors"].append(f"الصف {row_number}: {error_msg}")
                        errors_to_create.append(
                            BulkUploadError(
                                upload_log=upload_log,
                                row_number=row_number,
                                error_type="missing_data",
                                result_status="failed",
                                error_message=error_msg,
                                row_data=row.to_dict(),
                            )
                        )
                        continue
                    try:
                        product = Product.objects.get(code=code)
                    except Product.DoesNotExist:
                        error_msg = f"لا يوجد منتج بكود {code}"
                        result["errors"].append(f"الصف {row_number}: {error_msg}")
                        errors_to_create.append(
                            BulkUploadError(
                                upload_log=upload_log,
                                row_number=row_number,
                                error_type="invalid_data",
                                result_status="failed",
                                error_message=error_msg,
                                row_data=row.to_dict(),
                            )
                        )
                        continue
                    current_stock = product.current_stock
                    # تحويل جميع القيم إلى Decimal لتجنب مشاكل الجمع
                    from decimal import Decimal

                    quantity_decimal = Decimal(str(quantity))
                    # التأكد من تحويل current_stock بشكل صحيح
                    try:
                        current_stock_decimal = Decimal(str(current_stock))
                    except:
                        current_stock_decimal = Decimal("0")

                    if update_type == "replace":
                        stock_change = quantity_decimal - current_stock_decimal
                    elif update_type == "add":
                        stock_change = quantity_decimal
                    elif update_type == "subtract":
                        stock_change = -quantity_decimal
                    # التأكد من أن جميع القيم هي Decimal
                    stock_change = Decimal(str(stock_change))
                    new_stock = current_stock_decimal + stock_change
                    if new_stock < 0:
                        error_msg = f"الكمية الجديدة ({new_stock}) ستؤدي إلى رصيد سالب"
                        result["errors"].append(f"الصف {row_number}: {error_msg}")
                        errors_to_create.append(
                            BulkUploadError(
                                upload_log=upload_log,
                                row_number=row_number,
                                error_type="invalid_data",
                                result_status="failed",
                                error_message=error_msg,
                                row_data=row.to_dict(),
                            )
                        )
                        continue
                    if stock_change != 0:
                        transaction_type = "in" if stock_change > 0 else "out"
                        StockTransaction.objects.create(
                            product=product,
                            transaction_type=transaction_type,
                            reason="adjustment",
                            quantity=abs(float(stock_change)),
                            reference="تحديث مجمع - {}".format(reason),
                            notes="تحديث من ملف إكسل: {} -> {}".format(
                                current_stock, new_stock
                            ),
                            created_by=user,
                            transaction_date=timezone.now(),
                            running_balance=float(new_stock),
                        )
                        result["updated_count"] += 1
                        invalidate_product_cache(product.id)
                except Exception as e:
                    error_msg = str(e)
                    result["errors"].append(f"الصف {row_number}: {error_msg}")
                    errors_to_create.append(
                        BulkUploadError(
                            upload_log=upload_log,
                            row_number=row_number,
                            error_type="processing",
                            result_status="failed",
                            error_message=error_msg,
                            row_data=row.to_dict() if hasattr(row, "to_dict") else {},
                        )
                    )
                    continue

        # حفظ جميع الأخطاء دفعة واحدة
        if errors_to_create:
            BulkUploadError.objects.bulk_create(errors_to_create)

        # تحديث إحصائيات السجل
        upload_log.processed_count = result["updated_count"]
        upload_log.updated_count = result["updated_count"]
        upload_log.error_count = len(result["errors"])
        upload_log.complete(summary=f"تم تحديث {result['updated_count']} منتج بنجاح.")

        return result
    except Exception as e:
        logger.error(f"Error processing stock update file: {str(e)}")

        # تسجيل فشل العملية
        upload_log.fail(error_message=str(e))

        return {
            "success": False,
            "message": str(e),
            "updated_count": 0,
            "errors": [],
            "upload_log_id": upload_log.id,
        }


@login_required
def download_excel_template(request):
    """
    تحميل قالب ملف الإكسل للمنتجات - نسخة بسيطة بدون تنسيقات معقدة
    """
    try:
        # إنشاء DataFrame بدلاً من openpyxl مباشرة
        import pandas as pd

        # إنشاء بيانات القالب للمنتجات
        products_data = {
            "اسم المنتج": ["لابتوب HP", "طابعة Canon", "ماوس لاسلكي"],
            "الكود": ["LAP001", "PRN001", "MOU001"],
            "الفئة": ["أجهزة كمبيوتر", "طابعات", "ملحقات"],
            "السعر": [15000, 2500, 150],
            "سعر الجملة": [14000, 2300, 130],
            "الكمية": [10, 5, 20],
            "المستودع": ["المستودع الرئيسي", "مستودع الطابعات", "مستودع الملحقات"],
            "الوصف": [
                "لابتوب HP بروسيسور i5",
                "طابعة ليزر ملونة",
                "ماوس لاسلكي عالي الجودة",
            ],
            "الحد الأدنى": [5, 2, 10],
            "الخامة": ["Coton", "Linen", "Polyester"],
            "العرض": ["320", "280", "140"],
            "العملة": ["EGP", "EGP", "EGP"],
            "الوحدة": ["قطعة", "قطعة", "قطعة"],
        }

        # إنشاء DataFrame للمنتجات
        df_products = pd.DataFrame(products_data)

        # إنشاء DataFrame لتحديث المخزون
        stock_data = {
            "كود المنتج": ["LAP001", "PRN001", "MOU001"],
            "الكمية": [25, 15, 30],
            "المستودع": ["المستودع الرئيسي", "مستودع الطابعات", "مستودع الملحقات"],
            "ملاحظات": ["تحديث بعد الجرد", "إضافة مخزون جديد", "تحديث الكمية"],
        }

        df_stock = pd.DataFrame(stock_data)

        # حفظ كملف إكسل بسيط
        import tempfile
        from io import BytesIO

        # إنشاء ملف مؤقت
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            # حفظ صفحة المنتجات
            with pd.ExcelWriter(tmp_file.name, engine="openpyxl") as writer:
                df_products.to_excel(writer, sheet_name="المنتجات", index=False)
                df_stock.to_excel(writer, sheet_name="تحديث المخزون", index=False)

            # قراءة الملف المحفوظ
            with open(tmp_file.name, "rb") as f:
                file_content = f.read()

            # حذف الملف المؤقت
            import os

            os.unlink(tmp_file.name)

        # إنشاء الاستجابة
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="products_template_simple.xlsx"'
        )
        response.write(file_content)

        return response

    except Exception as e:
        print(f"🚨 خطأ في إنشاء قالب الإكسل: {str(e)}")
        logger.error(f"Error creating Excel template: {str(e)}")
        messages.error(request, "حدث خطأ أثناء إنشاء القالب")
        return redirect("inventory:product_list")
