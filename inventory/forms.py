from io import BytesIO

import openpyxl
import pandas as pd
from django import forms
from django.core.files.uploadedfile import UploadedFile
from django.forms import inlineformset_factory
from django.utils.translation import gettext_lazy as _

from .models import Category, Product, StockTransfer, StockTransferItem, Warehouse


class ProductExcelUploadForm(forms.Form):
    """
    نموذج لرفع المنتجات من ملف إكسل
    """

    excel_file = forms.FileField(
        label=_("ملف الإكسل"),
        help_text=_("يجب أن يكون الملف بصيغة .xlsx أو .xls"),
        widget=forms.ClearableFileInput(
            attrs={"class": "form-control", "accept": ".xlsx,.xls"}
        ),
    )

    warehouse = forms.ModelChoiceField(
        queryset=Warehouse.objects.none(),  # سيتم تعيينه في __init__
        label=_("المستودع الافتراضي"),
        required=False,
        empty_label=_("سيتم إنشاء المستودعات تلقائياً من الملف"),
        widget=forms.Select(attrs={"class": "form-select"}),
        help_text=_(
            'اختياري - إذا لم تحدد مستودع، سيتم إنشاء المستودعات تلقائياً من عمود "المستودع" في الملف'
        ),
    )

    UPLOAD_MODE_CHOICES = [
        (
            "smart_update",
            _("🔄 تحديث ذكي (يحدث البيانات وينقل للمستودع الصحيح) - موصى به"),
        ),
        (
            "merge_warehouses",
            _("🔀 دمج المخازن (يجمع الأصناف المكررة في مستودعاتها الصحيحة)"),
        ),
        ("add_only", _("➕ إضافة فقط (المنتجات الجديدة فقط، تجاهل الموجود)")),
        ("clean_start", _("⚠️ مسح وبدء من جديد (حذف الكل والبدء من الصفر) - خطر!")),
    ]

    upload_mode = forms.ChoiceField(
        choices=UPLOAD_MODE_CHOICES,
        label=_("وضع الرفع"),
        initial="smart_update",
        required=True,
        widget=forms.RadioSelect(attrs={"class": "form-check-input"}),
        help_text=_(
            "التحديث الذكي ينقل المنتجات للمستودعات الصحيحة تلقائياً ويمنع التكرار"
        ),
    )

    auto_delete_empty_warehouses = forms.BooleanField(
        label=_("🗑️ حذف المستودعات الفارغة تلقائياً"),
        required=False,
        initial=False,
        widget=forms.CheckboxInput(attrs={"class": "form-check-input"}),
        help_text=_(
            "تفعيل هذا الخيار سيحذف المستودعات التي لا تحتوي على أي مخزون بعد انتهاء عملية الرفع"
        ),
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # تحميل المستودعات النشطة فقط عند إنشاء النموذج
        self.fields["warehouse"].queryset = Warehouse.objects.filter(
            is_active=True
        ).order_by("name")

    def clean_excel_file(self):
        file = self.cleaned_data.get("excel_file")
        if not file:
            raise forms.ValidationError(_("يجب رفع ملف إكسل"))

        # التحقق من نوع الملف
        if not file.name.endswith((".xlsx", ".xls")):
            raise forms.ValidationError(
                _("نوع الملف غير مدعوم. يرجى رفع ملف .xlsx أو .xls")
            )

        # التحقق من حجم الملف (حد أقصى 10 ميجابايت)
        if file.size > 10 * 1024 * 1024:
            raise forms.ValidationError(
                _("حجم الملف كبير جداً. الحد الأقصى 10 ميجابايت")
            )

        try:
            # محاولة قراءة الملف للتأكد من صحته
            file_data = file.read()

            # محاولة قراءة الملف بطريقة آمنة
            try:
                df = pd.read_excel(
                    BytesIO(file_data), engine="openpyxl", keep_default_na=False
                )
            except Exception as e1:
                try:
                    df = pd.read_excel(BytesIO(file_data), engine="xlrd")
                except Exception as e2:
                    try:
                        df = pd.read_excel(BytesIO(file_data))
                    except Exception as e3:
                        # محاولة أخيرة مع openpyxl مع إعدادات خاصة
                        try:
                            from openpyxl import load_workbook

                            workbook = load_workbook(
                                BytesIO(file_data), data_only=True, read_only=True
                            )
                            sheet = workbook.active
                            data = []
                            for row in sheet.iter_rows(values_only=True):
                                if any(cell is not None for cell in row):
                                    data.append(row)

                            if data:
                                headers = data[0]
                                rows = data[1:]
                                df = pd.DataFrame(rows, columns=headers)
                            else:
                                raise Exception("الملف فارغ")
                        except Exception as e4:
                            raise Exception(
                                f"فشل في قراءة الملف: {str(e1)} | {str(e2)} | {str(e3)} | {str(e4)}"
                            )

            file.seek(0)  # إعادة تعيين مؤشر الملف

            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ["اسم المنتج", "السعر"]
            missing_columns = []

            for col in required_columns:
                if col not in df.columns:
                    missing_columns.append(col)

            if missing_columns:
                raise forms.ValidationError(
                    _("الأعمدة التالية مفقودة في الملف: {}").format(
                        ", ".join(missing_columns)
                    )
                )

            # التحقق من وجود بيانات
            if df.empty:
                raise forms.ValidationError(_("الملف فارغ أو لا يحتوي على بيانات"))

            # التحقق من وجود بيانات في الأعمدة المطلوبة
            df_filtered = df.dropna(subset=["اسم المنتج", "السعر"])
            if df_filtered.empty:
                raise forms.ValidationError(
                    _("لا توجد بيانات صحيحة في الأعمدة المطلوبة")
                )

        except Exception as e:
            if isinstance(e, forms.ValidationError):
                raise e
            raise forms.ValidationError(_("خطأ في قراءة الملف: {}").format(str(e)))

        return file


class BulkStockUpdateForm(forms.Form):
    """
    نموذج لتحديث المخزون بالجملة
    """

    excel_file = forms.FileField(
        label=_("ملف تحديث المخزون"),
        help_text=_("ملف إكسل يحتوي على أكواد المنتجات والكميات الجديدة"),
        widget=forms.ClearableFileInput(
            attrs={"class": "form-control", "accept": ".xlsx,.xls"}
        ),
    )

    warehouse = forms.ModelChoiceField(
        queryset=Warehouse.objects.none(),  # سيتم تعيينه في __init__
        label=_("المستودع"),
        required=True,
        empty_label=_("اختر المستودع"),
        widget=forms.Select(attrs={"class": "form-select"}),
    )

    update_type = forms.ChoiceField(
        label=_("نوع التحديث"),
        choices=[
            ("replace", _("استبدال الكمية الحالية")),
            ("add", _("إضافة للكمية الحالية")),
            ("subtract", _("خصم من الكمية الحالية")),
        ],
        initial="replace",
        widget=forms.Select(attrs={"class": "form-select"}),
    )

    reason = forms.CharField(
        label=_("سبب التحديث"),
        max_length=200,
        required=True,
        widget=forms.TextInput(
            attrs={"class": "form-control", "placeholder": _("مثال: جرد، تصحيح، إلخ")}
        ),
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # تحميل المستودعات النشطة فقط عند إنشاء النموذج
        self.fields["warehouse"].queryset = Warehouse.objects.filter(
            is_active=True
        ).order_by("name")

    def clean_excel_file(self):
        file = self.cleaned_data.get("excel_file")
        if not file:
            raise forms.ValidationError(_("يجب رفع ملف إكسل"))

        # التحقق من نوع الملف
        if not file.name.endswith((".xlsx", ".xls")):
            raise forms.ValidationError(
                _("نوع الملف غير مدعوم. يرجى رفع ملف .xlsx أو .xls")
            )

        # التحقق من حجم الملف (حد أقصى 10 ميجابايت)
        if file.size > 10 * 1024 * 1024:
            raise forms.ValidationError(
                _("حجم الملف كبير جداً. الحد الأقصى 10 ميجابايت")
            )

        try:
            # محاولة قراءة الملف للتأكد من صحته
            file_data = file.read()

            # محاولة قراءة الملف بطريقة آمنة
            try:
                df = pd.read_excel(
                    BytesIO(file_data), engine="openpyxl", keep_default_na=False
                )
            except Exception as e1:
                try:
                    df = pd.read_excel(BytesIO(file_data), engine="xlrd")
                except Exception as e2:
                    try:
                        df = pd.read_excel(BytesIO(file_data))
                    except Exception as e3:
                        # محاولة أخيرة مع openpyxl مع إعدادات خاصة
                        try:
                            from openpyxl import load_workbook

                            workbook = load_workbook(
                                BytesIO(file_data), data_only=True, read_only=True
                            )
                            sheet = workbook.active
                            data = []
                            for row in sheet.iter_rows(values_only=True):
                                if any(cell is not None for cell in row):
                                    data.append(row)

                            if data:
                                headers = data[0]
                                rows = data[1:]
                                df = pd.DataFrame(rows, columns=headers)
                            else:
                                raise Exception("الملف فارغ")
                        except Exception as e4:
                            raise Exception(
                                f"فشل في قراءة الملف: {str(e1)} | {str(e2)} | {str(e3)} | {str(e4)}"
                            )

            file.seek(0)  # إعادة تعيين مؤشر الملف

            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ["كود المنتج", "الكمية"]
            missing_columns = []

            for col in required_columns:
                if col not in df.columns:
                    missing_columns.append(col)

            if missing_columns:
                raise forms.ValidationError(
                    _("الأعمدة التالية مفقودة في الملف: {}").format(
                        ", ".join(missing_columns)
                    )
                )

            # التحقق من وجود بيانات
            if df.empty:
                raise forms.ValidationError(_("الملف فارغ أو لا يحتوي على بيانات"))

            # التحقق من وجود بيانات في الأعمدة المطلوبة
            df_filtered = df.dropna(subset=["كود المنتج", "الكمية"])
            if df_filtered.empty:
                raise forms.ValidationError(
                    _("لا توجد بيانات صحيحة في الأعمدة المطلوبة")
                )

        except Exception as e:
            if isinstance(e, forms.ValidationError):
                raise e
            raise forms.ValidationError(_("خطأ في قراءة الملف: {}").format(str(e)))

        return file


class ProductForm(forms.ModelForm):
    """
    نموذج لإضافة/تعديل منتج واحد
    ملاحظة: السعر يُدار الآن من نظام المتغيرات فقط لضمان تتبع جميع التغييرات
    """

    # إضافة حقول إضافية غير موجودة في النموذج
    warehouse = forms.ModelChoiceField(
        queryset=Warehouse.objects.none(),  # سيتم تعيينه في __init__
        label=_("المستودع"),
        required=True,
        empty_label=_("اختر المستودع"),
        widget=forms.Select(attrs={"class": "form-select"}),
        help_text=_("المستودع الذي سيتم إضافة المنتج إليه"),
    )

    initial_quantity = forms.DecimalField(
        label=_("الكمية الحالية"),
        required=False,
        initial=0,
        min_value=0,
        decimal_places=2,
        widget=forms.NumberInput(
            attrs={"class": "form-control", "step": "0.01", "placeholder": "0.00"}
        ),
        help_text=_("الكمية الحالية في المخزون (اختياري)"),
    )

    # حقل السعر للعرض فقط (للإضافة الجديدة)
    price = forms.DecimalField(
        label=_("السعر"),
        required=True,
        min_value=0,
        decimal_places=2,
        widget=forms.NumberInput(
            attrs={
                "class": "form-control",
                "step": "0.01",
                "placeholder": "0.00",
                "required": True,
            }
        ),
        help_text=_(
            "السعر للمنتج (مطلوب) - استخدم نموذج المنتجات الأساسية للتسعير الكامل"
        ),
        error_messages={
            "required": "يجب إدخال سعر المنتج",
            "invalid": "الرجاء إدخال سعر صحيح",
        },
    )

    # حقل سعر الجملة
    wholesale_price = forms.DecimalField(
        label=_("سعر الجملة"),
        required=True,
        min_value=0,
        decimal_places=2,
        widget=forms.NumberInput(
            attrs={
                "class": "form-control",
                "step": "0.01",
                "placeholder": "0.00",
                "required": True,
            }
        ),
        help_text=_("سعر الجملة للمنتج (مطلوب)"),
        error_messages={
            "required": "يجب إدخال سعر الجملة",
            "invalid": "الرجاء إدخال سعر صحيح",
        },
    )

    class Meta:
        model = Product
        fields = [
            "name",
            "code",
            "category",
            "description",
            "unit",
            "minimum_stock",
            "material",
            "width",
        ]
        widgets = {
            "name": forms.TextInput(attrs={"class": "form-control"}),
            "code": forms.TextInput(
                attrs={"class": "form-control", "placeholder": "أدخل كود المنتج"}
            ),
            "category": forms.Select(attrs={"class": "form-select"}),
            "description": forms.Textarea(attrs={"class": "form-control", "rows": 3}),
            "unit": forms.Select(attrs={"class": "form-select"}),
            "minimum_stock": forms.NumberInput(attrs={"class": "form-control"}),
            "material": forms.TextInput(
                attrs={
                    "class": "form-control",
                    "placeholder": "e.g. Linen, Cotton, Polyester",
                }
            ),
            "width": forms.TextInput(
                attrs={"class": "form-control", "placeholder": "e.g. 280 cm, 140 cm"}
            ),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # تحميل المستودعات النشطة
        self.fields["warehouse"].queryset = Warehouse.objects.filter(
            is_active=True
        ).order_by("name")

        self.fields["category"].queryset = Category.objects.all()
        self.fields["category"].empty_label = _("اختر الفئة")

        # إضافة خيارات الوحدات من النموذج
        self.fields["unit"].choices = Product.UNIT_CHOICES

        # إذا كان هذا تعديل منتج موجود
        if self.instance and self.instance.pk:
            # في حالة التعديل، إزالة الحقول الإضافية
            if "warehouse" in self.fields:
                del self.fields["warehouse"]
            if "initial_quantity" in self.fields:
                del self.fields["initial_quantity"]
            # إزالة حقول السعر - يُدار من نظام المتغيرات فقط (لإصدار الإضافة الجديد)
            # ولكن نبقيها للقراءة فقط إذا كانت موجودة (سيتم التعامل معها في القالب)
            if "price" in self.fields:
                self.fields["price"].required = False
            if "wholesale_price" in self.fields:
                self.fields["wholesale_price"].required = False

    def save(self, commit=True):
        instance = super().save(commit=False)
        # إذا كان منتج جديد، يجب إدخال الأسعار
        if not instance.pk:
            price = self.cleaned_data.get("price")
            wholesale_price = self.cleaned_data.get("wholesale_price")
            if price is None:
                raise forms.ValidationError("يجب إدخال سعر المنتج")
            if wholesale_price is None:
                raise forms.ValidationError("يجب إدخال سعر الجملة")
            instance.price = price
            instance.wholesale_price = wholesale_price
        if commit:
            instance.save()
        return instance


class StockTransferForm(forms.ModelForm):
    """نموذج التحويل المخزني المبسط"""

    class Meta:
        model = StockTransfer
        fields = ["from_warehouse", "to_warehouse", "reason", "notes"]
        widgets = {
            "from_warehouse": forms.Select(
                attrs={
                    "class": "form-select",
                    "required": True,
                    "id": "id_from_warehouse",
                }
            ),
            "to_warehouse": forms.Select(
                attrs={"class": "form-select", "required": True}
            ),
            "reason": forms.TextInput(
                attrs={
                    "class": "form-control",
                    "placeholder": "سبب التحويل (اختياري)...",
                }
            ),
            "notes": forms.Textarea(
                attrs={
                    "class": "form-control",
                    "rows": 2,
                    "placeholder": "ملاحظات (اختياري)...",
                }
            ),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # فقط المستودعات النشطة
        self.fields["from_warehouse"].queryset = Warehouse.objects.filter(
            is_active=True
        )
        self.fields["to_warehouse"].queryset = Warehouse.objects.filter(is_active=True)

        # تسميات الحقول
        self.fields["from_warehouse"].label = _("من مستودع")
        self.fields["to_warehouse"].label = _("إلى مستودع")
        self.fields["reason"].label = _("سبب التحويل")
        self.fields["notes"].label = _("ملاحظات")

        # جعل الحقول اختيارية
        self.fields["reason"].required = False
        self.fields["notes"].required = False

    def clean(self):
        cleaned_data = super().clean()
        from_warehouse = cleaned_data.get("from_warehouse")
        to_warehouse = cleaned_data.get("to_warehouse")

        # التحقق من أن المستودعين مختلفين
        if from_warehouse and to_warehouse and from_warehouse == to_warehouse:
            raise forms.ValidationError(_("لا يمكن التحويل من وإلى نفس المستودع"))

        return cleaned_data


class StockTransferItemForm(forms.ModelForm):
    """نموذج عنصر التحويل المخزني المبسط"""

    # حقل إضافي لنقل الصنف كاملاً
    transfer_all = forms.BooleanField(
        label=_("نقل الكل"),
        required=False,
        initial=False,
        widget=forms.CheckboxInput(
            attrs={"class": "form-check-input", "title": "نقل كامل المخزون"}
        ),
    )

    class Meta:
        model = StockTransferItem
        fields = ["product", "quantity", "notes"]
        widgets = {
            "product": forms.Select(
                attrs={"class": "form-select product-select", "required": True}
            ),
            "quantity": forms.NumberInput(
                attrs={
                    "class": "form-control quantity-input",
                    "min": "0.01",
                    "step": "0.01",
                    "required": True,
                    "placeholder": "الكمية",
                }
            ),
            "notes": forms.TextInput(
                attrs={"class": "form-control", "placeholder": "ملاحظات (اختياري)..."}
            ),
        }

    def __init__(self, *args, **kwargs):
        from_warehouse = kwargs.pop("from_warehouse", None)
        super().__init__(*args, **kwargs)

        # تسميات الحقول
        self.fields["product"].label = _("المنتج")
        self.fields["quantity"].label = _("الكمية")
        self.fields["notes"].label = _("ملاحظات")

        # جعل الملاحظات اختيارية
        self.fields["notes"].required = False

        # جميع المنتجات مرتبة حسب الاسم
        self.fields["product"].queryset = Product.objects.all().order_by("name")

    def clean_quantity(self):
        quantity = self.cleaned_data.get("quantity")
        if quantity and quantity <= 0:
            raise forms.ValidationError(_("الكمية يجب أن تكون أكبر من صفر"))
        return quantity


# Formset للتحويل المخزني
StockTransferItemFormSet = inlineformset_factory(
    StockTransfer,
    StockTransferItem,
    form=StockTransferItemForm,
    extra=3,
    can_delete=True,
    min_num=1,
    validate_min=True,
)


class StockTransferReceiveForm(forms.Form):
    """نموذج استلام التحويل المخزني"""

    def __init__(self, *args, **kwargs):
        transfer = kwargs.pop("transfer", None)
        super().__init__(*args, **kwargs)

        if transfer:
            for item in transfer.items.all():
                field_name = f"item_{item.id}_received"
                self.fields[field_name] = forms.DecimalField(
                    label=f"{item.product.name}",
                    initial=item.quantity,
                    max_digits=18,
                    decimal_places=2,
                    min_value=0,
                    widget=forms.NumberInput(
                        attrs={
                            "class": "form-control",
                            "step": "0.01",
                            "max": str(item.quantity),
                        }
                    ),
                )

                # حقل الملاحظات لكل عنصر
                notes_field_name = f"item_{item.id}_notes"
                self.fields[notes_field_name] = forms.CharField(
                    label=f"ملاحظات {item.product.name}",
                    required=False,
                    widget=forms.TextInput(
                        attrs={
                            "class": "form-control",
                            "placeholder": "ملاحظات الاستلام...",
                        }
                    ),
                )
