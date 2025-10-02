#!/usr/bin/env python
"""
اختبار الملف المحدد من المستخدم
"""
import os
import sys
import traceback
from io import BytesIO

import django
import openpyxl
import pandas as pd

# إعداد Django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "crm.settings")
django.setup()


def test_user_file(file_path):
    """
    اختبار شامل للملف المحدد من المستخدم
    """
    print("🔍 بدء اختبار الملف المحدد...")
    print("=" * 60)
    print(f"📁 مسار الملف: {file_path}")

    try:
        # معلومات الملف الأساسية
        if os.path.exists(file_path):
            file_size = os.path.getsize(file_path)
            print(f"📊 حجم الملف: {file_size:,} بايت")
        else:
            print(f"❌ الملف غير موجود: {file_path}")
            return False

        # اختبار قراءة الملف بـ openpyxl
        print("\n📈 اختبار قراءة الملف بـ openpyxl...")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            print("✅ نجح openpyxl")

            # تحليل الأوراق
            print(f"📋 عدد الأوراق: {len(wb.sheetnames)}")
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"  - {sheet_name}: {ws.max_row} صف، {ws.max_column} عمود")

                # تحليل البيانات في الورقة الأولى
                if sheet_name == wb.sheetnames[0]:
                    print(f"    📊 بيانات الصف الأول:")
                    for col in range(1, min(6, ws.max_column + 1)):
                        cell_value = ws.cell(row=1, column=col).value
                        print(f"      العمود {col}: {cell_value}")

                    # تحليل الصفوف الأخرى
                    print(f"    📊 عينة من البيانات:")
                    for row in range(2, min(6, ws.max_row + 1)):
                        row_data = []
                        for col in range(1, min(6, ws.max_column + 1)):
                            cell_value = ws.cell(row=row, column=col).value
                            row_data.append(
                                str(cell_value) if cell_value is not None else ""
                            )
                        print(f"      الصف {row}: {row_data}")

            wb.close()

        except Exception as e:
            print(f"❌ فشل openpyxl: {str(e)}")
            print(f"🔍 تفاصيل الخطأ: {traceback.format_exc()}")

        # اختبار قراءة الملف بـ pandas
        print("\n📊 اختبار قراءة الملف بـ pandas...")
        try:
            df = pd.read_excel(file_path, engine="openpyxl")
            print("✅ نجح pandas")
            print(f"📋 عدد الصفوف: {len(df)}")
            print(f"📝 الأعمدة: {list(df.columns)}")
            print(f"📊 عينة من البيانات:")
            print(df.head())

            # تحليل أنواع البيانات
            print(f"\n🔍 أنواع البيانات:")
            for col in df.columns:
                print(f"  - {col}: {df[col].dtype}")
                if df[col].dtype == "object":
                    sample_values = df[col].dropna().head(3).tolist()
                    print(f"    عينة: {sample_values}")
                elif df[col].dtype in ["int64", "float64"]:
                    print(f"    النطاق: {df[col].min()} - {df[col].max()}")

        except Exception as e:
            print(f"❌ فشل pandas: {str(e)}")
            print(f"🔍 تفاصيل الخطأ: {traceback.format_exc()}")

        # اختبار معالجة الملف في النظام
        print("\n🧪 اختبار معالجة الملف في النظام...")
        try:
            from inventory.views_bulk import safe_read_excel

            with open(file_path, "rb") as f:
                file_data = f.read()

            print(f"📊 حجم البيانات: {len(file_data):,} بايت")

            # محاولة القراءة
            df = safe_read_excel(file_data)
            print(f"✅ نجح قراءة الملف في النظام")
            print(f"📋 عدد الصفوف: {len(df)}")
            print(f"📝 الأعمدة: {list(df.columns)}")

            # اختبار معالجة البيانات
            print(f"\n🧪 اختبار معالجة البيانات...")
            success_count = 0
            error_count = 0

            for index, row in df.iterrows():
                try:
                    # محاكاة معالجة الصف كما في النظام
                    name = str(row.get("اسم المنتج", "")).strip()
                    code = (
                        str(row.get("الكود", "")).strip()
                        if pd.notna(row.get("الكود"))
                        else None
                    )
                    category_name = str(row.get("الفئة", "")).strip()
                    price = float(row.get("السعر", 0))

                    # معالجة الكميات الكبيرة بشكل آمن
                    quantity_raw = row.get("الكمية", 0)
                    if pd.notna(quantity_raw):
                        if isinstance(quantity_raw, str):
                            quantity = float(quantity_raw.replace(",", ""))
                        else:
                            quantity = float(quantity_raw)
                    else:
                        quantity = 0

                    description = str(row.get("الوصف", "")).strip()
                    minimum_stock = (
                        int(row.get("الحد الأدنى", 0))
                        if pd.notna(row.get("الحد الأدنى", 0))
                        else 0
                    )
                    currency = str(row.get("العملة", "EGP")).strip().upper()
                    unit = str(row.get("الوحدة", "piece")).strip()

                    # التحقق من صحة البيانات
                    if not name or price <= 0:
                        raise ValueError(
                            f"بيانات غير صحيحة: الاسم='{name}', السعر={price}"
                        )

                    if quantity < 0:
                        raise ValueError(f"الكمية سالبة: {quantity}")

                    success_count += 1
                    if index < 5:  # عرض أول 5 صفوف
                        print(
                            f"  ✅ الصف {index + 1}: {name} - السعر: {price} - الكمية: {quantity}"
                        )

                except Exception as e:
                    error_count += 1
                    print(f"  ❌ خطأ في الصف {index + 1}: {str(e)}")
                    if error_count <= 3:  # عرض أول 3 أخطاء فقط
                        print(f"    البيانات: {row.to_dict()}")

            print(f"\n📊 ملخص النتائج:")
            print(f"  ✅ نجح: {success_count} صف")
            print(f"  ❌ فشل: {error_count} صف")
            if success_count + error_count > 0:
                print(
                    f"  📈 نسبة النجاح: {(success_count/(success_count+error_count)*100):.1f}%"
                )

            return success_count > 0

        except Exception as e:
            print(f"❌ خطأ في معالجة الملف: {str(e)}")
            print(f"🔍 تفاصيل الخطأ: {traceback.format_exc()}")
            return False

    except Exception as e:
        print(f"❌ خطأ عام في الاختبار: {str(e)}")
        print(f"🔍 تفاصيل الخطأ: {traceback.format_exc()}")
        return False


def analyze_file_structure(file_path):
    """
    تحليل بنية الملف بالتفصيل
    """
    print("\n🔍 تحليل بنية الملف...")
    print("=" * 60)

    try:
        import zipfile

        # فحص الملف كـ ZIP
        with zipfile.ZipFile(file_path, "r") as zip_file:
            file_list = zip_file.namelist()
            print(f"📋 الملفات داخل ZIP: {len(file_list)}")
            for file_name in file_list:
                print(f"  - {file_name}")

                # فحص ملف styles.xml للتنسيقات
                if "styles.xml" in file_name:
                    try:
                        with zip_file.open(file_name) as style_file:
                            content = style_file.read()
                            if b"extLst" in content:
                                print("⚠️ تم العثور على 'extLst' في ملف التنسيقات")
                            if b"PatternFill" in content:
                                print("⚠️ تم العثور على 'PatternFill' في ملف التنسيقات")
                    except Exception as e:
                        print(f"❌ خطأ في قراءة {file_name}: {str(e)}")

    except Exception as e:
        print(f"❌ خطأ في تحليل البنية: {str(e)}")


if __name__ == "__main__":
    # مسار الملف المحدد
    file_path = "لينكس/products_template_simple.xlsx"

    print("🚀 بدء اختبار الملف المحدد...")
    print("=" * 60)

    # تحليل بنية الملف
    analyze_file_structure(file_path)

    # اختبار معالجة الملف
    success = test_user_file(file_path)

    print("\n" + "=" * 60)
    if success:
        print("🎉 الملف يعمل بشكل صحيح!")
    else:
        print("⚠️ الملف يحتاج إلى إصلاحات")

    print("�� انتهى الاختبار")
