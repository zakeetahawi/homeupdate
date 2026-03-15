"""
Factory Accounting Reports Views
عرض تقارير حسابات المصنع
"""

from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import CardMeasurementSplit, FactoryCard, ReadyCurtainEntry, Tailor


@login_required
@permission_required("factory_accounting.view_factory_reports", raise_exception=True)
def production_reports(request):
    """
    Production reports with filtering
    تقارير الإنتاج مع الفلترة
    """
    # Get filter parameters with default date: 25th of previous month
    today = timezone.now().date()

    # Calculate 25th of previous month
    # Calculate 25th of month logic
    if today.day >= 25:
        # If today is 25th or later, default to THIS month's 25th
        default_date_from = today.replace(day=25).strftime("%Y-%m-%d")
    else:
        # If before 25th, default to PREVIOUS month's 25th
        if today.month == 1:
            # If January, go to December of previous year
            default_date_from = today.replace(
                year=today.year - 1, month=12, day=25
            ).strftime("%Y-%m-%d")
        else:
            # Otherwise, go to previous month
            default_date_from = today.replace(month=today.month - 1, day=25).strftime(
                "%Y-%m-%d"
            )

    date_from = request.GET.get("date_from", default_date_from)
    date_to = request.GET.get("date_to")
    payment_status = request.GET.get("payment_status", "all")
    card_status = request.GET.get("card_status", "all")
    tailor_id = request.GET.get("tailor")
    cutter_id = request.GET.get("cutter")

    # Base queryset
    cards = FactoryCard.objects.select_related(
        "manufacturing_order",
        "manufacturing_order__production_line",  # Optimization for cutter name
        "manufacturing_order__order",
        "manufacturing_order__order__customer",
        "manufacturing_order__modification_request__tailor",  # Tailor for modification orders
    ).prefetch_related("splits__tailor")

    # Exclude cards without production date (old/incomplete records)
    cards = cards.exclude(production_date__isnull=True)

    # Exclude Accessories Line (ID 4)
    cards = cards.exclude(manufacturing_order__production_line_id=4)

    # Apply filters
    if date_from:
        cards = cards.filter(production_date__gte=date_from)
    if date_to:
        cards = cards.filter(production_date__lte=date_to)
    if card_status != "all":
        cards = cards.filter(status=card_status)

    # Filter by tailor
    if tailor_id:
        if tailor_id == "without":
            # cards without ANY splits AND without modification tailor
            cards = cards.filter(
                splits__isnull=True,
            ).exclude(
                manufacturing_order__order_type="modification",
                manufacturing_order__modification_request__tailor__isnull=False,
            )
        else:
            # Include cards with splits for this tailor OR modification orders with this tailor
            cards = cards.filter(
                Q(splits__tailor_id=tailor_id)
                | Q(
                    manufacturing_order__order_type="modification",
                    manufacturing_order__modification_request__tailor_id=tailor_id,
                )
            ).distinct()

    # Filter by cutter (production line)
    if cutter_id:
        cards = cards.filter(manufacturing_order__production_line_id=cutter_id)

    # Calculate totals DYNAMICALLY based on payment status
    # For unpaid items: use current settings
    # For paid items: preserve original prices
    # إجمالي الأمتار = الأمتار الفعلية (الخام) وليس تكلفة الخياطة
    total_meters = Decimal("0.00")
    total_cutter_cost = Decimal("0.00")
    for card in cards:
        total_meters += card.get_actual_meters()
        total_cutter_cost += card.get_current_cutter_cost()

    # Get splits for payment status filtering
    splits = CardMeasurementSplit.objects.filter(factory_card__in=cards)

    # IMPORTANT: If filtering by tailor, only show that tailor's splits and costs
    if tailor_id and tailor_id != "without":
        splits = splits.filter(tailor_id=tailor_id)

    if payment_status == "paid":
        splits = splits.filter(is_paid=True)
        # Also filter cards that are 'paid' for cutter calculation context
        # Note: If filtering solely by payment_status, cutter cost logic might need adjustment if separate
        # But per user request, we show totals based on current filter.
        cards = cards.filter(status="paid")
    elif payment_status == "unpaid":
        splits = splits.filter(is_paid=False)
        cards = cards.exclude(status="paid")

    # Calculate tailoring cost (based on TailoringTypePricing)
    total_amount = Decimal("0.00")
    if tailor_id and tailor_id.isdigit():
        # عند فلترة خياط: حصة الخياط من تكلفة التفصيل (share_amount = حصته من total_billable_meters = total_tailoring_cost)
        for split in splits:
            total_amount += split.share_amount or Decimal("0.00")
    else:
        # بدون فلتر خياط: نحسب من تكلفة البطاقة الكاملة
        for card in cards:
            total_amount += card.total_tailoring_cost or Decimal("0.00")

    paid_amount = Decimal("0.00")
    unpaid_amount = Decimal("0.00")

    # Get filter lists
    tailors = Tailor.objects.filter(is_active=True).order_by("name")

    # Import ProductionLine here to avoid circular import if placed at top
    from manufacturing.models import ProductionLine

    cutters = (
        ProductionLine.objects.filter(is_active=True).exclude(id=4).order_by("name")
    )

    # Safe conversion for integer ID
    filtered_tailor_id = None
    if tailor_id and tailor_id.isdigit():
        filtered_tailor_id = int(tailor_id)

    # Ready Curtains - ستائر جاهزة
    ready_curtains = ReadyCurtainEntry.objects.select_related(
        "tailor", "created_by"
    ).order_by("-production_date")

    # Apply same date filters
    if date_from:
        ready_curtains = ready_curtains.filter(production_date__gte=date_from)
    if date_to:
        ready_curtains = ready_curtains.filter(production_date__lte=date_to)

    # Filter by tailor
    if tailor_id and tailor_id.isdigit():
        ready_curtains = ready_curtains.filter(tailor_id=tailor_id)
    elif tailor_id == "without":
        ready_curtains = ready_curtains.none()

    # Filter by payment status
    if payment_status == "paid":
        ready_curtains = ready_curtains.filter(is_paid=True)
    elif payment_status == "unpaid":
        ready_curtains = ready_curtains.filter(is_paid=False)

    # Calculate ready curtain totals
    ready_curtain_total = ready_curtains.aggregate(
        total=Sum("total_cost")
    )["total"] or Decimal("0.00")

    # Add ready curtain cost to tailor total
    total_amount += ready_curtain_total

    context = {
        "cards": cards.order_by("-production_date"),  # Show all records
        "total_meters": total_meters,
        "total_amount": total_amount,
        "total_tailor_cost": total_amount,  # Alias for template compatibility
        "total_cutter_cost": total_cutter_cost,
        "paid_amount": paid_amount,
        "unpaid_amount": unpaid_amount,
        "tailors": tailors,
        "cutters": cutters,
        "ready_curtains": ready_curtains,
        "ready_curtain_total": ready_curtain_total,
        "filters": {
            "date_from": date_from,
            "date_to": date_to,
            "payment_status": payment_status,
            "card_status": card_status,
            "tailor": tailor_id,
            "cutter": cutter_id,
        },
        "filtered_tailor_id": filtered_tailor_id,  # For conditional display
    }

    return render(request, "factory_accounting/reports.html", context)


@login_required
@permission_required("factory_accounting.view_factory_reports", raise_exception=True)
def export_production_report(request):
    """
    Export production report to Excel - mirrors production_reports filters exactly
    تصدير تقرير الإنتاج إلى Excel - يطابق فلاتر صفحة التقارير بالكامل
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # ── 1. FILTERS (identical to production_reports) ──
    today = timezone.now().date()
    if today.day >= 25:
        default_date_from = today.replace(day=25).strftime("%Y-%m-%d")
    else:
        if today.month == 1:
            default_date_from = today.replace(
                year=today.year - 1, month=12, day=25
            ).strftime("%Y-%m-%d")
        else:
            default_date_from = today.replace(month=today.month - 1, day=25).strftime(
                "%Y-%m-%d"
            )

    date_from = request.GET.get("date_from", default_date_from)
    date_to = request.GET.get("date_to")
    payment_status = request.GET.get("payment_status", "all")
    card_status = request.GET.get("card_status", "all")
    tailor_id = request.GET.get("tailor")
    cutter_id = request.GET.get("cutter")

    # ── 2. CARDS QUERYSET (identical to production_reports) ──
    cards = FactoryCard.objects.select_related(
        "manufacturing_order",
        "manufacturing_order__production_line",
        "manufacturing_order__order",
        "manufacturing_order__order__customer",
        "manufacturing_order__modification_request__tailor",
    ).prefetch_related("splits__tailor")

    cards = cards.exclude(production_date__isnull=True)
    cards = cards.exclude(manufacturing_order__production_line_id=4)

    if date_from:
        cards = cards.filter(production_date__gte=date_from)
    if date_to:
        cards = cards.filter(production_date__lte=date_to)
    if card_status != "all":
        cards = cards.filter(status=card_status)
    if tailor_id:
        if tailor_id == "without":
            cards = cards.filter(
                splits__isnull=True,
            ).exclude(
                manufacturing_order__order_type="modification",
                manufacturing_order__modification_request__tailor__isnull=False,
            )
        else:
            cards = cards.filter(
                Q(splits__tailor_id=tailor_id)
                | Q(
                    manufacturing_order__order_type="modification",
                    manufacturing_order__modification_request__tailor_id=tailor_id,
                )
            ).distinct()
    if cutter_id:
        cards = cards.filter(manufacturing_order__production_line_id=cutter_id)

    # ── 3. SPLITS & PAYMENT STATUS (identical to production_reports) ──
    splits_qs = CardMeasurementSplit.objects.filter(factory_card__in=cards)
    if tailor_id and tailor_id != "without":
        splits_qs = splits_qs.filter(tailor_id=tailor_id)

    if payment_status == "paid":
        splits_qs = splits_qs.filter(is_paid=True)
        cards = cards.filter(status="paid")
    elif payment_status == "unpaid":
        splits_qs = splits_qs.filter(is_paid=False)
        cards = cards.exclude(status="paid")

    # Build a mapping: card_id -> list of splits (for per-row use)
    card_splits_map = {}
    for sp in splits_qs.select_related("tailor"):
        card_splits_map.setdefault(sp.factory_card_id, []).append(sp)

    # ── 4. TOTALS (identical to production_reports) ──
    # إجمالي الأمتار = الأمتار الفعلية (الخام)
    total_meters = Decimal("0.00")
    total_cutter_cost = Decimal("0.00")
    for card in cards:
        total_meters += card.get_actual_meters()
        total_cutter_cost += card.get_current_cutter_cost()

    # Tailoring cost - mirrors the view logic exactly
    total_tailor_cost = Decimal("0.00")
    if tailor_id and tailor_id.isdigit():
        for sp in splits_qs:
            total_tailor_cost += sp.share_amount or Decimal("0.00")
    else:
        for card in cards:
            total_tailor_cost += card.total_tailoring_cost or Decimal("0.00")

    # ── 5. READY CURTAINS (identical to production_reports) ──
    ready_curtains = ReadyCurtainEntry.objects.select_related(
        "tailor", "created_by"
    ).order_by("-production_date")

    if date_from:
        ready_curtains = ready_curtains.filter(production_date__gte=date_from)
    if date_to:
        ready_curtains = ready_curtains.filter(production_date__lte=date_to)
    if tailor_id and tailor_id.isdigit():
        ready_curtains = ready_curtains.filter(tailor_id=tailor_id)
    elif tailor_id == "without":
        ready_curtains = ready_curtains.none()
    if payment_status == "paid":
        ready_curtains = ready_curtains.filter(is_paid=True)
    elif payment_status == "unpaid":
        ready_curtains = ready_curtains.filter(is_paid=False)

    ready_curtain_total = ready_curtains.aggregate(
        total=Sum("total_cost")
    )["total"] or Decimal("0.00")

    # Grand total = tailoring + ready curtains (matches page)
    total_tailor_cost += ready_curtain_total

    # ── 6. BUILD FILTER DESCRIPTION (for title row) ──
    filter_parts = []
    if date_from:
        filter_parts.append(f"من {date_from}")
    if date_to:
        filter_parts.append(f"إلى {date_to}")
    if tailor_id:
        if tailor_id == "without":
            filter_parts.append("بدون خياط")
        elif tailor_id.isdigit():
            try:
                t = Tailor.objects.get(pk=tailor_id)
                filter_parts.append(f"خياط: {t.name}")
            except Tailor.DoesNotExist:
                pass
    if cutter_id:
        from manufacturing.models import ProductionLine
        try:
            pl = ProductionLine.objects.get(pk=cutter_id)
            filter_parts.append(f"قصاص: {pl.name}")
        except ProductionLine.DoesNotExist:
            pass
    if payment_status == "paid":
        filter_parts.append("مدفوع فقط")
    elif payment_status == "unpaid":
        filter_parts.append("غير مدفوع فقط")
    if card_status != "all":
        filter_parts.append(f"حالة البطاقة: {card_status}")

    filter_text = " | ".join(filter_parts) if filter_parts else "بدون فلترة"

    # ══════════════════════════════════════════════════════════
    # CREATE WORKBOOK
    # ══════════════════════════════════════════════════════════
    wb = Workbook()

    # --- STYLES ---
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_font = Font(size=11)
    row_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border_style = Side(style="thin", color="aaaaaa")
    thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    total_lbl_font = Font(bold=True, size=11, color="FFFFFF")
    total_lbl_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    total_val_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")

    card_head_font = Font(bold=True, size=12, color="FFFFFF")
    card_val_font = Font(bold=True, size=14, color="2c3e50")

    c1_head_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    c1_val_fill = PatternFill(start_color="ebf5fb", end_color="ebf5fb", fill_type="solid")
    c2_head_fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
    c2_val_fill = PatternFill(start_color="fdf2e9", end_color="fdf2e9", fill_type="solid")
    c3_head_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
    c3_val_fill = PatternFill(start_color="e9f7ef", end_color="e9f7ef", fill_type="solid")
    c4_head_fill = PatternFill(start_color="8e44ad", end_color="8e44ad", fill_type="solid")
    c4_val_fill = PatternFill(start_color="f4ecf7", end_color="f4ecf7", fill_type="solid")

    title_font = Font(bold=True, size=14, color="2c3e50")
    subtitle_font = Font(size=11, color="7f8c8d", italic=True)

    def apply_header_style(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    def apply_row_style(cell):
        cell.font = row_font
        cell.alignment = row_alignment
        cell.border = thin_border

    # ══════════════════════════════════════════════════════════
    # SHEET 1: تقرير الإنتاج (Production Report)
    # ══════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "تقرير الإنتاج"
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:K1")
    title_cell = ws.cell(row=1, column=1, value="تقرير الإنتاج")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Filter description row
    ws.merge_cells("A2:K2")
    filter_cell = ws.cell(row=2, column=1, value=f"الفلترة: {filter_text}")
    filter_cell.font = subtitle_font
    filter_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Table headers at row 4
    start_row = 4
    headers = [
        "رقم الأمر",
        "العميل",
        "تاريخ الإنتاج",
        "إجمالي الأمتار",
        "القصاص",
        "تكلفة القصاص",
        "الخياطين",
        "تكلفة الخياطين",
        "حالة الطلب",
        "حالة الدفع",
        "تاريخ الدفع",
    ]

    col_widths = {
        1: 20, 2: 28, 3: 16, 4: 16, 5: 22,
        6: 16, 7: 35, 8: 18, 9: 16, 10: 16, 11: 18,
    }

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(col, 15)

    ws.auto_filter.ref = f"A{start_row}:K{start_row}"

    # Table data
    sum_meters = Decimal("0.00")
    sum_cutter_cost = Decimal("0.00")
    sum_tailor_cost = Decimal("0.00")

    current_row = start_row + 1
    for card in cards.order_by("-production_date"):
        # Get splits for this card (from the pre-filtered map)
        row_splits = card_splits_map.get(card.id, [])

        # If no tailor filter, show all splits from prefetch
        if not tailor_id:
            row_splits = list(card.splits.all())

        # Build detailed tailor information
        tailor_lines = []
        for s in row_splits:
            tailor_lines.append(f"{s.tailor.name}: {float(s.share_amount):.2f}")

        # Add pricing breakdown
        breakdown = card.tailoring_cost_breakdown or {}
        if breakdown:
            tailor_lines.append("---")
            tailor_lines.append("تفاصيل التسعير:")
            for key, entry in breakdown.items():
                if entry.get('method') == 'per_piece':
                    tailor_lines.append(
                        f"  {entry['display']}: {int(entry.get('pieces', 0))} قطعة × {entry['rate']} = {entry['cost']}"
                    )
                else:
                    tailor_lines.append(
                        f"  {entry['display']}: {entry['meters']}م × {entry['rate']} = {entry['cost']}"
                    )

        tailors_details = "\n".join(tailor_lines) if tailor_lines else "-"

        # Row-level tailor cost: if tailor filtered, use sum of splits; else card total
        if tailor_id and tailor_id.isdigit():
            card_tailor_cost = sum(
                (s.share_amount or Decimal("0.00")) for s in row_splits
            )
        else:
            card_tailor_cost = card.total_tailoring_cost or Decimal("0.00")

        # Cutter
        cutter_name = "-"
        if (
            hasattr(card.manufacturing_order, "production_line")
            and card.manufacturing_order.production_line
        ):
            cutter_name = card.manufacturing_order.production_line.name

        # Payment status display
        if tailor_id and tailor_id.isdigit() and row_splits:
            # Show tailor-level payment status
            all_paid = all(s.is_paid for s in row_splits)
            pay_display = "مدفوع" if all_paid else "غير مدفوع"
            pay_date = "-"
            if all_paid and row_splits:
                last_paid = max(
                    (s.paid_date for s in row_splits if s.paid_date),
                    default=None,
                )
                pay_date = last_paid.strftime("%Y-%m-%d") if last_paid else "-"
        else:
            pay_display = "مدفوع" if card.status == "paid" else "غير مدفوع"
            pay_date = card.payment_date.strftime("%Y-%m-%d") if card.payment_date else "-"

        vals = [
            card.order_number,
            card.customer_name,
            card.production_date.strftime("%Y-%m-%d") if card.production_date else "-",
            float(card.get_actual_meters()),
            cutter_name,
            float(card.get_current_cutter_cost()),
            tailors_details,
            float(card_tailor_cost),
            card.manufacturing_order.get_status_display(),
            pay_display,
            pay_date,
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            apply_row_style(cell)

        sum_meters += card.get_actual_meters()
        sum_cutter_cost += card.get_current_cutter_cost()
        sum_tailor_cost += card_tailor_cost

        current_row += 1

    # Totals Row
    ws.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=3
    )
    total_label_cell = ws.cell(row=current_row, column=1, value="المجموع الكلي (بطاقات)")
    total_label_cell.font = total_lbl_font
    total_label_cell.fill = total_lbl_fill
    total_label_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_label_cell.border = thin_border

    for col in range(4, 12):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = total_val_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=11)

    ws.cell(row=current_row, column=4).value = float(sum_meters)
    ws.cell(row=current_row, column=6).value = float(sum_cutter_cost)
    ws.cell(row=current_row, column=8).value = float(sum_tailor_cost)

    current_row += 1

    # Ready curtains summary row (if any)
    if ready_curtain_total > 0:
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=3
        )
        rc_label = ws.cell(row=current_row, column=1, value="ستائر جاهزة")
        rc_label.font = total_lbl_font
        rc_label.fill = PatternFill(start_color="8e44ad", end_color="8e44ad", fill_type="solid")
        rc_label.alignment = Alignment(horizontal="center", vertical="center")
        rc_label.border = thin_border

        for col in range(4, 12):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = total_val_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=11)

        ws.cell(row=current_row, column=8).value = float(ready_curtain_total)
        current_row += 1

        # Grand total row
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=3
        )
        grand_label = ws.cell(row=current_row, column=1, value="الإجمالي الكلي (بطاقات + ستائر جاهزة)")
        grand_label.font = Font(bold=True, size=12, color="FFFFFF")
        grand_label.fill = PatternFill(start_color="1a252f", end_color="1a252f", fill_type="solid")
        grand_label.alignment = Alignment(horizontal="center", vertical="center")
        grand_label.border = thin_border

        for col in range(4, 12):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="d5dbdb", end_color="d5dbdb", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=12)

        ws.cell(row=current_row, column=4).value = float(sum_meters)
        ws.cell(row=current_row, column=6).value = float(sum_cutter_cost)
        ws.cell(row=current_row, column=8).value = float(sum_tailor_cost + ready_curtain_total)

    # SIDE CARDS
    ws.column_dimensions[get_column_letter(12)].width = 3  # Gap

    side_col = 13
    card_width_cols = 4

    for c in range(side_col, side_col + card_width_cols):
        ws.column_dimensions[get_column_letter(c)].width = 14

    def create_side_card(start_r, title, value, unit, h_fill, v_fill):
        ws.merge_cells(
            start_row=start_r, start_column=side_col,
            end_row=start_r, end_column=side_col + card_width_cols - 1,
        )
        head_cell = ws.cell(row=start_r, column=side_col, value=title)
        head_cell.font = card_head_font
        head_cell.fill = h_fill
        head_cell.alignment = Alignment(horizontal="center", vertical="center")

        val_r_start = start_r + 1
        val_r_end = start_r + 2
        ws.merge_cells(
            start_row=val_r_start, start_column=side_col,
            end_row=val_r_end, end_column=side_col + card_width_cols - 1,
        )

        final_val = f"{value:,.2f} {unit}" if unit else f"{value:,.2f}"
        val_cell = ws.cell(row=val_r_start, column=side_col, value=final_val)
        val_cell.font = card_val_font
        val_cell.fill = v_fill
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        return val_r_end + 2

    card_row = start_row
    card_row = create_side_card(card_row, "إجمالي الأمتار", float(total_meters), "متر", c1_head_fill, c1_val_fill)
    card_row = create_side_card(card_row, "تكلفة القصاص", float(total_cutter_cost), "", c2_head_fill, c2_val_fill)
    card_row = create_side_card(card_row, "تكلفة الخياطين", float(total_tailor_cost), "", c3_head_fill, c3_val_fill)
    if ready_curtain_total > 0:
        card_row = create_side_card(card_row, "ستائر جاهزة", float(ready_curtain_total), "", c4_head_fill, c4_val_fill)

    # ══════════════════════════════════════════════════════════
    # SHEET 2: ستائر جاهزة (Ready Curtains) - only if there are entries
    # ══════════════════════════════════════════════════════════
    if ready_curtains.exists():
        ws2 = wb.create_sheet(title="ستائر جاهزة")
        ws2.sheet_view.rightToLeft = True
        ws2.sheet_view.showGridLines = False

        # Title
        ws2.merge_cells("A1:H1")
        t_cell = ws2.cell(row=1, column=1, value="ستائر جاهزة")
        t_cell.font = title_font
        t_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws2.merge_cells("A2:H2")
        f_cell = ws2.cell(row=2, column=1, value=f"الفلترة: {filter_text}")
        f_cell.font = subtitle_font
        f_cell.alignment = Alignment(horizontal="center", vertical="center")

        rc_headers = [
            "الخياط",
            "الكمية",
            "سعر القطعة",
            "الإجمالي",
            "تاريخ الإنتاج",
            "الوصف",
            "حالة الدفع",
            "بواسطة",
        ]
        rc_col_widths = {1: 22, 2: 12, 3: 14, 4: 16, 5: 16, 6: 30, 7: 14, 8: 18}

        rc_start = 4
        for col, h in enumerate(rc_headers, 1):
            cell = ws2.cell(row=rc_start, column=col, value=h)
            apply_header_style(cell)
            ws2.column_dimensions[get_column_letter(col)].width = rc_col_widths.get(col, 14)

        ws2.auto_filter.ref = f"A{rc_start}:H{rc_start}"

        rc_row = rc_start + 1
        rc_sum = Decimal("0.00")
        for rc in ready_curtains:
            vals = [
                rc.tailor.name if rc.tailor else "-",
                rc.quantity,
                float(rc.price_per_piece),
                float(rc.total_cost),
                rc.production_date.strftime("%Y-%m-%d") if rc.production_date else "-",
                rc.description or "-",
                "مدفوع" if rc.is_paid else "غير مدفوع",
                rc.created_by.get_full_name() if rc.created_by else "-",
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(row=rc_row, column=ci, value=v)
                apply_row_style(cell)
            rc_sum += rc.total_cost
            rc_row += 1

        # Totals
        ws2.merge_cells(start_row=rc_row, start_column=1, end_row=rc_row, end_column=3)
        tl = ws2.cell(row=rc_row, column=1, value="المجموع")
        tl.font = total_lbl_font
        tl.fill = total_lbl_fill
        tl.alignment = Alignment(horizontal="center", vertical="center")
        tl.border = thin_border

        tc = ws2.cell(row=rc_row, column=4, value=float(rc_sum))
        tc.font = Font(bold=True, size=12)
        tc.fill = total_val_fill
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border = thin_border

        for col in [5, 6, 7, 8]:
            cell = ws2.cell(row=rc_row, column=col)
            cell.fill = total_val_fill
            cell.border = thin_border

    # ══════════════════════════════════════════════════════════
    # RESPONSE
    # ══════════════════════════════════════════════════════════
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename=production_report_{timezone.now().strftime("%Y%m%d")}.xlsx'
    )
    wb.save(response)

    return response
