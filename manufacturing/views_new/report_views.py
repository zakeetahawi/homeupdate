"""
Manufacturing Report Views - PDF and Excel exports
عروض تقارير التصنيع - تصدير PDF و Excel
"""

from typing import Any
from datetime import datetime
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse, HttpRequest
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.db.models import Count, Sum, Q

from manufacturing.models import ManufacturingOrder
from manufacturing.utils import get_material_summary_context


@login_required
@permission_required('manufacturing.view_manufacturingorder')
def generate_manufacturing_report(request: HttpRequest, pk: int) -> HttpResponse:
    """
    توليد تقرير PDF لأمر تصنيع
    
    Args:
        request: طلب HTTP
        pk: معرف أمر التصنيع
        
    Returns:
        HttpResponse: ملف PDF
    """
    try:
        from weasyprint import HTML
        import io
        
        order = get_object_or_404(
            ManufacturingOrder.objects.select_related('order', 'order__customer'),
            pk=pk
        )
        
        # الحصول على ملخص المواد
        material_summary = get_material_summary_context(order.order)
        
        # إعداد السياق
        context = {
            'order': order,
            'customer': order.order.customer,
            'items': order.items.select_related('product').all(),
            'material_summary': material_summary,
            'generated_at': datetime.now(),
        }
        
        # توليد HTML
        html_string = render_to_string('manufacturing/reports/order_report.html', context)
        
        # تحويل إلى PDF
        html = HTML(string=html_string)
        pdf_file = html.write_pdf()
        
        # إرجاع الاستجابة
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="manufacturing_order_{order.id}.pdf"'
        
        return response
    
    except Exception as e:
        return HttpResponse(f'خطأ في توليد التقرير: {str(e)}', status=500)


@login_required
@permission_required('manufacturing.view_manufacturingorder')
def export_to_excel(request: HttpRequest) -> HttpResponse:
    """
    تصدير أوامر التصنيع إلى Excel
    
    Returns:
        HttpResponse: ملف Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        # إنشاء workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "أوامر التصنيع"
        
        # العناوين
        headers = [
            'رقم الأمر',
            'رقم العقد',
            'العميل',
            'الحالة',
            'تاريخ الإنشاء',
            'تاريخ التسليم المتوقع',
            'الملاحظات'
        ]
        
        # تنسيق العناوين
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # البيانات
        orders = ManufacturingOrder.objects.select_related(
            'order',
            'order__customer'
        ).all()
        
        for row_num, order in enumerate(orders, 2):
            ws.cell(row=row_num, column=1, value=order.id)
            ws.cell(row=row_num, column=2, value=order.contract_number or '')
            ws.cell(row=row_num, column=3, value=order.order.customer.name)
            ws.cell(row=row_num, column=4, value=order.get_status_display())
            ws.cell(row=row_num, column=5, value=order.created_at.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=6, value=order.expected_delivery_date.strftime('%Y-%m-%d') if order.expected_delivery_date else '')
            ws.cell(row=row_num, column=7, value=order.notes or '')
        
        # ضبط عرض الأعمدة
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # حفظ في الذاكرة
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # إرجاع الاستجابة
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="manufacturing_orders_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
    
    except Exception as e:
        return HttpResponse(f'خطأ في التصدير: {str(e)}', status=500)


@login_required
@permission_required('manufacturing.view_manufacturingorder')
def generate_summary_report(request: HttpRequest) -> HttpResponse:
    """
    توليد تقرير ملخص شامل
    
    Returns:
        HttpResponse: ملف PDF
    """
    try:
        from weasyprint import HTML
        
        # إحصائيات شاملة
        stats = {
            'total_orders': ManufacturingOrder.objects.count(),
            'by_status': ManufacturingOrder.objects.values('status').annotate(
                count=Count('id')
            ),
            'this_month': ManufacturingOrder.objects.filter(
                created_at__month=datetime.now().month,
                created_at__year=datetime.now().year
            ).count(),
        }
        
        # السياق
        context = {
            'stats': stats,
            'generated_at': datetime.now(),
            'orders': ManufacturingOrder.objects.select_related(
                'order',
                'order__customer'
            ).order_by('-created_at')[:50]
        }
        
        # توليد HTML
        html_string = render_to_string('manufacturing/reports/summary_report.html', context)
        
        # تحويل إلى PDF
        html = HTML(string=html_string)
        pdf_file = html.write_pdf()
        
        # إرجاع الاستجابة
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="manufacturing_summary_{datetime.now().strftime("%Y%m%d")}.pdf"'
        
        return response
    
    except Exception as e:
        return HttpResponse(f'خطأ في توليد التقرير: {str(e)}', status=500)
