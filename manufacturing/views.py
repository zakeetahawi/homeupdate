import json
import logging
from datetime import timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import (
    login_required,
    permission_required,
    user_passes_test,
)
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models import Case, Count, F, IntegerField, Q, Sum, Value, When
from django.db.models.functions import ExtractMonth, ExtractYear, TruncDay, TruncMonth
from django.http import (
    HttpResponse,
    HttpResponseBadRequest,
    HttpResponseForbidden,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from weasyprint import HTML

from accounts.models import Department
from accounts.utils import apply_default_year_filter
from core.mixins import PaginationFixMixin
from notifications.models import Notification
from orders.models import Order

from .models import (
    FabricReceipt,
    FabricReceiptItem,
    ManufacturingOrder,
    ManufacturingOrderItem,
    ProductionLine,
)
from .utils import get_material_summary_context  # تم إضافة هذا السطر

# ...existing code...


logger = logging.getLogger(__name__)


class ManufacturingOrderListView(
    PaginationFixMixin, LoginRequiredMixin, PermissionRequiredMixin, ListView
):
    model = ManufacturingOrder
    template_name = "manufacturing/manufacturingorder_list.html"
    context_object_name = "manufacturing_orders"
    paginate_by = 25  # تفعيل Django pagination
    permission_required = "manufacturing.view_manufacturingorder"
    allow_empty = True  # السماح بالصفحات الفارغة دون رفع 404
    paginate_orphans = 0  # عدم دمج الصفحات الصغيرة

    def paginate_queryset(self, queryset, page_size):
        """Override pagination to handle invalid page numbers gracefully"""
        paginator = self.get_paginator(
            queryset,
            page_size,
            orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty(),
        )
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = int(page)
        except ValueError:
            page_number = 1

        # If page number is too high, redirect to last page
        if page_number > paginator.num_pages and paginator.num_pages > 0:
            page_number = paginator.num_pages
        elif page_number < 1:
            page_number = 1

        try:
            page = paginator.page(page_number)
            return (paginator, page, page.object_list, page.has_other_pages())
        except Exception:
            # If still fails, return first page
            page = paginator.page(1)
            return (paginator, page, page.object_list, page.has_other_pages())

    def get_paginate_by(self, queryset):
        try:
            page_size_str = self.request.GET.get("page_size", "25")
            if isinstance(page_size_str, list):
                page_size_str = page_size_str[0] if page_size_str else "25"
            page_size = int(page_size_str)
            if page_size > 100:
                page_size = 100
            elif page_size < 1:
                page_size = 25
            return page_size
        except Exception:
            return 25

    def get_queryset(self):
        from django.db.models import Case, F, Value, When
        from django.db.models.functions import Coalesce

        from core.monthly_filter_utils import apply_monthly_filter

        # Optimize the query by selecting related fields that exist
        queryset = ManufacturingOrder.objects.select_related(
            "order",  # This is the only direct foreign key in the model
            "order__customer",  # Access customer through order
            "production_line",  # Production line information
        )

        # استثناء طلبات المنتجات (products) من أوامر التصنيع - لا يجب أن تظهر هنا أبداً
        queryset = queryset.exclude(order__selected_types__contains=["products"])

        # استثناء طلبات التعديل افتراضياً (إلا إذا تم اختيارها يدوياً في فلتر نوع الطلب)
        order_type_filters = self.request.GET.getlist("order_type")
        order_type_filters = [f for f in order_type_filters if f and f.strip()]

        # إذا لم يكن هناك فلتر يدوي لنوع الطلب، استثني طلبات التعديل
        if not order_type_filters:
            queryset = queryset.exclude(order_type="modification")

        # تطبيق الفلترة الشهرية (بناءً على تاريخ الطلب)
        queryset, self.monthly_filter_context = apply_monthly_filter(
            queryset, self.request, "order_date"
        )

        # تطبيق إعدادات العرض للمستخدم الحالي
        queryset = self.apply_display_settings_filter(queryset)

        # Apply filters - دعم الفلاتر المتعددة مع معالجة محسنة
        # فلتر الحالات (اختيار متعدد)
        status_filters = self.request.GET.getlist("status")
        # تنظيف الفلاتر من القيم الفارغة
        status_filters = [f for f in status_filters if f and f.strip()]

        if status_filters:
            queryset = queryset.filter(status__in=status_filters)

        # فلتر الفروع (اختيار متعدد)
        branch_filters = self.request.GET.getlist("branch")
        branch_filters = [f for f in branch_filters if f and f.strip()]

        if branch_filters:
            branch_query = Q()
            for branch_id in branch_filters:
                try:
                    branch_id = int(branch_id)
                    branch_query |= Q(order__branch_id=branch_id)
                except (ValueError, TypeError):
                    continue
            if branch_query:
                queryset = queryset.filter(branch_query)

        # فلتر أنواع الطلبات (اختيار متعدد)
        # ملاحظة: تم نقل هذا الكود للأعلى قبل الفلترة الشهرية لاستخدامه في استثناء طلبات التعديل
        # order_type_filters تم تعريفه بالفعل في الأعلى

        if order_type_filters:
            queryset = queryset.filter(order_type__in=order_type_filters)

        # فلتر خطوط الإنتاج (اختيار متعدد)
        production_line_filters = self.request.GET.getlist("production_line")
        production_line_filters = [f for f in production_line_filters if f is not None]

        if production_line_filters:
            line_query = Q()
            for line_id in production_line_filters:
                if line_id == "":  # غير محدد
                    line_query |= Q(production_line__isnull=True)
                else:
                    try:
                        line_id = int(line_id)
                        line_query |= Q(production_line_id=line_id)
                    except (ValueError, TypeError):
                        continue
            if line_query:
                queryset = queryset.filter(line_query)

        # فلتر الطلبات المتأخرة
        overdue_filter = self.request.GET.get("overdue")
        if overdue_filter == "true":
            today = timezone.now().date()
            queryset = queryset.filter(
                expected_delivery_date__lt=today,
                status__in=["pending_approval", "pending", "in_progress"],
                order_type__in=["installation", "custom", "delivery"],
            )

        # فلتر حالة الأقمشة - تحسين الأداء باستخدام SQL Annotations
        fabric_status_filter = self.request.GET.get("fabric_status")
        if fabric_status_filter:
            # from django.db.models import Case, Count, F, IntegerField, Q, When (Removed to fix UnboundLocalError)

            from manufacturing.models import ManufacturingSettings

            # الحصول على المستودعات المحددة للعرض
            try:
                settings = ManufacturingSettings.get_settings()
                display_warehouses = list(settings.warehouses_for_display.all())
            except Exception:
                display_warehouses = []

            # إعداد فلاتر العد
            # إعداد فلاتر العد باستخدام items (OrderItem) من الطلب الأصلي
            # لأن ManufacturingOrderItem قد لا تكون موجودة للعناصر غير المقطوعة

            if display_warehouses:
                # في حالة تحديد مستودعات:

                # 1. Total: جميع عناصر الطلب المرتبطة بطلب تقطيع في هذه المستودعات
                total_filter = Q(
                    order__items__cutting_items__cutting_order__warehouse__in=display_warehouses
                )

                # تعريف حالة القطع الفعلي (مكتمل أو لديه بيانات استلام)
                is_actually_cut = Q(order__items__cutting_items__status="completed") | (
                    Q(order__items__cutting_items__receiver_name__isnull=False)
                    & ~Q(order__items__cutting_items__receiver_name="")
                    & Q(order__items__cutting_items__permit_number__isnull=False)
                    & ~Q(order__items__cutting_items__permit_number="")
                )

                # 2. Cut: عناصر من الـ Total ولكن تم قطعها فعلياً في المستودع المحدد
                cut_filter = total_filter & is_actually_cut

                # 3. Received: عناصر من الـ Total ولكن تم استلامها (يوجد ManufacturingOrderItem مرتبط ومستلم)
                # يجب تصفية ManufacturingOrderItem للتأكد من أنه مرتبط بنفس الطلب والمستودع؟
                # البادج في الموديل يفحص: items.filter(order_item=order_item, fabric_received=True)
                # وبما أننا داخل نفس الطلب، يكفي التحقق من manufacturing_items المرتبطة بـ OrderItem
                received_filter = total_filter & Q(
                    order__items__manufacturing_items__fabric_received=True
                )

            else:
                # بدون تحديد مستودعات:

                # 1. Total: جميع عناصر الطلب
                total_filter = Q()

                # تعريف حالة القطع الفعلي
                is_actually_cut = Q(order__items__cutting_items__status="completed") | (
                    Q(order__items__cutting_items__receiver_name__isnull=False)
                    & ~Q(order__items__cutting_items__receiver_name="")
                    & Q(order__items__cutting_items__permit_number__isnull=False)
                    & ~Q(order__items__cutting_items__permit_number="")
                )

                # 2. Cut: عناصر لديها cutting_item وحالته مكتملة
                cut_filter = (
                    Q(order__items__cutting_items__isnull=False) & is_actually_cut
                )

                # 3. Received: عناصر لديها manufacturing_item مستلم
                received_filter = Q(
                    order__items__manufacturing_items__fabric_received=True
                )

            # تطبيق Annotations للحساب
            queryset = queryset.annotate(
                # عدد العناصر الكلي (من OrderItem)
                calc_total=Count("order__items", filter=total_filter, distinct=True),
                # عدد العناصر المستلمة
                calc_received=Count(
                    "order__items", filter=received_filter, distinct=True
                ),
                # عدد العناصر المقطوعة
                calc_cut=Count("order__items", filter=cut_filter, distinct=True),
            )

            # تطبيق الفلاتر
            if fabric_status_filter == "needs_receipt":
                # بحاجة استلام: يوجد عناصر، وجميعها غير مستلمة
                # (calc_total > 0 AND calc_received = 0)
                # وأيضاً يجب أن يكون "مقطوعاً بالكامل" حسب طلب المستخدم
                # (calc_cut > 0 AND calc_cut = calc_total)
                queryset = queryset.filter(
                    calc_total__gt=0,
                    calc_received=0,
                    calc_cut__gt=0,
                    calc_cut=F(
                        "calc_total"
                    ),  # شرط عدم وجود عناصر غير مقطوعة (في حالة عدم تحديد مستودعات)
                )

            elif fabric_status_filter == "partial":
                # ناقص: يوجد عناصر، وبعضها مستلم (ليس الكل صفر وليس الكل كامل)
                # أو يوجد عناصر غير مقطوعة (calc_cut < calc_total)
                queryset = queryset.filter(
                    Q(calc_total__gt=0)
                    & (
                        Q(
                            calc_received__gt=0, calc_received__lt=F("calc_total")
                        )  # استلام جزئي
                        | Q(calc_cut__lt=F("calc_total"))  # وجود عناصر غير مقطوعة
                    )
                )

            elif fabric_status_filter == "complete":
                # كامل: يوجد عناصر، وجميعها مستلمة، وجميعها مقطوعة
                queryset = queryset.filter(
                    calc_total__gt=0,
                    calc_received=F("calc_total"),
                    calc_cut=F("calc_total"),
                )

            elif fabric_status_filter == "not_cut":
                # غير مقطوع: لا يوجد عناصر مقطوعة
                queryset = queryset.filter(calc_cut=0)
        search = self.request.GET.get("search")
        search_columns = self.request.GET.getlist("search_columns")

        # محاولة فك ترميز قوالب البحث إذا كانت مُرمّزة
        try:
            if search_columns and len(search_columns) == 1:
                search_columns = json.loads(
                    search_columns[0].replace("%27", "'").replace("%22", '"')
                )
                if not isinstance(search_columns, list):
                    search_columns = [search_columns]
        except:
            pass

        if search:
            # إذا لم يحدد المستخدم أعمدة أو اختار 'all'، نفذ البحث الشامل كما كان (يشمل جميع أرقام الفواتير والعقود)
            if not search_columns or "all" in search_columns:
                queryset = queryset.filter(
                    Q(order__id__icontains=search)
                    | Q(order__order_number__icontains=search)
                    | Q(contract_number__icontains=search)
                    | Q(order__contract_number_2__icontains=search)
                    | Q(order__contract_number_3__icontains=search)
                    | Q(invoice_number__icontains=search)
                    | Q(order__invoice_number_2__icontains=search)
                    | Q(order__invoice_number_3__icontains=search)
                    | Q(exit_permit_number__icontains=search)
                    | Q(order__customer__name__icontains=search)
                    | Q(order__salesperson__name__icontains=search)
                    | Q(order__branch__name__icontains=search)
                    | Q(notes__icontains=search)
                    | Q(order_type__icontains=search)
                    | Q(status__icontains=search)
                    | Q(order_date__icontains=search)
                    | Q(expected_delivery_date__icontains=search)
                )
            else:
                # خريطة الأعمدة المدعومة إلى شروط Q (مع دعم جميع أرقام الفواتير والعقود)
                column_map = {
                    "order_number": Q(order__order_number__icontains=search),
                    "customer_name": Q(order__customer__name__icontains=search),
                    "contract_number": Q(contract_number__icontains=search)
                    | Q(order__contract_number_2__icontains=search)
                    | Q(order__contract_number_3__icontains=search),
                    "invoice_number": Q(invoice_number__icontains=search)
                    | Q(order__invoice_number_2__icontains=search)
                    | Q(order__invoice_number_3__icontains=search),
                    "branch": Q(order__branch__name__icontains=search),
                }

                custom_q = Q()
                for col in search_columns:
                    if col in column_map:
                        custom_q |= column_map[col]

                if custom_q:
                    queryset = queryset.filter(custom_q)

        # فلتر التواريخ
        date_from = self.request.GET.get("date_from")
        if date_from:
            try:
                queryset = queryset.filter(order_date__gte=date_from)
            except:
                pass

        date_to = self.request.GET.get("date_to")
        if date_to:
            try:
                import datetime

                end_date = datetime.datetime.strptime(
                    date_to, "%Y-%m-%d"
                ) + datetime.timedelta(days=1)
                queryset = queryset.filter(order_date__lt=end_date)
            except:
                pass

        # تطبيق الترتيب
        sort_column = self.request.GET.get("sort")
        sort_direction = self.request.GET.get("direction", "asc")

        if sort_column and sort_direction != "none":
            sort_field = self.get_sort_field(sort_column)
            if sort_field:
                if sort_direction == "desc":
                    sort_field = f"-{sort_field}"
                queryset = queryset.order_by(sort_field)
            else:
                # الترتيب الافتراضي عند عدم تحديد عمود
                from django.db.models import Case, IntegerField, Value, When

                queryset = queryset.annotate(
                    priority=Case(
                        When(
                            status__in=["pending_approval", "pending", "in_progress"],
                            then=Value(1),
                        ),
                        default=Value(2),
                        output_field=IntegerField(),
                    )
                ).order_by("priority", "expected_delivery_date", "order_date")
        else:
            # الترتيب الافتراضي: الطلبات النشطة أولاً ثم حسب تاريخ التسليم
            from django.db.models import Case, IntegerField, Value, When

            queryset = queryset.annotate(
                priority=Case(
                    When(
                        status__in=["pending_approval", "pending", "in_progress"],
                        then=Value(1),
                    ),
                    default=Value(2),
                    output_field=IntegerField(),
                )
            ).order_by("priority", "expected_delivery_date", "order_date")

        return queryset

    def get_sort_field(self, sort_column):
        """تحديد حقل الترتيب المناسب"""
        sort_fields = {
            "id": "id",
            "order_number": "order__order_number",
            "order_type": "order_type",
            "contract_number": "contract_number",
            "production_line": "production_line__name",
            "invoice_number": "invoice_number",
            "customer": "order__customer__name",
            "salesperson": "order__salesperson__name",
            "branch": "order__branch__name",
            "order_date": "order_date",
            "expected_delivery_date": "expected_delivery_date",
            "status": "status",
        }
        return sort_fields.get(sort_column)

    def apply_display_settings_filter(self, queryset):
        """تطبيق فلترة بناءً على إعدادات العرض للمستخدم الحالي"""
        try:
            from .models import ManufacturingDisplaySettings

            # التحقق من وجود فلاتر يدوية في الطلب
            manual_filters = self.has_manual_filters()

            # إذا كان هناك فلاتر يدوية، لا نطبق إعدادات العرض التلقائية
            if manual_filters:
                return queryset

            # الحصول على إعدادات العرض للمستخدم الحالي
            display_settings = ManufacturingDisplaySettings.get_user_settings(
                self.request.user
            )

            if display_settings:
                # تطبيق فلترة الحالات
                if display_settings.allowed_statuses:
                    queryset = queryset.filter(
                        status__in=display_settings.allowed_statuses
                    )

                # تطبيق فلترة أنواع الطلبات
                if display_settings.allowed_order_types:
                    queryset = queryset.filter(
                        order_type__in=display_settings.allowed_order_types
                    )

            return queryset
        except Exception as e:
            # في حالة حدوث خطأ، إرجاع الـ queryset الأصلي
            logger.warning(f"خطأ في تطبيق إعدادات العرض: {e}")
            return queryset

    def has_manual_filters(self):
        """التحقق من وجود فلاتر يدوية في الطلب"""
        # قائمة بمعاملات الفلترة اليدوية
        filter_params = [
            "status",  # فلتر الحالة
            "order_type",  # فلتر نوع الطلب
            "search",  # البحث
            "date_from",  # تاريخ من
            "date_to",  # تاريخ إلى
            "customer",  # فلتر العميل
            "salesperson",  # فلتر البائع
            "branch",  # فلتر الفرع
            "production_line",  # فلتر خط الإنتاج
            "fabric_status",  # فلتر حالة القماش
            "contract_number",  # رقم العقد
            "invoice_number",  # رقم الفاتورة
        ]

        # التحقق من وجود أي من هذه المعاملات في الطلب
        for param in filter_params:
            if self.request.GET.get(param):
                return True

        return False

    def get_context_data(self, **kwargs):
        from core.monthly_filter_utils import get_available_years

        context = super().get_context_data(**kwargs)
        from django.contrib.auth import get_user_model
        from django.db.models import Count, Q

        from .models import ManufacturingDisplaySettings

        # Get manufacturing orders for statistics (مفلترة بالسنة الافتراضية)
        # استخدام نفس الفلتر المطبق على القائمة
        filtered_orders = ManufacturingOrder.objects.select_related(
            "order", "order__customer"
        )

        # تطبيق فلتر السنة الافتراضية على الإحصائيات
        from accounts.utils import apply_default_year_filter

        filtered_orders = apply_default_year_filter(
            filtered_orders, self.request, "order_date"
        )

        # Add necessary data for the form
        context.update(
            {
                "status_choices": dict(ManufacturingOrder.STATUS_CHOICES),
                "order_types": dict(ManufacturingOrder.ORDER_TYPE_CHOICES),
            }
        )

        # Add filter values to context - دعم الفلاتر المتعددة
        context.update(
            {
                "status_filters": self.request.GET.getlist("status"),
                "branch_filters": self.request.GET.getlist("branch"),
                "order_type_filters": self.request.GET.getlist("order_type"),
                "production_line_filters": self.request.GET.getlist("production_line"),
                "overdue_filter": self.request.GET.get("overdue", ""),
                "fabric_status": self.request.GET.get("fabric_status", ""),
                "search_query": self.request.GET.get("search", ""),
                "search_columns": self.request.GET.getlist("search_columns"),
                "date_from": self.request.GET.get("date_from", ""),
                "date_to": self.request.GET.get("date_to", ""),
                "page_size": self.request.GET.get("page_size", 25),
            }
        )

        # Add filter options
        from accounts.models import Branch

        from .models import ProductionLine

        context.update(
            {
                "branches": Branch.objects.filter(is_active=True).order_by("name"),
                "production_lines": ProductionLine.objects.filter(
                    is_active=True
                ).order_by("name"),
                "order_types": ManufacturingOrder.ORDER_TYPE_CHOICES,  # فقط أنواع التصنيع (بدون معاينات)
            }
        )

        # Add available display settings for filter dropdown
        try:
            context["available_display_settings"] = (
                ManufacturingDisplaySettings.objects.filter(is_active=True).order_by(
                    "-priority", "name"
                )
            )
        except Exception:
            context["available_display_settings"] = []

        # Add production lines for filter dropdown
        from .models import ProductionLine

        context["production_lines"] = ProductionLine.objects.filter(
            is_active=True
        ).order_by("-priority", "name")

        # Add order types for filter dropdown (إخفاء نوع "معاينات" من فلاتر المصنع)
        context["order_types"] = [
            ("installation", "تركيب"),
            ("custom", "تفصيل"),
            ("accessory", "إكسسوار"),
            ("delivery", "تسليم"),
        ]

        # Add current date for date picker max date
        from datetime import date

        context["today"] = date.today().isoformat()

        # Add function to get available statuses for each order
        context["get_available_statuses"] = (
            lambda current_status, order_type=None: self._get_available_statuses(
                self.request.user, current_status, order_type
            )
        )

        # Add page_size to context
        context["page_size"] = self.get_paginate_by(self.get_queryset())

        # معلومات حالة الفلترة
        has_manual_filters = self.has_manual_filters()
        active_display_settings = None

        if not has_manual_filters:
            # الحصول على إعدادات العرض النشطة
            active_display_settings = ManufacturingDisplaySettings.get_user_settings(
                self.request.user
            )

        context.update(
            {
                "has_manual_filters": has_manual_filters,
                "active_display_settings": active_display_settings,
                "display_settings_applied": not has_manual_filters
                and active_display_settings is not None,
            }
        )

        # Build human-readable display values for selected filters (Arabic labels)
        try:
            # status choices map code -> display
            status_choices_map = dict(ManufacturingOrder.STATUS_CHOICES)
        except Exception:
            status_choices_map = {}

        # branch id -> name map (branches already in context)
        branch_map = {str(b.id): b.name for b in context.get("branches", [])}

        # order types map from context 'order_types' (list of tuples)
        order_types_map = {
            str(code): name for code, name in context.get("order_types", [])
        }

        # column labels for search_columns
        column_labels = {
            "all": "الكل",
            "order_number": "رقم الطلب",
            "customer_name": "اسم العميل",
            "contract_number": "رقم العقد",
            "invoice_number": "رقم الفاتورة",
            "branch": "الفرع",
        }

        # compute display lists
        context["status_filters_display"] = [
            status_choices_map.get(s, s) for s in context.get("status_filters", [])
        ]
        context["branch_filters_display"] = [
            branch_map.get(bid, bid) for bid in context.get("branch_filters", [])
        ]
        context["order_type_filters_display"] = [
            order_types_map.get(code, code)
            for code in context.get("order_type_filters", [])
        ]
        context["search_columns_display"] = [
            column_labels.get(col, col) for col in context.get("search_columns", [])
        ]

        # إضافة القواميس للـ template
        context["status_choices_map"] = status_choices_map
        context["branch_map"] = branch_map
        context["order_types_map"] = order_types_map
        context["production_line_map"] = {
            str(line.id): line.name for line in context.get("production_lines", [])
        }

        # إضافة السنوات المتاحة والسياق الشهري (بناءً على تاريخ الطلب)
        context["available_years"] = get_available_years(
            ManufacturingOrder, "order_date"
        )
        if hasattr(self, "monthly_filter_context"):
            context.update(self.monthly_filter_context)

        # حساب الفلاتر النشطة للفلتر المضغوط
        active_filters = []
        if context.get("search_query"):
            active_filters.append("search")
        if context.get("status_filters"):
            active_filters.append("status")
        if context.get("production_line_filters"):
            active_filters.append("production_line")
        if context.get("priority_filters"):
            active_filters.append("priority")
        if context.get("date_from"):
            active_filters.append("date_from")
        if context.get("date_to"):
            active_filters.append("date_to")
        if hasattr(self, "monthly_filter_context") and self.monthly_filter_context.get(
            "selected_year"
        ):
            active_filters.append("year")
        if hasattr(self, "monthly_filter_context") and self.monthly_filter_context.get(
            "selected_month"
        ):
            active_filters.append("month")

        # الحصول على البيانات المطلوبة للفلاتر
        from manufacturing.models import ProductionLine

        production_lines = ProductionLine.objects.filter(is_active=True)

        # سياق الفلتر المضغوط
        context["has_active_filters"] = len(active_filters) > 0
        context["active_filters_count"] = len(active_filters)
        context["production_lines"] = production_lines
        context["status_filter"] = context.get("status_filters", [])
        context["production_line_filter"] = context.get("production_line_filters", [])
        context["priority_filter"] = context.get("priority_filters", [])

        return context

    def _get_available_statuses(self, user, current_status, order_type=None):
        """
        Get list of available statuses for the user based on current status and order type
        """
        if user.is_superuser:
            # Admin can see all statuses except current one, but follow the new logic
            all_statuses = [
                status
                for status in ManufacturingOrder.STATUS_CHOICES
                if status[0] != current_status
            ]

            # Apply the new logic for admin users too
            if current_status == "in_progress":
                if order_type == "installation":
                    return [("ready_install", "جاهز للتركيب")]
                elif order_type in ["custom", "accessory"]:
                    return [("completed", "مكتمل")]
                else:
                    return []
            elif current_status == "ready_install":
                return [("delivered", "تم التسليم")]
            elif current_status == "completed":
                return [("delivered", "تم التسليم")]
            else:
                return all_statuses

        available_statuses = []

        if current_status == "pending_approval":
            # Only approval users can change from pending_approval
            if user.has_perm("manufacturing.can_approve_orders"):
                available_statuses = [
                    ("pending", "قيد الانتظار"),
                    ("rejected", "مرفوض"),
                    ("cancelled", "ملغي"),
                ]
            else:
                available_statuses = []

        elif current_status == "pending":
            # Factory staff can see manufacturing progression
            if user.has_perm("manufacturing.change_manufacturingorder"):
                available_statuses = [
                    ("in_progress", "قيد التصنيع"),
                    ("cancelled", "ملغي"),
                ]
            else:
                available_statuses = []

        elif current_status == "in_progress":
            if user.has_perm("manufacturing.change_manufacturingorder"):
                if order_type == "installation":
                    # بعد قيد التنفيذ لطلبات التركيب: فقط جاهز للتركيب
                    available_statuses = [
                        ("ready_install", "جاهز للتركيب"),
                    ]
                elif order_type in ["custom", "accessory"]:
                    # بعد قيد التنفيذ لطلبات التفصيل أو الاكسسوار: فقط مكتمل
                    available_statuses = [
                        ("completed", "مكتمل"),
                    ]
                else:
                    available_statuses = []
            else:
                available_statuses = []

        elif current_status == "ready_install":
            # بعد جاهز للتركيب: فقط تم التسليم
            if user.has_perm("manufacturing.change_manufacturingorder"):
                available_statuses = [("delivered", "تم التسليم")]
            else:
                available_statuses = []

        elif current_status == "completed":
            # بعد مكتمل: فقط تم التسليم
            if user.has_perm("manufacturing.change_manufacturingorder"):
                available_statuses = [("delivered", "تم التسليم")]
            else:
                available_statuses = []

        elif current_status == "delivered":
            # Delivered orders are final for non-admin users
            available_statuses = []

        elif current_status in ["rejected", "cancelled"]:
            # Rejected/cancelled orders are final for non-admin users
            available_statuses = []

        return available_statuses


class VIPOrdersListView(
    PaginationFixMixin, LoginRequiredMixin, PermissionRequiredMixin, ListView
):
    """عرض طلبات VIP فقط"""

    model = ManufacturingOrder
    template_name = "manufacturing/vip_orders_list.html"
    context_object_name = "vip_orders"
    paginate_by = 25
    permission_required = "manufacturing.view_manufacturingorder"
    allow_empty = True  # السماح بالصفحات الفارغة دون رفع 404

    def get_queryset(self):
        """الحصول على طلبات VIP فقط"""
        queryset = (
            ManufacturingOrder.objects.select_related("order", "order__customer")
            .filter(order__status="vip")  # فلترة طلبات VIP فقط
            .order_by("-created_at", "expected_delivery_date")
        )

        # تطبيق فلترة السنة الافتراضية
        queryset = apply_default_year_filter(queryset, self.request, "order_date")

        # تطبيق فلاتر البحث إذا وجدت (يشمل جميع أرقام الفواتير والعقود)
        search_query = self.request.GET.get("search", "").strip()
        if search_query:
            from django.db.models import Q

            queryset = queryset.filter(
                Q(order__order_number__icontains=search_query)
                | Q(order__customer__name__icontains=search_query)
                | Q(contract_number__icontains=search_query)
                | Q(order__contract_number_2__icontains=search_query)
                | Q(order__contract_number_3__icontains=search_query)
                | Q(invoice_number__icontains=search_query)
                | Q(order__invoice_number_2__icontains=search_query)
                | Q(order__invoice_number_3__icontains=search_query)
            )

        # فلترة حسب الحالة
        status_filter = self.request.GET.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # فلترة حسب نوع الطلب
        order_type_filter = self.request.GET.get("order_type")
        if order_type_filter:
            queryset = queryset.filter(order_type=order_type_filter)

        # فلترة حسب إعداد العرض المحدد يدوياً
        display_setting_filter = self.request.GET.get("display_setting")
        if display_setting_filter:
            try:
                from .models import ManufacturingDisplaySettings

                display_setting = ManufacturingDisplaySettings.objects.get(
                    id=display_setting_filter, is_active=True
                )

                # تطبيق فلترة الحالات
                if display_setting.allowed_statuses:
                    queryset = queryset.filter(
                        status__in=display_setting.allowed_statuses
                    )

                # تطبيق فلترة أنواع الطلبات
                if display_setting.allowed_order_types:
                    queryset = queryset.filter(
                        order_type__in=display_setting.allowed_order_types
                    )

            except Exception as e:
                logger.warning(f"خطأ في تطبيق إعداد العرض: {e}")

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إحصائيات طلبات VIP
        vip_orders = self.get_queryset()

        # إحصائيات حسب الحالة
        status_stats = {}
        for status_code, status_display in ManufacturingOrder.STATUS_CHOICES:
            count = vip_orders.filter(status=status_code).count()
            if count > 0:
                status_stats[status_code] = {"count": count, "display": status_display}

        # إحصائيات حسب النوع
        type_stats = {}
        for type_code, type_display in ManufacturingOrder.ORDER_TYPE_CHOICES:
            count = vip_orders.filter(order_type=type_code).count()
            if count > 0:
                type_stats[type_code] = {"count": count, "display": type_display}

        # الطلبات المتأخرة
        today = timezone.now().date()
        # الطلبات المتأخرة هي التي لم تصل إلى "جاهز للتركيب" أو "مكتملة" أو "تم التسليم"
        overdue_count = vip_orders.filter(
            expected_delivery_date__lt=today,
            status__in=[
                "pending_approval",
                "pending",
                "in_progress",
            ],  # فقط هذه الحالات تعتبر متأخرة
        ).count()

        context.update(
            {
                "total_vip_orders": vip_orders.count(),
                "status_stats": status_stats,
                "type_stats": type_stats,
                "overdue_count": overdue_count,
                "search_query": self.request.GET.get("search", ""),
                "current_status_filter": self.request.GET.get("status", ""),
                "current_type_filter": self.request.GET.get("order_type", ""),
                "status_choices": ManufacturingOrder.STATUS_CHOICES,
                "order_type_choices": ManufacturingOrder.ORDER_TYPE_CHOICES,
            }
        )

        return context


class ManufacturingOrderDetailView(
    LoginRequiredMixin, PermissionRequiredMixin, DetailView
):
    model = ManufacturingOrder
    template_name = "manufacturing/manufacturingorder_detail.html"
    context_object_name = "manufacturing_order"
    permission_required = "manufacturing.view_manufacturingorder"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الحصول على إعدادات التصنيع
        from manufacturing.models import ManufacturingSettings

        settings = ManufacturingSettings.get_settings()

        # الحصول على المستودعات المحددة للعرض
        display_warehouses = list(settings.warehouses_for_display.all())
        meters_warehouses = list(settings.warehouses_for_meters_calculation.all())

        # إذا كانت القوائم فارغة، نستخدم None للدلالة على "الكل"
        display_warehouses = display_warehouses if display_warehouses else None
        meters_warehouses = meters_warehouses if meters_warehouses else None

        context["display_warehouses"] = display_warehouses
        context["meters_warehouses"] = meters_warehouses
        context["has_warehouse_filter"] = display_warehouses is not None

        # الحصول على جميع عناصر الطلب الأصلي
        if self.object.order:
            order_items = self.object.order.items.select_related("product").all()

            # إنشاء قاموس لبيانات التقطيع والاستلام
            items_data = []
            total_meters = 0  # لحساب إجمالي الأمتار من المستودعات المحددة

            for order_item in order_items:
                # البحث عن عنصر التصنيع المرتبط بهذا العنصر
                mfg_item = self.object.items.filter(order_item=order_item).first()

                # البحث عن عناصر التقطيع المرتبطة بهذا العنصر
                cutting_items = order_item.cutting_items.all()

                if mfg_item:
                    # العنصر له manufacturing item
                    cutting_item = mfg_item.cutting_item
                    warehouse = (
                        cutting_item.cutting_order.warehouse
                        if cutting_item and cutting_item.cutting_order
                        else None
                    )

                    # التحقق من أن المستودع ضمن المستودعات المحددة للعرض
                    should_display = True
                    if display_warehouses:
                        should_display = warehouse and warehouse in display_warehouses

                    if should_display:
                        item_data = {
                            "manufacturing_item": mfg_item,
                            "order_item": order_item,
                            "cutting_receiver_name": mfg_item.receiver_name
                            or (cutting_item.receiver_name if cutting_item else None),
                            "cutting_permit_number": mfg_item.permit_number
                            or (cutting_item.permit_number if cutting_item else None),
                            "cutting_date": mfg_item.cutting_date
                            or (cutting_item.cutting_date if cutting_item else None),
                            "cutting_status": mfg_item.cutting_status_display,
                            "cutting_status_color": mfg_item.get_cutting_status_color(),
                            "has_cutting_data": mfg_item.has_cutting_data,
                            "is_cut": mfg_item.is_cut,
                            "warehouse_name": warehouse.name if warehouse else None,
                            "warehouse": warehouse,
                            "fabric_received": mfg_item.fabric_received,
                            "bag_number": mfg_item.bag_number,
                            "fabric_received_date": mfg_item.fabric_received_date,
                            "fabric_receiver_name": (
                                mfg_item.fabric_received_by.get_full_name()
                                if mfg_item.fabric_received_by
                                else None
                            ),
                            "fabric_status": mfg_item.get_fabric_status_display(),
                            "fabric_status_color": mfg_item.get_fabric_status_color(),
                        }

                        items_data.append(item_data)

                        # حساب الأمتار
                        should_count = True
                        if meters_warehouses:
                            should_count = warehouse and warehouse in meters_warehouses

                        if should_count:
                            total_meters += order_item.quantity

                elif cutting_items.exists():
                    # العنصر تم تقطيعه لكن لا يوجد manufacturing item
                    for cutting_item in cutting_items:
                        warehouse = (
                            cutting_item.cutting_order.warehouse
                            if cutting_item.cutting_order
                            else None
                        )

                        # التحقق من الفلترة
                        should_display = True
                        if display_warehouses:
                            should_display = (
                                warehouse and warehouse in display_warehouses
                            )

                        if should_display:
                            # التحقق من وجود بيانات تقطيع حقيقية
                            has_actual_cutting_data = bool(
                                cutting_item.receiver_name
                                and cutting_item.permit_number
                            )

                            item_data = {
                                "manufacturing_item": None,
                                "order_item": order_item,
                                "cutting_receiver_name": cutting_item.receiver_name,
                                "cutting_permit_number": cutting_item.permit_number,
                                "cutting_date": cutting_item.cutting_date,
                                "cutting_status": cutting_item.get_status_display(),
                                "cutting_status_color": (
                                    "success"
                                    if cutting_item.status == "completed"
                                    else (
                                        "info"
                                        if cutting_item.status == "in_progress"
                                        else "secondary"
                                    )
                                ),
                                "has_cutting_data": has_actual_cutting_data,
                                "is_cut": cutting_item.status == "completed"
                                and has_actual_cutting_data,
                                "warehouse_name": warehouse.name if warehouse else None,
                                "warehouse": warehouse,
                                "fabric_received": False,
                                "bag_number": None,
                                "fabric_received_date": None,
                                "fabric_receiver_name": None,
                                "fabric_status": "لم يتم الاستلام",
                                "fabric_status_color": "warning",
                            }

                            items_data.append(item_data)

                            # حساب الأمتار
                            should_count = True
                            if meters_warehouses:
                                should_count = (
                                    warehouse and warehouse in meters_warehouses
                                )

                            if should_count:
                                total_meters += order_item.quantity
                else:
                    # العنصر لم يتم تقطيعه بعد - نعرضه دائماً ونحسب أمتاره إذا لم يكن هناك فلترة للمستودعات
                    item_data = {
                        "manufacturing_item": None,
                        "order_item": order_item,
                        "cutting_receiver_name": None,
                        "cutting_permit_number": None,
                        "cutting_date": None,
                        "cutting_status": "لم يتم التقطيع",
                        "cutting_status_color": "secondary",
                        "has_cutting_data": False,
                        "is_cut": False,
                        "warehouse_name": None,
                        "warehouse": None,
                        "fabric_received": False,
                        "bag_number": None,
                        "fabric_received_date": None,
                        "fabric_receiver_name": None,
                        "fabric_status": "لم يتم التقطيع",
                        "fabric_status_color": "secondary",
                    }

                    items_data.append(item_data)

                    # حساب الأمتار للعناصر غير المقطوعة فقط إذا لم يكن هناك فلترة للمستودعات
                    # (لأن العناصر غير المقطوعة ليس لها مستودع بعد)
                    if not meters_warehouses:
                        total_meters += order_item.quantity

            # --- إضافة الأقمشة الخارجية (External Fabrics) ---
            try:
                from cutting.models import CuttingOrderItem

                external_cutting_items = CuttingOrderItem.objects.filter(
                    cutting_order__order=self.object.order, is_external=True
                )

                for cutting_item in external_cutting_items:
                    # التحقق من وجود manufacturing item
                    mfg_item = self.object.items.filter(
                        cutting_item=cutting_item
                    ).first()

                    warehouse = (
                        cutting_item.cutting_order.warehouse
                        if cutting_item.cutting_order
                        else None
                    )

                    # التحقق من الفلترة
                    should_display = True
                    if display_warehouses:
                        should_display = warehouse and warehouse in display_warehouses

                    if should_display:
                        # تجهيز كائن وهمي يحاكي order_item لاستخدامه في القالب
                        # القالب يتوقع: order_item.product.name, order_item.quantity, etc.
                        class FakeProduct:
                            def __init__(self, name):
                                self.name = name
                                self.code = ""

                        class FakeOrderItem:
                            def __init__(self, item):
                                self.product = FakeProduct(
                                    f"{item.external_fabric_name} (خارجي)"
                                )
                                self.quantity = item.quantity
                                self.fabric_type = "قماش خارجي"
                                self.fabric_color = ""
                                self.fabric_meters = item.quantity

                        fake_order_item = FakeOrderItem(cutting_item)

                        if mfg_item:
                            item_data = {
                                "manufacturing_item": mfg_item,
                                "order_item": fake_order_item,
                                "cutting_receiver_name": mfg_item.receiver_name
                                or cutting_item.receiver_name,
                                "cutting_permit_number": mfg_item.permit_number
                                or cutting_item.permit_number,
                                "cutting_date": mfg_item.cutting_date
                                or cutting_item.cutting_date,
                                "cutting_status": mfg_item.cutting_status_display,
                                "cutting_status_color": mfg_item.get_cutting_status_color(),
                                "has_cutting_data": mfg_item.has_cutting_data,
                                "is_cut": mfg_item.is_cut,
                                "warehouse_name": warehouse.name if warehouse else None,
                                "warehouse": warehouse,
                                "fabric_received": mfg_item.fabric_received,
                                "bag_number": mfg_item.bag_number,
                                "fabric_received_date": mfg_item.fabric_received_date,
                                "fabric_receiver_name": (
                                    mfg_item.fabric_received_by.get_full_name()
                                    if mfg_item.fabric_received_by
                                    else None
                                ),
                                "fabric_status": mfg_item.get_fabric_status_display(),
                                "fabric_status_color": mfg_item.get_fabric_status_color(),
                            }
                        else:
                            # لم يتم إنشاء عنصر تصنيع بعد (أو لم يكتمل التقطيع)
                            has_actual_cutting_data = bool(
                                cutting_item.receiver_name
                                and cutting_item.permit_number
                            )

                            item_data = {
                                "manufacturing_item": None,
                                "order_item": fake_order_item,
                                "cutting_receiver_name": cutting_item.receiver_name,
                                "cutting_permit_number": cutting_item.permit_number,
                                "cutting_date": cutting_item.cutting_date,
                                "cutting_status": cutting_item.get_status_display(),
                                "cutting_status_color": (
                                    "success"
                                    if cutting_item.status == "completed"
                                    else (
                                        "info"
                                        if cutting_item.status == "in_progress"
                                        else "secondary"
                                    )
                                ),
                                "has_cutting_data": has_actual_cutting_data,
                                "is_cut": cutting_item.status == "completed"
                                and has_actual_cutting_data,
                                "warehouse_name": warehouse.name if warehouse else None,
                                "warehouse": warehouse,
                                "fabric_received": False,
                                "bag_number": None,
                                "fabric_received_date": None,
                                "fabric_receiver_name": None,
                                "fabric_status": "لم يتم الاستلام",
                                "fabric_status_color": "warning",
                            }

                        items_data.append(item_data)

                        # حساب الأمتار
                        should_count = True
                        if meters_warehouses:
                            should_count = warehouse and warehouse in meters_warehouses

                        if should_count:
                            total_meters += cutting_item.quantity

            except Exception as e:
                logger.error(
                    f"Error processing external fabrics in manufacturing detail: {e}"
                )

            context["items_data"] = items_data

            # حساب إجمالي الكمية من المستودعات المحددة فقط
            context["total_quantity"] = total_meters

            # حساب إجمالي الكمية من جميع عناصر الطلب (للمقارنة)
            context["all_items_quantity"] = sum(item.quantity for item in order_items)

            # إضافة إحصائيات العناصر
            context["total_items"] = len(items_data)
            context["cut_items"] = sum(1 for item in items_data if item["is_cut"])
            context["received_items"] = sum(
                1 for item in items_data if item["fabric_received"]
            )
            context["pending_items"] = (
                context["total_items"] - context["received_items"]
            )
        else:
            context["items_data"] = []
            context["total_quantity"] = 0
            context["all_items_quantity"] = 0
            context["total_items"] = 0
            context["cut_items"] = 0
            context["received_items"] = 0
            context["pending_items"] = 0

        return context


class ManufacturingOrderCreateView(
    LoginRequiredMixin, PermissionRequiredMixin, CreateView
):
    model = ManufacturingOrder
    template_name = "manufacturing/manufacturingorder_form.html"
    fields = [
        "order",
        "order_type",
        "contract_number",
        "invoice_number",
        "order_date",
        "expected_delivery_date",
        "notes",
        "contract_file",
        "inspection_file",
    ]
    permission_required = "manufacturing.add_manufacturingorder"

    def get_success_url(self):
        messages.success(self.request, "تم إنشاء أمر التصنيع بنجاح")
        return reverse("manufacturing:order_list")

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class ChangeProductionLineView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """تبديل خط الإنتاج لأمر تصنيع"""

    permission_required = "manufacturing.change_manufacturingorder"

    def post(self, request, pk):
        """تبديل خط الإنتاج"""
        try:
            # الحصول على أمر التصنيع
            manufacturing_order = get_object_or_404(ManufacturingOrder, pk=pk)

            # الحصول على البيانات من JSON أو POST
            if request.content_type == "application/json":
                import json

                data = json.loads(request.body)
                new_production_line_id = data.get("production_line_id")
                reason = data.get("reason", "")
            else:
                new_production_line_id = request.POST.get("production_line_id")
                reason = request.POST.get("reason", "")

            if not new_production_line_id:
                return JsonResponse(
                    {"success": False, "message": "يجب اختيار خط إنتاج"}
                )

            # التحقق من وجود خط الإنتاج
            from .models import ProductionLine

            try:
                new_production_line = ProductionLine.objects.get(
                    id=new_production_line_id, is_active=True
                )
            except ProductionLine.DoesNotExist:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "خط الإنتاج المحدد غير موجود أو غير نشط",
                    }
                )

            # حفظ خط الإنتاج القديم للسجل
            old_production_line = manufacturing_order.production_line
            old_line_name = (
                old_production_line.name if old_production_line else "غير محدد"
            )

            # تحديث خط الإنتاج
            manufacturing_order.production_line = new_production_line
            manufacturing_order.save()

            # تسجيل العملية في السجل
            from django.contrib.admin.models import CHANGE, LogEntry
            from django.contrib.contenttypes.models import ContentType

            LogEntry.objects.create(
                user_id=request.user.id,
                content_type_id=ContentType.objects.get_for_model(
                    ManufacturingOrder
                ).pk,
                object_id=str(manufacturing_order.pk),
                object_repr=str(manufacturing_order),
                action_flag=CHANGE,
                change_message=f'تم تبديل خط الإنتاج من "{old_line_name}" إلى "{new_production_line.name}"',
            )

            return JsonResponse(
                {
                    "success": True,
                    "message": f'تم تبديل خط الإنتاج بنجاح إلى "{new_production_line.name}"',
                    "new_line_name": new_production_line.name,
                    "old_line_name": old_line_name,
                }
            )

        except Exception as e:
            return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


@require_http_methods(["GET"])
def get_production_lines_api(request):
    """API لجلب خطوط الإنتاج النشطة"""
    try:
        from .models import ProductionLine

        lines = ProductionLine.objects.filter(is_active=True).order_by(
            "-priority", "name"
        )

        lines_data = []
        for line in lines:
            lines_data.append(
                {
                    "id": line.id,
                    "name": line.name,
                    "description": line.description or "",
                    "capacity_per_day": line.capacity_per_day,
                    "priority": line.priority,
                }
            )

        return JsonResponse(lines_data, safe=False)

    except Exception as e:
        return JsonResponse({"error": f"حدث خطأ: {str(e)}"}, status=500)


class ProductionLinePrintView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """طباعة طلبات خط إنتاج محدد مع فلاتر متقدمة وصفحات"""

    model = ManufacturingOrder
    template_name = "manufacturing/production_line_print.html"
    context_object_name = "orders"
    paginate_by = 100  # عرض 100 طلب في كل صفحة
    permission_required = "manufacturing.view_manufacturingorder"

    def get_queryset(self):
        """الحصول على طلبات خط الإنتاج مع تطبيق الفلاتر"""
        import datetime

        from django.db.models import Q

        from accounts.models import Branch

        from .models import ProductionLine

        # الحصول على خط الإنتاج
        line_id = self.kwargs.get("line_id")
        self.production_line = get_object_or_404(
            ProductionLine, id=line_id, is_active=True
        )

        # الحصول على طلبات خط الإنتاج
        queryset = ManufacturingOrder.objects.select_related(
            "order", "order__customer", "order__branch", "order__salesperson"
        ).filter(production_line=self.production_line)

        # تطبيق الترتيب
        sort_column = self.request.GET.get("sort")
        sort_direction = self.request.GET.get("direction", "asc")

        if sort_column and sort_direction != "none":
            # تحديد حقل الترتيب
            sort_field = self.get_sort_field(sort_column)
            if sort_field:
                if sort_direction == "desc":
                    sort_field = f"-{sort_field}"
                queryset = queryset.order_by(sort_field)
            else:
                queryset = queryset.order_by("-created_at")
        else:
            queryset = queryset.order_by("-created_at")

        # فلتر الحالات (اختيار متعدد)
        status_filters = self.request.GET.getlist("status")
        if status_filters:
            queryset = queryset.filter(status__in=status_filters)

        # فلتر الفروع (اختيار متعدد)
        branch_filters = self.request.GET.getlist("branch")
        if branch_filters:
            branch_query = Q()
            for branch_id in branch_filters:
                try:
                    branch_id = int(branch_id)
                    branch_query |= Q(order__branch_id=branch_id) | Q(
                        order__customer__branch_id=branch_id
                    )
                except (ValueError, TypeError):
                    continue
            if branch_query:
                queryset = queryset.filter(branch_query)

        # فلتر نوع الطلب (اختيار متعدد)
        order_type_filters = self.request.GET.getlist("order_type")
        if order_type_filters:
            queryset = queryset.filter(order_type__in=order_type_filters)

        # فلتر الطلبات المتأخرة
        overdue_filter = self.request.GET.get("overdue")
        if overdue_filter == "true":
            today = timezone.now().date()
            queryset = queryset.filter(
                expected_delivery_date__lt=today,
                status__in=["pending_approval", "pending", "in_progress"],
                order_type__in=["installation", "custom", "delivery"],
            )

        # فلتر التواريخ
        date_from = self.request.GET.get("date_from")
        if date_from:
            try:
                queryset = queryset.filter(order_date__gte=date_from)
            except:
                pass

        date_to = self.request.GET.get("date_to")
        if date_to:
            try:
                end_date = datetime.datetime.strptime(
                    date_to, "%Y-%m-%d"
                ) + datetime.timedelta(days=1)
                queryset = queryset.filter(order_date__lt=end_date)
            except:
                pass

        return queryset

    def get_sort_field(self, sort_column):
        """تحديد حقل الترتيب المناسب"""
        sort_fields = {
            "id": "id",
            "manufacturing_code": "manufacturing_code",
            "contract_number": "contract_number",
            "invoice_number": "invoice_number",
            "customer": "order__customer__name",
            "order_type": "order_type",
            "status": "status",
            "order_date": "order_date",
            "expected_delivery_date": "expected_delivery_date",
            "branch": "order__branch__name",
            "salesperson": "order__salesperson__name",
        }
        return sort_fields.get(sort_column)

    def get_context_data(self, **kwargs):
        """إضافة البيانات الإضافية للسياق"""
        context = super().get_context_data(**kwargs)
        from accounts.models import Branch

        # إحصائيات خط الإنتاج
        all_orders = self.get_queryset()
        total_orders = all_orders.count()
        active_orders = all_orders.filter(
            status__in=["pending_approval", "pending", "in_progress"]
        ).count()
        completed_orders = all_orders.filter(
            status__in=["ready_install", "completed", "delivered"]
        ).count()
        overdue_orders = all_orders.filter(
            expected_delivery_date__lt=timezone.now().date(),
            status__in=[
                "pending_approval",
                "pending",
                "in_progress",
            ],  # فقط هذه الحالات تعتبر متأخرة
            order_type__in=["installation", "custom", "delivery"],
        ).count()

        # الفلاتر المطبقة
        applied_filters = {
            "status_filters": self.request.GET.getlist("status"),
            "branch_filters": self.request.GET.getlist("branch"),
            "order_type_filters": self.request.GET.getlist("order_type"),
            "overdue_filter": self.request.GET.get("overdue"),
            "date_from": self.request.GET.get("date_from"),
            "date_to": self.request.GET.get("date_to"),
        }

        # خيارات الفلاتر
        # الحصول على أنواع الطلبات المدعومة في هذا الخط
        supported_order_types = []
        if (
            hasattr(self.production_line, "supported_order_types")
            and self.production_line.supported_order_types
        ):
            for order_type in self.production_line.supported_order_types:
                for choice_code, choice_name in ManufacturingOrder.ORDER_TYPE_CHOICES:
                    if choice_code == order_type:
                        supported_order_types.append((choice_code, choice_name))
        else:
            # إذا لم تكن محددة، عرض جميع الأنواع
            supported_order_types = ManufacturingOrder.ORDER_TYPE_CHOICES

        filter_options = {
            "status_choices": ManufacturingOrder.STATUS_CHOICES,
            "order_type_choices": supported_order_types,
            "branches": Branch.objects.filter(is_active=True).order_by("name"),
        }

        context.update(
            {
                "production_line": self.production_line,
                "total_orders": total_orders,
                "active_orders": active_orders,
                "completed_orders": completed_orders,
                "overdue_orders": overdue_orders,
                "applied_filters": applied_filters,
                "filter_options": filter_options,
                "print_date": timezone.now(),
                "filtered_count": (
                    context["orders"].count() if context.get("orders") else 0
                ),
            }
        )

        return context


class ProductionLinePrintTemplateView(ProductionLinePrintView):
    """طباعة تقرير خط الإنتاج بقالب بسيط"""

    template_name = "manufacturing/production_line_print_template.html"
    paginate_by = None  # إزالة pagination للطباعة

    def get_context_data(self, **kwargs):
        # الحصول على جميع الطلبات المفلترة بدون pagination
        queryset = self.get_queryset()

        context = {
            "orders": queryset,
            "production_line": self.production_line,
            "total_orders": queryset.count(),
            "now": timezone.now(),
        }

        # إضافة معلومات الفلاتر المطبقة للعرض
        applied_filters = {}

        # فلاتر الحالة
        status_filters = self.request.GET.getlist("status")
        if status_filters:
            applied_filters["status"] = status_filters

        # فلاتر الفروع
        branch_filters = self.request.GET.getlist("branch")
        if branch_filters:
            applied_filters["branch"] = branch_filters

        # فلاتر نوع الطلب
        order_type_filters = self.request.GET.getlist("order_type")
        if order_type_filters:
            applied_filters["order_type"] = order_type_filters

        # فلتر التواريخ
        date_from = self.request.GET.get("date_from")
        if date_from:
            applied_filters["date_from"] = date_from

        date_to = self.request.GET.get("date_to")
        if date_to:
            applied_filters["date_to"] = date_to

        # فلتر الطلبات المتأخرة
        overdue_filter = self.request.GET.get("overdue")
        if overdue_filter == "true":
            applied_filters["overdue"] = True

        context["applied_filters"] = applied_filters
        context["has_filters"] = bool(applied_filters)

        return context


class ProductionLinePDFView(ProductionLinePrintTemplateView):
    """تحميل تقرير خط الإنتاج كملف PDF"""

    def setup(self, request, *args, **kwargs):
        """إعداد المتغيرات المطلوبة"""
        super().setup(request, *args, **kwargs)
        # تهيئة production_line من get_queryset
        self.get_queryset()

    def get(self, request, *args, **kwargs):
        try:
            import io

            from django.http import HttpResponse
            from django.template.loader import render_to_string
            from weasyprint import CSS, HTML

            # الحصول على البيانات
            context = self.get_context_data()

            # رندر HTML
            html_string = render_to_string(self.template_name, context, request=request)

            # تحويل إلى PDF
            html = HTML(string=html_string, base_url=request.build_absolute_uri())

            # CSS إضافي للطباعة
            css = CSS(
                string="""
                @page {
                    size: A4;
                    margin: 1.5cm;
                }
                body {
                    font-family: 'Arial', sans-serif;
                    -webkit-print-color-adjust: exact;
                    print-color-adjust: exact;
                }
            """
            )

            # إنشاء PDF
            pdf_file = html.write_pdf(stylesheets=[css])

            # إنشاء الاستجابة
            response = HttpResponse(pdf_file, content_type="application/pdf")
            filename = f"production_line_{self.production_line.name}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except ImportError:
            # إذا لم تكن weasyprint مثبتة، استخدم طريقة بديلة
            from django.http import JsonResponse

            return JsonResponse(
                {
                    "error": "مكتبة PDF غير متوفرة. يرجى استخدام خيار الطباعة العادي.",
                    "message": "PDF library not available. Please use regular print option.",
                },
                status=500,
            )

        except Exception as e:
            from django.http import JsonResponse

            return JsonResponse(
                {
                    "error": f"خطأ في إنشاء PDF: {str(e)}",
                    "message": "Error generating PDF",
                },
                status=500,
            )

    def get(self, request, *args, **kwargs):
        """معالجة طلبات GET مع دعم تصدير PDF"""
        try:
            # تحديد نوع الاستجابة (HTML أو PDF)
            if request.GET.get("format") == "pdf":
                return self.generate_pdf()

            return super().get(request, *args, **kwargs)

        except Exception as e:
            messages.error(request, f"حدث خطأ: {str(e)}")
            return redirect("manufacturing:dashboard")

    def generate_pdf(self):
        """إنتاج ملف PDF للطلبات المفلترة"""
        try:
            from django.template.loader import render_to_string
            from weasyprint import HTML

            # الحصول على جميع الطلبات المفلترة (بدون صفحات للـ PDF)
            all_filtered_orders = self.get_queryset()

            context = self.get_context_data()
            context.update(
                {
                    "orders": all_filtered_orders,  # جميع الطلبات المفلترة للـ PDF
                    "is_pdf": True,
                    "user": self.request.user,
                }
            )

            html_string = render_to_string(
                "manufacturing/production_line_print_pdf.html", context
            )
            html = HTML(string=html_string)
            pdf = html.write_pdf()

            response = HttpResponse(pdf, content_type="application/pdf")
            filename = f"production_line_{self.production_line.name}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            messages.error(self.request, f"خطأ في إنتاج PDF: {str(e)}")
            return redirect("manufacturing:dashboard")


class ManufacturingOrderDeleteView(
    LoginRequiredMixin, PermissionRequiredMixin, DeleteView
):
    model = ManufacturingOrder
    template_name = "manufacturing/manufacturingorder_confirm_delete.html"
    success_url = reverse_lazy("manufacturing:order_list")
    permission_required = "manufacturing.delete_manufacturingorder"

    def delete(self, request, *args, **kwargs):
        import json

        from django.db import transaction

        from orders.models import ManufacturingDeletionLog

        # الحصول على أمر التصنيع قبل الحذف
        manufacturing_order = self.get_object()
        order = manufacturing_order.order

        # حفظ بيانات أمر التصنيع قبل الحذف
        manufacturing_data = {
            "id": manufacturing_order.id,
            "status": manufacturing_order.status,
            "status_display": manufacturing_order.get_status_display(),
            "order_type": manufacturing_order.order_type,
            "contract_number": manufacturing_order.contract_number,
            "invoice_number": manufacturing_order.invoice_number,
            "order_date": (
                manufacturing_order.order_date.isoformat()
                if manufacturing_order.order_date
                else None
            ),
            "expected_delivery_date": (
                manufacturing_order.expected_delivery_date.isoformat()
                if manufacturing_order.expected_delivery_date
                else None
            ),
            "notes": manufacturing_order.notes,
            "created_at": (
                manufacturing_order.created_at.isoformat()
                if manufacturing_order.created_at
                else None
            ),
            "created_by": (
                manufacturing_order.created_by.username
                if manufacturing_order.created_by
                else None
            ),
        }

        with transaction.atomic():
            # إنشاء سجل الحذف
            ManufacturingDeletionLog.objects.create(
                order=order,
                manufacturing_order_id=manufacturing_order.id,
                deleted_by=request.user,
                reason=f"تم حذف أمر التصنيع بواسطة {request.user.get_full_name() or request.user.username}",
                manufacturing_order_data=manufacturing_data,
            )

            # تحديث حالة الطلب
            order.order_status = "manufacturing_deleted"
            order.tracking_status = "processing"  # إعادة الطلب لحالة المعالجة
            order.save(update_fields=["order_status", "tracking_status"])

            # حذف أمر التصنيع
            result = super().delete(request, *args, **kwargs)

            messages.success(
                request,
                f'تم حذف أمر التصنيع #{manufacturing_order.id} بنجاح وتم تحديث حالة الطلب #{order.order_number} إلى "أمر تصنيع محذوف"',
            )

            return result


import logging

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from .models import ManufacturingOrder

# Configure logger
logger = logging.getLogger(__name__)


@require_http_methods(["POST"])
@login_required
@permission_required("manufacturing.change_manufacturingorder", raise_exception=True)
def update_order_status(request, pk):
    """
    API endpoint to update the status of a manufacturing order.
    محمي بـ CSRF و login_required

    Expected POST data: {'status': 'new_status'}
    """
    # logger.info(f"[update_order_status] Starting update for order {pk}")  # معطل لتجنب الرسائل الكثيرة
    # logger.debug(f"[update_order_status] Request user: {request.user}")  # معطل لتجنب الرسائل الكثيرة
    # logger.debug(f"[update_order_status] POST data: {request.POST}")  # معطل لتجنب الرسائل الكثيرة

    try:
        # Check authentication
        if not request.user.is_authenticated:
            logger.warning("[update_order_status] Unauthenticated access attempt")
            return JsonResponse(
                {"success": False, "error": "يجب تسجيل الدخول أولاً"}, status=401
            )

        # Get the order
        try:
            order = ManufacturingOrder.objects.get(pk=pk)
            logger.debug(f"[update_order_status] Found order: {order}")
        except ManufacturingOrder.DoesNotExist:
            # logger.error(f"[update_order_status] Order {pk} not found")  # معطل لتجنب الرسائل الكثيرة
            return JsonResponse(
                {"success": False, "error": "لم يتم العثور على أمر التصنيع"}, status=404
            )

        # Check if order is in pending_approval status
        if order.status == "pending_approval":
            return JsonResponse(
                {
                    "success": False,
                    "error": "لا يمكن تغيير حالة الطلب قبل الموافقة عليه. يرجى استخدام أزرار الموافقة أو الرفض.",
                },
                status=403,
            )

        # Check permission for manufacturing status changes
        if not request.user.has_perm("manufacturing.change_manufacturingorder"):
            logger.warning(
                f"[update_order_status] Permission denied for user {request.user}"
            )
            return JsonResponse(
                {"success": False, "error": "ليس لديك صلاحية لتحديث حالة الطلب"},
                status=403,
            )

        # Get and validate status
        new_status = request.POST.get("status")
        logger.debug(f"[update_order_status] Requested status: {new_status}")

        if not new_status:
            logger.warning("[update_order_status] No status provided")
            return JsonResponse(
                {"success": False, "error": "لم يتم تحديد الحالة الجديدة"}, status=400
            )

        # التحقق من نوع الطلب الأصلي والحالة المطلوبة
        order_types = (
            order.order.get_selected_types_list()
            if hasattr(order.order, "get_selected_types_list")
            else []
        )

        # التحقق من نوع الطلب والحالة المطلوبة
        if new_status == "ready_install" and "installation" not in order_types:
            return JsonResponse(
                {
                    "success": False,
                    "error": 'لا يمكن تعيين حالة "جاهز للتركيب" إلا لأوامر التصنيع من نوع تركيب (installation) فقط.',
                },
                status=400,
            )

        # منع تعيين حالة "مكتمل" لطلبات التركيب
        if new_status == "completed" and "installation" in order_types:
            return JsonResponse(
                {
                    "success": False,
                    "error": 'لا يمكن تعيين حالة "مكتمل" لطلبات التركيب. طلبات التركيب تصبح "جاهزة للتركيب" ثم "تم التسليم" فقط.',
                },
                status=400,
            )

        # منع تعيين حالة "جاهز للتركيب" لطلبات التفصيل والاكسسوار
        if new_status == "ready_install" and (
            "tailoring" in order_types or "accessory" in order_types
        ):
            return JsonResponse(
                {
                    "success": False,
                    "error": 'لا يمكن تعيين حالة "جاهز للتركيب" لطلبات التفصيل أو الاكسسوار. هذه الطلبات تصبح "مكتملة" فقط.',
                },
                status=400,
            )

        # منع الرفض والإلغاء بعد الحالات المكتملة
        if order.status in ["completed", "ready_install"] and new_status in [
            "rejected",
            "cancelled",
        ]:
            return JsonResponse(
                {
                    "success": False,
                    "error": "لا يمكن رفض أو إلغاء الطلب بعد أن يصبح مكتملاً أو جاهزاً للتركيب. يمكن فقط تسليمه.",
                },
                status=400,
            )

        # Prevent going back to pending_approval unless user is superuser
        if new_status == "pending_approval" and not request.user.is_superuser:
            return JsonResponse(
                {
                    "success": False,
                    "error": 'لا يمكن العودة إلى حالة "قيد الموافقة" إلا من قبل مدير النظام',
                },
                status=403,
            )

        # Define status hierarchy to prevent going backwards (unless superuser)
        status_hierarchy = {
            "pending_approval": 0,
            "pending": 1,
            "in_progress": 2,
            "ready_install": 3,
            "completed": 4,
            "delivered": 5,
            "rejected": -1,
            "cancelled": -2,
        }

        current_level = status_hierarchy.get(order.status, 0)
        new_level = status_hierarchy.get(new_status, 0)

        # Prevent going backwards unless user is superuser
        if (
            new_level < current_level
            and not request.user.is_superuser
            and new_status not in ["rejected", "cancelled"]
        ):
            return JsonResponse(
                {
                    "success": False,
                    "error": "لا يمكن العودة إلى حالة سابقة إلا من قبل مدير النظام",
                },
                status=403,
            )

        valid_statuses = dict(ManufacturingOrder.STATUS_CHOICES).keys()
        if new_status not in valid_statuses:
            logger.warning(
                f"[update_order_status] Invalid status: {new_status}. Valid statuses: {list(valid_statuses)}"
            )
            return JsonResponse(
                {
                    "success": False,
                    "error": f"حالة غير صالحة. الحالات المتاحة: {list(valid_statuses)}",
                    "received_status": new_status,
                    "valid_statuses": list(valid_statuses),
                },
                status=400,
            )

        # Handle delivery status - require delivery fields
        if new_status == "delivered":
            delivery_permit_number = request.POST.get(
                "delivery_permit_number", ""
            ).strip()
            delivery_recipient_name = request.POST.get(
                "delivery_recipient_name", ""
            ).strip()

            if not delivery_permit_number:
                return JsonResponse(
                    {
                        "success": False,
                        "error": 'رقم إذن التسليم مطلوب عند تغيير الحالة إلى "تم التسليم"',
                    },
                    status=400,
                )

            if not delivery_recipient_name:
                return JsonResponse(
                    {
                        "success": False,
                        "error": 'اسم المستلم مطلوب عند تغيير الحالة إلى "تم التسليم"',
                    },
                    status=400,
                )

        # Update order status
        old_status = order.status
        # logger.info(f"[update_order_status] Updating order {pk} status from {old_status} to {new_status}")  # معطل لتجنب الرسائل الكثيرة

        order.status = new_status
        order.updated_at = timezone.now()

        # إنشاء سجل تغيير الحالة - نسجل جميع تغييرات حالة التصنيع
        try:
            from orders.models import OrderStatusLog

            # إنشاء سجل لتغيير حالة التصنيع (نستخدم حالات التصنيع مباشرة)
            OrderStatusLog.objects.create(
                order=order.order,
                old_status=old_status,
                new_status=new_status,
                changed_by=request.user,
                change_type="manufacturing",
                notes=f'تم تغيير حالة أمر التصنيع من {dict(ManufacturingOrder.STATUS_CHOICES).get(old_status, "")} إلى {dict(ManufacturingOrder.STATUS_CHOICES).get(new_status, "")}',
            )
            logger.debug(f"[update_order_status] Created status log for order {pk}")
        except Exception as e:
            logger.error(
                f"[update_order_status] Error creating status log: {str(e)}",
                exc_info=True,
            )
            # Continue even if status logging fails

        # Handle delivery fields
        if new_status == "delivered":
            order.delivery_permit_number = request.POST.get(
                "delivery_permit_number", ""
            ).strip()
            order.delivery_recipient_name = request.POST.get(
                "delivery_recipient_name", ""
            ).strip()
            order.delivery_date = timezone.now()
            # logger.debug(f"[update_order_status] Set delivery fields for order {pk}")  # معطل لتجنب الرسائل الكثيرة

        # Set completion date if status is completed or ready for installation
        if (
            new_status in ["completed", "ready_install", "delivered"]
            and not order.completion_date
        ):
            order.completion_date = timezone.now()
            # logger.debug(f"[update_order_status] Set completion date to {order.completion_date}")  # معطل لتجنب الرسائل الكثيرة

        # Set the user who changed the status for the signal handler
        order._changed_by = request.user

        # Save the order with transaction to ensure integrity
        try:
            with transaction.atomic():
                save_fields = ["status", "updated_at", "completion_date"]
                if new_status == "delivered":
                    save_fields.extend(
                        [
                            "delivery_permit_number",
                            "delivery_recipient_name",
                            "delivery_date",
                        ]
                    )

                # Check for cached signal issues by explicitly setting _changed_by
                if not hasattr(order, "_changed_by"):
                    order._changed_by = request.user

                order.save(update_fields=save_fields)

                # Log the status change in Django admin within the same transaction
                try:
                    from django.contrib.admin.models import CHANGE, LogEntry
                    from django.contrib.contenttypes.models import ContentType

                    LogEntry.objects.create(
                        user_id=request.user.id,
                        content_type_id=ContentType.objects.get_for_model(order).pk,
                        object_id=order.id,
                        object_repr=str(order),
                        action_flag=CHANGE,
                        change_message=f'تم تغيير الحالة من {dict(ManufacturingOrder.STATUS_CHOICES).get(old_status, "")} إلى {dict(ManufacturingOrder.STATUS_CHOICES).get(new_status, "")}',
                    )
                except Exception as log_error:
                    logger.error(
                        f"[update_order_status] Error adding admin log entry: {str(log_error)}",
                        exc_info=True,
                    )
                    # Continue even if admin logging fails

        except Exception as e:
            logger.error(
                f"[update_order_status] Error saving order {pk}: {str(e)}",
                exc_info=True,
            )
            return JsonResponse(
                {
                    "success": False,
                    "error": "حدث خطأ أثناء حفظ التغييرات",
                    "details": str(e),
                    "exception_type": e.__class__.__name__,
                },
                status=500,
            )

        # Verify the update
        order.refresh_from_db()
        if order.status != new_status:
            logger.warning(
                f"[update_order_status] Status Mismatch! Expected {new_status}, got {order.status} for order {pk}"
            )

        # Prepare success response
        response_data = {
            "success": True,
            "status": order.status,
            "status_display": order.get_status_display(),
            "updated_at": order.updated_at.strftime("%Y-%m-%d %H:%M"),
            "message": "تم تحديث حالة الطلب بنجاح",
        }

        response = JsonResponse(response_data)
        response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response["Pragma"] = "no-cache"
        return response
        # logger.info(f"[update_order_status] Success response: {response_data}")  # معطل لتجنب الرسائل الكثيرة

    except Exception as e:
        import traceback

        error_trace = traceback.format_exc()
        logger.critical(
            f"[update_order_status] Unexpected error: {str(e)}\n{error_trace}"
        )

        return JsonResponse(
            {
                "success": False,
                "error": "حدث خطأ غير متوقع",
                "exception_type": e.__class__.__name__,
                "details": str(e),
            },
            status=500,
        )


@require_http_methods(["POST"])
@login_required
def update_exit_permit(request, pk):
    """محمي بـ CSRF و login_required"""
    if not request.user.has_perm("manufacturing.change_manufacturingorder"):
        return JsonResponse(
            {"success": False, "error": "ليس لديك صلاحية لتحديث إذن الخروج"}, status=403
        )

    try:
        order = get_object_or_404(ManufacturingOrder, pk=pk)
        exit_permit_number = request.POST.get("exit_permit_number", "").strip()

        if not exit_permit_number:
            return JsonResponse(
                {"success": False, "error": "يرجى إدخال رقم إذن الخروج"}, status=400
            )

        order.exit_permit_number = exit_permit_number
        order.updated_at = timezone.now()
        order.save()

        return JsonResponse(
            {
                "success": True,
                "exit_permit_number": order.exit_permit_number,
                "updated_at": order.updated_at.strftime("%Y-%m-%d %H:%M"),
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


def create_from_order(request, order_id):
    if not request.user.has_perm("manufacturing.add_manufacturingorder"):
        messages.error(request, "ليس لديك صلاحية لإنشاء أمر تصنيع")
        return redirect("manufacturing:order_list")

    try:
        order = get_object_or_404(Order, pk=order_id)

        # Check if manufacturing order already exists for this order
        if ManufacturingOrder.objects.filter(order=order).exists():
            messages.warning(request, "يوجد أمر تصنيع مسبقاً لهذا الطلب")
            return redirect("manufacturing:order_list")

        # Create manufacturing order
        manufacturing_order = ManufacturingOrder.objects.create(
            order=order,
            order_type=(
                "installation" if order.order_type == "installation" else "detail"
            ),
            contract_number=order.contract_number,
            order_date=(
                order.order_date.date() if order.order_date else timezone.now().date()
            ),
            expected_delivery_date=order.expected_delivery_date,
            created_by=request.user,
        )

        # تم حذف نظام الإشعارات

        # Add order items to manufacturing order
        for item in order.items.all():
            ManufacturingOrderItem.objects.create(
                manufacturing_order=manufacturing_order,
                product_name=item.product.name if item.product else f"منتج #{item.id}",
                quantity=item.quantity or 1,
                specifications=getattr(item, "specifications", None)
                or getattr(item, "notes", None)
                or "لا توجد مواصفات",
            )

        messages.success(request, "تم إنشاء أمر التصنيع بنجاح")
        return redirect("manufacturing:order_update", pk=manufacturing_order.pk)

    except Exception as e:
        messages.error(request, f"حدث خطأ أثناء إنشاء أمر التصنيع: {str(e)}")
        return redirect("orders:order_detail", pk=order_id)


def print_manufacturing_order(request, pk):
    """Generate a PDF for the manufacturing order"""
    manufacturing_order = get_object_or_404(ManufacturingOrder, pk=pk)

    # Get company name from settings
    from django.conf import settings

    company_name = getattr(settings, "COMPANY_NAME", "شركتنا")

    # Render the HTML template
    html_string = render_to_string(
        "manufacturing/print/manufacturing_order.html",
        {
            "manufacturing_order": manufacturing_order,
            "items": manufacturing_order.items.all(),
            "now": timezone.now(),
            "company_name": company_name,
        },
    )

    # Create PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri("/"))
    pdf = html.write_pdf()

    # Create HTTP response with PDF
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = (
        f"filename=manufacturing_order_{manufacturing_order.id}.pdf"
    )
    return response


class DashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = "manufacturing/dashboard.html"
    permission_required = "manufacturing.view_manufacturingorder"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get all manufacturing orders (بدون فلترة افتراضية)
        orders = ManufacturingOrder.objects.all().select_related(
            "order", "order__customer"
        )

        # Calculate status counts
        status_counts = (
            orders.values("status").annotate(count=Count("status")).order_by("status")
        )

        # Prepare status data for the chart
        status_data = {
            "labels": [],
            "data": [],
            "colors": [],
        }

        status_colors = {
            "pending": "#ffc107",  # Yellow
            "in_progress": "#0dcaf0",  # Blue
            "completed": "#198754",  # Green
            "delivered": "#0d6efd",  # Primary blue
            "cancelled": "#dc3545",  # Red
        }

        for status in status_counts:
            status_display = dict(ManufacturingOrder.STATUS_CHOICES).get(
                status["status"], status["status"]
            )
            status_data["labels"].append(status_display)
            status_data["data"].append(status["count"])
            status_data["colors"].append(status_colors.get(status["status"], "#6c757d"))

        # Get monthly order counts
        monthly_orders = (
            orders.annotate(
                month=ExtractMonth("order_date"), year=ExtractYear("order_date")
            )
            .values("year", "month")
            .annotate(total=Count("id"))
            .order_by("year", "month")
        )

        # Prepare monthly data for the chart
        monthly_data = {
            "labels": [],
            "data": [],
        }

        # Get month names in Arabic
        month_names = [
            "يناير",
            "فبراير",
            "مارس",
            "أبريل",
            "مايو",
            "يونيو",
            "يوليو",
            "أغسطس",
            "سبتمبر",
            "أكتوبر",
            "نوفمبر",
            "ديسمبر",
        ]

        for item in monthly_orders:
            month_name = f"{month_names[item['month']-1]} {item['year']}"
            monthly_data["labels"].append(month_name)
            monthly_data["data"].append(item["total"])

        # Get recent orders
        recent_orders = orders.order_by("-order_date")[:5]

        # Get orders by type
        orders_by_type = orders.values("order_type").annotate(
            count=Count("id"), total=Sum("order__total_amount")
        )

        context.update(
            {
                "status_data": json.dumps(status_data),
                "monthly_data": json.dumps(monthly_data),
                "recent_orders": recent_orders,
                "orders_by_type": orders_by_type,
                "total_orders": orders.count(),
                "pending_orders": orders.filter(status="pending").count(),
                "in_progress_orders": orders.filter(status="in_progress").count(),
                "completed_orders": orders.filter(
                    status__in=["ready_install", "completed"]
                ).count(),
                "delivered_orders": orders.filter(status="delivered").count(),
                "cancelled_orders": orders.filter(status="cancelled").count(),
                "total_revenue": sum(
                    order.order.total_amount
                    for order in orders
                    if order.order and order.order.total_amount
                ),
            }
        )

        return context


def dashboard_data(request):
    """
    API endpoint to get dashboard data for AJAX requests
    """
    if not request.user.has_perm("manufacturing.view_manufacturingorder"):
        return JsonResponse({"error": "Permission denied"}, status=403)

    # Get date range from request or use default (last 30 days)
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    # Get all manufacturing orders
    orders = ManufacturingOrder.objects.filter(
        order_date__range=(start_date, end_date)
    ).select_related("order", "order__customer")

    # Calculate status counts
    status_counts = (
        orders.values("status").annotate(count=Count("status")).order_by("status")
    )

    # Prepare response data
    data = {
        "success": True,
        "total_orders": orders.count(),
        "pending_orders": orders.filter(status="pending").count(),
        "in_progress_orders": orders.filter(status="in_progress").count(),
        "completed_orders": orders.filter(
            status__in=["ready_install", "completed"]
        ).count(),
        "delivered_orders": orders.filter(status="delivered").count(),
        "cancelled_orders": orders.filter(status="cancelled").count(),
        "status_data": {item["status"]: item["count"] for item in status_counts},
        "last_updated": timezone.now().isoformat(),
    }

    return JsonResponse(data)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

from .models import ManufacturingOrder


@require_POST
def update_approval_status(request, pk):
    """
    API endpoint to approve or reject manufacturing orders
    """
    import logging

    logger = logging.getLogger(__name__)

    try:
        # Check authentication
        if not request.user.is_authenticated:
            logger.warning("Unauthenticated approval attempt")
            return JsonResponse(
                {"success": False, "error": "يجب تسجيل الدخول أولاً"}, status=401
            )

        # Check approval permission - only users with specific permission can approve
        if not (
            request.user.has_perm("manufacturing.can_approve_orders")
            or request.user.is_superuser
        ):
            logger.warning(f"Permission denied for user {request.user.username}")
            return JsonResponse(
                {
                    "success": False,
                    "error": "ليس لديك صلاحية الموافقة على أوامر التصنيع",
                },
                status=403,
            )

        # Get the manufacturing order
        try:
            order = ManufacturingOrder.objects.select_related(
                "order", "order__created_by"
            ).get(pk=pk)
        except ManufacturingOrder.DoesNotExist:
            logger.error(f"Manufacturing order {pk} not found")
            return JsonResponse(
                {"success": False, "error": "لم يتم العثور على الطلب."}, status=404
            )

        # Parse request data
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            logger.error("Invalid JSON in request body")
            return JsonResponse(
                {"success": False, "error": "بيانات غير صالحة"}, status=400
            )

        action = data.get("action")

        # Validate action
        if action not in ["approve", "reject"]:
            return JsonResponse(
                {"success": False, "error": "إجراء غير صالح."}, status=400
            )

        # Check if order is still pending approval (allow rejection from other statuses)
        if order.status != "pending_approval" and action == "approve":
            return JsonResponse(
                {
                    "success": False,
                    "error": f'لا يمكن الموافقة على الطلب. الحالة الحالية: {order.get_status_display()}. يمكن الموافقة فقط على الطلبات في حالة "قيد الموافقة".',
                },
                status=400,
            )

        # For rejection, check if order is in a state that allows rejection
        if action == "reject":
            # يُسمح بالرفض فقط من حالات معينة
            allowed_rejection_statuses = ["pending_approval", "pending"]

            if order.status == "rejected":
                return JsonResponse(
                    {"success": False, "error": "هذا الطلب مرفوض بالفعل."}, status=400
                )

            if order.status not in allowed_rejection_statuses:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "لا يمكن رفض الطلب بعد دخوله مرحلة التنفيذ. يُسمح بالرفض فقط قبل بدء التصنيع.",
                    },
                    status=400,
                )

        with transaction.atomic():
            if action == "approve":
                old_status = order.status
                order.status = (
                    "pending"  # The manufacturing process itself is now 'pending'
                )
                order.rejection_reason = None
                # Set the user who changed the status for the signal handler
                order._changed_by = request.user
                order.save()

                # إنشاء سجل تغيير الحالة
                try:
                    from orders.models import OrderStatusLog

                    OrderStatusLog.objects.create(
                        order=order.order,
                        old_status=old_status,
                        new_status="pending",
                        changed_by=request.user,
                        change_type="manufacturing",
                        notes=f"تمت الموافقة على أمر التصنيع",
                    )
                except Exception as e:
                    logger.error(f"Error creating status log: {str(e)}")

                # تم حذف نظام الإشعارات

                # Update order tracking status to in_progress
                if order.order:
                    order.order.tracking_status = "in_progress"
                    order.order.save(update_fields=["tracking_status"])

                logger.info(f"Order {pk} approved by {request.user.username}")
                return JsonResponse(
                    {"success": True, "message": "تمت الموافقة على الطلب وبدء التصنيع."}
                )

            elif action == "reject":
                reason = data.get("reason", "").strip()
                if not reason:
                    return JsonResponse(
                        {"success": False, "error": "سبب الرفض مطلوب."}, status=400
                    )

                old_status = order.status
                order.status = "rejected"
                order.rejection_reason = reason
                # Set the user who changed the status for the signal handler
                order._changed_by = request.user
                order.save()

                # إنشاء سجل تغيير الحالة
                try:
                    from orders.models import OrderStatusLog

                    OrderStatusLog.objects.create(
                        order=order.order,
                        old_status=old_status,
                        new_status="rejected",
                        changed_by=request.user,
                        change_type="manufacturing",
                        notes=f"تم رفض أمر التصنيع - السبب: {reason}",
                    )
                except Exception as e:
                    logger.error(f"Error creating status log: {str(e)}")

                # Revert original order status if it was set to 'factory'
                original_order = order.order
                if (
                    hasattr(original_order, "tracking_status")
                    and original_order.tracking_status == "factory"
                ):
                    original_order.tracking_status = "processing"  # Or 'pending'
                    original_order.save(update_fields=["tracking_status"])

                # تم حذف نظام الإشعارات

                # Update order tracking status to rejected
                if order.order:
                    order.order.tracking_status = "rejected"
                    order.order.save(update_fields=["tracking_status"])

                logger.info(
                    f"Order {pk} rejected by {request.user.username}, reason: {reason}"
                )
                return JsonResponse(
                    {"success": True, "message": "تم رفض الطلب وإرسال إشعار للمستخدم."}
                )

    except Exception as e:
        logger.error(f"Unexpected error in update_approval_status: {e}")
        return JsonResponse(
            {"success": False, "error": "حدث خطأ غير متوقع. يرجى المحاولة مرة أخرى."},
            status=500,
        )


def get_order_details(request, pk):
    """
    Get manufacturing order details including rejection reply
    """
    try:
        if not request.user.is_authenticated:
            return JsonResponse(
                {"success": False, "error": "يجب تسجيل الدخول أولاً"}, status=401
            )

        order = ManufacturingOrder.objects.select_related(
            "order", "order__created_by"
        ).get(pk=pk)

        # Check permission
        if not (
            request.user.has_perm("manufacturing.view_manufacturingorder")
            or request.user.is_superuser
        ):
            return JsonResponse(
                {"success": False, "error": "ليس لديك صلاحية لعرض تفاصيل هذا الطلب"},
                status=403,
            )

        # Prepare order data
        order_data = {
            "id": order.id,
            "status": order.status,
            "status_display": order.get_status_display(),
            "rejection_reason": order.rejection_reason,
            "rejection_reply": order.rejection_reply,
            "rejection_reply_date": (
                order.rejection_reply_date.isoformat()
                if order.rejection_reply_date
                else None
            ),
            "has_rejection_reply": order.has_rejection_reply,
        }

        return JsonResponse({"success": True, "order": order_data})

    except ManufacturingOrder.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "لم يتم العثور على أمر التصنيع"}, status=404
        )
    except Exception as e:
        logger.error(f"Error getting order details: {e}")
        return JsonResponse(
            {"success": False, "error": "حدث خطأ غير متوقع"}, status=500
        )


@require_POST
@login_required
def send_reply(request, pk):
    """
    Send reply to rejection notification - محمي بـ CSRF و login_required
    """
    from django.utils import timezone

    try:
        # Check authentication
        if not request.user.is_authenticated:
            return JsonResponse(
                {"success": False, "error": "يجب تسجيل الدخول أولاً"}, status=401
            )

        order = ManufacturingOrder.objects.select_related(
            "order", "order__created_by"
        ).get(pk=pk)

        # Debug logging
        logger.info(f"User {request.user.username} trying to reply to order {pk}")
        logger.info(
            f"Order created_by: {order.order.created_by if order.order else None}"
        )
        logger.info(f"User is_superuser: {request.user.is_superuser}")
        logger.info(f"User is_staff: {request.user.is_staff}")

        # Check if user is the order creator, has permission, or is staff
        can_reply = (
            request.user.is_superuser
            or request.user.is_staff
            or (order.order and order.order.created_by == request.user)
            or request.user.has_perm("manufacturing.can_approve_orders")
            or request.user.has_perm("orders.change_order")
        )

        logger.info(f"Can reply: {can_reply}")

        if not can_reply:
            return JsonResponse(
                {"success": False, "error": "ليس لديك صلاحية للرد على هذا الطلب"},
                status=403,
            )

        # Parse JSON request data
        try:
            data = json.loads(request.body)
            reply_message = data.get("reply_message", "").strip()
        except json.JSONDecodeError:
            return JsonResponse(
                {"success": False, "error": "بيانات غير صالحة"}, status=400
            )

        if not reply_message:
            return JsonResponse(
                {"success": False, "error": "نص الرد مطلوب"}, status=400
            )

        # Save reply to the order
        order.rejection_reply = reply_message
        order.rejection_reply_date = timezone.now()
        order.has_rejection_reply = True
        order.save(
            update_fields=[
                "rejection_reply",
                "rejection_reply_date",
                "has_rejection_reply",
            ]
        )

        # إرسال إشعار للمدير أو المستخدمين المخولين بالموافقة
        from django.contrib.auth import get_user_model
        from django.db import models

        User = get_user_model()

        # العثور على المستخدمين المخولين بالموافقة
        approval_users = User.objects.filter(
            models.Q(is_superuser=True)
            | models.Q(user_permissions__codename="can_approve_orders")
        ).distinct()

        # إرسال إشعار لكل مستخدم مخول
        for user in approval_users:
            try:
                customer_name = (
                    order.order.created_by.get_full_name()
                    or order.order.created_by.username
                )
                message = (
                    f"رد من {customer_name}:\n\n{reply_message}\n\n"
                    f"أمر التصنيع: #{order.id}\n"
                    f"سبب الرفض الأصلي: {order.rejection_reason}"
                )

                Notification.objects.create(
                    recipient=user,
                    title=f"رد على رفض أمر التصنيع #{order.id}",
                    message=message,
                    priority="medium",
                    link=order.get_absolute_url(),
                )
                logger.info(f"Reply notification sent to {user.username}")
            except Exception as e:
                logger.error(
                    f"Failed to create reply notification for " f"{user.username}: {e}"
                )

        return JsonResponse({"success": True, "message": "تم إرسال الرد بنجاح للإدارة"})

    except ManufacturingOrder.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "لم يتم العثور على أمر التصنيع"}, status=404
        )
    except Exception as e:
        logger.error(f"Unexpected error in send_reply: {e}")
        return JsonResponse(
            {"success": False, "error": f"حدث خطأ غير متوقع: {str(e)}"}, status=500
        )


@require_POST
def send_reply_to_rejection_log(request, log_id):
    """
    إرسال رد على سجل رفض محدد
    """
    from django.utils import timezone

    from .models import ManufacturingRejectionLog

    try:
        # التحقق من المصادقة
        if not request.user.is_authenticated:
            return JsonResponse(
                {"success": False, "error": "يجب تسجيل الدخول أولاً"}, status=401
            )

        # الحصول على سجل الرفض
        rejection_log = ManufacturingRejectionLog.objects.select_related(
            "manufacturing_order",
            "manufacturing_order__order",
            "manufacturing_order__order__created_by",
        ).get(pk=log_id)

        order = rejection_log.manufacturing_order.order

        # التحقق من صلاحية الرد
        can_reply = (
            request.user.is_superuser
            or request.user.is_staff
            or (order and order.created_by == request.user)
            or request.user.has_perm("manufacturing.can_approve_orders")
            or request.user.has_perm("orders.change_order")
        )

        if not can_reply:
            return JsonResponse(
                {"success": False, "error": "ليس لديك صلاحية للرد على هذا الطلب"},
                status=403,
            )

        # التحقق من عدم وجود رد سابق
        if rejection_log.reply_message:
            return JsonResponse(
                {"success": False, "error": "تم الرد على هذا الرفض مسبقاً"}, status=400
            )

        # تحليل البيانات
        try:
            data = json.loads(request.body)
            reply_message = data.get("reply_message", "").strip()
        except json.JSONDecodeError:
            return JsonResponse(
                {"success": False, "error": "بيانات غير صالحة"}, status=400
            )

        if not reply_message:
            return JsonResponse(
                {"success": False, "error": "نص الرد مطلوب"}, status=400
            )

        # حفظ الرد
        rejection_log.reply_message = reply_message
        rejection_log.replied_at = timezone.now()
        rejection_log.replied_by = request.user
        rejection_log.save(update_fields=["reply_message", "replied_at", "replied_by"])

        # إرسال إشعار للمدير أو المستخدمين المخولين
        from django.contrib.auth import get_user_model
        from django.db import models

        User = get_user_model()

        approval_users = User.objects.filter(
            models.Q(is_superuser=True)
            | models.Q(user_permissions__codename="can_approve_orders")
        ).distinct()

        for user in approval_users:
            try:
                customer_name = (
                    (order.created_by.get_full_name() or order.created_by.username)
                    if order.created_by
                    else "غير معروف"
                )
                message = (
                    f"رد من {customer_name}:\n\n{reply_message}\n\n"
                    f"أمر التصنيع: {rejection_log.manufacturing_order.manufacturing_code}\n"
                    f"سبب الرفض الأصلي: {rejection_log.rejection_reason}"
                )

                Notification.objects.create(
                    recipient=user,
                    title=f"رد على رفض أمر التصنيع {rejection_log.manufacturing_order.manufacturing_code}",
                    message=message,
                    priority="medium",
                    link=rejection_log.manufacturing_order.get_absolute_url(),
                )
                logger.info(f"Reply notification sent to {user.username}")
            except Exception as e:
                logger.error(
                    f"Failed to create reply notification for {user.username}: {e}"
                )

        return JsonResponse({"success": True, "message": "تم إرسال الرد بنجاح للإدارة"})

    except ManufacturingRejectionLog.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "لم يتم العثور على سجل الرفض"}, status=404
        )
    except Exception as e:
        logger.error(f"Unexpected error in send_reply_to_rejection_log: {e}")
        return JsonResponse(
            {"success": False, "error": f"حدث خطأ غير متوقع: {str(e)}"}, status=500
        )


@require_POST
def re_approve_after_reply(request, pk):
    """
    Re-approve manufacturing order after reply to rejection
    """
    import logging

    from django.db import transaction

    logger = logging.getLogger(__name__)

    try:
        # Check authentication
        if not request.user.is_authenticated:
            return JsonResponse(
                {"success": False, "error": "يجب تسجيل الدخول أولاً"}, status=401
            )

        # Check approval permission
        if not (
            request.user.has_perm("manufacturing.can_approve_orders")
            or request.user.is_superuser
        ):
            return JsonResponse(
                {
                    "success": False,
                    "error": "ليس لديك صلاحية الموافقة على أوامر التصنيع",
                },
                status=403,
            )

        order = ManufacturingOrder.objects.select_related(
            "order", "order__created_by"
        ).get(pk=pk)

        # Check if order was rejected and has a reply
        if order.status != "rejected" or not order.has_rejection_reply:
            return JsonResponse(
                {
                    "success": False,
                    "error": "يمكن الموافقة فقط على الطلبات المرفوضة التي تم الرد عليها",
                },
                status=400,
            )

        with transaction.atomic():
            # Reset rejection status and approve
            old_status = order.status
            order.status = "pending"
            # Set the user who changed the status for the signal handler
            order._changed_by = request.user
            order.save(update_fields=["status"])

            # إنشاء سجل تغيير الحالة
            try:
                from orders.models import OrderStatusLog

                OrderStatusLog.objects.create(
                    order=order.order,
                    old_status=old_status,
                    new_status="pending",
                    changed_by=request.user,
                    change_type="manufacturing",
                    notes=f"تمت الموافقة على أمر التصنيع بعد الرد على الرفض",
                )
            except Exception as e:
                logger.error(f"Error creating status log: {str(e)}")

            # Update order tracking status to in_progress
            if order.order:
                order.order.tracking_status = "factory"
                order.order.save(update_fields=["tracking_status"])

            # Create notification for the order creator
            if order.order and order.order.created_by:
                try:
                    customer_name = order.order.customer.name
                    title = f"تمت الموافقة على طلبك بعد المراجعة - {customer_name}"
                    message = (
                        f"تمت الموافقة على أمر التصنيع للعميل "
                        f"{customer_name} - "
                        f"الطلب #{order.order.order_number}\n"
                        f"دخل الطلب مرحلة التصنيع. "
                        f"رقم أمر التصنيع #{order.pk}."
                    )

                    Notification.objects.create(
                        recipient=order.order.created_by,
                        title=title,
                        message=message,
                        priority="high",
                        link=order.get_absolute_url(),
                    )
                    logger.info(
                        f"Re-approval notification sent to "
                        f"{order.order.created_by.username}"
                    )
                except Exception as e:
                    logger.error(f"Failed to create re-approval notification: {e}")

            logger.info(f"Order {pk} re-approved by {request.user.username}")
            return JsonResponse(
                {
                    "success": True,
                    "message": "تمت الموافقة على الطلب وبدء التصنيع بعد المراجعة.",
                }
            )

    except ManufacturingOrder.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "لم يتم العثور على أمر التصنيع"}, status=404
        )
    except Exception as e:
        logger.error(f"Unexpected error in re_approve_after_reply: {e}")
        return JsonResponse(
            {"success": False, "error": f"حدث خطأ غير متوقع: {str(e)}"}, status=500
        )


# Views للوصول بالكود بدلاً من ID
from django.contrib.auth.decorators import login_required


@login_required
def manufacturing_order_detail_by_code(request, manufacturing_code):
    """عرض تفاصيل أمر التصنيع باستخدام كود التصنيع"""
    # البحث بطريقة محسنة للأداء
    if "-M" in manufacturing_code:
        order_number = manufacturing_code.replace("-M", "")
        manufacturing_order = get_object_or_404(
            ManufacturingOrder.objects.select_related("order", "order__customer"),
            order__order_number=order_number,
        )
    else:
        # للأكواد القديمة
        manufacturing_id = manufacturing_code.replace("#", "").replace("-M", "")
        manufacturing_order = get_object_or_404(
            ManufacturingOrder.objects.select_related("order", "order__customer"),
            id=manufacturing_id,
        )

    return ManufacturingOrderDetailView.as_view()(request, pk=manufacturing_order.pk)


@login_required
def manufacturing_order_detail_redirect(request, pk):
    """إعادة توجيه من ID إلى كود التصنيع"""
    manufacturing_order = get_object_or_404(ManufacturingOrder, pk=pk)
    return redirect(
        "manufacturing:order_detail_by_code",
        manufacturing_code=manufacturing_order.manufacturing_code,
    )


# نظام استلام الأقمشة في المصنع - تم نقله للأسفل


@login_required
def receive_fabric_item(request, item_id):
    """استلام قماش عنصر واحد"""
    if request.method == "POST":
        try:
            item = get_object_or_404(ManufacturingOrderItem, pk=item_id)

            data = json.loads(request.body)

            # طباعة البيانات المستلمة للتشخيص
            print(f"DEBUG - Received data: {data}")

            bag_number = data.get("bag_number", "").strip()
            use_previous_bag = data.get("use_previous_bag", False)
            notes = data.get("notes", "").strip()

            print(f"DEBUG - bag_number: '{bag_number}'")
            print(f"DEBUG - use_previous_bag: {use_previous_bag}")

            # إذا كان المستخدم يريد استخدام شنطة سابقة
            if use_previous_bag:
                previous_bag_number = data.get("previous_bag_number", "").strip()
                print(f"DEBUG - previous_bag_number: '{previous_bag_number}'")
                if not previous_bag_number:
                    return JsonResponse(
                        {"success": False, "message": "يجب اختيار رقم الشنطة السابقة"}
                    )
                bag_number = previous_bag_number
            else:
                # استخدام رقم الشنطة المُرسل (المُولد تلقائياً)
                if not bag_number:
                    print(f"DEBUG - bag_number is empty!")
                    return JsonResponse(
                        {"success": False, "message": "رقم الشنطة مطلوب"}
                    )

            print(f"DEBUG - Final bag_number before save: '{bag_number}'")

            # تعيين العنصر كمستلم
            item.mark_fabric_received(
                bag_number=bag_number, user=request.user, notes=notes
            )

            # التحقق من الحفظ
            item.refresh_from_db()
            print(f"DEBUG - bag_number after save: '{item.bag_number}'")

            return JsonResponse(
                {
                    "success": True,
                    "message": f"تم استلام القماش بنجاح - رقم الشنطة: {bag_number}",
                    "bag_number": bag_number,
                }
            )

        except Exception as e:
            import traceback

            print(f"DEBUG - Exception: {str(e)}")
            print(f"DEBUG - Traceback: {traceback.format_exc()}")
            return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})

    return JsonResponse({"success": False, "message": "طريقة غير مدعومة"})


@login_required
def get_bag_number_data(request):
    """الحصول على رقم الشنطة التالي وقائمة الأرقام المتاحة لإعادة الاستخدام"""
    try:
        next_bag = ManufacturingOrderItem.get_next_bag_number()
        available_bags = ManufacturingOrderItem.get_available_bag_numbers()

        return JsonResponse(
            {
                "success": True,
                "next_bag_number": next_bag,
                "available_bags": available_bags,
            }
        )
    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


@login_required
def bulk_receive_fabric(request, order_id):
    """استلام جميع أقمشة الطلب بنفس رقم الشنطة"""
    if request.method == "POST":
        try:
            manufacturing_order = get_object_or_404(ManufacturingOrder, pk=order_id)

            data = json.loads(request.body)
            bag_number = data.get("bag_number", "").strip()
            notes = data.get("notes", "").strip()

            if not bag_number:
                return JsonResponse({"success": False, "message": "رقم الشنطة مطلوب"})

            # استلام جميع العناصر الجاهزة للاستلام
            items_to_receive = manufacturing_order.items.filter(
                receiver_name__isnull=False,
                permit_number__isnull=False,
                fabric_received=False,
            )

            received_count = 0
            for item in items_to_receive:
                item.mark_fabric_received(
                    bag_number=bag_number, user=request.user, notes=notes
                )
                received_count += 1

            return JsonResponse(
                {"success": True, "message": f"تم استلام {received_count} عنصر بنجاح"}
            )

        except Exception as e:
            return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})

    return JsonResponse({"success": False, "message": "طريقة غير مدعومة"})


@login_required
def fabric_receipt_status_api(request, order_id):
    """API للحصول على حالة استلام الأقمشة لأمر تصنيع"""
    try:
        manufacturing_order = get_object_or_404(ManufacturingOrder, pk=order_id)

        items_data = []
        for item in manufacturing_order.items.all():
            items_data.append(
                {
                    "id": item.id,
                    "product_name": item.product_name,
                    "quantity": str(item.quantity),
                    "receiver_name": item.receiver_name,
                    "permit_number": item.permit_number,
                    "cutting_date": (
                        item.cutting_date.isoformat() if item.cutting_date else None
                    ),
                    "delivery_date": (
                        item.delivery_date.isoformat() if item.delivery_date else None
                    ),
                    "bag_number": item.bag_number,
                    "fabric_received": item.fabric_received,
                    "fabric_received_date": (
                        item.fabric_received_date.isoformat()
                        if item.fabric_received_date
                        else None
                    ),
                    "fabric_status": item.get_fabric_status_display(),
                    "fabric_status_color": item.get_fabric_status_color(),
                    "has_cutting_data": item.has_cutting_data,
                }
            )

        return JsonResponse(
            {
                "success": True,
                "order": {
                    "id": manufacturing_order.id,
                    "customer_name": (
                        manufacturing_order.order.customer.name
                        if manufacturing_order.order
                        else ""
                    ),
                    "contract_number": (
                        manufacturing_order.order.contract_number
                        if manufacturing_order.order
                        else ""
                    ),
                    "total_items": manufacturing_order.total_items_count,
                    "received_items": manufacturing_order.received_items_count,
                    "pending_items": manufacturing_order.pending_items_count,
                    "completion_percentage": manufacturing_order.items_completion_percentage,
                    "is_all_received": manufacturing_order.is_all_items_received,
                    "status_display": manufacturing_order.get_items_status_display(),
                    "status_color": manufacturing_order.get_items_status_color(),
                },
                "items": items_data,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


class FabricReceiptView(LoginRequiredMixin, TemplateView):
    """صفحة استلام الأقمشة في المصنع - عرض العناصر الفردية المقطوعة"""

    template_name = "manufacturing/fabric_receipt.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from django.db.models import Q

        from cutting.models import CuttingOrderItem

        # الحصول على العناصر المقطوعة الفردية الجاهزة للاستلام
        # بغض النظر عن حالة أمر التقطيع الكامل
        cutting_items = (
            CuttingOrderItem.objects.select_related(
                "cutting_order",
                "cutting_order__order",
                "cutting_order__order__customer",
                "cutting_order__warehouse",
                "order_item",
                "order_item__product",
            )
            .filter(
                # العناصر المكتملة فقط
                status="completed",
                # لم يتم استلامها بعد (في نظام التقطيع)
                fabric_received=False,
            )
            .exclude(
                # استبعاد العناصر بدون اسم مستلم أو رقم إذن
                receiver_name__isnull=True
            )
            .exclude(receiver_name="")
            .exclude(permit_number__isnull=True)
            .exclude(permit_number="")
            .exclude(
                # استبعاد العناصر التي تم استلامها بالفعل في نظام التصنيع
                manufacturing_items__fabric_received=True
            )
            .filter(
                # تضمين جميع أنواع الطلبات
                Q(cutting_order__order__selected_types__icontains="manufacturing")
                | Q(cutting_order__order__selected_types__icontains="installation")
                | Q(cutting_order__order__selected_types__icontains="tailoring")
                | Q(cutting_order__order__selected_types__icontains="accessory")
            )
            .distinct()
            .order_by("-cutting_date", "-cutting_order__created_at")
        )

        # البحث (يشمل جميع أرقام الفواتير والعقود ورقم الإذن)
        search = self.request.GET.get("search", "").strip()
        if search:
            cutting_items = cutting_items.filter(
                Q(cutting_order__cutting_code__icontains=search)
                | Q(cutting_order__order__contract_number__icontains=search)
                | Q(cutting_order__order__contract_number_2__icontains=search)
                | Q(cutting_order__order__contract_number_3__icontains=search)
                | Q(cutting_order__order__invoice_number__icontains=search)
                | Q(cutting_order__order__invoice_number_2__icontains=search)
                | Q(cutting_order__order__invoice_number_3__icontains=search)
                | Q(cutting_order__order__customer__name__icontains=search)
                | Q(order_item__product__name__icontains=search)
                | Q(permit_number__icontains=search)
                | Q(receiver_name__icontains=search)
            )

        # تجميع العناصر حسب أمر التقطيع لعرض أفضل
        orders_dict = {}
        for item in cutting_items:
            order_id = item.cutting_order.id
            if order_id not in orders_dict:
                orders_dict[order_id] = {
                    "cutting_order": item.cutting_order,
                    "items": [],
                }
            orders_dict[order_id]["items"].append(item)

        context["cutting_orders_with_items"] = list(orders_dict.values())
        context["total_available_items"] = cutting_items.count()
        context["total_available_orders"] = len(orders_dict)
        context["search_query"] = search

        return context


@login_required
def receive_cutting_order_for_manufacturing(request, cutting_order_id):
    """استلام أمر تقطيع للتصنيع عبر AJAX"""
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "طريقة غير مدعومة"})

    try:
        from cutting.models import CuttingOrder

        cutting_order = get_object_or_404(CuttingOrder, pk=cutting_order_id)

        # التحقق من وجود عناصر جاهزة للاستلام
        ready_items = cutting_order.items.filter(
            status="completed",
            receiver_name__isnull=False,
            permit_number__isnull=False,
            fabric_received=False,
        )

        if not ready_items.exists():
            return JsonResponse(
                {
                    "success": False,
                    "message": "لا توجد عناصر جاهزة للاستلام في هذا الأمر",
                }
            )

        data = json.loads(request.body)
        bag_number = data.get("bag_number", "")
        received_by_name = data.get("received_by_name", "")
        notes = data.get("notes", "")

        if not received_by_name:
            return JsonResponse({"success": False, "message": "يجب إدخال اسم المستلم"})

        # إنشاء استلام الأقمشة
        fabric_receipt = FabricReceipt.objects.create(
            order=cutting_order.order,
            cutting_order=cutting_order,
            receipt_type="cutting_order",
            permit_number=ready_items.first().permit_number,
            bag_number=bag_number,
            received_by_name=received_by_name,
            received_by=request.user,
            notes=notes,
        )

        # إنشاء عناصر الاستلام وتحديث حالة العناصر
        items_count = 0
        for item in ready_items:
            FabricReceiptItem.objects.create(
                fabric_receipt=fabric_receipt,
                cutting_item=item,
                order_item=item.order_item,
                quantity_received=item.order_item.quantity + item.additional_quantity,
                item_notes=item.notes or "",
            )

            # تحديث حالة العنصر
            item.fabric_received = True
            item.save()
            items_count += 1

        return JsonResponse(
            {
                "success": True,
                "message": f"تم استلام {items_count} عنصر بنجاح",
                "receipt_id": fabric_receipt.id,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


@login_required
def get_cutting_order_data(request, cutting_order_id):
    """جلب بيانات أمر التقطيع للاستلام"""
    try:
        from cutting.models import CuttingOrder

        cutting_order = get_object_or_404(CuttingOrder, pk=cutting_order_id)

        # الحصول على العناصر المكتملة والجاهزة للاستلام
        completed_items = cutting_order.items.filter(
            status="completed",
            receiver_name__isnull=False,
            permit_number__isnull=False,
            fabric_received=False,
        )

        if completed_items.exists():
            # جلب البيانات من أول عنصر (عادة تكون نفس البيانات لجميع العناصر في نفس الأمر)
            first_item = completed_items.first()

            return JsonResponse(
                {
                    "success": True,
                    "permit_number": first_item.permit_number,
                    "receiver_name": first_item.receiver_name,
                    "cutting_code": cutting_order.cutting_code,
                    "customer_name": cutting_order.order.customer.name,
                    "items_count": completed_items.count(),
                }
            )
        else:
            return JsonResponse(
                {
                    "success": False,
                    "message": "لا توجد عناصر جاهزة للاستلام في هذا الأمر",
                }
            )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


# تم حذف الـ view المكرر - استخدم الـ view في السطر 2290


@require_http_methods(["POST"])
def receive_all_fabric_items(request, order_id):
    """استلام جميع عناصر أمر التصنيع الجاهزة"""
    try:
        order = get_object_or_404(ManufacturingOrder, id=order_id)

        # الحصول على العناصر الجاهزة للاستلام
        ready_items = order.items.filter(
            receiver_name__isnull=False,
            permit_number__isnull=False,
            fabric_received=False,
        )

        if not ready_items.exists():
            return JsonResponse(
                {"success": False, "message": "لا توجد عناصر جاهزة للاستلام"}
            )

        # استلام جميع العناصر
        received_count = 0
        for item in ready_items:
            item.fabric_received = True
            item.fabric_received_date = timezone.now()
            item.fabric_received_by = request.user
            item.save()
            received_count += 1

        # إرسال إشعار
        try:
            from notifications.signals import create_notification

            create_notification(
                title="استلام أقمشة جماعي",
                message=f"تم استلام {received_count} عنصر من أمر التصنيع #{order.id}",
                notification_type="fabric_received",
                related_object=order,
                created_by=request.user,
            )
        except:
            pass

        return JsonResponse(
            {
                "success": True,
                "message": f"تم استلام {received_count} عنصر بنجاح",
                "received_count": received_count,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


@require_http_methods(["GET"])
def recent_fabric_receipts(request):
    """الحصول على آخر عمليات استلام الأقمشة"""
    try:
        recent_items = (
            ManufacturingOrderItem.objects.filter(fabric_received=True)
            .select_related("fabric_received_by")
            .order_by("-fabric_received_date")[:10]
        )

        receipts = []
        for item in recent_items:
            receipts.append(
                {
                    "product_name": item.product_name,
                    "receiver_name": item.receiver_name,
                    "permit_number": item.permit_number,
                    "received_date": (
                        item.fabric_received_date.strftime("%d/%m/%Y %H:%M")
                        if item.fabric_received_date
                        else ""
                    ),
                    "received_by": (
                        item.fabric_received_by.get_full_name()
                        if item.fabric_received_by
                        else "غير محدد"
                    ),
                }
            )

        return JsonResponse({"success": True, "receipts": receipts})

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


@require_http_methods(["POST"])
@login_required
def receive_cutting_order(request, cutting_order_id):
    """استلام أمر تقطيع كامل في المصنع"""
    try:
        from cutting.models import CuttingOrder

        cutting_order = get_object_or_404(CuttingOrder, id=cutting_order_id)

        # التحقق من أن أمر التقطيع مكتمل
        if cutting_order.status != "completed":
            return JsonResponse(
                {"success": False, "message": "أمر التقطيع غير مكتمل بعد"}
            )

        # التحقق من أن الطلب ليس نوع "منتجات"
        if "products" in cutting_order.order.get_selected_types_list():
            return JsonResponse(
                {
                    "success": False,
                    "message": "طلبات المنتجات لا تحتاج استلام في المصنع - تكتمل تلقائياً بعد التقطيع",
                }
            )

        data = json.loads(request.body)
        bag_number = data.get("bag_number", "").strip()
        notes = data.get("notes", "").strip()

        if not bag_number:
            return JsonResponse({"success": False, "message": "رقم الشنطة مطلوب"})

        # إنشاء سجل استلام الأقمشة
        fabric_receipt = FabricReceipt.objects.create(
            receipt_type="cutting_order",
            order=cutting_order.order,
            cutting_order=cutting_order,
            bag_number=bag_number,
            received_by=request.user,
            notes=notes,
        )

        # البحث عن أمر تصنيع موجود أو إنشاء جديد
        manufacturing_order, created = ManufacturingOrder.objects.get_or_create(
            order=cutting_order.order,
            defaults={
                "order_date": timezone.now().date(),
                "expected_delivery_date": timezone.now().date() + timedelta(days=7),
                "notes": f"تم إنشاؤه من أمر التقطيع {cutting_order.cutting_code}. {notes}".strip(),
            },
        )

        # ربط استلام الأقمشة بأمر التصنيع
        fabric_receipt.manufacturing_order = manufacturing_order
        fabric_receipt.save()

        # إنشاء عناصر التصنيع وعناصر الاستلام
        created_items = 0
        for cutting_item in cutting_order.items.all():
            # إنشاء عنصر التصنيع
            manufacturing_item = ManufacturingOrderItem.objects.create(
                manufacturing_order=manufacturing_order,
                order_item=cutting_item.order_item,
                product_name=(
                    cutting_item.order_item.product.name
                    if cutting_item.order_item.product
                    else "منتج غير محدد"
                ),
                quantity=cutting_item.order_item.quantity,
                bag_number=bag_number,
                fabric_received=True,
                fabric_received_date=timezone.now(),
                fabric_received_by=request.user,
                fabric_notes=f"مستلم من أمر التقطيع {cutting_order.cutting_code}",
            )

            # إنشاء عنصر الاستلام
            FabricReceiptItem.objects.create(
                fabric_receipt=fabric_receipt,
                order_item=cutting_item.order_item,
                cutting_item=cutting_item,
                product_name=(
                    cutting_item.order_item.product.name
                    if cutting_item.order_item.product
                    else "منتج غير محدد"
                ),
                quantity_received=cutting_item.order_item.quantity,
                item_notes=f"مستلم من عنصر التقطيع",
            )

            created_items += 1

        # تحديث حالة أمر التقطيع لتجنب الاستلام المكرر
        cutting_order.notes = f"{cutting_order.notes}\nتم استلامه في المصنع - أمر تصنيع #{manufacturing_order.id}".strip()
        cutting_order.save()

        # إرسال إشعار
        try:
            from notifications.signals import create_notification

            create_notification(
                title="استلام أمر تقطيع",
                message=f"تم استلام أمر التقطيع {cutting_order.cutting_code} في المصنع وإنشاء أمر تصنيع #{manufacturing_order.id}",
                notification_type="cutting_received",
                related_object=manufacturing_order,
                created_by=request.user,
            )
        except:
            pass

        return JsonResponse(
            {
                "success": True,
                "message": f"تم استلام أمر التقطيع بنجاح وإنشاء أمر تصنيع #{manufacturing_order.id} مع {created_items} عنصر",
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


class FabricReceiptDetailView(LoginRequiredMixin, DetailView):
    """عرض تفاصيل استلام الأقمشة"""

    model = None  # سيتم تحديده في get_object
    template_name = "manufacturing/fabric_receipt_detail.html"
    context_object_name = "fabric_receipt"

    def get_object(self, queryset=None):
        receipt_id = self.kwargs.get("receipt_id")
        return get_object_or_404(FabricReceipt, id=receipt_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fabric_receipt = self.get_object()

        # إضافة العناصر
        context["receipt_items"] = fabric_receipt.items.all().select_related(
            "order_item", "cutting_item"
        )

        # إضافة معلومات إضافية
        context["total_items"] = fabric_receipt.items.count()
        context["total_quantity"] = (
            fabric_receipt.items.aggregate(total=models.Sum("quantity_received"))[
                "total"
            ]
            or 0
        )

        return context


class FabricReceiptListView(LoginRequiredMixin, ListView):
    """عرض قائمة استلامات الأقمشة - العناصر المستلمة مع فلاتر شاملة"""

    template_name = "manufacturing/fabric_receipt_list.html"
    context_object_name = "received_items"
    paginate_by = 50

    def get_queryset(self):
        # عرض جميع عناصر التصنيع المستلمة مع تحسين الأداء
        queryset = (
            ManufacturingOrderItem.objects.filter(fabric_received=True)
            .select_related(
                "manufacturing_order",
                "manufacturing_order__order",
                "manufacturing_order__order__customer",
                "fabric_received_by",
            )
            .only(
                # حقول ManufacturingOrderItem
                "id",
                "product_name",
                "quantity",
                "permit_number",
                "receiver_name",
                "bag_number",
                "fabric_received_date",
                "fabric_received",
                "manufacturing_order_id",
                "fabric_received_by_id",
                # حقول ManufacturingOrder
                "manufacturing_order__id",
                # حقول Order
                "manufacturing_order__order__id",
                "manufacturing_order__order__order_number",
                # حقول Customer
                "manufacturing_order__order__customer__id",
                "manufacturing_order__order__customer__name",
                # حقول User
                "fabric_received_by__id",
                "fabric_received_by__first_name",
                "fabric_received_by__last_name",
                "fabric_received_by__username",
            )
        )

        # تطبيق الفلاتر
        # فلتر التاريخ من
        date_from = self.request.GET.get("date_from")
        if date_from:
            queryset = queryset.filter(fabric_received_date__date__gte=date_from)

        # فلتر التاريخ إلى
        date_to = self.request.GET.get("date_to")
        if date_to:
            queryset = queryset.filter(fabric_received_date__date__lte=date_to)

        # فلتر العميل (ID)
        customer_id = self.request.GET.get("customer")
        if customer_id:
            queryset = queryset.filter(
                manufacturing_order__order__customer_id=customer_id
            )

        # فلتر اسم العميل (بحث نصي)
        customer_name = self.request.GET.get("customer_name")
        if customer_name:
            queryset = queryset.filter(
                manufacturing_order__order__customer__name__icontains=customer_name
            )

        # فلتر المنتج (بحث جزئي)
        product_name = self.request.GET.get("product_name")
        if product_name:
            queryset = queryset.filter(product_name__icontains=product_name)

        # فلتر رقم الشنطة
        bag_number = self.request.GET.get("bag_number")
        if bag_number:
            queryset = queryset.filter(bag_number=bag_number)

        # فلتر رقم الإذن
        permit_number = self.request.GET.get("permit_number")
        if permit_number:
            queryset = queryset.filter(permit_number__icontains=permit_number)

        # فلتر المستلم بالتقطيع
        receiver_name = self.request.GET.get("receiver_name")
        if receiver_name:
            queryset = queryset.filter(receiver_name__icontains=receiver_name)

        # فلتر المستلم بالمصنع
        received_by = self.request.GET.get("received_by")
        if received_by:
            queryset = queryset.filter(fabric_received_by_id=received_by)

        # فلتر أمر التصنيع (بحث في رقم الطلب الأصلي)
        manufacturing_code = self.request.GET.get("manufacturing_code")
        if manufacturing_code:
            # إزالة -M من نهاية الكود إذا وجد
            search_code = manufacturing_code.replace("-M", "").replace("-m", "")
            queryset = queryset.filter(
                manufacturing_order__order__order_number__icontains=search_code
            )

        return queryset.order_by("-fabric_received_date")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الحصول على الـ queryset المفلتر
        filtered_queryset = self.get_queryset()

        # إحصائيات عامة (بدون فلترة) - استخدام aggregate لتحسين الأداء
        from django.db.models import Count, Q

        today = timezone.now().date()
        week_ago = today - timedelta(days=7)

        stats = ManufacturingOrderItem.objects.filter(fabric_received=True).aggregate(
            total=Count("id"),
            today=Count("id", filter=Q(fabric_received_date__date=today)),
            this_week=Count("id", filter=Q(fabric_received_date__date__gte=week_ago)),
        )

        context["total_receipts"] = stats["total"] or 0
        context["today_receipts"] = stats["today"] or 0
        context["this_week_receipts"] = stats["this_week"] or 0
        context["daily_average"] = (
            round(stats["this_week"] / 7, 1) if stats["this_week"] > 0 else 0
        )

        # إحصائيات مفلترة
        from django.db.models import Count, Sum

        filtered_stats = filtered_queryset.aggregate(
            total_quantity=Sum("quantity"), total_items=Count("id")
        )
        context["filtered_total_quantity"] = filtered_stats["total_quantity"] or 0
        context["filtered_total_items"] = filtered_stats["total_items"] or 0

        # قائمة المستلمين بالمصنع للفلتر (فقط الذين استلموا فعلياً)
        from accounts.models import User

        context["factory_receivers"] = User.objects.filter(
            id__in=ManufacturingOrderItem.objects.filter(
                fabric_received=True, fabric_received_by__isnull=False
            )
            .values_list("fabric_received_by_id", flat=True)
            .distinct()
        ).order_by("first_name")

        # قائمة أرقام الشنطات المتاحة (محدودة لتحسين الأداء)
        context["bag_numbers"] = (
            ManufacturingOrderItem.objects.filter(
                fabric_received=True, bag_number__isnull=False
            )
            .exclude(bag_number="")
            .values_list("bag_number", flat=True)
            .distinct()[:100]
        )  # أول 100 رقم فقط

        # قائمة خطوط الإنتاج النشطة
        context["production_lines"] = ProductionLine.objects.filter(
            is_active=True
        ).order_by("name")

        # الاحتفاظ بقيم الفلاتر
        context["filters"] = {
            "date_from": self.request.GET.get("date_from", ""),
            "date_to": self.request.GET.get("date_to", ""),
            "customer": self.request.GET.get("customer", ""),
            "customer_name": self.request.GET.get("customer_name", ""),
            "product_name": self.request.GET.get("product_name", ""),
            "bag_number": self.request.GET.get("bag_number", ""),
            "permit_number": self.request.GET.get("permit_number", ""),
            "receiver_name": self.request.GET.get("receiver_name", ""),
            "received_by": self.request.GET.get("received_by", ""),
            "manufacturing_code": self.request.GET.get("manufacturing_code", ""),
        }

        return context


@require_http_methods(["POST"])
@login_required
def deliver_to_production_line(request):
    """تسليم عنصر أو جميع عناصر طلب لخط الإنتاج"""
    try:
        import json

        data = json.loads(request.body)

        item_id = data.get("item_id")
        order_id = data.get("order_id")
        production_line_id = data.get("production_line_id")
        delivery_type = data.get("delivery_type", "single")  # single أو all
        notes = data.get("notes", "")

        if not all([item_id, production_line_id]):
            return JsonResponse({"success": False, "message": "بيانات غير كاملة"})

        # التحقق من خط الإنتاج
        try:
            production_line = ProductionLine.objects.get(
                id=production_line_id, is_active=True
            )
        except ProductionLine.DoesNotExist:
            return JsonResponse({"success": False, "message": "خط الإنتاج غير موجود"})

        # تحديد العناصر المراد تسليمها
        if delivery_type == "all" and order_id:
            # تسليم جميع عناصر الطلب المستلمة وغير المسلمة لخطوط إنتاج
            items = ManufacturingOrderItem.objects.filter(
                manufacturing_order_id=order_id,
                fabric_received=True,
                delivered_to_production=False,
            )
        else:
            # تسليم العنصر المحدد فقط
            items = ManufacturingOrderItem.objects.filter(
                id=item_id, fabric_received=True, delivered_to_production=False
            )

        if not items.exists():
            return JsonResponse(
                {"success": False, "message": "لا توجد عناصر متاحة للتسليم"}
            )

        # تحديث العناصر
        from django.utils import timezone

        delivered_count = items.update(
            production_line=production_line,
            delivered_to_production=True,
            production_delivery_date=timezone.now(),
            production_delivered_by=request.user,
            production_delivery_notes=notes,
        )

        return JsonResponse(
            {
                "success": True,
                "message": f"تم تسليم {delivered_count} عنصر بنجاح",
                "delivered_count": delivered_count,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"خطأ: {str(e)}"})


@require_http_methods(["POST"])
@login_required
def cleanup_products_manufacturing_orders(request):
    """حذف أوامر التصنيع الخاطئة لطلبات المنتجات"""
    try:
        # البحث عن أوامر التصنيع لطلبات المنتجات
        products_manufacturing_orders = ManufacturingOrder.objects.filter(
            order__selected_types__contains=["products"]
        )

        deleted_count = products_manufacturing_orders.count()

        # حذف أوامر التصنيع
        products_manufacturing_orders.delete()

        return JsonResponse(
            {
                "success": True,
                "message": f"تم حذف {deleted_count} أمر تصنيع خاطئ لطلبات المنتجات",
                "deleted_count": deleted_count,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


@require_http_methods(["POST"])
@login_required
def fix_manufacturing_order_items(request):
    """إصلاح العناصر المفقودة في أوامر التصنيع"""
    try:
        # البحث عن أوامر التصنيع بدون عناصر
        orders_without_items = ManufacturingOrder.objects.filter(
            items__isnull=True
        ).distinct()

        fixed_count = 0
        for manufacturing_order in orders_without_items:
            # البحث عن أمر التقطيع المرتبط
            cutting_orders = manufacturing_order.order.get_cutting_orders()

            if cutting_orders:
                cutting_order = cutting_orders.first()

                # إنشاء العناصر من أمر التقطيع
                for cutting_item in cutting_order.items.all():
                    ManufacturingOrderItem.objects.create(
                        manufacturing_order=manufacturing_order,
                        order_item=cutting_item.order_item,
                        product_name=(
                            cutting_item.order_item.product.name
                            if cutting_item.order_item.product
                            else "منتج غير محدد"
                        ),
                        quantity=cutting_item.order_item.quantity,
                        fabric_received=False,  # لم يتم الاستلام بعد
                        fabric_notes=f"تم إنشاؤه من أمر التقطيع {cutting_order.cutting_code}",
                    )

                fixed_count += 1

        return JsonResponse(
            {
                "success": True,
                "message": f"تم إصلاح {fixed_count} أمر تصنيع",
                "fixed_count": fixed_count,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


@require_http_methods(["POST"])
@login_required
def create_manufacturing_receipt(request):
    """إنشاء استلام أقمشة من المصنع"""
    try:
        manufacturing_order_id = request.POST.get("manufacturing_order_id")
        bag_number = request.POST.get("bag_number")
        received_by_name = request.POST.get("received_by_name")
        notes = request.POST.get("notes", "")

        if not all([manufacturing_order_id, bag_number, received_by_name]):
            return JsonResponse({"success": False, "message": "جميع الحقول مطلوبة"})

        # الحصول على أمر التصنيع
        manufacturing_order = ManufacturingOrder.objects.get(id=manufacturing_order_id)

        # التحقق من عدم وجود استلام سابق
        if FabricReceipt.objects.filter(
            manufacturing_order=manufacturing_order
        ).exists():
            return JsonResponse(
                {"success": False, "message": "تم استلام هذا الأمر من قبل"}
            )

        # الحصول على بيانات التقطيع المرتبطة
        cutting_orders = manufacturing_order.order.cutting_orders.filter(
            status="completed"
        )
        cutting_permit_number = None
        cutting_receiver_name = None

        if cutting_orders.exists():
            # أخذ البيانات من أول أمر تقطيع مكتمل
            cutting_order = cutting_orders.first()
            cutting_items = cutting_order.items.filter(
                receiver_name__isnull=False, permit_number__isnull=False
            )

            if cutting_items.exists():
                cutting_item = cutting_items.first()
                cutting_permit_number = cutting_item.permit_number
                cutting_receiver_name = cutting_item.receiver_name

        # إنشاء استلام الأقمشة
        fabric_receipt = FabricReceipt.objects.create(
            order=manufacturing_order.order,
            manufacturing_order=manufacturing_order,
            receipt_type="manufacturing_order",
            permit_number=cutting_permit_number
            or manufacturing_order.manufacturing_code,
            bag_number=bag_number,
            received_by_name=cutting_receiver_name or received_by_name,
            received_by=request.user,
            notes=notes,
        )

        return JsonResponse(
            {
                "success": True,
                "message": f"تم إنشاء استلام الأقمشة بنجاح",
                "receipt_id": fabric_receipt.id,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


@require_http_methods(["GET"])
@login_required
def get_cutting_data(request, manufacturing_order_id):
    """جلب بيانات التقطيع المرتبطة بأمر التصنيع"""
    try:
        manufacturing_order = ManufacturingOrder.objects.get(id=manufacturing_order_id)

        # الحصول على بيانات التقطيع المرتبطة
        cutting_orders = manufacturing_order.order.cutting_orders.filter(
            status="completed"
        )
        cutting_permit_number = None
        cutting_receiver_name = None

        if cutting_orders.exists():
            cutting_order = cutting_orders.first()
            cutting_items = cutting_order.items.filter(
                receiver_name__isnull=False, permit_number__isnull=False
            )

            if cutting_items.exists():
                cutting_item = cutting_items.first()
                cutting_permit_number = cutting_item.permit_number
                cutting_receiver_name = cutting_item.receiver_name

        return JsonResponse(
            {
                "success": True,
                "cutting_permit_number": cutting_permit_number,
                "cutting_receiver_name": cutting_receiver_name,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


class ProductReceiptView(LoginRequiredMixin, TemplateView):
    """صفحة استلام المنتجات من المخازن"""

    template_name = "manufacturing/product_receipt.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الحصول على أوامر التقطيع المكتملة لطلبات المنتجات
        # جميع أوامر التقطيع المكتملة للمنتجات (إصلاح JSONField)
        from django.db.models import Q

        from cutting.models import CuttingOrder

        # الحصول على أوامر التقطيع المكتملة لطلبات المنتجات (حل جذري مبسط)
        from .models import ProductReceipt

        cutting_orders_ready = (
            CuttingOrder.objects.filter(status="completed")
            .filter(Q(order__selected_types__icontains="products"))
            .select_related("order", "order__customer", "warehouse")
            .prefetch_related("items")
            .order_by("-created_at")
        )

        # استبعاد المستلمة بالفعل
        received_cutting_order_ids = ProductReceipt.objects.values_list(
            "cutting_order_id", flat=True
        )
        cutting_orders_ready = cutting_orders_ready.exclude(
            id__in=received_cutting_order_ids
        )

        # Debug info
        print(
            f"DEBUG: أوامر التقطيع المكتملة للمنتجات: {CuttingOrder.objects.filter(status='completed').filter(Q(order__selected_types__icontains='products')).count()}"
        )
        print(f"DEBUG: أوامر التقطيع المستلمة: {len(received_cutting_order_ids)}")
        print(f"DEBUG: أوامر التقطيع الجاهزة للاستلام: {cutting_orders_ready.count()}")

        # إحصائيات
        total_pending_items = cutting_orders_ready.count()
        received_today = ProductReceipt.objects.filter(
            receipt_date__date=timezone.now().date()
        ).count()

        # آخر الاستلامات
        recent_receipts = ProductReceipt.objects.select_related(
            "cutting_order",
            "cutting_order__order",
            "cutting_order__order__customer",
            "cutting_order__warehouse",
        ).order_by("-receipt_date")[:10]

        context.update(
            {
                "cutting_orders_ready": cutting_orders_ready,
                "total_pending_items": total_pending_items,
                "received_today": received_today,
                "total_cutting_ready": cutting_orders_ready.count(),
                "recent_receipts": recent_receipts,
            }
        )

        return context


class ProductReceiptsListView(TemplateView):
    """صفحة قائمة استلامات المنتجات"""

    template_name = "manufacturing/product_receipts_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الحصول على جميع الاستلامات مع التفاصيل
        from .models import ProductReceipt

        receipts = (
            ProductReceipt.objects.select_related(
                "cutting_order",
                "cutting_order__order",
                "cutting_order__order__customer",
                "cutting_order__warehouse",
            )
            .prefetch_related("cutting_order__items")
            .order_by("-receipt_date")
        )

        # تجميع الاستلامات حسب العميل
        receipts_by_customer = {}
        for receipt in receipts:
            customer_name = receipt.cutting_order.order.customer.name
            if customer_name not in receipts_by_customer:
                receipts_by_customer[customer_name] = []
            receipts_by_customer[customer_name].append(receipt)

        # إحصائيات
        total_receipts = receipts.count()
        today_receipts = receipts.filter(
            receipt_date__date=timezone.now().date()
        ).count()

        context.update(
            {
                "receipts_by_customer": receipts_by_customer,
                "total_receipts": total_receipts,
                "today_receipts": today_receipts,
            }
        )

        return context


@require_http_methods(["POST"])
@login_required
def create_product_receipt(request):
    """إنشاء استلام منتج من المخزن"""
    try:
        cutting_order_id = request.POST.get("cutting_order_id")
        bag_number = request.POST.get("bag_number")
        received_by_name = request.POST.get("received_by_name")
        notes = request.POST.get("notes", "")

        if not all([cutting_order_id, bag_number, received_by_name]):
            return JsonResponse({"success": False, "message": "جميع الحقول مطلوبة"})

        # الحصول على أمر التقطيع
        from cutting.models import CuttingOrder

        cutting_order = CuttingOrder.objects.get(id=cutting_order_id)

        # التحقق من أن الأمر مكتمل
        if cutting_order.status != "completed":
            return JsonResponse(
                {"success": False, "message": "أمر التقطيع غير مكتمل بعد"}
            )

        # التحقق من أن الطلب نوع "منتجات"
        if "products" not in cutting_order.order.get_selected_types_list():
            return JsonResponse(
                {"success": False, "message": "هذا الطلب ليس من نوع المنتجات"}
            )

        # التحقق من عدم وجود استلام سابق
        from .models import ProductReceipt

        if ProductReceipt.objects.filter(cutting_order=cutting_order).exists():
            return JsonResponse(
                {"success": False, "message": "تم استلام هذا الأمر من قبل"}
            )

        # إنشاء استلام المنتج
        product_receipt = ProductReceipt.objects.create(
            order=cutting_order.order,
            cutting_order=cutting_order,
            permit_number=cutting_order.cutting_code,
            bag_number=bag_number,
            received_by_name=received_by_name,
            received_by_user=request.user,
            notes=notes,
        )

        return JsonResponse(
            {
                "success": True,
                "message": f"تم إنشاء استلام المنتج بنجاح",
                "receipt_id": product_receipt.id,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"حدث خطأ: {str(e)}"})


class ManufacturingItemStatusReportView(
    LoginRequiredMixin, PermissionRequiredMixin, TemplateView
):
    """
    عرض تقرير حالة العناصر في أوامر التصنيع
    يعرض:
    - الأوامر التي تم استلام جميع عناصرها
    - الأوامر التي بها عناصر ناقصة
    - تفصيل العناصر المفقودة
    """

    template_name = "manufacturing/item_status_report.html"
    permission_required = "manufacturing.view_manufacturingorder"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الحصول على جميع أوامر التصنيع
        manufacturing_orders = (
            ManufacturingOrder.objects.select_related(
                "order", "order__customer", "order__salesperson"
            )
            .prefetch_related(
                "items",
                "items__cutting_item",
                "items__order_item",
                "items__order_item__product",
            )
            .all()
        )

        # تطبيق فلتر السنة
        manufacturing_orders = apply_default_year_filter(
            manufacturing_orders, "order_date"
        )

        # تصنيف الأوامر
        complete_orders = []
        incomplete_orders = []

        for order in manufacturing_orders:
            total_items = order.total_items_count
            received_items = order.received_items_count

            order_data = {
                "order": order,
                "total_items": total_items,
                "cut_items": order.cut_items_count,
                "received_items": received_items,
                "pending_items": order.pending_items_count,
                "missing_items": [],
            }

            # إذا كان هناك عناصر معلقة، نجمع تفاصيلها
            if received_items < total_items:
                for item in order.items.all():
                    if not item.fabric_received:
                        order_data["missing_items"].append(
                            {
                                "product_name": item.product_name,
                                "quantity": item.quantity,
                                "cutting_status": item.cutting_status_display,
                                "is_cut": item.is_cut,
                            }
                        )
                incomplete_orders.append(order_data)
            else:
                complete_orders.append(order_data)

        # الفلاتر
        filter_type = self.request.GET.get("filter", "all")

        if filter_type == "complete":
            context["orders"] = complete_orders
            context["showing"] = "complete"
        elif filter_type == "incomplete":
            context["orders"] = incomplete_orders
            context["showing"] = "incomplete"
        else:
            context["orders"] = complete_orders + incomplete_orders
            context["showing"] = "all"

        # الإحصائيات
        context["total_orders"] = len(complete_orders) + len(incomplete_orders)
        context["complete_orders_count"] = len(complete_orders)
        context["incomplete_orders_count"] = len(incomplete_orders)
        context["filter_type"] = filter_type

        return context


@login_required
def export_manufacturing_orders(request):
    """
    تصدير أوامر التصنيع إلى ملف Excel مع بطاقات ملخص وتنسيق احترافي
    """
    # 1. التحقق من الصلاحية
    if not hasattr(request.user, "can_export") or not request.user.can_export:
        from django.http import HttpResponseForbidden

        return HttpResponseForbidden("ليس لديك صلاحية لتصدير البيانات")

    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # 2. الحصول على البيانات المفلترة باستخدام View Class الموجود
    # نستخدم نفس منطق ManufacturingOrderListView لضمان تطابق النتائج
    view = ManufacturingOrderListView()
    view.request = request
    view.kwargs = {}

    # الحصول على QuerySet المفلتر
    queryset = view.get_queryset()

    # تحسين الأداء لحساب الأمتار
    from django.db.models import Prefetch

    queryset = queryset.prefetch_related(
        "items__cutting_item__cutting_order__warehouse"
    )

    # إعدادات حساب الأمتار
    from manufacturing.models import ManufacturingSettings

    try:
        settings = ManufacturingSettings.get_settings()
        warehouses_for_meters = list(settings.warehouses_for_meters_calculation.all())
    except Exception:
        warehouses_for_meters = []

    # 3. حساب الإحصائيات للبطاقات
    total_count = queryset.count()

    # عدد المتأخرات داخل النتائج المفلترة
    from django.utils import timezone

    today = timezone.now().date()
    overdue_count = queryset.filter(
        expected_delivery_date__lt=today,
        status__in=["pending_approval", "pending", "in_progress", "under_execution"],
    ).count()

    # حساب إجمالي الأمتار
    grand_total_meters = 0
    # سنقوم بحساب الأمتار لكل طلب وتخزينها لاستخدامها لاحقاً
    # هذا لتجنب إعادة الحساب مرتين
    orders_with_meters = []

    for order in queryset:
        order_meters = 0
        # حساب أمتار كل طلب
        for item in order.items.all():
            quantity = item.quantity or 0
            if not warehouses_for_meters:
                # إذا لم تكن هناك مستودعات محددة، نجمع الكل
                order_meters += quantity
            else:
                # التحقق من المستودع
                # نتأكد من وجود سلسلة العلاقات كاملة
                if (
                    item.cutting_item
                    and item.cutting_item.cutting_order
                    and item.cutting_item.cutting_order.warehouse
                    in warehouses_for_meters
                ):
                    order_meters += quantity

        grand_total_meters += order_meters
        # تخزين الطلب مع أمتاره
        order.calculated_meters = order_meters
        orders_with_meters.append(order)

    # تحديد اسم الفلتر النشط
    active_filter_name = "الكل"
    status_filter = request.GET.getlist("status")
    if status_filter:
        status_map = dict(ManufacturingOrder.STATUS_CHOICES)
        filter_names = [str(status_map.get(s, s)) for s in status_filter if s]
        if filter_names:
            active_filter_name = ", ".join(filter_names)

    # 4. إعداد ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "أوامر التصنيع"
    ws.sheet_view.rightToLeft = True  # اتجاه الورقة من اليمين لليسار

    # --- تنسيقات ---
    # حدود الخلايا
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # أنماط البطاقات
    card_title_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    card_value_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    card_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ألوان البطاقات
    # الأزرق (Total)
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # الأحمر (Overdue)
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    # الأخضر (Active Filter)
    green_fill = PatternFill(
        start_color="548235", end_color="548235", fill_type="solid"
    )
    # البنفسجي (Total Meters)
    purple_fill = PatternFill(
        start_color="7030A0", end_color="7030A0", fill_type="solid"
    )

    # --- رسم البطاقات (الصف 2-4) ---

    # بطاقة الإجمالي (A2:B4)
    ws.merge_cells("A2:B4")
    cell = ws["A2"]
    cell.value = f"إجمالي الأوامر\n{total_count}"
    cell.fill = blue_fill
    cell.font = card_value_font
    cell.alignment = card_alignment

    # بطاقة المتأخرات (C2:E4)
    ws.merge_cells("C2:E4")
    cell = ws["C2"]
    cell.value = f"المتأخرات\n{overdue_count}"
    cell.fill = red_fill
    cell.font = card_value_font
    cell.alignment = card_alignment

    # بطاقة الفلتر النشط (F2:H4)
    ws.merge_cells("F2:H4")
    cell = ws["F2"]
    cell.value = f"الفلتر الحالي\n{active_filter_name}"
    cell.fill = green_fill
    cell.font = card_value_font
    cell.alignment = card_alignment

    # بطاقة إجمالي الأمتار (I2:K4)
    ws.merge_cells("I2:K4")
    cell = ws["I2"]
    # تنسيق الرقم بفاصلة عشرية واحدة
    formatted_meters = "{:,.1f}".format(grand_total_meters)
    # إزالة .0 اذا كان رقماً صحيحاً
    if formatted_meters.endswith(".0"):
        formatted_meters = formatted_meters[:-2]

    cell.value = f"إجمالي الأمتار\n{formatted_meters} م"
    cell.fill = purple_fill
    cell.font = card_value_font
    cell.alignment = card_alignment

    # --- جدول البيانات (يبدأ من الصف 6) ---
    headers = [
        "رقم الطلب",
        "العميل",
        "خط الإنتاج",
        "التاريخ",
        "تاريخ التسليم",
        "النوع",
        "الفرع",
        "الحالة",
        "البائع",
        "حالة القماش",
        "الأمتار",
    ]

    header_row = 6
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="808080", end_color="808080", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # تعبئة البيانات
    data_row = 7
    # نستخدم القائمة المحسوبة مسبقاً بدلاً من queryset
    for order in orders_with_meters:
        # ترجمة الحالة
        status_display = order.get_status_display()
        type_display = order.get_order_type_display()

        # حالة القماش (من Bages) - استخدام الخصائص الموجودة في الموديل أو القيم المحسوبة
        # نستخدم القيم المحسوبة إذا كانت متوفرة لضمان تطابق الفلتر، وإلا نستخدم الخصائص

        fabric_status = "غير محدد"
        if hasattr(order, "calc_total"):
            # استخدام القيم المحسوبة من الـ annotation
            total = order.calc_total
            received = order.calc_received
            cut = order.calc_cut

            if total > 0 and received == 0 and cut == total:
                fabric_status = "بحاجة استلام"
            elif total > 0 and (received < total or cut < total):
                fabric_status = "ناقص"
            elif total > 0 and received == total and cut == total:
                fabric_status = "كامل"
            elif cut == 0:
                fabric_status = "غير مقطوع"
        else:
            # استخدام خصائص الموديل
            total = order.total_items_count
            received = order.received_items_count
            cut = order.cut_items_count

            if total > 0 and received == 0 and cut == total:
                fabric_status = "بحاجة استلام"
            elif total > 0 and (received < total or cut < total):
                fabric_status = "ناقص"
            elif total > 0 and received == total and cut == total:
                fabric_status = "كامل"
            elif cut == 0:
                fabric_status = "غير مقطوع"

        row_data = [
            order.order.order_number if order.order else "-",
            order.order.customer.name if order.order.customer else "-",
            order.production_line.name if order.production_line else "-",
            order.order_date.strftime("%Y-%m-%d") if order.order_date else "-",
            (
                order.expected_delivery_date.strftime("%Y-%m-%d")
                if order.expected_delivery_date
                else "-"
            ),
            type_display,
            order.order.branch.name if order.order.branch else "-",
            status_display,
            order.order.salesperson.name if order.order.salesperson else "-",
            fabric_status,
            order.calculated_meters,  # استخدام القيمة المحسوبة مسبقاً
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=data_row, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            # تلوين صف المتأخرات
            is_overdue = False
            if order.expected_delivery_date and order.expected_delivery_date < today:
                if order.status in ["pending_approval", "pending", "in_progress"]:
                    is_overdue = True

            if is_overdue:
                cell.fill = PatternFill(
                    start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                )

        data_row += 1

    # ضبط عرض الأعمدة تلقائياً
    column_widths = [15, 25, 20, 15, 15, 15, 15, 15, 20, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # إعداد الاستجابة
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"manufacturing_orders_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def material_summary_view(request, pk):
    """
    عرض بطاقة عناصر الطلب (Material Summary)
    """
    manufacturing_order = get_object_or_404(ManufacturingOrder, pk=pk)
    order = manufacturing_order.order

    # Check if partial render requested
    is_partial = request.GET.get("partial") == "true"

    # استخدام الدالة المساعدة لتوليد البيانات
    summary_context = get_material_summary_context(order)

    # الحصول على أسماء الخياطين من البطاقة المصنعية
    tailors_list = []
    try:
        if hasattr(manufacturing_order, "factory_card"):
            card = manufacturing_order.factory_card
            # Get tailor names from split distributions
            splits = card.splits.all().select_related("tailor")
            tailors_list = [split.tailor.name for split in splits]
    except Exception:
        pass

    tailors_display = "، ".join(tailors_list) if tailors_list else ""

    context = {
        "order": order,
        "materials_summary": summary_context["materials_summary"],
        "grand_total_quantity": summary_context["grand_total_quantity"],
        "grand_total_sewing": summary_context["grand_total_sewing"],
        "production_line_name": (
            manufacturing_order.production_line.name
            if manufacturing_order.production_line
            else ""
        ),
        "tailors_display": tailors_display,  # تم إضافة أسماء الخياطين
        "double_meter_types": summary_context.get(
            "double_meter_types", []
        ),  # تمرير أنواع الضعف للقالب
    }

    # Check if partial render requested (for Modal)
    if is_partial:
        return render(
            request, "manufacturing/partials/material_summary_content.html", context
        )

    return render(request, "manufacturing/material_summary_print.html", context)
