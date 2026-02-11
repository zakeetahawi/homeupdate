"""
أدوات التصدير - Excel & PDF
Export Utilities
"""

import io
import logging
from datetime import datetime

from django.http import HttpResponse

logger = logging.getLogger('accounting')


def export_to_excel(data, columns, filename="export", sheet_name="Sheet1"):
    """
    تصدير البيانات إلى ملف Excel

    Args:
        data: قائمة من القواميس أو tuples تحتوي على البيانات
        columns: قائمة من القواميس {"header": "عنوان العمود", "key": "مفتاح_البيانات", "width": 20}
        filename: اسم الملف (بدون امتداد)
        sheet_name: اسم الورقة

    Returns:
        HttpResponse مع ملف Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl غير مثبت. قم بتثبيته: pip install openpyxl")
        return HttpResponse("مكتبة openpyxl غير مثبتة", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True  # RTL للعربية

    # أنماط
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    number_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # عنوان التقرير
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=filename)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # تاريخ التصدير
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    date_cell = ws.cell(
        row=2, column=1, value=f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    date_cell.alignment = Alignment(horizontal="center")

    # رؤوس الأعمدة
    header_row = 4
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col["header"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col.get("width", 18)

    # البيانات
    for row_idx, row_data in enumerate(data, header_row + 1):
        for col_idx, col in enumerate(columns, 1):
            if isinstance(row_data, dict):
                value = row_data.get(col["key"], "")
            else:
                value = row_data[col_idx - 1] if col_idx - 1 < len(row_data) else ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # محاذاة الأرقام
            if isinstance(value, (int, float)):
                cell.alignment = number_alignment
                cell.number_format = "#,##0.00"
            else:
                cell.alignment = cell_alignment

        # تلوين الصفوف بالتناوب
        if (row_idx - header_row) % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                )

    # إنشاء الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    safe_filename = filename.replace(" ", "_")
    response["Content-Disposition"] = f'attachment; filename="{safe_filename}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response


def export_queryset_to_excel(queryset, fields, filename="export"):
    """
    تصدير queryset مباشرة إلى Excel

    Args:
        queryset: Django QuerySet
        fields: قائمة من tuples (field_name, header_display, width)
        filename: اسم الملف

    Returns:
        HttpResponse
    """
    columns = [{"header": f[1], "key": f[0], "width": f[2] if len(f) > 2 else 18} for f in fields]

    data = []
    for obj in queryset:
        row = {}
        for field_name, _, *_ in fields:
            value = obj
            for part in field_name.split("__"):
                if value is None:
                    break
                value = getattr(value, part, None)
            if callable(value):
                value = value()
            row[field_name] = value if value is not None else ""
        data.append(row)

    return export_to_excel(data, columns, filename)


def export_trial_balance_excel(trial_data, total_debit, total_credit, currency_symbol="ج.م"):
    """
    تصدير ميزان المراجعة إلى Excel
    """
    columns = [
        {"header": "كود الحساب", "key": "code", "width": 15},
        {"header": "اسم الحساب", "key": "name", "width": 35},
        {"header": "نوع الحساب", "key": "type", "width": 20},
        {"header": f"مدين ({currency_symbol})", "key": "debit", "width": 18},
        {"header": f"دائن ({currency_symbol})", "key": "credit", "width": 18},
    ]

    data = []
    for item in trial_data:
        data.append(
            {
                "code": item["account"].code,
                "name": item["account"].name,
                "type": item["account"].get_account_type_display() if hasattr(item["account"], "get_account_type_display") else str(item["account"].account_type),
                "debit": float(item["debit_balance"]) if item["debit_balance"] else 0,
                "credit": float(item["credit_balance"]) if item["credit_balance"] else 0,
            }
        )

    # إضافة صف الإجمالي
    data.append(
        {
            "code": "",
            "name": "الإجمالي",
            "type": "",
            "debit": float(total_debit),
            "credit": float(total_credit),
        }
    )

    return export_to_excel(data, columns, "ميزان_المراجعة")


def export_customer_balances_excel(customers_data, currency_symbol="ج.م"):
    """
    تصدير أرصدة العملاء إلى Excel
    """
    columns = [
        {"header": "كود العميل", "key": "code", "width": 15},
        {"header": "اسم العميل", "key": "name", "width": 30},
        {"header": "الهاتف", "key": "phone", "width": 18},
        {"header": f"إجمالي الطلبات ({currency_symbol})", "key": "total_orders", "width": 20},
        {"header": f"إجمالي المدفوع ({currency_symbol})", "key": "total_paid", "width": 20},
        {"header": f"الرصيد ({currency_symbol})", "key": "balance", "width": 20},
        {"header": "الحالة", "key": "status", "width": 15},
    ]

    data = []
    for item in customers_data:
        summary = item.get("summary")
        customer = item.get("customer")
        if summary and customer:
            data.append(
                {
                    "code": customer.code or "",
                    "name": customer.name,
                    "phone": customer.phone or "",
                    "total_orders": float(summary.total_orders_amount),
                    "total_paid": float(summary.total_paid),
                    "balance": float(summary.total_debt),
                    "status": summary.get_financial_status_display() if hasattr(summary, "get_financial_status_display") else summary.financial_status,
                }
            )

    return export_to_excel(data, columns, "أرصدة_العملاء")


def export_general_ledger_excel(ledger_entries, account, currency_symbol="ج.م"):
    """
    تصدير دفتر الأستاذ إلى Excel
    """
    columns = [
        {"header": "التاريخ", "key": "date", "width": 14},
        {"header": "رقم القيد", "key": "transaction_number", "width": 18},
        {"header": "البيان", "key": "description", "width": 40},
        {"header": f"مدين ({currency_symbol})", "key": "debit", "width": 18},
        {"header": f"دائن ({currency_symbol})", "key": "credit", "width": 18},
        {"header": f"الرصيد ({currency_symbol})", "key": "balance", "width": 18},
    ]

    data = []
    for entry in ledger_entries:
        data.append(
            {
                "date": str(entry["date"]),
                "transaction_number": entry["transaction"].transaction_number,
                "description": entry["description"],
                "debit": float(entry["debit"]),
                "credit": float(entry["credit"]),
                "balance": float(entry["balance"]),
            }
        )

    account_name = f"دفتر_الأستاذ_{account.code}_{account.name}"
    return export_to_excel(data, columns, account_name)
