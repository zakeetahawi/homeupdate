import logging
from datetime import date, timedelta

from django.contrib import messages
from django.db.models import Count, Prefetch, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import (
    CreateView,
    DetailView,
    FormView,
    ListView,
    TemplateView,
    UpdateView,
)

from accounting.export_utils import export_to_excel
from core.utils.general import convert_arabic_numbers_to_english
from customers.models import Customer

from .forms import (
    ContactLogForm,
    DecoratorEngineerProfileForm,
    EngineerExcelImportForm,
    EngineerProfileEditForm,
    LinkCustomerForm,
    LinkOrderForm,
    MaterialInterestForm,
)
from .mixins import DecoratorDeptRequiredMixin, DecoratorManagerRequiredMixin
from .models import (
    DecoratorEngineerProfile,
    EngineerContactLog,
    EngineerLinkedCustomer,
    EngineerLinkedOrder,
    EngineerMaterialInterest,
)
from .utils import compute_engineer_analytics

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
#  DASHBOARD
# ═══════════════════════════════════════════════════════════════
class DecoratorDashboardView(DecoratorDeptRequiredMixin, TemplateView):
    template_name = "external_sales/decorator/dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        inactive_threshold = today - timedelta(days=60)

        inactive_qs = (
            DecoratorEngineerProfile.objects.filter(
                Q(last_contact_date__lt=inactive_threshold)
                | Q(last_contact_date__isnull=True)
            )
            .select_related("customer", "assigned_staff")
            .order_by("last_contact_date")
        )

        ctx["inactive_engineers"] = list(inactive_qs)

        # Engineers needing attention: no contact in 30 days AND no linked orders/customers
        attention_threshold = today - timedelta(days=30)
        ctx["needs_attention_engineers"] = list(
            DecoratorEngineerProfile.objects.annotate(
                orders_count=Count("linked_orders", distinct=True),
                customers_count=Count("linked_customers", distinct=True),
            )
            .filter(
                Q(last_contact_date__lt=attention_threshold) | Q(last_contact_date__isnull=True),
                orders_count=0,
                customers_count=0,
            )
            .select_related("customer", "assigned_staff")
            .order_by("last_contact_date")
        )

        inactive_list = ctx["inactive_engineers"]
        ctx["stats"] = {
            "total": DecoratorEngineerProfile.objects.count(),
            "active": DecoratorEngineerProfile.objects.exclude(priority="cold").count(),
            "inactive_60d": len(inactive_list),
            "this_month": DecoratorEngineerProfile.objects.filter(
                created_at__date__gte=today.replace(day=1)
            ).count(),
        }

        # Upcoming follow-ups for the next 7 days
        ctx["upcoming_followups"] = (
            EngineerContactLog.objects.filter(
                next_followup_date__gte=today,
                next_followup_date__lte=today + timedelta(days=7),
            )
            .select_related("engineer__customer", "created_by")
            .order_by("next_followup_date")[:10]
        )

        # Recent activity
        ctx["recent_contacts"] = (
            EngineerContactLog.objects.select_related(
                "engineer__customer", "created_by"
            ).order_by("-contact_date")[:10]
        )

        return ctx


# ═══════════════════════════════════════════════════════════════
#  ENGINEER LIST (with filtering + pagination + export)
# ═══════════════════════════════════════════════════════════════
class EngineerListView(DecoratorDeptRequiredMixin, ListView):
    model = DecoratorEngineerProfile
    template_name = "external_sales/decorator/engineer_list.html"
    context_object_name = "engineers"
    paginate_by = 25

    def get(self, request, *args, **kwargs):
        """إذا كان البحث كوداً دقيقاً → انتقل مباشرةً لصفحة المهندس"""
        search = request.GET.get("search", "").strip()
        if search:
            exact = DecoratorEngineerProfile.objects.filter(
                designer_code__iexact=search
            ).first()
            if exact:
                return redirect(
                    "external_sales:engineer_detail", pk=exact.pk
                )
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = DecoratorEngineerProfile.objects.select_related(
            "customer", "customer__branch", "assigned_staff"
        ).order_by("-created_at")

        # فلتر خاص: بدون تواصل +60 يوم
        if self.request.GET.get("inactive"):
            threshold = date.today() - timedelta(days=60)
            qs = qs.filter(
                Q(last_contact_date__lt=threshold) | Q(last_contact_date__isnull=True)
            ).order_by("last_contact_date")
            return qs

        search = self.request.GET.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(customer__name__icontains=search)
                | Q(customer__phone__icontains=search)
                | Q(customer__code__icontains=search)
                | Q(city__icontains=search)
                | Q(designer_code__icontains=search)
            )

        priority = self.request.GET.get("priority")
        if priority:
            qs = qs.filter(priority=priority)

        city = self.request.GET.get("city")
        if city:
            qs = qs.filter(city=city)

        branch = self.request.GET.get("branch")
        if branch:
            qs = qs.filter(customer__branch_id=branch)

        staff = self.request.GET.get("staff")
        if staff:
            qs = qs.filter(assigned_staff_id=staff)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from accounts.models import User

        ctx["is_inactive_filter"] = bool(self.request.GET.get("inactive"))
        from accounts.models import Branch
        ctx["branches"] = Branch.objects.filter(is_active=True).order_by("name")
        ctx["cities"] = (
            DecoratorEngineerProfile.objects.exclude(city="")
            .values_list("city", flat=True)
            .distinct()
            .order_by("city")
        )
        ctx["staff_list"] = User.objects.filter(
            Q(is_decorator_dept_staff=True)
            | Q(is_decorator_dept_manager=True)
        ).filter(is_active=True)
        return ctx


# ═══════════════════════════════════════════════════════════════
#  ENGINEER DETAIL (6-tab page)
# ═══════════════════════════════════════════════════════════════
class EngineerDetailView(DecoratorDeptRequiredMixin, DetailView):
    model = DecoratorEngineerProfile
    template_name = "external_sales/decorator/engineer_detail.html"
    context_object_name = "profile"

    def get_queryset(self):
        return DecoratorEngineerProfile.objects.select_related(
            "customer",
            "customer__branch",
            "customer__category",
            "assigned_staff",
        ).prefetch_related(
            Prefetch(
                "linked_customers",
                queryset=EngineerLinkedCustomer.objects.filter(is_active=True)
                .select_related("customer", "customer__branch", "linked_by"),
            ),
            Prefetch(
                "linked_orders",
                queryset=EngineerLinkedOrder.objects.select_related(
                    "order", "order__customer", "linked_by"
                ).order_by("-linked_at"),
            ),
            Prefetch(
                "contact_logs",
                queryset=EngineerContactLog.objects.select_related("created_by")
                .order_by("-contact_date"),
            ),
            "material_interests",
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        profile = self.object
        ctx["analytics"] = compute_engineer_analytics(profile)
        ctx["recent_contacts"] = list(profile.contact_logs.all()[:10])
        ctx["pending_commissions"] = [
            lo for lo in profile.linked_orders.all()
            if lo.commission_status == "pending"
        ]
        ctx["upcoming_followups"] = (
            EngineerContactLog.objects.filter(
                engineer=profile, next_followup_date__gte=date.today()
            )
            .order_by("next_followup_date")[:5]
        )
        ctx["today"] = date.today()
        return ctx


# ═══════════════════════════════════════════════════════════════
#  CREATE / EDIT PROFILE
# ═══════════════════════════════════════════════════════════════
class CreateEngineerProfileView(DecoratorDeptRequiredMixin, CreateView):
    model = DecoratorEngineerProfile
    form_class = DecoratorEngineerProfileForm
    template_name = "external_sales/decorator/engineer_form.html"

    def get_initial(self):
        initial = super().get_initial()
        customer_code = self.kwargs.get("customer_code")
        if customer_code and customer_code != "NEW":
            try:
                customer = Customer.objects.get(code=customer_code)
                initial["customer"] = customer.pk
            except Customer.DoesNotExist:
                pass
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "إنشاء ملف مهندس ديكور جديد"
        ctx["is_new"] = True
        return ctx

    def form_valid(self, form):
        from accounts.middleware.current_user import get_current_user

        form.instance.created_by = get_current_user()
        messages.success(self.request, "تم إنشاء ملف المهندس بنجاح")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            "external_sales:engineer_detail", kwargs={"pk": self.object.pk}
        )


class EngineerProfileEditView(DecoratorDeptRequiredMixin, UpdateView):
    model = DecoratorEngineerProfile
    form_class = EngineerProfileEditForm
    template_name = "external_sales/decorator/engineer_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = f"تعديل ملف: {self.object.customer.name}"
        ctx["is_new"] = False
        return ctx

    def form_valid(self, form):
        from accounts.middleware.current_user import get_current_user

        form.instance.updated_by = get_current_user()
        messages.success(self.request, "تم تحديث الملف بنجاح")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            "external_sales:engineer_detail", kwargs={"pk": self.object.pk}
        )


# ═══════════════════════════════════════════════════════════════
#  CONTACT LOG
# ═══════════════════════════════════════════════════════════════
class AddContactLogView(DecoratorDeptRequiredMixin, CreateView):
    model = EngineerContactLog
    form_class = ContactLogForm
    template_name = "external_sales/decorator/contact_log_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["profile"] = get_object_or_404(
            DecoratorEngineerProfile.objects.select_related("customer"),
            pk=self.kwargs["pk"],
        )
        return ctx

    def form_valid(self, form):
        profile = get_object_or_404(DecoratorEngineerProfile, pk=self.kwargs["pk"])
        form.instance.engineer = profile
        form.instance.created_by = self.request.user
        messages.success(self.request, "تم تسجيل التواصل بنجاح")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            "external_sales:engineer_detail", kwargs={"pk": self.kwargs["pk"]}
        )


class ContactLogListView(DecoratorDeptRequiredMixin, ListView):
    model = EngineerContactLog
    template_name = "external_sales/decorator/contact_log_list.html"
    context_object_name = "logs"
    paginate_by = 30

    def get_queryset(self):
        return (
            EngineerContactLog.objects.filter(engineer_id=self.kwargs["pk"])
            .select_related("created_by")
            .order_by("-contact_date")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["profile"] = get_object_or_404(
            DecoratorEngineerProfile.objects.select_related("customer"),
            pk=self.kwargs["pk"],
        )
        return ctx


# ═══════════════════════════════════════════════════════════════
#  LINKING — CUSTOMERS
# ═══════════════════════════════════════════════════════════════
class LinkCustomerView(DecoratorDeptRequiredMixin, CreateView):
    model = EngineerLinkedCustomer
    form_class = LinkCustomerForm
    template_name = "external_sales/decorator/link_customer_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["profile"] = get_object_or_404(
            DecoratorEngineerProfile.objects.select_related("customer"),
            pk=self.kwargs["pk"],
        )
        return ctx

    def form_valid(self, form):
        profile = get_object_or_404(DecoratorEngineerProfile, pk=self.kwargs["pk"])
        form.instance.engineer = profile
        form.instance.linked_by = self.request.user
        messages.success(self.request, "تم ربط العميل بنجاح")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            "external_sales:engineer_detail", kwargs={"pk": self.kwargs["pk"]}
        )


class UnlinkCustomerView(DecoratorDeptRequiredMixin, View):
    def post(self, request, pk, link_id):
        link = get_object_or_404(EngineerLinkedCustomer, pk=link_id, engineer_id=pk)
        link.is_active = False
        link.save(update_fields=["is_active"])
        messages.success(request, "تم فك ربط العميل بنجاح")
        return redirect("external_sales:engineer_detail", pk=pk)

    def get(self, request, pk, link_id):
        return self.post(request, pk, link_id)


# ═══════════════════════════════════════════════════════════════
#  LINKING — ORDERS
# ═══════════════════════════════════════════════════════════════
class LinkOrderView(DecoratorDeptRequiredMixin, CreateView):
    model = EngineerLinkedOrder
    form_class = LinkOrderForm
    template_name = "external_sales/decorator/link_order_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["profile"] = get_object_or_404(
            DecoratorEngineerProfile.objects.select_related("customer"),
            pk=self.kwargs["pk"],
        )
        return ctx

    def form_valid(self, form):
        profile = get_object_or_404(DecoratorEngineerProfile, pk=self.kwargs["pk"])
        form.instance.engineer = profile
        form.instance.linked_by = self.request.user
        obj = form.save(commit=False)
        obj.calculate_commission()
        obj.save()
        # Auto-link the order's customer if not already linked
        order_customer = obj.order.customer
        if order_customer:
            EngineerLinkedCustomer.objects.get_or_create(
                engineer=profile,
                customer=order_customer,
                defaults={
                    "linked_by": self.request.user,
                    "relationship_type": "referred_client",
                },
            )
        messages.success(self.request, "تم ربط الطلب والعميل بنجاح")
        return redirect("external_sales:engineer_detail", pk=self.kwargs["pk"])


class UnlinkOrderView(DecoratorManagerRequiredMixin, View):
    def post(self, request, pk, link_id):
        link = get_object_or_404(EngineerLinkedOrder, pk=link_id, engineer_id=pk)
        link.delete()
        messages.success(request, "تم فك ربط الطلب بنجاح")
        return redirect("external_sales:engineer_detail", pk=pk)

    def get(self, request, pk, link_id):
        return self.post(request, pk, link_id)


# ═══════════════════════════════════════════════════════════════
#  MATERIALS
# ═══════════════════════════════════════════════════════════════
class MaterialInterestView(DecoratorDeptRequiredMixin, CreateView):
    model = EngineerMaterialInterest
    form_class = MaterialInterestForm
    template_name = "external_sales/decorator/material_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        profile = get_object_or_404(
            DecoratorEngineerProfile.objects.select_related("customer"),
            pk=self.kwargs["pk"],
        )
        ctx["profile"] = profile
        ctx["existing_materials"] = profile.material_interests.all()
        return ctx

    def form_valid(self, form):
        profile = get_object_or_404(DecoratorEngineerProfile, pk=self.kwargs["pk"])
        form.instance.engineer = profile
        messages.success(self.request, "تمت إضافة اهتمام بالخامة بنجاح")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(
            "external_sales:materials", kwargs={"pk": self.kwargs["pk"]}
        )


# ═══════════════════════════════════════════════════════════════
#  DECORATOR ORDERS (filtered view)
# ═══════════════════════════════════════════════════════════════
class DecoratorOrdersView(DecoratorDeptRequiredMixin, ListView):
    template_name = "external_sales/decorator/decorator_orders.html"
    context_object_name = "linked_orders"
    paginate_by = 30

    def get_queryset(self):
        return (
            EngineerLinkedOrder.objects.select_related(
                "engineer__customer", "order", "order__customer", "linked_by"
            )
            .order_by("-linked_at")
        )


# ═══════════════════════════════════════════════════════════════
#  COMMISSIONS
# ═══════════════════════════════════════════════════════════════
class CommissionsView(DecoratorDeptRequiredMixin, ListView):
    template_name = "external_sales/decorator/commissions.html"
    context_object_name = "commissions"
    paginate_by = 30

    def get_queryset(self):
        qs = EngineerLinkedOrder.objects.select_related(
            "engineer__customer", "order", "order__customer"
        ).order_by("-linked_at")

        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(commission_status=status)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["summary"] = EngineerLinkedOrder.objects.aggregate(
            pending=Sum(
                "commission_value", filter=Q(commission_status="pending")
            ),
            approved=Sum(
                "commission_value", filter=Q(commission_status="approved")
            ),
            paid=Sum(
                "commission_value", filter=Q(commission_status="paid")
            ),
        )
        return ctx


class ApproveCommissionView(DecoratorManagerRequiredMixin, View):
    def post(self, request, pk):
        commission = get_object_or_404(
            EngineerLinkedOrder, pk=pk, commission_status="pending"
        )
        commission.commission_status = "approved"
        commission.save(update_fields=["commission_status"])
        messages.success(request, "تم اعتماد العمولة بنجاح")
        return redirect("external_sales:commissions")

    def get(self, request, pk):
        return self.post(request, pk)


class MarkCommissionPaidView(DecoratorManagerRequiredMixin, View):
    def post(self, request, pk):
        commission = get_object_or_404(
            EngineerLinkedOrder, pk=pk, commission_status="approved"
        )
        commission.commission_status = "paid"
        commission.commission_paid_at = timezone.now()
        commission.commission_paid_by = request.user
        commission.save(
            update_fields=[
                "commission_status",
                "commission_paid_at",
                "commission_paid_by",
            ]
        )
        messages.success(request, "تم تسجيل دفع العمولة بنجاح")
        return redirect("external_sales:commissions")

    def get(self, request, pk):
        return self.post(request, pk)


# ═══════════════════════════════════════════════════════════════
#  AJAX ENDPOINTS
# ═══════════════════════════════════════════════════════════════
class CustomerSearchAjax(DecoratorDeptRequiredMixin, View):
    """AJAX endpoint for Select2 customer search."""

    def get(self, request):
        from customers.models import Customer

        q = request.GET.get("q", "").strip()
        if len(q) < 2:
            return JsonResponse({"results": []})
        customers = (
            Customer.objects.filter(
                Q(name__icontains=q)
                | Q(phone__icontains=q)
                | Q(code__icontains=q)
            )
            .select_related("branch")
            .only("pk", "name", "phone", "code", "branch__name")
            .order_by("name")[:20]
        )
        results = [
            {
                "id": c.pk,
                "text": f"{c.name} — {c.phone or ''} ({c.branch.name if c.branch else ''})",
            }
            for c in customers
        ]
        return JsonResponse({"results": results})


class DesignerCustomerSearchAjax(DecoratorDeptRequiredMixin, View):
    """AJAX search for designer customers without profiles (for create-profile)."""

    def get(self, request):
        from customers.models import Customer

        q = request.GET.get("q", "").strip()
        if len(q) < 2:
            return JsonResponse({"results": []})
        # Only designer customers who don't already have a profile
        customers = (
            Customer.objects.filter(
                customer_type="designer",
            )
            .exclude(decorator_profile__isnull=False)
            .filter(
                Q(name__icontains=q)
                | Q(phone__icontains=q)
                | Q(code__icontains=q)
            )
            .select_related("branch")
            .only("pk", "name", "phone", "code", "branch__name")
            .order_by("name")[:20]
        )
        results = [
            {
                "id": c.pk,
                "text": f"{c.name} — {c.phone or ''} ({c.branch.name if c.branch else ''})",
            }
            for c in customers
        ]
        return JsonResponse({"results": results})


class OrderSearchAjax(DecoratorDeptRequiredMixin, View):
    """AJAX endpoint for Select2 order search."""


class AllUpcomingFollowupsView(DecoratorDeptRequiredMixin, TemplateView):
    """صفحة مستقلة لجميع المتابعات القادمة"""

    template_name = "external_sales/decorator/all_followups.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        ctx["today"] = today
        ctx["followups"] = (
            EngineerContactLog.objects.filter(
                next_followup_date__gte=today,
                next_followup_date__lte=today + timedelta(days=30),
            )
            .select_related("engineer__customer", "engineer__assigned_staff", "created_by")
            .order_by("next_followup_date")
        )
        ctx["overdue_followups"] = (
            EngineerContactLog.objects.filter(
                next_followup_date__lt=today,
            )
            .select_related("engineer__customer", "engineer__assigned_staff", "created_by")
            .order_by("next_followup_date")
        )
        return ctx


class AllContactLogsView(DecoratorDeptRequiredMixin, ListView):
    """صفحة كل سجلات التواصل مع المهندسين"""

    model = EngineerContactLog
    template_name = "external_sales/decorator/all_contacts.html"
    context_object_name = "contact_logs"
    paginate_by = 50

    def get_queryset(self):
        return EngineerContactLog.objects.select_related(
            "engineer__customer", "created_by"
        ).order_by("-contact_date")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["today"] = date.today()
        return ctx


class FindEngineerByCodeAjax(DecoratorDeptRequiredMixin, View):
    """AJAX: البحث عن مهندس بكود المصمم — مثل find_customer_by_phone تماماً"""

    def get(self, request, *args, **kwargs):
        code = request.GET.get("code", "").strip()
        if not code:
            return JsonResponse({"found": False, "error": "كود المصمم مطلوب"}, status=400)

        code = convert_arabic_numbers_to_english(code)
        profiles = (
            DecoratorEngineerProfile.objects.filter(
                Q(designer_code__iexact=code) | Q(designer_code__icontains=code)
            )
            .select_related("customer", "customer__branch", "assigned_staff")
            .order_by("designer_code")
        )

        if profiles.exists():
            data = []
            for p in profiles:
                data.append(
                    {
                        "pk": p.pk,
                        "name": p.customer.name,
                        "designer_code": p.designer_code,
                        "phone": p.customer.phone or "",
                        "city": p.city or "",
                        "priority": p.get_priority_display(),
                        "company": p.company_office_name or "",
                        "assigned_staff": p.assigned_staff.get_full_name() if p.assigned_staff else "",
                        "url": f"/external-sales/decorator/engineers/{p.pk}/",
                        "is_exact": p.designer_code.upper() == code.upper(),
                    }
                )
            return JsonResponse({"found": True, "engineers": data, "count": len(data)})

        return JsonResponse({"found": False})


    def get(self, request):
        from orders.models import Order

        q = request.GET.get("q", "").strip()
        if len(q) < 2:
            return JsonResponse({"results": []})
        orders = (
            Order.objects.filter(
                Q(order_number__icontains=q)
                | Q(customer__name__icontains=q)
                | Q(customer__phone__icontains=q)
            )
            .select_related("customer")
            .only("pk", "order_number", "total_amount", "order_date", "customer__name")
            .order_by("-order_date")[:20]
        )
        results = [
            {
                "id": o.pk,
                "text": f"{o.order_number} — {o.customer.name if o.customer else ''} ({o.total_amount} ر.س)",
            }
            for o in orders
        ]
        return JsonResponse({"results": results})


class EngineerSearchAjax(DecoratorDeptRequiredMixin, View):
    def get(self, request):
        q = request.GET.get("q", "").strip()
        if len(q) < 2:
            return JsonResponse({"results": []})

        profiles = (
            DecoratorEngineerProfile.objects.filter(
                Q(customer__name__icontains=q)
                | Q(designer_code__icontains=q)
                | Q(customer__phone__icontains=q)
            )
            .select_related("customer")
            .values("pk", "designer_code", "customer__name", "customer__phone")[:10]
        )
        return JsonResponse({"results": list(profiles)})


class AvailableOrdersAjax(DecoratorDeptRequiredMixin, View):
    def get(self, request, customer_code):
        from orders.models import Order

        orders = (
            Order.objects.filter(customer__code=customer_code)
            .exclude(engineer_link__isnull=False)
            .values("pk", "order_number", "total_amount", "order_date")[:20]
        )
        result = []
        for o in orders:
            result.append(
                {
                    "pk": o["pk"],
                    "order_number": o["order_number"],
                    "total_amount": str(o["total_amount"]),
                    "order_date": str(o["order_date"])[:10] if o["order_date"] else "",
                }
            )
        return JsonResponse({"results": result})


# ═══════════════════════════════════════════════════════════════
#  CHART AJAX ENDPOINTS (§22)
# ═══════════════════════════════════════════════════════════════
class ChartTopByRevenueAjax(DecoratorDeptRequiredMixin, View):
    def get(self, request):
        data = list(
            DecoratorEngineerProfile.objects.annotate(
                total_value=Sum("linked_orders__order__total_amount")
            )
            .filter(total_value__gt=0)
            .order_by("-total_value")[:10]
            .values("customer__name", "designer_code", "total_value")
        )
        return JsonResponse(
            {
                "labels": [r["customer__name"] for r in data],
                "values": [float(r["total_value"] or 0) for r in data],
            }
        )


class ChartTopByOrdersAjax(DecoratorDeptRequiredMixin, View):
    def get(self, request):
        data = list(
            DecoratorEngineerProfile.objects.annotate(
                order_cnt=Count("linked_orders")
            )
            .filter(order_cnt__gt=0)
            .order_by("-order_cnt")[:10]
            .values("customer__name", "designer_code", "order_cnt")
        )
        return JsonResponse(
            {
                "labels": [r["customer__name"] for r in data],
                "values": [r["order_cnt"] for r in data],
            }
        )


class ChartTopMaterialsAjax(DecoratorDeptRequiredMixin, View):
    def get(self, request):
        from orders.models import OrderItem

        # أكثر الخامات/المنتجات طلباً من خلال الطلبات المرتبطة بمهندسين
        data = list(
            OrderItem.objects.filter(
                order__engineer_link__isnull=False,
                product_name_snapshot__isnull=False,
            )
            .exclude(product_name_snapshot="")
            .values("product_name_snapshot")
            .annotate(total=Count("id"))
            .order_by("-total")[:10]
        )
        return JsonResponse(
            {
                "labels": [r["product_name_snapshot"] for r in data],
                "values": [r["total"] for r in data],
            }
        )


class ChartMonthlyActivityAjax(DecoratorDeptRequiredMixin, View):
    def get(self, request):
        six_months_ago = date.today() - timedelta(days=180)

        contacts_map = {
            str(r["month"])[:7]: r["cnt"]
            for r in EngineerContactLog.objects.filter(contact_date__gte=six_months_ago)
            .annotate(month=TruncMonth("contact_date"))
            .values("month")
            .annotate(cnt=Count("id"))
        }
        orders_map = {
            str(r["month"])[:7]: r["cnt"]
            for r in EngineerLinkedOrder.objects.filter(
                order__order_date__gte=six_months_ago
            )
            .annotate(month=TruncMonth("order__order_date"))
            .values("month")
            .annotate(cnt=Count("id"))
        }

        # بناء محور شهري كامل للستة أشهر الماضية
        all_months = []
        d = (date.today() - timedelta(days=180)).replace(day=1)
        today_m = date.today().replace(day=1)
        while d <= today_m:
            all_months.append(str(d)[:7])
            d = d.replace(month=d.month + 1) if d.month < 12 else d.replace(year=d.year + 1, month=1)

        return JsonResponse(
            {
                "months": all_months,
                "contacts": [contacts_map.get(m, 0) for m in all_months],
                "orders": [orders_map.get(m, 0) for m in all_months],
            }
        )


# ═══════════════════════════════════════════════════════════════
#  EXCEL EXPORT (§26)
# ═══════════════════════════════════════════════════════════════
class ExportEngineersExcelView(DecoratorDeptRequiredMixin, View):
    def get(self, request):
        profiles = DecoratorEngineerProfile.objects.select_related(
            "customer", "customer__branch", "assigned_staff"
        ).order_by("designer_code")

        data = []
        for p in profiles:
            data.append(
                {
                    "كود المهندس": p.designer_code,
                    "اسم المهندس": p.customer.name,
                    "الهاتف": p.customer.phone,
                    "الفرع": p.customer.branch.name if p.customer.branch else "",
                    "المدينة": p.city,
                    "الأولوية": p.get_priority_display(),
                    "المكتب": p.company_office_name,
                    "عدد العملاء": p.total_clients_count,
                    "عدد الطلبات": p.total_orders_count,
                    "آخر تواصل": str(p.last_contact_date or ""),
                    "موظف المتابعة": (
                        p.assigned_staff.get_full_name() if p.assigned_staff else ""
                    ),
                }
            )

        columns = [
            {"header": h, "key": h}
            for h in [
                "كود المهندس", "اسم المهندس", "الهاتف", "الفرع", "المدينة",
                "الأولوية", "المكتب", "عدد العملاء", "عدد الطلبات",
                "آخر تواصل", "موظف المتابعة",
            ]
        ]
        return export_to_excel(
            data, columns, filename="decorator_engineers", sheet_name="مهندسو الديكور"
        )


class ExportEngineerFullDataView(DecoratorDeptRequiredMixin, View):
    def get(self, request, pk):
        profile = get_object_or_404(
            DecoratorEngineerProfile.objects.select_related("customer"), pk=pk
        )
        analytics = compute_engineer_analytics(profile)

        # Export contact logs
        logs_data = []
        for log in profile.contact_logs.select_related("created_by").order_by(
            "-contact_date"
        ):
            logs_data.append(
                {
                    "التاريخ": str(log.contact_date)[:16],
                    "نوع التواصل": log.get_contact_type_display(),
                    "النتيجة": log.get_outcome_display(),
                    "الملاحظات": log.notes[:100] if log.notes else "",
                    "بواسطة": (
                        log.created_by.get_full_name() if log.created_by else ""
                    ),
                }
            )

        columns = [
            {"header": h, "key": h}
            for h in ["التاريخ", "نوع التواصل", "النتيجة", "الملاحظات", "بواسطة"]
        ]
        return export_to_excel(
            logs_data,
            columns,
            filename=f"engineer_{profile.designer_code}",
            sheet_name=profile.customer.name[:30],
        )


class ExportCommissionsExcelView(DecoratorDeptRequiredMixin, View):
    def get(self, request):
        commissions = EngineerLinkedOrder.objects.select_related(
            "engineer__customer", "order", "order__customer"
        ).order_by("-linked_at")

        data = []
        for c in commissions:
            data.append(
                {
                    "كود المهندس": c.engineer.designer_code,
                    "اسم المهندس": c.engineer.customer.name,
                    "رقم الطلب": c.order.order_number,
                    "العميل": c.order.customer.name if c.order.customer else "",
                    "إجمالي الطلب": str(c.order.total_amount or 0),
                    "نسبة العمولة": str(c.commission_rate),
                    "قيمة العمولة": str(c.commission_value),
                    "الحالة": c.get_commission_status_display(),
                    "تاريخ الربط": str(c.linked_at)[:10],
                }
            )

        columns = [
            {"header": h, "key": h}
            for h in [
                "كود المهندس", "اسم المهندس", "رقم الطلب", "العميل",
                "إجمالي الطلب", "نسبة العمولة", "قيمة العمولة", "الحالة", "تاريخ الربط",
            ]
        ]
        return export_to_excel(
            data, columns, filename="commissions", sheet_name="العمولات"
        )


# ═══════════════════════════════════════════════════════════════
#  EXCEL IMPORT (§26)
# ═══════════════════════════════════════════════════════════════
class ImportEngineersExcelView(DecoratorManagerRequiredMixin, FormView):
    template_name = "external_sales/decorator/import_engineers.html"
    form_class = EngineerExcelImportForm
    success_url = reverse_lazy("external_sales:engineer_list")

    def form_valid(self, form):
        import openpyxl

        file = form.cleaned_data["file"]

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            created_count = 0
            errors = []

            for i, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue
                customer_name = str(row[0]).strip()
                phone = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if phone:
                    phone = convert_arabic_numbers_to_english(phone)
                city = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                priority = str(row[3]).strip() if len(row) > 3 and row[3] else "regular"
                company = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                # Find or create customer
                try:
                    customer = Customer.objects.filter(
                        Q(name=customer_name) | Q(phone=phone)
                    ).first()
                    if not customer:
                        errors.append(f"سطر {i}: العميل '{customer_name}' غير موجود")
                        continue

                    if hasattr(customer, "decorator_profile"):
                        errors.append(
                            f"سطر {i}: العميل '{customer_name}' لديه ملف مهندس بالفعل"
                        )
                        continue

                    DecoratorEngineerProfile.objects.create(
                        customer=customer,
                        city=city,
                        priority=priority if priority in ["vip", "active", "regular", "cold"] else "regular",
                        company_office_name=company,
                        created_by=self.request.user,
                    )
                    created_count += 1
                except Exception as e:
                    errors.append(f"سطر {i}: خطأ — {e}")

            wb.close()

            if created_count:
                messages.success(
                    self.request, f"تم استيراد {created_count} مهندس بنجاح"
                )
            if errors:
                messages.warning(
                    self.request,
                    f"حدثت {len(errors)} أخطاء أثناء الاستيراد:\n"
                    + "\n".join(errors[:10]),
                )

        except Exception as e:
            messages.error(self.request, f"خطأ في قراءة الملف: {e}")

        return super().form_valid(form)


class DownloadImportTemplateView(DecoratorDeptRequiredMixin, View):
    def get(self, request):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "قالب استيراد المهندسين"
        ws.sheet_view.rightToLeft = True

        headers = [
            "اسم المهندس (مطابق للعميل)",
            "رقم الهاتف (اختياري)",
            "المدينة",
            "الأولوية (vip/active/regular/cold)",
            "اسم المكتب / الشركة",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = 25

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="import_template.xlsx"'
        )
        wb.save(response)
        return response
