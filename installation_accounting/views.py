import json
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from installations.models import Technician

from .models import InstallationAccountingSettings, InstallationCard, TechnicianShare


@login_required
def installation_reports(request):
    """
    نتائج تقارير التركيبات (بناءً على الجدول الزمني)
    يعرض جميع التركيبات المجدولة في يوم معين بغض النظر عن حالة الإكمال
    """
    today = timezone.now().date()

    # تحديد التاريخ الافتراضي
    if today.day >= 25:
        default_date_from = today.replace(day=25).strftime("%Y-%m-%d")
    else:
        if today.month == 1:
            default_date_from = today.replace(
                year=today.year - 1, month=12, day=25
            ).strftime("%Y-%m-%d")
        else:
            default_date_from = today.replace(month=today.month - 1, day=25).strftime(
                "%Y-%m-%d"
            )

    date_from = request.GET.get("date_from", default_date_from)
    date_to = request.GET.get("date_to")
    payment_status = request.GET.get("payment_status", "all")
    completion_status = request.GET.get("completion_status", "all")
    technician_id = request.GET.get("technician")

    # Base queryset - Schedule based
    from installations.models import InstallationSchedule

    schedules = (
        InstallationSchedule.objects.select_related(
            "order",
            "order__customer",
            "accounting_card",  # Left join with accounting card
        )
        .prefetch_related(
            "technicians",
            "accounting_card__shares",
            "accounting_card__shares__technician",
        )
        .exclude(status="cancelled")
    )  # Exclude cancelled (optional based on user request "full schedule", but usually cancelled means removed)

    # Filter by SCHEDULED DATE
    if date_from:
        schedules = schedules.filter(scheduled_date__gte=date_from)
    if date_to:
        schedules = schedules.filter(scheduled_date__lte=date_to)

    # Filter by TECHNICIAN (if assigned in schedule)
    if technician_id:
        schedules = schedules.filter(technicians__id=technician_id).distinct()

    # Filter by COMPLETION STATUS (Strictly Completed Only as per user request)
    schedules = schedules.filter(status__in=["completed", "modification_completed"])

    # Filter by PAYMENT STATUS (requires accounting card)
    if payment_status == "paid":
        schedules = schedules.filter(accounting_card__status="paid")
    elif payment_status == "unpaid":
        schedules = schedules.exclude(accounting_card__status="paid")

    # Calculate Totals
    # Note: Totals should account for filtered list
    total_windows = 0
    total_cost = Decimal("0.00")
    total_orders_count = schedules.count()
    technician_total = Decimal("0.00")

    # We iterate to sum values because some might not have cards yet
    # Or use aggregation if possible, but cleaner logic here:

    # For technician total, we need to be careful if filtering by technician

    for schedule in schedules:
        # Windows count (from schedule or card)
        # Use card's windows count if exists (final), else schedule's
        if hasattr(schedule, "accounting_card"):
            card = schedule.accounting_card
            cost = card.total_cost

            # Tech shares
            shares = card.shares.all()
            if technician_id:
                shares = shares.filter(technician_id=technician_id)
                # الشبابيك الفعلية للفني المحدد فقط (assigned_windows)
                wins = sum(
                    share.assigned_windows for share in shares
                ) if shares else 0
            else:
                wins = card.windows_count

            # Payment status logic for tech total
            if payment_status == "paid":
                shares = shares.filter(is_paid=True)
            elif payment_status == "unpaid":
                shares = shares.filter(is_paid=False)

            tech_sum = sum(share.amount for share in shares)
        else:
            wins = schedule.windows_count or 0
            cost = Decimal("0.00")  # No accounting yet
            tech_sum = Decimal("0.00")

        total_windows += wins
        total_cost += cost
        technician_total += tech_sum

    # Dropdowns
    technicians = Technician.objects.filter(
        is_active=True, department="installation"
    ).order_by("name")

    context = {
        "schedules": schedules.order_by("scheduled_date"),  # Sort by scheduled date
        "total_windows": total_windows,
        "total_cost": total_cost,
        "total_orders_count": total_orders_count,
        "technician_total": technician_total,
        "technicians": technicians,
        "filters": {
            "date_from": date_from,
            "date_to": date_to,
            "payment_status": payment_status,
            "completion_status": completion_status,
            "technician": technician_id,
        },
    }

    return render(request, "installation_accounting/reports.html", context)


@login_required
def export_installation_report(request):
    """
    تصدير تقرير التركيبات إلى Excel بناءً على الفلتر الحالي
    """
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    payment_status = request.GET.get("payment_status", "all")
    technician_id = request.GET.get("technician")

    from installations.models import InstallationSchedule

    schedules = (
        InstallationSchedule.objects.select_related(
            "order",
            "order__customer",
            "accounting_card",
        )
        .prefetch_related(
            "technicians",
            "accounting_card__shares",
            "accounting_card__shares__technician",
        )
        .exclude(status="cancelled")
        .filter(status__in=["completed", "modification_completed"])
    )

    if date_from:
        schedules = schedules.filter(scheduled_date__gte=date_from)
    if date_to:
        schedules = schedules.filter(scheduled_date__lte=date_to)
    if technician_id:
        schedules = schedules.filter(technicians__id=technician_id).distinct()
    if payment_status == "paid":
        schedules = schedules.filter(accounting_card__status="paid")
    elif payment_status == "unpaid":
        schedules = schedules.exclude(accounting_card__status="paid")

    # احسب الإجماليات للبطاقات الجانبية
    total_windows_export = 0
    total_cost_export = Decimal("0.00")
    technician_total_export = Decimal("0.00")

    for schedule in schedules:
        if hasattr(schedule, "accounting_card"):
            card = schedule.accounting_card
            shares = card.shares.all()
            if technician_id:
                shares = shares.filter(technician_id=technician_id)
                wins = sum(share.assigned_windows for share in shares) if shares else 0
            else:
                wins = card.windows_count
            if payment_status == "paid":
                shares = shares.filter(is_paid=True)
            elif payment_status == "unpaid":
                shares = shares.filter(is_paid=False)
            tech_sum = sum(share.amount for share in shares)
            cost = card.total_cost
        else:
            wins = schedule.windows_count or 0
            cost = Decimal("0.00")
            tech_sum = Decimal("0.00")
        total_windows_export += wins
        total_cost_export += cost
        technician_total_export += tech_sum

    # --- بناء ملف Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير التركيبات"
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_font = Font(size=11)
    row_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style="thin", color="aaaaaa")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    total_lbl_font = Font(bold=True, size=11, color="FFFFFF")
    total_lbl_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    total_val_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
    card_head_font = Font(bold=True, size=12, color="FFFFFF")
    card_val_font = Font(bold=True, size=14, color="1a3c5e")

    start_row = 3
    headers = ["التاريخ المجدول", "رقم الطلب", "العميل", "الشبابيك", "سعر الشباك",
               "الفنيين (الحصص)", "المستحقات", "حالة الدفع", "تاريخ الدفع"]
    col_widths = {1: 15, 2: 18, 3: 25, 4: 12, 5: 15, 6: 35, 7: 15, 8: 15, 9: 15}

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col)].width = col_widths.get(col, 15)

    ws.auto_filter.ref = f"A{start_row}:I{start_row}"

    sum_windows = 0
    sum_tech = Decimal("0.00")
    current_row = start_row + 1

    for schedule in schedules:
        card = getattr(schedule, "accounting_card", None)
        if card:
            shares = card.shares.all()
            if technician_id:
                shares = shares.filter(technician_id=technician_id)
                wins = sum(share.assigned_windows for share in shares) if shares else 0
            else:
                wins = card.windows_count
            if payment_status == "paid":
                shares = shares.filter(is_paid=True)
            elif payment_status == "unpaid":
                shares = shares.filter(is_paid=False)
            techs_detail = "\n".join(
                [f"{s.technician.name}: {float(s.assigned_windows)} شباك" for s in shares]
            ) or "-"
            tech_total = sum(share.amount for share in shares)
            pay_status_str = "مدفوع" if card.status == "paid" else "غير مدفوع"
            pay_date_str = card.payment_date.strftime("%Y-%m-%d") if card.payment_date else "-"
            price_per_win = float(card.price_per_window)
        else:
            wins = schedule.windows_count or 0
            techs_detail = ", ".join([t.name for t in schedule.technicians.all()]) or "-"
            tech_total = Decimal("0.00")
            pay_status_str = "غير جاهز"
            pay_date_str = "-"
            price_per_win = "-"

        vals = [
            schedule.scheduled_date.strftime("%Y-%m-%d") if schedule.scheduled_date else "-",
            schedule.order.order_number,
            schedule.order.customer.name,
            float(wins),
            price_per_win,
            techs_detail,
            float(tech_total),
            pay_status_str,
            pay_date_str,
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = row_font
            cell.alignment = row_align
            cell.border = thin_border

        sum_windows += wins
        sum_tech += tech_total
        current_row += 1

    # صف الإجمالي
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    lbl = ws.cell(row=current_row, column=1, value="المجموع الكلي")
    lbl.font = total_lbl_font
    lbl.fill = total_lbl_fill
    lbl.alignment = Alignment(horizontal="center", vertical="center")
    lbl.border = thin_border
    for col in range(4, 10):
        c = ws.cell(row=current_row, column=col)
        c.fill = total_val_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, size=11)
    ws.cell(row=current_row, column=4).value = float(sum_windows)
    ws.cell(row=current_row, column=7).value = float(sum_tech)

    # البطاقات الجانبية
    side_col = 11
    card_w = 4
    for c in range(side_col, side_col + card_w):
        col_letter = chr(64 + c) if c <= 26 else f"{chr(64 + (c-1)//26)}{chr(64 + (c-1)%26+1)}"
        ws.column_dimensions[col_letter].width = 13

    fills = [
        (PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid"),
         PatternFill(start_color="e8f0fe", end_color="e8f0fe", fill_type="solid")),
        (PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid"),
         PatternFill(start_color="fdf2e9", end_color="fdf2e9", fill_type="solid")),
        (PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid"),
         PatternFill(start_color="e9f7ef", end_color="e9f7ef", fill_type="solid")),
    ]
    cards_data = [
        ("🪟 إجمالي الشبابيك", float(total_windows_export), "شباك"),
        ("💰 إجمالي تكلفة التركيبات", float(total_cost_export), ""),
        ("👷 إجمالي مستحقات الفنيين", float(technician_total_export), ""),
    ]
    card_row = start_row
    for (title, val, unit), (hf, vf) in zip(cards_data, fills):
        ws.merge_cells(start_row=card_row, start_column=side_col,
                       end_row=card_row, end_column=side_col + card_w - 1)
        h = ws.cell(row=card_row, column=side_col, value=title)
        h.font = card_head_font
        h.fill = hf
        h.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=card_row+1, start_column=side_col,
                       end_row=card_row+2, end_column=side_col + card_w - 1)
        display = f"{val} {unit}".strip() if unit else str(val)
        v = ws.cell(row=card_row+1, column=side_col, value=display)
        v.font = card_val_font
        v.fill = vf
        v.alignment = Alignment(horizontal="center", vertical="center")
        card_row += 5

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename=installation_report_{timezone.now().strftime("%Y%m%d")}.xlsx'
    )
    wb.save(response)
    return response


@login_required
def api_bulk_pay_installations(request):
    """
    إتمام الدفع لمجموعة من التركيبات (حصص الفنيين)
    """
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Method not allowed"})

    try:
        data = json.loads(request.body)
        card_ids = data.get("card_ids", [])
        technician_id = data.get("technician_id")

        shares_query = TechnicianShare.objects.filter(
            card_id__in=card_ids, is_paid=False
        )
        if technician_id:
            shares_query = shares_query.filter(technician_id=technician_id)

        count = shares_query.count()
        shares_query.update(is_paid=True, paid_date=timezone.now())

        # Check if all shares for a card are paid to update card status
        for card_id in card_ids:
            card = InstallationCard.objects.get(id=card_id)
            if not card.shares.filter(is_paid=False).exists():
                card.status = "paid"
                card.payment_date = timezone.now()
                card.save(update_fields=["status", "payment_date"])

        return JsonResponse(
            {"success": True, "message": f"تم إتمام الدفع لـ {count} حصة بنجاح"}
        )
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})
